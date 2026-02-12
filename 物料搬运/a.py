import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import time
import re
from datetime import datetime, timedelta

work_order_dict = {
    "KLKT": "IKHZKLK",
    "OLTT": "IKOLTT",
    "DGAH": "IKDGAH",
    "DGQG": "IKDGQG",
    "HQDZ": "ICKHQDZ",
    "FXDZ": "IKFXYCL",
    "JYZZ": "IKJYZZCL",
    "AHEM": "IKDGAH",
    "HQEM": "ICKHQDZ",
    "JYEM": "IKJYZZCL",
    "OLEM": "IKOLTT",
    "XMXC": "IKXCYCL",
    "QGEM": "IKDGQG",
    "KLEM": "IKHZKLK",
    "FXLK": "ICKFXLKYCL"
}
wxbm={"IKOLTT":	10002561,
"IKDGAH":10008889,
"IKJYZZCL":10009828,
"ICKHQDZ":10000687,
"IKHZKLK":10000989,
"IKDGQG":10000757
}
def calculate_all_formulas(sheet):
    """计算工作表中的所有公式"""
    # 获取最大行和列
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # 遍历所有单元格
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # 如果单元格包含公式
            if cell.data_type == 'f':
                try:
                    # 获取公式
                    formula = cell.value
                    
                    # 计算并存储结果（如果公式有效）
                    if formula and formula.startswith('='):
                        # 这里只是标记公式已计算，实际计算由Excel完成
                        # openpyxl不直接计算公式，而是依赖Excel或其他软件
                        pass
                    
                except Exception as e:
                    print(f"计算单元格 {get_column_letter(col_idx)}{row_idx} 公式时出错: {str(e)}")
kucun=['IKLBYCL','IKPCYCL','IKZCYCL']
result=[]
big_plan_path="E:\\供应链\\周欠料\\需求数据--2025.06.28 vs 0626版大计划-1.xlsx"
processed_order_list=[]
def process_excel_files(file1,file2,file3,path,project_list): #每天更新
    # 记录开始时间
    processed_order_list.extend(project_list)
    print(processed_order_list)
    print(project_list)
    start_time = time.time()
    current_day = datetime.now().strftime('%Y%m%d')

    try:
        # 文件路径
        base_dir = os.path.dirname(os.path.abspath(__file__))
        file1_path = os.path.join(base_dir, file1)
        file2_path = os.path.join(base_dir, file2)
        file3_path = os.path.join(base_dir, file3)
        file4_path = os.path.join(base_dir, path+"\\电子料总欠料表"+current_day+".xlsx")

        # 检查文件是否存在
        for file_path in [file1_path, file2_path, file3_path]:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
        
        print("开始处理文件...")
        
        # ===== 处理电子料总欠料表.xlsx =====
        print("处理电子料总欠料表.xlsx...")
        # 打开工作簿
      
        
        # ===== 处理BHSC_库存收发存报表.xls（作为HTML文件处理） =====
        print("处理BHSC_库存收发存报表.xls...")
        # 读取HTML文件
        df2 = read_html_file(file2_path, skiprows=10)
        
        # 删除指定列（5,7,8,10,11,13,14,16,18列，注意pandas列索引从0开始）
        columns_to_delete = [0,1,2,4, 6, 7,8, 9, 10, 12, 13, 15, 17]
        df2 = df2.drop(df2.columns[columns_to_delete], axis=1)
        #print(df2)
        # 确保数据框不为空
       
        
        # ===== 处理BHSC_外协厂投料明细表_250425.xls（作为HTML文件处理） =====
        print("处理BHSC_外协厂投料明细表_250425.xls...")
        # 读取HTML文件
        df3 = read_html_file(file3_path, skiprows=1)
        target_list = ['ICK', 'SZK','EMS']  # 指定的字符串列表
        first_col = df3.columns[0]                 # 获取第一列的列名
        sencond_col=df3.columns[1]   
        # 筛选第一列值存在于列表中的行
        df3 = df3[df3[first_col].isin(target_list)]
        df3 = df3[df3[sencond_col].isin(processed_order_list)]

        # 删除指定列（3,11列，注意pandas列索引从0开始）
        columns_to_delete = [2,4,5,10]
        df3 = df3.drop(df3.columns[columns_to_delete], axis=1)


        # print(df3)
        #print(df3.columns)

        result_dict = {}

        # 遍历 DataFrame 的每一行
        for _, row in df2.iterrows():
            key1 = row.iloc[0]  # 第3列作为外层键
            key2 = row.iloc[3]  # 第5列作为内层键
            value = row.iloc[5] # 第6列作为值
            
            # 如果外层键不存在，则创建一个新的内层字典
            if key1 not in result_dict:
                result_dict[key1[2:-1]] = {}
            
            # 设置内层键值对
            result_dict[key1[2:-1]][key2] = value
        print(result_dict)
        print(df3)
        current_date = datetime.now()

        # 格式化为YYYY/MM/DD
        formatted_date = current_date.strftime('%Y/%m/%d')
        for _, row in df3.iterrows():
            if(row.iloc[7] not in result_dict):continue
            value = row.iloc[12] # 第6列作为值
            print(value,row.iloc[2],row.iloc[7])
            # break
            # if()
            print(result_dict[row.iloc[7]])
            if(work_order_dict[row.iloc[2]] in result_dict):

                value-=min(value,result_dict[row.iloc[7]][work_order_dict[row.iloc[2]]])
            for i in kucun:
                print(value)
                if i in result_dict[row.iloc[7]] and result_dict[row.iloc[7]][i]>0 and value > 0:
                    ans=min(result_dict[row.iloc[7]][i],value)

                    result.append((row.iloc[0],row.iloc[7],ans,formatted_date,work_order_dict[row.iloc[2]],i,wxbm[work_order_dict[row.iloc[2]]]))
                    value-=ans
                    result_dict[row.iloc[7]][i]-=ans
                # 计算处理时间
        end_time = time.time()
        print(f"处理完成，耗时: {end_time - start_time:.2f}秒")
        print(result)
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")



def adjust_formula(formula, old_row, new_row):
    """调整公式中的相对引用"""
    # 这是一个简化的实现，处理A1样式的相对引用
    # 例如：=SUM(A1:B5) 会被调整为 =SUM(A2:B6)
    import re
    
    # 匹配Excel单元格引用（如A1, B2, AA10）
    cell_ref_pattern = r'([A-Z]+)(\d+)'
    
    def replace_ref(match):
        col = match.group(1)
        row = int(match.group(2))
        
        # 如果是绝对引用（如$A$1），不做修改
        if '$' in match.group(0):
            return match.group(0)
        
        # 计算新的行号
        new_row_num = row + (new_row - old_row)
        return f"{col}{new_row_num}"
    
    # 替换公式中的所有单元格引用
    return re.sub(cell_ref_pattern, replace_ref, formula)
def read_html_file(file_path, skiprows=0):
    """读取HTML文件并返回DataFrame"""
    try:
        # 读取HTML文件内容
        with open(file_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # 使用pandas的read_html解析表格
        dfs = pd.read_html(html_content, skiprows=skiprows)
        
        # 通常第一个表格是我们需要的，但也可能需要根据实际情况调整
        if not dfs:
            raise ValueError(f"在{file_path}中未找到表格")
        
        # 返回第一个表格
        return dfs[0]
    
    except Exception as e:
        print(f"读取HTML文件时发生错误: {str(e)}")
        raise
def merge_excel_data(source_file, target_file, source_sheet='Sheet1', target_sheet='电源欠料', 
                     material_code_col='物料编码', source_owe_col='合并欠料', target_owe_col='总欠料'):
    """
    合并两个Excel文件中特定工作表的数据
    
    参数:
    source_file: 源Excel文件路径
    target_file: 目标Excel文件路径
    source_sheet: 源工作表名称
    target_sheet: 目标工作表名称
    material_code_col: 物料编码列名
    source_owe_col: 源文件中欠料列名
    target_owe_col: 目标文件中欠料列名
    """
    # 检查文件是否存在
    if not os.path.exists(source_file):
        raise FileNotFoundError(f"源文件不存在: {source_file}")
    if not os.path.exists(target_file):
        raise FileNotFoundError(f"目标文件不存在: {target_file}")
    
    try:
        # 读取源Excel文件
        print(f"正在读取源文件: {source_file}")
        source_df = pd.read_excel(source_file, sheet_name=source_sheet,header=4)
        
        # 检查源DataFrame是否包含必要的列
        print(source_df.columns)
        if material_code_col not in source_df.columns:
            raise ValueError(f"源工作表 '{source_sheet}' 中不存在列 '{material_code_col}'")
        if source_owe_col not in source_df.columns:
            raise ValueError(f"源工作表 '{source_sheet}' 中不存在列 '{source_owe_col}'")
        
        # 创建物料编码到合并欠料的映射
        print(source_df)
        material_to_owe = dict(zip(source_df[material_code_col], source_df[source_owe_col]))
        print(material_to_owe)
        print(f"已从源工作表中提取 {len(material_to_owe)} 条物料数据")
        
        # 使用openpyxl加载目标Excel文件
        print(f"正在加载目标文件: {target_file}")
        wb = openpyxl.load_workbook(target_file)
        
        # 检查目标工作表是否存在
        if target_sheet not in wb.sheetnames:
            raise ValueError(f"目标文件中不存在工作表 '{target_sheet}'")
        
        ws = wb[target_sheet]
        
        # 找到物料编码列和总欠料列的索引
        header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        material_col_idx = None
        target_owe_col_idx = None
        
        for idx, cell_value in enumerate(header_row, 1):
            print(cell_value)
            if cell_value == material_code_col:
                material_col_idx = idx
            if cell_value == target_owe_col:
                target_owe_col_idx = idx
        
        if material_col_idx is None:
            raise ValueError(f"目标工作表 '{target_sheet}' 中不存在列 '{material_code_col}'")
        if target_owe_col_idx is None:
            raise ValueError(f"目标工作表 '{target_sheet}' 中不存在列 '{target_owe_col}'")
        
        # 确定要插入的列位置（总欠料列的下一列）
        insert_col_idx = target_owe_col_idx + 1
        
        # 插入新列并设置标题
        ws.insert_cols(insert_col_idx)
        ws.cell(row=2, column=insert_col_idx, value=f"{source_owe_col}（合并）")
        
        # 填充数据
        matched_count = 0
        total_rows = ws.max_row
        print("+++++++++++")
        for row_idx in range(3, total_rows + 1):
            material_code = ws.cell(row=row_idx, column=material_col_idx).value
            if material_code in material_to_owe:
                print(material_to_owe[material_code])
                ws.cell(row=row_idx, column=insert_col_idx).value=material_to_owe[material_code]
                matched_count += 1
            print(ws.cell(row=row_idx, column=insert_col_idx).value)
            # 每处理1000行显示进度
            if row_idx % 1000 == 0:
                print(f"已处理 {row_idx}/{total_rows} 行")
        
        print(f"共匹配到 {matched_count} 条物料数据")
        
        # 保存修改后的文件
        file_dir, file_name = os.path.split(target_file)
        new_file_name = os.path.splitext(file_name)[0] + '_合并版' + os.path.splitext(file_name)[1]
        new_file_path = os.path.join(file_dir, new_file_name)
        
        print(f"正在保存结果到: {new_file_path}")
        wb.save(new_file_path)
        print("数据合并完成!")
        
        return new_file_path
        
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        raise

if __name__ == "__main__":
    process_excel_files("E:\\供应链\\总表\\110725\\电子料总欠料表20250711.xlsx", "E:\\供应链\\总表\\110725\\BHSC_库存收发存报表(不_110725.xls", "E:\\供应链\\总表\\110725\\BHSC_外协厂投料明细表(_110725.xls","E:\\供应链\\总表\\110725",['OLTT20250702003-ZB'])
    # 
    # updatebig_plan("E:\\供应链\\周欠料\\需求数据--2025.06.28 vs 0626版大计划-1.xlsx","E:\\供应链\\总表\\110725\电子料总欠料表20250702.xlsx",6)    

    # from openpyxl import load_workbook

    # wb = load_workbook('交期_电子料总欠料表20250427.xlsx')
    # ws = wb['欠料汇总(查询替换) (2)']
    # print(ws)
    # for row_idx in range(6,10):
    #     print(ws.cell(row=row_idx, column=20).value)  # 应输出 =[{'2025-04-28': 600.0}]
    #     print(ws.cell(row=row_idx, column=20).number_format)  # 应输出 '@'
    #update_schedue("E:\\供应链\\总表\\110725\\0425交期回复 (1)秋梅.xlsx","电子料总欠料表20250704.xlsx","E:\\供应链\\总表\\110725\\")    
    # convert_formulas_to_values("交期_电子料总欠料表20250508.xlsx")
    # merge_excel_data("交期_电子料总欠料表20250427.xlsx","群光-世纪云芯电源欠料明细表0428.xlsx","欠料汇总(查询替换) (2)","电源欠料表" ,'物料编码', '合并欠料','总欠料')