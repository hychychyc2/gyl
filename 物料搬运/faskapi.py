import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import time
import re
from datetime import datetime, timedelta

work_order_dict = {
    "KLKT": "IKHZKLK",
    "OLTT": "IKOLTT",
    "DGAH": "IKDGAH",
    "DGQG": "IKDGQG",
    "HQDZ": "ICKHQDZ",
    "FXDZ": "IKFXYCL",
    "JYZZ": "IKJYZZCL",
    "AHEM": "IKDGAH",
    "HQEM": "ICKHQDZ",
    "JYEM": "IKJYZZCL",
    "OLEM": "IKOLTT",
    "XMXC": "IKXCYCL",
    "QGEM": "IKDGQG",
    "KLEM": "IKHZKLK",
    "FXLK": "ICKFXLKYCL"
}
wxbm={"IKOLTT":	10002561,
"IKDGAH":10008889,
"IKJYZZCL":10009828,
"ICKHQDZ":10000687,
"IKHZKLK":10000989,
"IKDGQG":10000757
}
def calculate_all_formulas(sheet):
    """计算工作表中的所有公式"""
    # 获取最大行和列
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # 遍历所有单元格
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # 如果单元格包含公式
            if cell.data_type == 'f':
                try:
                    # 获取公式
                    formula = cell.value
                    
                    # 计算并存储结果（如果公式有效）
                    if formula and formula.startswith('='):
                        # 这里只是标记公式已计算，实际计算由Excel完成
                        # openpyxl不直接计算公式，而是依赖Excel或其他软件
                        pass
                    
                except Exception as e:
                    print(f"计算单元格 {get_column_letter(col_idx)}{row_idx} 公式时出错: {str(e)}")
kucun=['IKLBYCL','IKPCYCL','IKZCYCL']
big_plan_path="E:\\供应链\\周欠料\\需求数据--2025.06.28 vs 0626版大计划-1.xlsx"
processed_order_list=[]
group={}
def process_excel_files(file2,file3,project_list): #每天更新
    # 记录开始时间
    result=[]
    processed_order_list.extend(project_list)
    print(processed_order_list)
    start_time = time.time()
    current_day = datetime.now().strftime('%Y%m%d')
    source_df = pd.read_excel(big_plan_path, sheet_name='量产欠料 ', header=15)
    next_id=-1
    for row_idx in range(len(source_df)):
        if(next_id!=source_df.iloc[row_idx, 1]):
            tmp_id=row_idx
            tmp_list=[]
            while tmp_id<len(source_df) and source_df.iloc[tmp_id, 1]==source_df.iloc[row_idx, 1]:
                tmp_list.append(source_df.iloc[tmp_id, 2])
                tmp_id+=1
            group[source_df.iloc[row_idx, 2]]=tmp_list
        else:
            group[source_df.iloc[row_idx, 2]]=group[source_df.iloc[row_idx-1, 2]]

        next_id=source_df.iloc[row_idx, 1]
    try:
        # 文件路径
        base_dir = os.path.dirname(os.path.abspath(__file__))
        file2_path = os.path.join(base_dir, file2)
        file3_path = os.path.join(base_dir, file3)

        file2_path=file2
        file3_path=file3
        print("开始处理文件...")
        

        
        # ===== 处理BHSC_库存收发存报表.xls（作为HTML文件处理） =====
        print("处理BHSC_库存收发存报表.xls...")
        # 读取HTML文件
        df2 = read_html_file(file2_path, skiprows=10)
        
        # 删除指定列（5,7,8,10,11,13,14,16,18列，注意pandas列索引从0开始）
        columns_to_delete = [0,1,2,4, 6, 7,8, 9, 10, 12, 13, 15, 17]
        df2 = df2.drop(df2.columns[columns_to_delete], axis=1)
        #print(df2)
        # 确保数据框不为空
       
        
        # ===== 处理BHSC_外协厂投料明细表_250425.xls（作为HTML文件处理） =====
        print("处理BHSC_外协厂投料明细表_250425.xls...")
        # 读取HTML文件
        df3 = read_html_file(file3_path, skiprows=1)
        target_list = ['ICK', 'SZK','EMS']  # 指定的字符串列表
        first_col = df3.columns[0]                 # 获取第一列的列名
        sencond_col=df3.columns[1]   
        # 筛选第一列值存在于列表中的行
        df3 = df3[df3[first_col].isin(target_list)]
        df3 = df3[df3[sencond_col].isin(processed_order_list)]

        # 删除指定列（3,11列，注意pandas列索引从0开始）
        columns_to_delete = [2,4,5,10]
        df3 = df3.drop(df3.columns[columns_to_delete], axis=1)


        # print(df3)
        #print(df3.columns)

        result_dict = {}

        # 遍历 DataFrame 的每一行
        for _, row in df2.iterrows():
            key1 = row.iloc[0]  # 第3列作为外层键
            key2 = row.iloc[3]  # 第5列作为内层键
            value = row.iloc[5] # 第6列作为值
            
            # 如果外层键不存在，则创建一个新的内层字典
            if key1[2:-1] not in result_dict:
                result_dict[key1[2:-1]] = {}
            
            # 设置内层键值对
            result_dict[key1[2:-1]][key2] = value
        # print(result_dict)
        # print(df3)
        current_date = datetime.now()

        # 格式化为YYYY/MM/DD
        formatted_date = current_date.strftime('%Y/%m/%d')
        for _, row in df3.iterrows():
            print(row.iloc[7])
            if(row.iloc[7] not in group):continue
            flag=0
            for j in group[row.iloc[7]]:

                if(j in result_dict):
                    flag=1
            print(row.iloc[7])
            if flag==0:continue
            value = row.iloc[12] # 第6列作为值
            # print(value,row.iloc[2],row.iloc[7])
            # break
            # if()
            # print(result_dict[row.iloc[7]])
            # print(result_dict[row.iloc[7]])
            for j in group[row.iloc[7]]:
                if(j not in result_dict):continue

                if(work_order_dict[row.iloc[2]] in result_dict[j]):

                    value-=min(value,result_dict[j][work_order_dict[row.iloc[2]]])
                for i in kucun:
                    # print(value)
                    if i in result_dict[j] and result_dict[j][i]>0 and value > 0:
                        ans=min(result_dict[j][i],value)

                        result.append((row.iloc[0],j,ans,formatted_date,work_order_dict[row.iloc[2]],i,wxbm[work_order_dict[row.iloc[2]]]))
                        value-=ans
                        result_dict[j][i]-=ans
                # 计算处理时间
        end_time = time.time()
        print(f"处理完成，耗时: {end_time - start_time:.2f}秒")
        # print(result)
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
    return result


def adjust_formula(formula, old_row, new_row):
    """调整公式中的相对引用"""
    # 这是一个简化的实现，处理A1样式的相对引用
    # 例如：=SUM(A1:B5) 会被调整为 =SUM(A2:B6)
    import re
    
    # 匹配Excel单元格引用（如A1, B2, AA10）
    cell_ref_pattern = r'([A-Z]+)(\d+)'
    
    def replace_ref(match):
        col = match.group(1)
        row = int(match.group(2))
        
        # 如果是绝对引用（如$A$1），不做修改
        if '$' in match.group(0):
            return match.group(0)
        
        # 计算新的行号
        new_row_num = row + (new_row - old_row)
        return f"{col}{new_row_num}"
    
    # 替换公式中的所有单元格引用
    return re.sub(cell_ref_pattern, replace_ref, formula)
def read_html_file(file_path, skiprows=0):
    """读取HTML文件并返回DataFrame"""
    try:
        # 读取HTML文件内容
        with open(file_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # 使用pandas的read_html解析表格
        dfs = pd.read_html(html_content, skiprows=skiprows)
        
        # 通常第一个表格是我们需要的，但也可能需要根据实际情况调整
        if not dfs:
            raise ValueError(f"在{file_path}中未找到表格")
        
        # 返回第一个表格
        return dfs[0]
    
    except Exception as e:
        print(f"读取HTML文件时发生错误: {str(e)}")
        raise
def merge_excel_data(source_file, target_file, source_sheet='Sheet1', target_sheet='电源欠料', 
                     material_code_col='物料编码', source_owe_col='合并欠料', target_owe_col='总欠料'):
    """
    合并两个Excel文件中特定工作表的数据
    
    参数:
    source_file: 源Excel文件路径
    target_file: 目标Excel文件路径
    source_sheet: 源工作表名称
    target_sheet: 目标工作表名称
    material_code_col: 物料编码列名
    source_owe_col: 源文件中欠料列名
    target_owe_col: 目标文件中欠料列名
    """
    # 检查文件是否存在
    if not os.path.exists(source_file):
        raise FileNotFoundError(f"源文件不存在: {source_file}")
    if not os.path.exists(target_file):
        raise FileNotFoundError(f"目标文件不存在: {target_file}")
    
    try:
        # 读取源Excel文件
        print(f"正在读取源文件: {source_file}")
        source_df = pd.read_excel(source_file, sheet_name=source_sheet,header=4)
        
        # 检查源DataFrame是否包含必要的列
        print(source_df.columns)
        if material_code_col not in source_df.columns:
            raise ValueError(f"源工作表 '{source_sheet}' 中不存在列 '{material_code_col}'")
        if source_owe_col not in source_df.columns:
            raise ValueError(f"源工作表 '{source_sheet}' 中不存在列 '{source_owe_col}'")
        
        # 创建物料编码到合并欠料的映射
        print(source_df)
        material_to_owe = dict(zip(source_df[material_code_col], source_df[source_owe_col]))
        print(material_to_owe)
        print(f"已从源工作表中提取 {len(material_to_owe)} 条物料数据")
        
        # 使用openpyxl加载目标Excel文件
        print(f"正在加载目标文件: {target_file}")
        wb = openpyxl.load_workbook(target_file)
        
        # 检查目标工作表是否存在
        if target_sheet not in wb.sheetnames:
            raise ValueError(f"目标文件中不存在工作表 '{target_sheet}'")
        
        ws = wb[target_sheet]
        
        # 找到物料编码列和总欠料列的索引
        header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        material_col_idx = None
        target_owe_col_idx = None
        
        for idx, cell_value in enumerate(header_row, 1):
            print(cell_value)
            if cell_value == material_code_col:
                material_col_idx = idx
            if cell_value == target_owe_col:
                target_owe_col_idx = idx
        
        if material_col_idx is None:
            raise ValueError(f"目标工作表 '{target_sheet}' 中不存在列 '{material_code_col}'")
        if target_owe_col_idx is None:
            raise ValueError(f"目标工作表 '{target_sheet}' 中不存在列 '{target_owe_col}'")
        
        # 确定要插入的列位置（总欠料列的下一列）
        insert_col_idx = target_owe_col_idx + 1
        
        # 插入新列并设置标题
        ws.insert_cols(insert_col_idx)
        ws.cell(row=2, column=insert_col_idx, value=f"{source_owe_col}（合并）")
        
        # 填充数据
        matched_count = 0
        total_rows = ws.max_row
        print("+++++++++++")
        for row_idx in range(3, total_rows + 1):
            material_code = ws.cell(row=row_idx, column=material_col_idx).value
            if material_code in material_to_owe:
                print(material_to_owe[material_code])
                ws.cell(row=row_idx, column=insert_col_idx).value=material_to_owe[material_code]
                matched_count += 1
            print(ws.cell(row=row_idx, column=insert_col_idx).value)
            # 每处理1000行显示进度
            if row_idx % 1000 == 0:
                print(f"已处理 {row_idx}/{total_rows} 行")
        
        print(f"共匹配到 {matched_count} 条物料数据")
        
        # 保存修改后的文件
        file_dir, file_name = os.path.split(target_file)
        new_file_name = os.path.splitext(file_name)[0] + '_合并版' + os.path.splitext(file_name)[1]
        new_file_path = os.path.join(file_dir, new_file_name)
        
        print(f"正在保存结果到: {new_file_path}")
        wb.save(new_file_path)
        print("数据合并完成!")
        
        return new_file_path
        
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        raise

# if __name__ == "__main__":
#     process_excel_files("E:\\供应链\\总表\\110725\\BHSC_库存收发存报表(不_110725.xls", "E:\\供应链\\总表\\110725\\BHSC_外协厂投料明细表(_110725.xls",['OLTT20250702003-ZB'])
    

import os
from datetime import datetime
from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import xlwings as xw
import threading
import uuid
from queue import Queue

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'xlsx', 'xlsm','xls'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# 用于存储任务状态（线程安全队列）
task_queue = Queue()
task_status = {}  # {task_id: {'status': 'pending', 'progress': 0, 'error': None, 'result': None}}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_excel_worker(task_id, xlsx_path,xlsx_path2, xlsm_path):
    # try:
        result=process_excel_files(xlsx_path,xlsx_path2,['OLTT20250702003-ZB'])
        task_status[task_id]['status'] = 'processing'
        current_month = datetime.now().strftime('%Y%m')
        current_day = datetime.now().strftime('%Y%m%d')
       
        wb = xw.Book(xlsm_path)
        sheet = wb.sheets['WebADI']
        header_row = 3
        data_start_row = header_row + 2
        
        # col_mapping = {}
        # for col_idx in range(1,sheet.api.Columns.Count):
        #     header_value = sheet.cells(header_row, col_idx + 1).value
        #     if(header_value==None):break
        #     if header_value in ['订单编号', '订单日期', '发运说明', '客户PO', '物料编码', '数量']:
        #         col_mapping[header_value] = col_idx + 1
        
        row_idx2 = data_start_row
        total_rows=len(result)
        p=0
        for i in result:
            
            # if data['外协'] != subject:
            #     if data['外协'] == None:
            #         break
            # # 填充订单编号
            #     continue
            # 更新进度（每处理10%触发一次更新）
            progress = int((p / total_rows) * 100)
            if progress > task_status[task_id]['progress']:
                task_status[task_id]['progress'] = progress
            p+=1
            # 业务逻辑...（保持原有处理逻辑）
            print(i)
            sheet.cells(row_idx2, 3).value = i[0]
            sheet.cells(row_idx2, 4).value = i[1]
            sheet.cells(row_idx2, 5).value = i[2]
            sheet.cells(row_idx2, 6).value = i[3]
            sheet.cells(row_idx2, 7).value = i[4]
            sheet.cells(row_idx2, 8).value = i[5]
            sheet.cells(row_idx2, 9).value = i[6]

            row_idx2 += 1
        
        # 删除多余行
        if row_idx2 <= sheet.api.UsedRange.Rows.Count:
            sheet.api.Rows(f"{row_idx2}:{sheet.api.UsedRange.Rows.Count}").Delete()
        
        output_path = os.path.join(OUTPUT_FOLDER, "搬运单" + '.xlsm')
        print(output_path)
        wb.save(output_path)
        wb.close()
        
        task_status[task_id]['status'] = 'complete'
        task_status[task_id]['result'] = 'D:\\供应链\\'+output_path
    # except Exception as e:
    #      task_status[task_id]['status'] = 'error'
    #      task_status[task_id]['error'] = str(e)
    # finally:
    #     # 清理临时文件
    #     os.remove(xlsx_path)
    #     os.remove(xlsm_path)

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        xlsx_file = request.files.get('xlsx_file')
        xlsx_file2 = request.files.get('xlsx_file2')

        xlsm_file = request.files.get('xlsm_file')
        
        if not xlsx_file or not xlsm_file or not xlsx_file2:
            return "请上传3个文件", 400
        if not (allowed_file(xlsx_file.filename) and allowed_file(xlsm_file.filename) and allowed_file(xlsx_file2.filename)):
            return "文件类型不允许", 400
        
        # 生成任务ID
        task_id = str(uuid.uuid4())
        xlsx_path = os.path.join(UPLOAD_FOLDER, xlsx_file.filename)
        xlsx_path2 = os.path.join(UPLOAD_FOLDER, xlsx_file2.filename)

        xlsm_path = os.path.join(UPLOAD_FOLDER, xlsm_file.filename)
        print(xlsx_path)
        xlsx_file.save(xlsx_path)
        xlsx_file2.save(xlsx_path2)

        xlsm_file.save(xlsm_path)
        
        # 初始化任务状态
        task_status[task_id] = {
            'status': 'pending',
            'progress': 0,
            'error': None,
            'result': None
        }
        
        # 启动后台线程处理任务
        threading.Thread(
            target=process_excel_worker,
            args=(task_id, xlsx_path,xlsx_path2, xlsm_path),
            daemon=True
        ).start()
        print(task_id)
        return jsonify({'task_id': task_id}), 202  # 返回任务ID
    
    return render_template('index.html')

@app.route('/get_progress/<task_id>')
def get_progress(task_id):
    status = task_status.get(task_id, {'status': 'not_found'})
    return jsonify(status)

@app.route('/download/<task_id>')
def download_file(task_id):
    status = task_status.get(task_id)
    if status and status['status'] == 'complete':
        output_path = status['result']
        return send_file(output_path, as_attachment=True)
    return "文件处理未完成", 404


if __name__ == '__main__':
     app.run(debug=True)