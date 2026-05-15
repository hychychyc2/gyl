import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import time
import re
from datetime import datetime, timedelta
import uuid
from flask import Flask, render_template, request, send_file, jsonify
import xlwings as xw
import threading
from queue import Queue

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'xlsx', 'xlsm', 'xls'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# 全局变量
task_queue = Queue()
task_status = {}  # {task_id: {'status': 'pending', 'progress': 0, 'error': None, 'result': None}}
processed_order_list = []  # 订单列表（仅在文件上传时清空）
last_uploaded_files = None  # 记录上次上传的文件标识


# 原有映射关系保留
work_order_dict = {
    "KLKT": "IKHZKLK",
    "OLTT": "IKOLTT",
    "DGAH": "IKDGAH",
    "DGQG": "IKDGQG",
    "HQDZ": "ICKHQDZ",
    "FXDZ": "IKFXYCL",
    "JYZZ": "IKJYZZCL",
    "AHEM": "IKDGAH",
    "HQEM": "ICKHQDZ",
    "JYEM": "IKJYZZCL",
    "OLEM": "IKOLTT",
    "XMXC": "IKXCYCL",
    "QGEM": "IKDGQG",
    "KLEM": "IKHZKLK",
    "FXLK": "ICKFXLKYCL"
}

wxbm = {
    "IKOLTT": 10002561,
    "IKDGAH": 10008889,
    "IKJYZZCL": 10009828,
    "ICKHQDZ": 10000687,
    "IKHZKLK": 10000989,
    "IKDGQG": 10000757
}

kucun = ['IKLBYCL', 'IKPCYCL', 'IKZCYCL']


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def calculate_all_formulas(sheet):
    max_row = sheet.max_row
    max_col = sheet.max_column
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.data_type == 'f':
                try:
                    formula = cell.value
                    if formula and formula.startswith('='):
                        pass
                except Exception as e:
                    print(f"计算单元格 {get_column_letter(col_idx)}{row_idx} 公式时出错: {str(e)}")


def read_html_file(file_path, skiprows=0):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        dfs = pd.read_html(html_content, skiprows=skiprows)
        return dfs[0] if dfs else pd.DataFrame()
    except Exception as e:
        print(f"读取HTML文件错误: {str(e)}")
        raise


def process_excel_files(file2, file3, project_list, big_plan_path):
    result = []
    start_time = time.time()
    current_day = datetime.now().strftime('%Y%m%d')
    
    if not os.path.exists(big_plan_path):
        raise FileNotFoundError(f"大计划文件不存在: {big_plan_path}")

    source_df = pd.read_excel(big_plan_path, sheet_name='量产欠料 ', header=15)
    group = {}
    next_id = -1

    for row_idx in range(len(source_df)):
        if next_id != source_df.iloc[row_idx, 1]:
            tmp_id = row_idx
            tmp_list = []
            while tmp_id < len(source_df) and source_df.iloc[tmp_id, 1] == source_df.iloc[row_idx, 1]:
                tmp_list.append(source_df.iloc[tmp_id, 2])
                tmp_id += 1
            group[source_df.iloc[row_idx, 2]] = tmp_list
        else:
            group[source_df.iloc[row_idx, 2]] = group[source_df.iloc[row_idx-1, 2]]
        next_id = source_df.iloc[row_idx, 1]

    try:
        df2 = read_html_file(file2, skiprows=10)
        columns_to_delete = [0,1,2,4,6,7,8,9,10,12,13,15,17]
        df2 = df2.drop(df2.columns[columns_to_delete], axis=1)

        df3 = read_html_file(file3, skiprows=1)
        target_list = ['ICK', 'SZK', 'EMS']
        first_col = df3.columns[0]
        second_col = df3.columns[1]
        df3 = df3[df3[first_col].isin(target_list)]
        df3 = df3[df3[second_col].isin(project_list)]
        order_to_index = {order: idx for idx, order in enumerate(project_list)}
    
        # 添加排序键列并排序
        df3['sort_key'] = df3[second_col].map(order_to_index)
        df3 = df3.sort_values('sort_key')
        df3 = df3.drop('sort_key', axis=1) 
        columns_to_delete = [2,4,5,10]
        df3 = df3.drop(df3.columns[columns_to_delete], axis=1)

        result_dict = {}
        for _, row in df2.iterrows():
            key1 = row.iloc[0]
            key2 = row.iloc[3]
            value = row.iloc[5]
            if key1[2:-1] not in result_dict:
                result_dict[key1[2:-1]] = {}
            result_dict[key1[2:-1]][key2] = value

        current_date = datetime.now().strftime('%Y/%m/%d')
        for _, row in df3.iterrows():
            if row.iloc[7] not in group:
                continue
            flag = 0
            for j in group[row.iloc[7]]:
                if j in result_dict:
                    flag = 1
            if flag == 0:
                continue

            value = row.iloc[12]
            for j in group[row.iloc[7]]:
                if j not in result_dict:
                    continue
                if work_order_dict[row.iloc[2]] in result_dict[j]:
                    value -= min(value, result_dict[j][work_order_dict[row.iloc[2]]])
                for i in kucun:
                    if i in result_dict[j] and result_dict[j][i] > 0 and value > 0:
                        ans = min(result_dict[j][i], value)
                        result.append((
                            row.iloc[0], j, ans, current_date,
                            work_order_dict[row.iloc[2]], i,
                            wxbm[work_order_dict[row.iloc[2]]]
                        ))
                        value -= ans
                        result_dict[j][i] -= ans

        end_time = time.time()
        print(f"处理完成，耗时: {end_time - start_time:.2f}秒")
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        raise
    return result


def adjust_formula(formula, old_row, new_row):
    cell_ref_pattern = r'([A-Z]+)(\d+)'
    def replace_ref(match):
        col = match.group(1)
        row = int(match.group(2))
        if '$' in match.group(0):
            return match.group(0)
        new_row_num = row + (new_row - old_row)
        return f"{col}{new_row_num}"
    return re.sub(cell_ref_pattern, replace_ref, formula)


def merge_excel_data(source_file, target_file, source_sheet='Sheet1', target_sheet='电源欠料', 
                     material_code_col='物料编码', source_owe_col='合并欠料', target_owe_col='总欠料'):
    if not os.path.exists(source_file):
        raise FileNotFoundError(f"源文件不存在: {source_file}")
    if not os.path.exists(target_file):
        raise FileNotFoundError(f"目标文件不存在: {target_file}")

    try:
        source_df = pd.read_excel(source_file, sheet_name=source_sheet, header=4)
        if material_code_col not in source_df.columns:
            raise ValueError(f"源工作表缺少列 '{material_code_col}'")
        if source_owe_col not in source_df.columns:
            raise ValueError(f"源工作表缺少列 '{source_owe_col}'")

        material_to_owe = dict(zip(source_df[material_code_col], source_df[source_owe_col]))
        wb = openpyxl.load_workbook(target_file)
        if target_sheet not in wb.sheetnames:
            raise ValueError(f"目标工作表不存在: '{target_sheet}'")

        ws = wb[target_sheet]
        header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        material_col_idx = None
        target_owe_col_idx = None

        for idx, cell_value in enumerate(header_row, 1):
            if cell_value == material_code_col:
                material_col_idx = idx
            if cell_value == target_owe_col:
                target_owe_col_idx = idx

        if material_col_idx is None or target_owe_col_idx is None:
            raise ValueError("未找到物料编码或总欠料列")

        insert_col_idx = target_owe_col_idx + 1
        ws.insert_cols(insert_col_idx)
        ws.cell(row=2, column=insert_col_idx, value=f"{source_owe_col}（合并）")

        matched_count = 0
        total_rows = ws.max_row
        for row_idx in range(3, total_rows + 1):
            material_code = ws.cell(row=row_idx, column=material_col_idx).value
            if material_code in material_to_owe:
                ws.cell(row=row_idx, column=insert_col_idx).value = material_to_owe[material_code]
                matched_count += 1

        file_dir, file_name = os.path.split(target_file)
        new_file_name = f"{os.path.splitext(file_name)[0]}_合并版{os.path.splitext(file_name)[1]}"
        new_file_path = os.path.join(file_dir, new_file_name)
        wb.save(new_file_path)
        return new_file_path

    except Exception as e:
        print(f"合并数据错误: {str(e)}")
        raise


def process_excel_worker(task_id, xlsx_path, xlsx_path2, xlsm_path, order_list, big_plan_path, file_hash):
    """处理任务的工作线程（新增file_hash参数用于判断文件是否变更）"""
    global processed_order_list, last_uploaded_files
    try:
        # 核心变更：仅当文件发生变化时才清空订单列表
        if file_hash != last_uploaded_files:
            processed_order_list = []  # 清空列表
            last_uploaded_files = file_hash  # 更新文件标识
        
        # 添加新订单（不重复添加）
        for order in order_list:
            if order not in processed_order_list:
                processed_order_list.append(order)
        
        # 处理Excel文件
        result = process_excel_files(xlsx_path, xlsx_path2, processed_order_list, big_plan_path)
        task_status[task_id]['status'] = 'processing'

        # 写入模板文件
        wb = xw.Book(xlsm_path)
        sheet = wb.sheets['WebADI']
        header_row = 3
        data_start_row = header_row + 2
        row_idx2 = data_start_row
        total_rows = len(result)
        p = 0

        for item in result:
            progress = int((p / total_rows) * 100)
            if progress > task_status[task_id]['progress']:
                task_status[task_id]['progress'] = progress
            p += 1

            sheet.cells(row_idx2, 3).value = item[0]
            sheet.cells(row_idx2, 4).value = item[1]
            sheet.cells(row_idx2, 5).value = item[2]
            sheet.cells(row_idx2, 6).value = item[3]
            sheet.cells(row_idx2, 7).value = item[4]
            sheet.cells(row_idx2, 8).value = item[5]
            sheet.cells(row_idx2, 9).value = item[6]
            row_idx2 += 1

        if row_idx2 <= sheet.api.UsedRange.Rows.Count:
            sheet.api.Rows(f"{row_idx2}:{sheet.api.UsedRange.Rows.Count}").Delete()

        output_filename = f"搬运单_{uuid.uuid4().hex}.xlsm"
        output_path = os.path.join(OUTPUT_FOLDER, output_filename)
        wb.save(output_path)
        wb.close()

        task_status[task_id]['status'] = 'complete'
        task_status[task_id]['result'] = 'D:\\供应链\\'+output_path

    except Exception as e:
        task_status[task_id]['status'] = 'error'
        task_status[task_id]['error'] = str(e)
    finally:
        for path in [xlsx_path, xlsx_path2, xlsm_path]:
            if os.path.exists(path):
                try:
                    os.remove(path)
                except:
                    pass


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        xlsx_file = request.files.get('xlsx_file')
        xlsx_file2 = request.files.get('xlsx_file2')
        xlsm_file = request.files.get('xlsm_file')
        order_list_str = request.form.get('order_list', '')
        big_plan_path = request.form.get('big_plan_path', '')

        if not all([xlsx_file, xlsx_file2, xlsm_file]) or not allowed_file(xlsx_file.filename) or not allowed_file(xlsx_file2.filename) or not allowed_file(xlsm_file.filename):
            return jsonify({'error': '文件类型不允许'}), 400

        if not order_list_str or not big_plan_path:
            return jsonify({'error': '请输入订单列表和大计划路径'}), 400

        order_list = [item.strip() for item in order_list_str.split(',') if item.strip()]
        if not order_list:
            return jsonify({'error': '订单列表解析失败'}), 400
        existing_orders = [order for order in order_list if order in processed_order_list]
        if existing_orders:
            return jsonify({'error': f'以下订单已在处理列表中: {", ".join(existing_orders)}'}), 400
        # 生成文件唯一标识（用于判断文件是否变更）
        file_hash = f"{xlsx_file.filename}_{xlsx_file2.filename}_{xlsm_file.filename}_{xlsx_file.content_length}_{xlsx_file2.content_length}_{xlsm_file.content_length}"

        task_id = str(uuid.uuid4())
        xlsx_path = os.path.join(UPLOAD_FOLDER, f"{task_id}_1.{xlsx_file.filename.split('.')[-1]}")
        xlsx_path2 = os.path.join(UPLOAD_FOLDER, f"{task_id}_2.{xlsx_file2.filename.split('.')[-1]}")
        xlsm_path = os.path.join(UPLOAD_FOLDER, f"{task_id}_3.{xlsm_file.filename.split('.')[-1]}")

        xlsx_file.save(xlsx_path)
        xlsx_file2.save(xlsx_path2)
        xlsm_file.save(xlsm_path)

        task_status[task_id] = {
            'status': 'pending',
            'progress': 0,
            'error': None,
            'result': None
        }

        # 启动处理线程（传入file_hash）
        threading.Thread(
            target=process_excel_worker,
            args=(task_id, xlsx_path, xlsx_path2, xlsm_path, order_list, big_plan_path, file_hash),
            daemon=True
        ).start()

        return jsonify({'task_id': task_id}), 202

    return render_template('index.html')


@app.route('/get_progress/<task_id>')
def get_progress(task_id):
    status = task_status.get(task_id, {'status': 'not_found'})
    return jsonify(status)


@app.route('/download/<task_id>')
def download_file(task_id):
    status = task_status.get(task_id)
    if status and status['status'] == 'complete' and status['result'] and os.path.exists(status['result']):
        return send_file(status['result'], as_attachment=True)
    return jsonify({'error': '文件不存在或未处理完成'}), 404


if __name__ == '__main__':
    app.run(debug=True)