"""
芯片齐套管理系统 - HTTP服务器
Python内置http.server + JSON API
"""
import os, sys, json, time, re, urllib.parse, io, threading
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn

class ThreadingHTTPServer(ThreadingMixIn, HTTPServer):
    """多线程HTTP服务器"""
    daemon_threads = True
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from database import (
    init_db, insert, insert_many, update, delete, delete_where,
    query, count, raw, generate_batch_id, write_lock,
    encrypt_password, decrypt_password
)
from email_collector import (
    fetch_all, download_email_attachments, process_shipping_detail,
    process_osat_inventory, process_hold_inventory, process_model_mapping,
    process_mix_bin, process_order_allocation, parse_excel, safe_int
)

FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")
TEMP_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "temp")
os.makedirs(TEMP_DIR, exist_ok=True)

# ============ 定时任务 ============
_scheduler_running = False

def start_scheduler():
    """启动定时邮件采集（早晚各一次）"""
    global _scheduler_running
    if _scheduler_running:
        return
    _scheduler_running = True

    def _run():
        import time as _time
        while _scheduler_running:
            now = datetime.now()
            # 早上9点和晚上21点执行
            if now.hour in (9, 21) and now.minute == 0:
                try:
                    print(f"⏰ 定时采集开始 ({now.strftime('%H:%M')})")
                    fetch_all(TEMP_DIR)
                    print(f"⏰ 定时采集完成")
                except Exception as e:
                    print(f"⏰ 定时采集失败: {e}")
                _time.sleep(60)  # 避免同一分钟重复执行
            _time.sleep(30)

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    print("⏰ 定时任务已启动（每天9:00和21:00）")

# ============ 工具函数 ============
def parse_post_body(handler):
    content_length = int(handler.headers.get('Content-Length', 0))
    body = handler.rfile.read(content_length)
    return json.loads(body.decode('utf-8')) if body else {}

def parse_multipart(handler):
    content_type = handler.headers.get('Content-Type', '')
    if 'multipart/form-data' not in content_type:
        return {}, {}
    boundary = content_type.split('boundary=')[1].strip()
    body = handler.rfile.read(int(handler.headers.get('Content-Length', 0)))
    parts = body.split(f'--{boundary}'.encode())
    fields, files = {}, {}
    for part in parts:
        if b'Content-Disposition' not in part: continue
        header_end = part.find(b'\r\n\r\n')
        if header_end == -1: continue
        headers_raw = part[:header_end].decode('utf-8', errors='ignore')
        content = part[header_end + 4:]
        if content.endswith(b'\r\n'): content = content[:-2]
        name_match = re.search(r'name="([^"]+)"', headers_raw)
        if not name_match: continue
        name = name_match.group(1)
        if 'filename=' in headers_raw:
            fn_match = re.search(r'filename="([^"]*)"', headers_raw)
            files[name] = (fn_match.group(1) if fn_match else 'unknown', content)
        else:
            fields[name] = content.decode('utf-8', errors='ignore')
    return fields, files

def send_json(handler, data, status=200):
    try:
        handler.send_response(status)
        handler.send_header('Content-Type', 'application/json; charset=utf-8')
        handler.send_header('Access-Control-Allow-Origin', '*')
        handler.end_headers()
        handler.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))
    except (BrokenPipeError, ConnectionResetError):
        pass

def send_file(handler, path, content_type=None):
    if not os.path.exists(path):
        handler.send_error(404)
        return
    handler.send_response(200)
    if content_type:
        handler.send_header('Content-Type', content_type)
    handler.send_header('Access-Control-Allow-Origin', '*')
    handler.end_headers()
    with open(path, 'rb') as f:
        handler.wfile.write(f.read())

# ============ 解析 MES/ERP 格式上传 ============
def parse_mes_inventory(file_path: str, warehouse_name: str) -> list:
    """解析MES格式（SZKXYCL等）"""
    rows = parse_excel(file_path, header_row=1)
    if not rows: return []

    batch_id = generate_batch_id()
    result = []
    for row in rows:
        if not row.get('device', '').strip():
            continue

        # 解析生产日期/批次字段: "2529/T1GX25BH13/BIN2/BM1366_F1V24B3C1"
        batch_raw = row.get('batch_raw', '') or row.get('生产日期/批次', '') or ''
        parts = batch_raw.split('/') if batch_raw else []
        date_code = parts[0] if len(parts) > 0 else ''
        marking = parts[1] if len(parts) > 1 else ''
        b = ''
        tp = ''
        for p in parts[2:]:
            if p.startswith('BIN'):
                b = p
            elif p and not p.startswith('BIN'):
                tp = p

        device = row.get('device', '')
        dpb = f"{device}{tp}{b}" if device and tp and b else ''

        result.append({
            'device': device,
            'marking': marking,
            'qty': safe_int(row.get('qty', 0)),
            'bin': b,
            'test_program': tp,
            'warehouse_type': 'other',
            'warehouse_name': warehouse_name,
            'material_code': row.get('material_code', '') or row.get('产品编码', ''),
            'product_desc': row.get('product_desc', '') or row.get('产品描述', ''),
            'batch': batch_raw,
            'date_code': date_code,
            'status': row.get('status', '正常') or row.get('批号状态', '正常'),
            'location_area': row.get('location_area', ''),
            'device_prog_bin': dpb,
            'import_batch': batch_id,
        })
    return result

def parse_erp_inventory(file_path: str, warehouse_name: str) -> list:
    """解析ERP格式（QHBS等）"""
    rows = parse_excel(file_path, header_row=3)
    if not rows: return []

    batch_id = generate_batch_id()
    result = []
    for row in rows:
        batch_raw = row.get('batch', '') or ''
        if not batch_raw: continue

        # 解析批次: "2603/P1ZX26AC34/BIN1/BM1370P2_F2V02B1C2"
        parts = batch_raw.split('/')
        marking = parts[1] if len(parts) > 1 else ''
        b = ''
        tp = ''
        for p in parts[2:]:
            if p.startswith('BIN'):
                b = p
            elif p and not p.startswith('BIN'):
                tp = p

        device = row.get('device', '') or parts[0] if len(parts) > 0 and not parts[0].startswith('P') else ''
        # 如果 device 为空，尝试从物料编码查
        if not device:
            mat_code = row.get('material_code', '') or row.get('物料编码', '')
            if mat_code:
                dev_row = query('material_device', where='erp_code=?', params=(mat_code,))
                if dev_row:
                    device = dev_row[0].get('device', '')

        dpb = f"{device}{tp}{b}" if device and tp and b else ''

        result.append({
            'device': device,
            'marking': marking,
            'qty': safe_int(row.get('qty', 0)),
            'bin': b,
            'test_program': tp,
            'warehouse_type': 'bonded',
            'warehouse_name': warehouse_name,
            'material_code': row.get('material_code', '') or row.get('物料编码', ''),
            'product_desc': row.get('description', '') or row.get('物料说明', ''),
            'batch': batch_raw,
            'status': '正常',
            'sub_inventory': row.get('sub_inventory', '') or row.get('子库存', ''),
            'location': row.get('location', '') or row.get('货位', ''),
            'org': row.get('org', '') or row.get('库存组织', ''),
            'device_prog_bin': dpb,
            'import_batch': batch_id,
        })
    return result

def parse_ems_inventory(file_path: str, warehouse_name: str) -> list:
    """解析EMS库存（芯片结存统计格式）"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True, read_only=True)
    sheet = '各外EMS外协库存明细' if '各外EMS外协库存明细' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    batch_id = generate_batch_id()
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row): continue
        v = [clean_text(x) for x in row]
        if len(v) < 8: continue
        device = v[3] if len(v) > 3 else ''
        b = v[7] if len(v) > 7 else ''
        tp = v[6] if len(v) > 6 else ''
        dpb = f"{device}{tp}{b}" if device and tp and b else ''
        result.append({
            'device': device,
            'marking': v[5] if len(v) > 5 else '',
            'qty': safe_int(v[4]),
            'bin': b,
            'test_program': tp,
            'warehouse_type': 'ems',
            'warehouse_name': warehouse_name or (v[0] if len(v) > 0 else ''),
            'material_code': v[1] if len(v) > 1 else '',
            'batch': v[5] if len(v) > 5 else '',
            'status': '正常',
            'device_prog_bin': dpb,
            'import_batch': batch_id,
        })
    wb.close()
    return result

def clean_text(t):
    if t is None: return ''
    if isinstance(t, (int, float)): return str(int(t)) if t == int(t) else str(t)
    return str(t).replace('\r', '').replace('\n', '').replace('\t', '').strip()

# ============ HTTP Handler ============
class ChipKitHandler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET,POST,PUT,DELETE,OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_GET(self):
        p = urllib.parse.urlparse(self.path)
        path = p.path
        qs = urllib.parse.parse_qs(p.query)

        if path == '/api/dashboard':
            return self._dashboard()
        elif path == '/api/login':
            return self._login(qs)
        elif path.startswith('/api/query/'):
            return self._generic_query(path, qs)
        elif path == '/api/query':
            return self._generic_query(path, qs)
        elif path == '/api/inventory/summary':
            return self._inv_summary()
        elif path == '/api/inventory/pivot':
            return self._inv_pivot(qs)
        elif path == '/api/inventory/with_model':
            return self._inv_with_model(qs)
        elif path.startswith('/api/shipping/'):
            return self._shipping(path, qs)
        elif path.startswith('/api/model/'):
            return self._model(path, qs)
        elif path == '/api/mixbin/list':
            return self._mixbin()
        elif path == '/api/kit/completion':
            return self._kit(qs)
        elif path.startswith('/api/mapping/'):
            return self._mapping_get(path, qs)
        elif path == '/api/email_configs':
            return self._email_configs()
        elif path == '/api/usage/list':
            return self._usage()
        elif path == '/api/users':
            return self._users()
        elif path == '/api/logs':
            return self._logs(qs)
        elif path == '/api/export/excel':
            return self._export_excel()
        elif path == '/api/raw':
            return self._raw(qs)
        elif path == '/api/email/fetch_all':
            return self._fetch_all()
        elif path == '/api/email/fetch/':
            return self._fetch_one(path)
        elif path == '/api/email/import_json':
            return self._import_email_json(body)

        # 静态文件
        if path == '/' or path == '/index.html':
            return send_file(self, os.path.join(FRONTEND_DIR, 'index.html'), 'text/html')
        filepath = os.path.join(FRONTEND_DIR, path.lstrip('/'))
        if os.path.exists(filepath) and os.path.isfile(filepath):
            ct = 'text/html' if path.endswith('.html') else 'text/javascript' if path.endswith('.js') else 'text/css' if path.endswith('.css') else None
            return send_file(self, filepath, ct)

        send_json(self, {'ok': False, 'error': 'Not found'}, 404)

    def do_POST(self):
        p = urllib.parse.urlparse(self.path)
        path = p.path
        content_type = self.headers.get('Content-Type', '')

        if 'multipart/form-data' in content_type:
            fields, files = parse_multipart(self)
            return self._handle_upload(path, fields, files)

        body = parse_post_body(self)

        if path == '/api/login':
            return self._login_post(body)
        elif path == '/api/query':
            return self._query_post(path, body)
        elif path == '/api/insert/':
            return self._insert(path, body)
        elif path == '/api/insert_many/':
            return self._insert_many(path, body)
        elif path == '/api/kit/calculate_shortage':
            return self._calc_shortage(body)
        elif path == '/api/shipping/auto_plan':
            return self._auto_plan()
        elif path.startswith('/api/email_configs'):
            return self._email_config_post(path, body)
        elif path == '/api/email/import_json':
            return self._import_email_json(body)
        elif path.startswith('/api/mapping/'):
            return self._mapping_post(path, body)
        elif path == '/api/usage':
            return self._usage_post(body)
        elif path == '/api/users':
            return self._users_post(body)

        send_json(self, {'ok': False, 'error': 'Unknown route'}, 404)

    def do_PUT(self):
        p = urllib.parse.urlparse(self.path)
        path = p.path
        body = parse_post_body(self)

        if path.startswith('/api/update/'):
            return self._update(path, body)
        elif path.startswith('/api/mapping/'):
            return self._mapping_put(path, body)
        elif path.startswith('/api/usage/'):
            return self._usage_put(path, body)
        elif path.startswith('/api/email_configs/'):
            return self._update(path, body)

        send_json(self, {'ok': False, 'error': 'Unknown route'}, 404)

    def do_DELETE(self):
        p = urllib.parse.urlparse(self.path)
        path = p.path
        parts = path.split('/')
        try:
            # /api/delete/{table}/{id}
            if len(parts) >= 4 and parts[2] == 'delete':
                ok = delete(parts[3], int(parts[4]))
            # /api/mapping/{table}/{id}
            elif len(parts) >= 4 and parts[2] == 'mapping':
                ok = delete(parts[3], int(parts[4]))
            # /api/email_configs/{id}
            elif len(parts) >= 3 and parts[2] == 'email_configs':
                ok = delete('email_config', int(parts[3]))
            # /api/usage/{id}
            elif len(parts) >= 3 and parts[2] == 'usage':
                ok = delete('usage_mapping', int(parts[3]))
            else:
                return send_json(self, {'ok': False}, 400)
            send_json(self, {'ok': ok})
        except Exception as e:
            send_json(self, {'ok': False, 'error': str(e)})

    # ============ 登录 ============
    def _login(self, qs):
        send_json(self, {'ok': False, 'error': '请使用POST请求'})

    def _login_post(self, body):
        email_addr = body.get('email', '')
        password = body.get('password', '')
        user = query('users', where='email=?', params=(email_addr,))
        if not user:
            return send_json(self, {'ok': False, 'error': '用户不存在，请先运行 migrate.py'})
        user = user[0]
        if not user.get('active'):
            return send_json(self, {'ok': False, 'error': '账号已禁用'})
        if user.get('password_hash') == password:
            return send_json(self, {'ok': True, 'user': {'id': user['id'], 'email': user['email'], 'name': user['name'], 'role': user['role']}})
        return send_json(self, {'ok': False, 'error': '密码错误'})

    # ============ 仪表盘 ============
    def _dashboard(self):
        data = {
            'total_shipping': count('shipping_detail'),
            'total_inventory': count('inventory'),
            'total_models': count('model_mapping'),
            'total_kit': count('kit_completion'),
            'inventory_by_type': raw("SELECT warehouse_type, SUM(qty) as total_qty FROM inventory GROUP BY warehouse_type"),
            'recent_shipping': query('shipping_detail', order_by='ship_date DESC', limit=10),
            'low_stock': raw("SELECT device, SUM(qty) as total_qty FROM inventory WHERE warehouse_type='osat' GROUP BY device HAVING total_qty < 1000 ORDER BY total_qty"),
        }
        send_json(self, {'ok': True, 'data': data})

    # ============ 通用查询 ============
    def _generic_query(self, path, qs):
        parts = path.split('/')
        table = parts[3]
        where = qs.get('where', [''])[0]
        params = json.loads(qs.get('params', ['[]'])[0])
        try:
            rows = query(table, where=where, params=tuple(params), order_by=qs.get('order_by', [''])[0],
                         limit=int(qs.get('limit', ['0'])[0]), offset=int(qs.get('offset', ['0'])[0]))
            total = count(table, where=where, params=tuple(params))
            send_json(self, {'ok': True, 'data': rows, 'total': total})
        except Exception as e:
            send_json(self, {'ok': False, 'error': str(e)})

    def _query_post(self, path, body):
        parts = path.split('/')
        table = body.get('table', parts[3] if len(parts) > 3 else '')
        try:
            rows = query(table, where=body.get('where', ''), params=tuple(body.get('params', [])),
                         order_by=body.get('order_by', ''), limit=body.get('limit', 0),
                         offset=body.get('offset', 0), columns=body.get('columns', '*'))
            total = count(table, where=body.get('where', ''), params=tuple(body.get('params', [])))
            send_json(self, {'ok': True, 'data': rows, 'total': total})
        except Exception as e:
            send_json(self, {'ok': False, 'error': str(e)})

    def _insert(self, path, body):
        table = path.split('/')[3]
        try:
            rid = insert(table, body)
            send_json(self, {'ok': True, 'id': rid})
        except Exception as e:
            send_json(self, {'ok': False, 'error': str(e)})

    def _insert_many(self, path, body):
        table = path.split('/')[3]
        try:
            cnt = insert_many(table, body)
            send_json(self, {'ok': True, 'count': cnt})
        except Exception as e:
            send_json(self, {'ok': False, 'error': str(e)})

    def _update(self, path, body):
        parts = path.split('/')
        if len(parts) >= 5:
            try:
                ok = update(parts[3], int(parts[4]), body)
                send_json(self, {'ok': ok})
            except Exception as e:
                send_json(self, {'ok': False, 'error': str(e)})

    # ============ 库存模块 ============
    def _inv_summary(self):
        rows = raw("SELECT warehouse_type, warehouse_name, SUM(qty) as total_qty FROM inventory GROUP BY warehouse_type, warehouse_name ORDER BY warehouse_type, total_qty DESC")
        send_json(self, {'ok': True, 'data': rows})

    def _inv_pivot(self, qs):
        """库存透视：按机型显示各库存类型的芯片数量"""
        model = qs.get('model', [''])[0]
        filter_type = qs.get('warehouse_type', [''])[0]

        sql = """
        SELECT m.model1, i.device, i.device_prog_bin, i.bin, i.test_program,
               i.warehouse_type, i.warehouse_name,
               SUM(i.qty) as total_qty,
               MAX(u.usage_qty) as usage_qty,
               CASE WHEN MAX(u.usage_qty) > 0 THEN SUM(i.qty) * 1.0 / MAX(u.usage_qty) ELSE 0 END as machine_count
        FROM model_mapping m
        LEFT JOIN inventory i ON m.device_prog_bin = i.device_prog_bin
        LEFT JOIN usage_mapping u ON m.device LIKE u.device || '%' AND m.model1 = u.project
        WHERE m.model1 != '' AND i.qty > 0
        """
        params = []
        if model:
            sql += " AND m.model1 LIKE ?"
            params.append(f"%{model}%")
        if filter_type:
            sql += " AND i.warehouse_type = ?"
            params.append(filter_type)
        sql += " GROUP BY m.model1, m.device_prog_bin, i.warehouse_type, i.warehouse_name ORDER BY m.model1, total_qty DESC"

        rows = raw(sql, tuple(params))
        send_json(self, {'ok': True, 'data': rows})

    def _inv_with_model(self, qs):
        """库存关联机型"""
        device = qs.get('device', [''])[0]
        wh_type = qs.get('warehouse_type', [''])[0]
        where = []; params = []
        if device:
            where.append('i.device LIKE ?')
            params.append(f'{device}%')
        if wh_type:
            where.append('i.warehouse_type = ?')
            params.append(wh_type)
        where_clause = 'WHERE ' + ' AND '.join(where) if where else ''
        sql = f"""
        SELECT i.device, i.device_prog_bin, i.bin, i.test_program,
               i.warehouse_name, i.warehouse_type, i.status, i.id, i.version,
               i.source_email, i.source_file, i.source_time,
               SUM(i.qty) as total_qty,
               m.model1, m.model2, m.model3,
               u.usage_qty,
               CASE WHEN u.usage_qty > 0 THEN SUM(i.qty) * 1.0 / u.usage_qty ELSE 0 END as machine_count
        FROM inventory i
        LEFT JOIN model_mapping m ON i.device_prog_bin = m.device_prog_bin
        LEFT JOIN usage_mapping u ON m.device LIKE u.device || '%' AND m.model1 = u.project
        {where_clause}
        AND i.device != '' AND i.device != '#N/A'
        GROUP BY i.device, i.device_prog_bin, i.warehouse_name, i.warehouse_type
        ORDER BY i.warehouse_type, total_qty DESC
        LIMIT 500
        """
        rows = raw(sql, tuple(params))
        send_json(self, {'ok': True, 'data': rows})

    # ============ 出货明细 ============
    def _shipping(self, path, qs):
        if path == '/api/shipping/expired':
            rows = raw("SELECT * FROM shipping_detail WHERE shipped_qty=0 AND ship_date<date('now','-180 days') ORDER BY ship_date DESC LIMIT 200")
            send_json(self, {'ok': True, 'data': rows})
        elif path == '/api/shipping/summary':
            rows = raw("SELECT osat, device_pn, bin, test_program, COUNT(*) as cnt, SUM(good_qty) as total FROM shipping_detail GROUP BY osat, device_pn, bin, test_program ORDER BY total DESC LIMIT 200")
            send_json(self, {'ok': True, 'data': rows})
        else:
            send_json(self, {'ok': False}, 404)

    # ============ 机型对照 ============
    def _model(self, path, qs):
        if path == '/api/model/mapping':
            where = []; params = []
            for k in ['device','model1']:
                if qs.get(k, [''])[0]:
                    where.append(f"{k} LIKE ?"); params.append(f"%{qs[k][0]}%")
            sql = "SELECT * FROM model_mapping"
            if where: sql += " WHERE " + " AND ".join(where)
            sql += " ORDER BY device, bin LIMIT 500"
            send_json(self, {'ok': True, 'data': raw(sql, tuple(params))})
        elif path == '/api/model/with_stock':
            rows = raw("""
            SELECT m.*, COALESCE(i.total_qty, 0) as stock_qty, u.usage_qty,
                   CASE WHEN u.usage_qty>0 THEN COALESCE(i.total_qty,0)/u.usage_qty ELSE 0 END as machine_count
            FROM model_mapping m
            LEFT JOIN (SELECT device_prog_bin, SUM(qty) as total_qty FROM inventory GROUP BY device_prog_bin) i ON m.device_prog_bin=i.device_prog_bin
            LEFT JOIN usage_mapping u ON m.device LIKE u.device || '%' AND m.model1 = u.project
            ORDER BY m.device, m.bin LIMIT 500
            """)
            send_json(self, {'ok': True, 'data': rows})
        elif path == '/api/model/exclusive':
            rows = query('model_mapping', where='exclusive_bin=1', order_by='device, bin')
            send_json(self, {'ok': True, 'data': rows})

    def _mixbin(self):
        send_json(self, {'ok': True, 'data': query('mix_bin', order_by='device, bin')})

    # ============ 齐套达成 ============
    def _kit(self, qs):
        rows = query('kit_completion', order_by='region, device, subcontractor')
        for r in rows:
            for f in ['month_plan','shortage','actual_ship','planned_arrival']:
                try: r[f] = json.loads(r.get(f,'{}') or '{}')
                except: r[f] = {}
        send_json(self, {'ok': True, 'data': rows})

    def _calc_shortage(self, body):
        region = body.get('region', '')
        where = 'region=?' if region else ''
        params = (region,) if region else ()
        rows = query('kit_completion', where=where, params=params)
        for r in rows:
            mp = json.loads(r.get('month_plan','{}') or '{}')
            usage = r.get('usage_per_unit',0) or 0
            stock = r.get('initial_stock',0) or 0
            cum = stock
            shortage = {}
            for m, plan in sorted(mp.items()):
                cum -= int(plan) * usage
                shortage[m] = cum
            update('kit_completion', r['id'], {
                'shortage': json.dumps(shortage, ensure_ascii=False),
                'version': r.get('version',1)
            })
        send_json(self, {'ok': True, 'data': rows})

    def _auto_plan(self):
        kits = query('kit_completion')
        plans = []
        for kit in kits:
            shortage = json.loads(kit.get('shortage','{}') or '{}')
            total = sum(abs(v) for v in shortage.values() if v < 0)
            if total <= 0: continue
            device = kit.get('device','')
            stocks = query('inventory', where='device=? AND warehouse_type IN (?,?,?)',
                          params=(device,'osat','bonded','other'),
                          order_by='CASE warehouse_type WHEN "osat" THEN 1 WHEN "bonded" THEN 2 ELSE 3 END')
            remaining = total
            for s in stocks:
                if remaining <= 0: break
                qty = min(s.get('qty',0), remaining)
                if qty <= 0: continue
                plans.append({
                    'plan_date': datetime.now().strftime('%Y-%m-%d'),
                    'osat': s.get('warehouse_name',''), 'device': device,
                    'bin': s.get('bin',''), 'qty': qty,
                    'from_warehouse': s.get('warehouse_name',''),
                    'warehouse_type': s.get('warehouse_type',''),
                    'ship_to': kit.get('subcontractor',''),
                    'model_name': kit.get('model_name',''),
                    'project': kit.get('project',''), 'status': '待确认',
                })
                remaining -= qty
        if plans:
            delete_where('shipping_plan')
            insert_many('shipping_plan', plans)
        send_json(self, {'ok': True, 'data': plans, 'total': len(plans)})

    # ============ 映射管理 ============
    def _mapping_get(self, path, qs):
        table = path.split('/')[3]
        if table not in ('subcontractor_mapping','logistics_time','material_device'):
            return send_json(self, {'ok': False}, 400)
        search = qs.get('search',[''])[0]
        where = ''; params = ()
        if search:
            if table == 'subcontractor_mapping':
                where = 'short_name LIKE ? OR internal_code LIKE ?'
                params = (f'%{search}%',)*2
            elif table == 'material_device':
                where = 'erp_code LIKE ? OR device LIKE ?'
                params = (f'%{search}%',)*2
        send_json(self, {'ok': True, 'data': query(table, where=where, params=params)})

    def _mapping_post(self, path, body):
        table = path.split('/')[3]
        rid = insert(table, body)
        send_json(self, {'ok': True, 'id': rid})

    def _mapping_put(self, path, body):
        parts = path.split('/')
        ok = update(parts[3], int(parts[4]), body)
        send_json(self, {'ok': ok})

    # ============ 邮件配置 ============
    def _email_configs(self):
        rows = query('email_config', order_by='purpose')
        for r in rows:
            if r.get('password_encrypted'):
                r['password_encrypted'] = '********'
        send_json(self, {'ok': True, 'data': rows})

    def _email_config_post(self, path, body):
        if len(path.split('/')) > 3:
            # /api/email_configs/{id}/fetch
            rid = int(path.split('/')[2])
            cfg = dict(query('email_config', where='id=?', params=(rid,))[0])
            result = download_email_attachments(cfg, TEMP_DIR)
            if not result: return send_json(self, {'ok': False, 'error': '未找到附件'})
            fp, source_info = result[0], result[1:]
            count = 0
            p = cfg.get('purpose','')
            if p == 'shipping_detail': count = process_shipping_detail(fp, cfg, source_info)
            elif p == 'osat_inventory': count = process_osat_inventory(fp, cfg, source_info)
            elif p == 'hold_inventory': count = process_hold_inventory(fp, cfg, source_info)
            elif p == 'model_mapping': count = process_model_mapping(fp, cfg, source_info)
            elif p == 'mix_bin': count = process_mix_bin(fp, cfg, source_info)
            elif p == 'order_allocation': count = process_order_allocation(fp, cfg, source_info)
            update('email_config', rid, {'last_fetch': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'version': cfg.get('version',1)})
            try: os.remove(fp)
            except: pass
            return send_json(self, {'ok': True, 'count': count})
        else:
            # 新增配置，加密密码
            if 'password_encrypted' in body and body['password_encrypted'] and body['password_encrypted'] != '********':
                body['password_encrypted'] = encrypt_password(body['password_encrypted'])
            rid = insert('email_config', body)
            return send_json(self, {'ok': True, 'id': rid})

    def _fetch_all(self):
        """手动触发全部采集"""
        configs = query('email_config', where='active=1')
        if not configs:
            return send_json(self, {'ok': False, 'error': '没有配置邮箱，请先在📧邮件配置中添'})
        results = fetch_all(TEMP_DIR)
        total = sum(v for v in results.values() if v)
        if total == 0:
            return send_json(self, {'ok': False, 'error': f'采集完成但未获取到数据，请检查邮箱配置和邮件内容'})
        send_json(self, {'ok': True, 'results': results, 'total': total})

    def _import_email_json(self, body):
        """从JSON批量导入邮件配置"""
        configs = body if isinstance(body, list) else [body]
        imported = 0
        for cfg in configs:
            if 'password_encrypted' in cfg and cfg['password_encrypted']:
                cfg['password_encrypted'] = encrypt_password(cfg['password_encrypted'])
            insert('email_config', cfg)
            imported += 1
        return send_json(self, {'ok': True, 'count': imported})

    def _fetch_one(self, path):
        rid = int(path.split('/')[-1])
        cfg = dict(query('email_config', where='id=?', params=(rid,))[0])
        if not cfg.get('account'):
            return send_json(self, {'ok': False, 'error': '邮箱配置不完整，请填写邮箱账号密码'})
        result = download_email_attachments(cfg, TEMP_DIR)
        if not result:
            return send_json(self, {'ok': False, 'error': '未找到匹配附件。请检查：1)邮箱密码是否正确 2)匹配关键词是否匹配 3)附件后缀是否正确'})
        fp, source_info = result[0], result[1:]
        count = 0
        p = cfg.get('purpose','')
        if p == 'shipping_detail': count = process_shipping_detail(fp, cfg)
        elif p == 'osat_inventory': count = process_osat_inventory(fp, cfg)
        elif p == 'hold_inventory': count = process_hold_inventory(fp, cfg)
        elif p == 'model_mapping': count = process_model_mapping(fp, cfg)
        elif p == 'mix_bin': count = process_mix_bin(fp, cfg)
        elif p == 'order_allocation': count = process_order_allocation(fp, cfg)
        update('email_config', rid, {'last_fetch': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'version': cfg.get('version',1)})
        try: os.remove(fp)
        except: pass
        send_json(self, {'ok': True, 'count': count})

    # ============ 文件上传 ============
    def _handle_upload(self, path, fields, files):
        if 'file' not in files: return send_json(self, {'ok': False, 'error': 'No file'})

        filename, content = files['file']
        fp = os.path.join(TEMP_DIR, f"upload_{int(time.time())}_{filename}")
        with open(fp, 'wb') as f: f.write(content)

        try:
            if path == '/api/upload/inventory':
                wh_type = fields.get('warehouse_type', 'other')
                wh_name = fields.get('warehouse_name', '')
                format_type = fields.get('format_type', 'mes')

                if format_type == 'mes':
                    rows = parse_mes_inventory(fp, wh_name)
                elif format_type == 'erp':
                    rows = parse_erp_inventory(fp, wh_name)
                elif format_type == 'ems':
                    rows = parse_ems_inventory(fp, wh_name)
                else:
                    return send_json(self, {'ok': False, 'error': '未知格式类型'})

                if not rows:
                    return send_json(self, {'ok': False, 'error': '解析无数据'})

                # 覆盖该仓库类型+名称的旧数据
                delete_where('inventory', warehouse_type=wh_type, warehouse_name=wh_name)
                cnt = insert_many('inventory', rows)
                return send_json(self, {'ok': True, 'count': cnt, 'warehouse_type': wh_type, 'warehouse_name': wh_name})

            elif path == '/api/upload/shipping':
                rows = parse_excel(fp, header_row=3)
                if not rows: return send_json(self, {'ok': False, 'error': '解析无数据'})
                batch_id = generate_batch_id()
                ship_rows = []
                for row in rows:
                    if not row.get('device_pn','').strip() and not row.get('entity','').strip(): continue
                    ship_rows.append({
                        'entity': row.get('entity',''), 'ship_date': row.get('ship_date',''),
                        'device_pn': row.get('device_pn',''), 'wafer_lot_id': row.get('wafer_lot_id',''),
                        'marking': row.get('marking',''), 'good_qty': safe_int(row.get('good_qty',0)),
                        'bin': row.get('bin',''), 'invoice_no': row.get('invoice_no',''),
                        'test_program': row.get('test_program',''), 'osat': row.get('osat',''),
                        'ship_to': row.get('ship_to',''), 'test_wo': row.get('test_wo',''),
                        'po': row.get('po','暂无'), 'source': 'upload', 'import_batch': batch_id,
                    })
                cnt = insert_many('shipping_detail', ship_rows)
                return send_json(self, {'ok': True, 'count': cnt})

            elif path == '/api/upload/erp_inventory':
                rows = parse_erp_inventory(fp, fields.get('warehouse_name','QHBS'))
                cnt = insert_many('erp_inventory', rows)
                return send_json(self, {'ok': True, 'count': cnt})

        finally:
            try: os.remove(fp)
            except: pass

        send_json(self, {'ok': False}, 404)

    # ============ 其他 ============
    def _usage(self):
        send_json(self, {'ok': True, 'data': query('usage_mapping', order_by='device, model_name')})

    def _usage_post(self, body):
        send_json(self, {'ok': True, 'id': insert('usage_mapping', body)})

    def _usage_put(self, path, body):
        ok = update('usage_mapping', int(path.split('/')[3]), body)
        send_json(self, {'ok': ok})

    def _users(self):
        send_json(self, {'ok': True, 'data': query('users', columns='id,email,name,role,active,created_at')})

    def _users_post(self, body):
        send_json(self, {'ok': True, 'id': insert('users', body)})

    def _logs(self, qs):
        limit = int(qs.get('limit',['100'])[0])
        send_json(self, {'ok': True, 'data': query('operation_log', order_by='created_at DESC', limit=limit)})

    def _export(self, path):
        table = path.split('/')[3]
        rows = query(table, limit=50000)
        filepath = os.path.join(os.path.dirname(__file__), '..', 'exports', f"{table}_{datetime.now().strftime('%Y%m%d')}.json")
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
        send_file(self, filepath, 'application/json')

    def _export_excel(self):
        """导出Excel格式"""
        from export_excel import export_inventory
        try:
            filepath = export_inventory()
            send_file(self, filepath, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            send_json(self, {'ok': False, 'error': str(e)})

    def _raw(self, qs):
        sql = qs.get('sql',[''])[0]
        params = json.loads(qs.get('params',['[]'])[0])
        try:
            rows = raw(sql, tuple(params), fetch=qs.get('fetch',['true'])[0]=='true')
            send_json(self, {'ok': True, 'data': rows} if isinstance(rows, list) else {'ok': True, 'count': rows})
        except Exception as e:
            send_json(self, {'ok': False, 'error': str(e)})

    def log_message(self, format, *args):
        pass

def main():
    init_db()
    start_scheduler()
    port = 8765
    server = ThreadingHTTPServer(('0.0.0.0', port), ChipKitHandler)
    print(f"🦞 芯片齐套管理系统: http://localhost:{port}")
    print(f"⏰ 定时采集: 每天 9:00 和 21:00")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        server.shutdown()

if __name__ == '__main__':
    main()