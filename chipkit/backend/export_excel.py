"""
芯片齐套管理系统 - Excel导出模块
支持导出为 .xlsx 格式，字段对齐源Excel
"""
import os, sys, json
from datetime import datetime
sys.path.insert(0, os.path.dirname(__file__))
from database import query, raw, count

EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

def safe_val(v, default=''):
    return v if v is not None else default

def export_inventory(filename=None):
    """导出库存为Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not filename:
        filename = os.path.join(EXPORTS_DIR, f"库存导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    wb = openpyxl.Workbook()

    # 1. 库存总览
    ws = wb.active
    ws.title = "库存总览"
    headers = ['芯片型号', 'Marking', '数量', 'BIN', '测试程序', '仓库类型', '仓库名称',
               '物料编码', '批次', '状态', '可做机型', '单机用量', '可做台数', 'device_prog_bin']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    rows = raw("""
        SELECT i.device, i.marking, SUM(i.qty) as total_qty, i.bin, i.test_program,
               i.warehouse_type, i.warehouse_name, i.material_code, i.batch, i.status,
               m.model1, u.usage_qty,
               CASE WHEN u.usage_qty > 0 THEN SUM(i.qty) / u.usage_qty ELSE 0 END as machine_count,
               i.device_prog_bin
        FROM inventory i
        LEFT JOIN model_mapping m ON i.device_prog_bin = m.device_prog_bin
        LEFT JOIN usage_mapping u ON i.device LIKE u.device || '%' AND m.model1 = u.project
        GROUP BY i.device_prog_bin, i.warehouse_type, i.warehouse_name
        ORDER BY i.warehouse_type, i.device, total_qty DESC
    """)

    for r_idx, row in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=safe_val(row['device']))
        ws.cell(row=r_idx, column=2, value=safe_val(row['marking']))
        ws.cell(row=r_idx, column=3, value=safe_val(row['total_qty'], 0))
        ws.cell(row=r_idx, column=4, value=safe_val(row['bin']))
        ws.cell(row=r_idx, column=5, value=safe_val(row['test_program']))
        ws.cell(row=r_idx, column=6, value=safe_val(row['warehouse_type']))
        ws.cell(row=r_idx, column=7, value=safe_val(row['warehouse_name']))
        ws.cell(row=r_idx, column=8, value=safe_val(row['material_code']))
        ws.cell(row=r_idx, column=9, value=safe_val(row['batch']))
        ws.cell(row=r_idx, column=10, value=safe_val(row['status']))
        ws.cell(row=r_idx, column=11, value=safe_val(row['model1']))
        ws.cell(row=r_idx, column=12, value=safe_val(row['usage_qty'], 0))
        ws.cell(row=r_idx, column=13, value=round(safe_val(row['machine_count'], 0), 1))
        ws.cell(row=r_idx, column=14, value=safe_val(row['device_prog_bin']))

    ws.auto_filter.ref = ws.dimensions
    for col_letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']:
        ws.column_dimensions[col_letter].width = 18

    # 2. 出货明细
    ws2 = wb.create_sheet("出货明细")
    headers2 = ['主体', '出货日期', '芯片型号', 'waferLotId', 'Marking', '良品数量', 'BIN',
                'InvoiceNo', '测试程序', 'OSAT', '收货地址', '测试工单号', 'DATE CODE', '采购PO', '销售SO', '机型']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    rows2 = query('shipping_detail', order_by='ship_date DESC', limit=50000)
    for r_idx, row in enumerate(rows2, 2):
        ws2.cell(row=r_idx, column=1, value=safe_val(row['entity']))
        ws2.cell(row=r_idx, column=2, value=safe_val(row['ship_date']))
        ws2.cell(row=r_idx, column=3, value=safe_val(row['device_pn']))
        ws2.cell(row=r_idx, column=4, value=safe_val(row['wafer_lot_id']))
        ws2.cell(row=r_idx, column=5, value=safe_val(row['marking']))
        ws2.cell(row=r_idx, column=6, value=safe_val(row['good_qty'], 0))
        ws2.cell(row=r_idx, column=7, value=safe_val(row['bin']))
        ws2.cell(row=r_idx, column=8, value=safe_val(row['invoice_no']))
        ws2.cell(row=r_idx, column=9, value=safe_val(row['test_program']))
        ws2.cell(row=r_idx, column=10, value=safe_val(row['osat']))
        ws2.cell(row=r_idx, column=11, value=safe_val(row['ship_to']))
        ws2.cell(row=r_idx, column=12, value=safe_val(row['test_wo']))
        ws2.cell(row=r_idx, column=13, value=safe_val(row['date_code']))
        ws2.cell(row=r_idx, column=14, value=safe_val(row['po']))
        ws2.cell(row=r_idx, column=15, value=safe_val(row['so']))
        ws2.cell(row=r_idx, column=16, value=safe_val(row['model1']))

    ws2.auto_filter.ref = ws2.dimensions
    for col_letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']:
        ws2.column_dimensions[col_letter].width = 16

    # 3. 机型对照
    ws3 = wb.create_sheet("机型对照")
    headers3 = ['芯片型号', 'ATE程式', 'BIN', '机型1', '机型2', '机型3', '项目', '独占BIN', 'device_prog_bin']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')

    rows3 = query('model_mapping', order_by='device, bin')
    for r_idx, row in enumerate(rows3, 2):
        ws3.cell(row=r_idx, column=1, value=safe_val(row['device']))
        ws3.cell(row=r_idx, column=2, value=safe_val(row['test_program']))
        ws3.cell(row=r_idx, column=3, value=safe_val(row['bin']))
        ws3.cell(row=r_idx, column=4, value=safe_val(row['model1']))
        ws3.cell(row=r_idx, column=5, value=safe_val(row['model2']))
        ws3.cell(row=r_idx, column=6, value=safe_val(row['model3']))
        ws3.cell(row=r_idx, column=7, value=safe_val(row['project']))
        ws3.cell(row=r_idx, column=8, value=safe_val(row['exclusive_bin']))
        ws3.cell(row=r_idx, column=9, value=safe_val(row['device_prog_bin']))

    # 4. 库存透视汇总
    ws4 = wb.create_sheet("库存透视")
    headers4 = ['机型', '芯片', '库存类型', '仓库名称', '芯片数量', '单机用量', '可做台数']
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='8B5CF6', end_color='8B5CF6', fill_type='solid')

    rows4 = raw("""
        SELECT m.model1, i.device, i.warehouse_type, i.warehouse_name,
               SUM(i.qty) as total_qty, u.usage_qty,
               CASE WHEN u.usage_qty > 0 THEN SUM(i.qty) / u.usage_qty ELSE 0 END as machine_count
        FROM inventory i
        LEFT JOIN model_mapping m ON i.device_prog_bin = m.device_prog_bin
        LEFT JOIN usage_mapping u ON i.device LIKE u.device || '%' AND m.model1 = u.project
        WHERE m.model1 != ''
        GROUP BY m.model1, i.device, i.warehouse_type, i.warehouse_name
        ORDER BY m.model1, total_qty DESC
    """)

    for r_idx, row in enumerate(rows4, 2):
        ws4.cell(row=r_idx, column=1, value=safe_val(row['model1']))
        ws4.cell(row=r_idx, column=2, value=safe_val(row['device']))
        ws4.cell(row=r_idx, column=3, value=safe_val(row['warehouse_type']))
        ws4.cell(row=r_idx, column=4, value=safe_val(row['warehouse_name']))
        ws4.cell(row=r_idx, column=5, value=safe_val(row['total_qty'], 0))
        ws4.cell(row=r_idx, column=6, value=safe_val(row['usage_qty'], 0))
        ws4.cell(row=r_idx, column=7, value=round(safe_val(row['machine_count'], 0), 1))

    wb.save(filename)
    print(f"✅ 导出完成: {filename}")
    return filename

if __name__ == '__main__':
    export_inventory()