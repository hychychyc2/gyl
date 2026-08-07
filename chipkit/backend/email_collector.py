"""
芯片齐套管理系统 - 邮件采集引擎
支持多邮箱配置，每种用途独立规则
"""
import imaplib
import email
import os
import re
import json
import base64
from email.header import decode_header
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
import openpyxl
import xlrd

from database import (
    insert_many, delete_where, query, update, generate_batch_id,
    write_lock, decrypt_password
)

FILE_MAGIC = {
    b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1": "xls",
    b"\xd0\xcf\x11\xe0": "xls",
    b"PK\x03\x04": "xlsx",
    b"Rar!\x1a\x07\x00": "rar",
}

def safe_int(v, d=0):
    try: return int(float(str(v)))
    except: return d

def safe_float(v, d=0.0):
    try: return float(str(v))
    except: return d

def clean_text(t):
    if t is None: return ''
    if isinstance(t, bytes):
        try: return t.decode('gbk')
        except: return t.decode('utf-8', errors='ignore')
    if isinstance(t, (int, float)):
        return str(int(t)) if t == int(t) else str(t)
    return str(t).replace('\r', '').replace('\n', '').replace('\t', '').strip()

def decode_email_header(header):
    if not header: return ""
    parts = []
    for part, encoding in decode_header(header):
        if isinstance(part, bytes):
            try: parts.append(part.decode(encoding or 'gbk'))
            except:
                try: parts.append(part.decode('utf-8'))
                except: parts.append(part.decode('latin-1'))
        else: parts.append(str(part))
    return ''.join(parts).strip()

def detect_format(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ['.xlsx', '.xls', '.rar']:
        with open(file_path, 'rb') as f:
            header = f.read(8)
        for magic, fmt in FILE_MAGIC.items():
            if header.startswith(magic):
                return fmt
    return ext.lstrip('.')

def get_imap_date_str(d: datetime) -> str:
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    return f"{d.day:02d}-{months[d.month-1]}-{d.year}"

def imap_utf7_decode(s: str) -> str:
    """IMAP UTF-7 解码"""
    if not s or "&" not in s:
        return s
    pattern = re.compile(r'&([^-]+)-')
    def decode_match(match):
        encoded_part = match.group(1).replace(",", "/")
        if not encoded_part:
            return "&"
        try:
            return base64.b64decode(encoded_part + "==", altchars=b"+/").decode("utf-16be")
        except:
            return match.group(0)
    return pattern.sub(decode_match, s)

def imap_utf7_encode(s: str) -> str:
    """IMAP UTF-7 编码"""
    if not s or all(ord(c) < 128 for c in s):
        return s
    result = []
    buffer = []
    for c in s:
        if ord(c) < 128:
            if buffer:
                b64 = base64.b64encode(''.join(buffer).encode("utf-16be")).decode("ascii").rstrip("=").replace("/", ",")
                result.append(f"&{b64}-")
                buffer = []
            result.append(c)
        else:
            buffer.append(c)
    if buffer:
        b64 = base64.b64encode(''.join(buffer).encode("utf-16be")).decode("ascii").rstrip("=").replace("/", ",")
        result.append(f"&{b64}-")
    return ''.join(result)

# ============ 邮件下载 ============
def download_email_attachments(config: Dict, temp_dir: str) -> Optional[tuple]:
    """从邮箱下载匹配的附件，返回 (文件路径, 发件人, 邮件主题, 邮件日期)"""
    account = config.get('account', '')
    password_enc = config.get('password_encrypted', '')
    password = decrypt_password(password_enc)
    imap_server = config.get('imap_server', 'imap.appia.vip')
    root_folder = config.get('root_folder', 'INBOX')
    match_key = config.get('match_key', '')
    suffix = config.get('suffix', '.xlsx')

    if not account or not password:
        print(f"  ⚠️ 邮箱配置不完整: account={account[:5]}...")
        return None

    try:
        print(f"  📧 连接 {imap_server}...")
        mail = imaplib.IMAP4_SSL(imap_server)
        mail.login(account, password)
        print(f"  ✅ 登录成功")

        # 搜索当天邮件（和V8.py完全一致）
        today = datetime.now().date()
        tomorrow = today + timedelta(days=1)
        criteria = f'SINCE "{get_imap_date_str(datetime(today.year, today.month, today.day))}" BEFORE "{get_imap_date_str(datetime(tomorrow.year, tomorrow.month, tomorrow.day))}"'
        print(f"  🔍 搜索条件: {criteria}")

        # 选择文件夹：用 IMAP UTF-7 编码处理中文文件夹名
        # 先列出所有子文件夹
        encoded_root = imap_utf7_encode(root_folder)
        sub_folders = [encoded_root]
        
        try:
            status, folder_list = mail.list()
            if status == 'OK':
                for f in folder_list:
                    if f:
                        fname = f.decode('latin-1') if isinstance(f, bytes) else str(f)
                        # 提取文件夹路径
                        parts = fname.split(' "/" ')
                        if len(parts) > 1:
                            folder_path = parts[1].strip('"')
                            if folder_path.startswith(encoded_root + '/') and folder_path != encoded_root:
                                sub_folders.append(folder_path)
        except Exception as e:
            print(f"  ⚠️ 列出子文件夹失败: {e}")
        
        print(f"  📁 搜索 {len(sub_folders)} 个文件夹")

        all_attachments = []
        for folder_name in sub_folders:
            try:
                mail.select(f'"{folder_name}"', readonly=True)
            except:
                continue
            
            status, messages = mail.search(None, criteria)
            if status != 'OK' or not messages[0]:
                continue

            email_ids = messages[0].split()
            print(f"  📁 {folder_name.replace(encoded_root, '') or '/'}: {len(email_ids)} 封邮件")

            for eid in reversed(email_ids[:50]):  # 每个文件夹最多处理50封
                try:
                    status, msg_data = mail.fetch(eid, '(RFC822)')
                    if status != 'OK': continue
                    msg = email.message_from_bytes(msg_data[0][1])
                    subject = decode_email_header(msg['Subject'])

                    for part in msg.walk():
                        if part.get_content_maintype() == 'multipart': continue
                        if not part.get('Content-Disposition'): continue
                        filename = decode_email_header(part.get_filename())
                        if not filename: continue
                        if not filename.lower().endswith(suffix.lower()): continue
                        if match_key and match_key not in filename: continue

                        payload = part.get_payload(decode=True)
                        all_attachments.append((filename, payload, subject))
                        print(f"  📎 找到: {filename}")
                except Exception as e:
                    continue

            if all_attachments:
                break  # 找到附件了就停止

        mail.close(); mail.logout()

        if not all_attachments:
            return None

        filename, payload, subject = all_attachments[0]
        # 获取发件人
        sender = decode_email_header(msg['From'])
        mail_date = decode_email_header(msg['Date'])

        safe_name = re.sub(r'[\\/:*?"<>|]', '_', filename)
        save_path = os.path.join(temp_dir, safe_name)
        os.makedirs(temp_dir, exist_ok=True)
        with open(save_path, 'wb') as f:
            f.write(payload)
        print(f"  ✅ 保存: {save_path}")
        print(f"  📧 发件人: {sender} | 日期: {mail_date}")
        return (save_path, sender, subject, mail_date)

    except Exception as e:
        print(f"  ❌ 邮件下载失败: {e}")
        return None

# ============ Excel 解析 ============
def parse_excel(file_path: str, sheet_name: str = None, header_row: int = 1,
                col_mapping: Dict = None, max_col: int = None) -> List[Dict]:
    """通用Excel解析，返回 [{col_name: value}, ...]"""
    fmt = detect_format(file_path)
    if fmt not in ['xlsx', 'xls']:
        print(f"  ❌ 不支持格式: {fmt}")
        return []

    try:
        if fmt == 'xlsx':
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb[wb.sheetnames[0]]
            max_row = ws.max_row or 0
            max_col_real = max_col or ws.max_column or 0
        else:
            wb = xlrd.open_workbook(file_path, encoding_override='gbk')
            if sheet_name and sheet_name in wb.sheet_names():
                ws = wb.sheet_by_name(sheet_name)
            else:
                ws = wb.sheet_by_index(0)
            max_row = ws.nrows
            max_col_real = max_col or ws.ncols

        if max_row <= header_row:
            wb.close()
            return []

        # 读表头
        headers = {}
        if fmt == 'xlsx':
            for col in range(1, max_col_real + 1):
                headers[col] = clean_text(ws.cell(row=header_row, column=col).value)
        else:
            for col in range(max_col_real):
                headers[col + 1] = clean_text(ws.cell_value(header_row - 1, col))

        # 建立列映射
        if col_mapping:
            col_indexes = {}
            for target_col, source_col in col_mapping.items():
                try:
                    # 尝试数字列号
                    col_indexes[target_col] = int(source_col)
                except:
                    # 尝试字母列号 (A=1, B=2, ...)
                    if isinstance(source_col, str) and len(source_col) == 1 and source_col.isalpha():
                        col_idx = ord(source_col.upper()) - ord('A') + 1
                        col_indexes[target_col] = col_idx
                    else:
                        # 按表头名匹配
                        for idx, name in headers.items():
                            if name == source_col:
                                col_indexes[target_col] = idx
                                break
            if not col_indexes:
                col_indexes = {str(k): k for k in headers}
        else:
            col_indexes = {str(k): k for k in headers}

        # 读数据
        rows = []
        for row_num in range(header_row + 1, max_row + 1):
            row_data = {}
            if fmt == 'xlsx':
                for target_col, col_idx in col_indexes.items():
                    val = ws.cell(row=row_num, column=col_idx).value
                    row_data[target_col] = clean_text(val) if val is not None else ''
            else:
                for target_col, col_idx in col_indexes.items():
                    col_0 = col_idx - 1
                    if col_0 < max_col_real:
                        val = ws.cell_value(row_num - 1, col_0)
                        row_data[target_col] = clean_text(val) if val is not None else ''

            if any(v for v in row_data.values()):
                rows.append(row_data)

        wb.close()
        return rows

    except Exception as e:
        print(f"  ❌ Excel解析失败: {e}")
        import traceback
        traceback.print_exc()
        return []

# ============ 邮件采集处理函数 ============
def process_shipping_detail(file_path: str, config: Dict, source_info: tuple = None) -> int:
    """处理出货明细 - 从OSAT邮件获取shipping list"""
    mapping = config.get('mapping_config', {})
    if isinstance(mapping, str):
        mapping = json.loads(mapping)

    rows = parse_excel(
        file_path,
        sheet_name=mapping.get('sheet', ''),
        header_row=mapping.get('header_row', 1),
        col_mapping=mapping.get('col_mapping', {})
    )
    if not rows:
        print("  ⚠️ 出货明细解析为空")
        return 0

    batch_id = generate_batch_id()
    src_email = source_info[0] if source_info else ''
    src_file = source_info[1] + ' ' + source_info[2] if source_info else ''
    src_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    ship_rows = []
    for row in rows:
        ship_rows.append({
            'entity': row.get('entity', ''),
            'ship_date': row.get('ship_date', ''),
            'device_pn': row.get('device_pn', ''),
            'wafer_lot_id': row.get('wafer_lot_id', ''),
            'marking': row.get('marking', ''),
            'good_qty': safe_int(row.get('good_qty', 0)),
            'bin': row.get('bin', ''),
            'invoice_no': row.get('invoice_no', ''),
            'test_program': row.get('test_program', ''),
            'osat': row.get('osat', ''),
            'ship_to': row.get('ship_to', ''),
            'test_wo': row.get('test_wo', ''),
            'date_code': row.get('date_code', ''),
            'po': row.get('po', '暂无'),
            'source': 'email',
            'import_batch': batch_id,
            'source_email': src_email,
            'source_file': src_file,
            'source_time': src_time,
        })
    cnt = insert_many('shipping_detail', ship_rows)
    print(f"  ✅ 出货明细: {cnt} 条")
    return cnt

def process_osat_inventory(file_path: str, config: Dict, source_info: tuple = None) -> int:
    """处理OSAT库存 - 覆盖式导入"""
    mapping = config.get('mapping_config', {})
    if isinstance(mapping, str):
        mapping = json.loads(mapping)

    rows = parse_excel(
        file_path,
        sheet_name=mapping.get('sheet', ''),
        header_row=mapping.get('header_row', 3),
        col_mapping=mapping.get('col_mapping', {})
    )
    if not rows:
        print("  ⚠️ OSAT库存解析为空")
        return 0

    # 从描述中提取仓库名
    wn = mapping.get('warehouse_name', '') or config.get('description', '').replace('库存', '').replace('出货明细', '').strip()
    if wn:
        delete_where('inventory', warehouse_type='osat', warehouse_name=wn)

    batch_id = generate_batch_id()
    src_email = source_info[0] if source_info else ''
    src_file = source_info[1] + ' ' + source_info[2] if source_info else ''
    src_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    inv_rows = []
    for row in rows:
        device = row.get('device', '')
        marking = row.get('marking', '')
        b = row.get('bin', '')
        tp = row.get('test_program', '')
        dpb = f"{device}{tp}{b}" if device and tp and b else ''
        inv_rows.append({
            'device': device, 'marking': marking,
            'qty': safe_int(row.get('qty', 0)),
            'bin': b, 'test_program': tp,
            'warehouse_type': 'osat',
            'warehouse_name': wn or row.get('warehouse_name', ''),
            'status': '正常',
            'device_prog_bin': dpb,
            'import_batch': batch_id,
            'source_email': src_email,
            'source_file': src_file,
            'source_time': src_time,
        })
    cnt = insert_many('inventory', inv_rows)
    print(f"  ✅ OSAT库存: {cnt} 条")
    return cnt

def process_hold_inventory(file_path: str, config: Dict, source_info: tuple = None) -> int:
    """处理Hold库存"""
    mapping = config.get('mapping_config', {})
    if isinstance(mapping, str):
        mapping = json.loads(mapping)

    rows = parse_excel(
        file_path,
        sheet_name=mapping.get('sheet', ''),
        header_row=mapping.get('header_row', 2),
        col_mapping=mapping.get('col_mapping', {})
    )
    if not rows:
        return 0

    # 从描述中提取仓库名
    wn = mapping.get('warehouse_name', '') or config.get('description', '').replace('Hold库存', '').replace('库存', '').strip()
    if wn:
        delete_where('inventory', warehouse_type='hold', warehouse_name=wn)

    batch_id = generate_batch_id()
    src_email = source_info[0] if source_info else ''
    src_file = source_info[1] + ' ' + source_info[2] if source_info else ''
    src_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    inv_rows = []
    for row in rows:
        device = row.get('device', '')
        b = row.get('bin', '')
        tp = row.get('test_program', '')
        dpb = f"{device}{tp}{b}" if device and tp and b else ''
        inv_rows.append({
            'device': device, 'marking': row.get('marking', ''),
            'qty': safe_int(row.get('qty', 0)),
            'bin': b, 'test_program': tp,
            'warehouse_type': 'hold',
            'warehouse_name': wn or row.get('warehouse_name', ''),
            'status': 'hold',
            'device_prog_bin': dpb,
            'import_batch': batch_id,
            'source_email': src_email,
            'source_file': src_file,
            'source_time': src_time,
        })
    cnt = insert_many('inventory', inv_rows)
    print(f"  ✅ Hold库存: {cnt} 条")
    return cnt

def process_model_mapping(file_path: str, config: Dict) -> int:
    """处理机型对照表"""
    rows = parse_excel(file_path, header_row=1)
    if not rows:
        return 0

    model_rows = []
    for row in rows:
        device = row.get('device', '')
        tp = row.get('test_program', '')
        b = row.get('bin', '')
        dpb = f"{device}{tp}{b}" if device and tp and b else ''
        model_rows.append({
            'device': device, 'test_program': tp, 'bin': b,
            'device_prog_bin': dpb,
            'model1': row.get('model1', ''),
            'model2': row.get('model2', ''),
            'model3': row.get('model3', ''),
            'model4': row.get('model4', ''),
            'model5': row.get('model5', ''),
            'product': row.get('product', ''),
            'osat_model': row.get('osat_model', ''),
            'project': row.get('project', ''),
            'exclusive_bin': safe_int(row.get('exclusive_bin', 0)),
        })
    cnt = insert_many('model_mapping', model_rows)
    print(f"  ✅ 机型对照: {cnt} 条")
    return cnt

def process_mix_bin(file_path: str, config: Dict) -> int:
    """处理混BIN关系 - 覆盖"""
    rows = parse_excel(file_path, header_row=1)
    if not rows:
        return 0

    delete_where('mix_bin')
    mix_rows = []
    for row in rows:
        if not row.get('device_prog_bin', '').strip():
            continue
        mix_rows.append({
            'device_prog_bin': row.get('device_prog_bin', ''),
            'material_code': row.get('material_code', ''),
            'device': row.get('device', ''),
            'test_program': row.get('test_program', ''),
            'bin': row.get('bin', ''),
            'col': row.get('col', ''),
            'model_name': row.get('model_name', ''),
            'mix_group': row.get('mix_group', ''),
            'stock_qty': safe_int(row.get('stock_qty', 0)),
            'chips_per_unit': safe_int(row.get('chips_per_unit', 0)),
            'convertible_qty': safe_float(row.get('convertible_qty', 0)),
            'summary_actual': safe_int(row.get('summary_actual', 0)),
            'is_exclusive': safe_int(row.get('is_exclusive', 0)),
        })
    cnt = insert_many('mix_bin', mix_rows)
    print(f"  ✅ 混BIN: {cnt} 条")
    return cnt

def process_order_allocation(file_path: str, config: Dict) -> int:
    """处理订单分配（张胜文邮件）- 更新齐套达成"""
    rows = parse_excel(file_path, header_row=1)
    if not rows:
        return 0

    for row in rows:
        region = row.get('region', '')
        sub = row.get('subcontractor', '')
        if not sub: continue

        # 解析月份
        month_plan = {}
        for k, v in row.items():
            if k.startswith('month_') or k.startswith('plan_'):
                month = k.replace('month_', '').replace('plan_', '')
                month_plan[month] = safe_int(v)

        dev = row.get('device', '')
        model = row.get('model_name', '')
        project = row.get('project', '')
        existing = query('kit_completion',
                        where='region=? AND device=? AND model_name=? AND project=? AND subcontractor=?',
                        params=(region, dev, model, project, sub))

        data = {
            'region': region,
            'location': row.get('location', ''),
            'device': dev,
            'model_name': model,
            'project': project,
            'usage_per_unit': safe_int(row.get('usage_per_unit', 0)),
            'subcontractor': sub,
            'sub_code': row.get('sub_code', ''),
            'month_plan': json.dumps(month_plan, ensure_ascii=False),
            'remark': row.get('remark', ''),
        }

        if existing:
            update('kit_completion', existing[0]['id'], data)
        else:
            insert('kit_completion', data)

    print(f"  ✅ 订单分配: {len(rows)} 条")
    return len(rows)

# ============ 批量采集 ============
def fetch_all(temp_dir: str) -> Dict:
    """根据所有活跃的邮件配置批量采集"""
    configs = query('email_config', where='active=1')
    results = {}
    for cfg in configs:
        cfg = dict(cfg)
        purpose = cfg.get('purpose', '')
        desc = cfg.get('description', purpose)
        print(f"\n📧 [{desc}]")

        file_path = download_email_attachments(cfg, temp_dir)
        if not file_path:
            print(f"  ⚠️ 未找到附件")
            results[purpose] = 0
            continue

        fp, source_info = file_path[0], file_path[1:]

        update('email_config', cfg['id'], {
            'last_fetch': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'version': cfg.get('version', 1)
        })

        count = 0
        if purpose == 'shipping_detail':
            count = process_shipping_detail(fp, cfg, source_info)
        elif purpose == 'osat_inventory':
            count = process_osat_inventory(fp, cfg, source_info)
        elif purpose == 'hold_inventory':
            count = process_hold_inventory(fp, cfg, source_info)
        elif purpose == 'model_mapping':
            count = process_model_mapping(fp, cfg, source_info)
        elif purpose == 'mix_bin':
            count = process_mix_bin(fp, cfg, source_info)
        elif purpose == 'order_allocation':
            count = process_order_allocation(fp, cfg, source_info)

        results[purpose] = count
        try: os.remove(file_path)
        except: pass

    return results

if __name__ == "__main__":
    from database import init_db
    init_db()
    fetch_all("/tmp/chipkit_email")