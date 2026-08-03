"""
库存迁移 - 仅导入库存sheet
"""
import os, sys, re
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))
from database import init_db, insert_many, delete_where, count as get_count
import openpyxl

WORKSPACE = os.path.join(os.path.dirname(__file__), "..", "..")
EXCEL = os.path.join(WORKSPACE, "芯片齐套表_lastest (78).xlsx")

def safe_int(v, d=0):
    try: return int(float(str(v)))
    except: return d
def safe_str(v):
    if v is None: return ''
    return str(v).replace('\r','').replace('\n','').strip()

def migrate_inventory():
    print("📦 库存数据迁移")
    init_db()

    all_rows = []
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)

    # MES 库存 sheets
    mes_sheets = [
        ('SZKXYCL', 'SZKXYCL', 'other'),
        ('HSJXYCL', 'HSJXYCL', 'other'),
        ('SZKYCGL', 'SZKYCGL', 'other'),
        ('HSJYCGL', 'HSJYCGL', 'other'),
    ]

    for sheet_name, wh_name, wh_type in mes_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️ {sheet_name} 不存在")
            continue
        ws = wb[sheet_name]
        cnt = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row): continue
            v = [safe_str(x) for x in row]
            if len(v) < 10: continue
            device = v[19] if len(v)>19 else ''
            mark = v[22] if len(v)>22 else ''
            b = v[23] if len(v)>23 else ''
            tp = v[24] if len(v)>24 else ''
            dpb = f"{device}{tp}{b}" if device and tp and b else ''
            all_rows.append({
                'device': device, 'marking': mark, 'qty': safe_int(v[9]),
                'bin': b, 'test_program': tp, 'location_code': wh_name,
                'warehouse_type': wh_type, 'warehouse_name': wh_name,
                'batch': mark, 'date_code': v[20] if len(v)>20 else '',
                'material_code': v[1] if len(v)>1 else '',
                'status': v[12] if len(v)>12 else '正常',
                'device_prog_bin': dpb, 'import_batch': 'MIGRATION',
            })
            cnt += 1
        print(f"  ✅ {sheet_name}: {cnt} 条")

    # QHBS 保税仓
    if 'QHBS' in wb.sheetnames:
        ws = wb['QHBS']
        cnt = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not any(row): continue
            v = [safe_str(x) for x in row]
            if len(v) < 13: continue
            device = v[19] if len(v)>19 else ''
            b = v[18] if len(v)>18 else ''
            tp = v[17] if len(v)>17 else ''
            dpb = f"{device}{tp}{b}" if device and tp and b else ''
            all_rows.append({
                'device': device, 'marking': v[12] if len(v)>12 else '',
                'qty': safe_int(v[13]) if len(v)>13 else 0,
                'bin': b, 'test_program': tp, 'location_code': 'QHBS',
                'warehouse_type': 'bonded', 'warehouse_name': 'QHBS',
                'batch': v[12] if len(v)>12 else '',
                'material_code': v[1] if len(v)>1 else '',
                'status': '正常', 'device_prog_bin': dpb, 'import_batch': 'MIGRATION',
            })
            cnt += 1
        print(f"  ✅ QHBS: {cnt} 条")

    # osat库存
    if 'osat库存' in wb.sheetnames:
        ws = wb['osat库存']
        cnt = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not any(row): continue
            v = [safe_str(x) for x in row]
            if len(v) < 6: continue
            device = v[0]; mark = v[1]; b = v[3]; tp = v[4]
            dpb = f"{device}{tp}{b}" if device and tp and b else ''
            all_rows.append({
                'device': device, 'marking': mark, 'qty': safe_int(v[2]),
                'bin': b, 'test_program': tp, 'location_code': v[5] if len(v)>5 else '',
                'warehouse_type': 'osat', 'warehouse_name': v[5] if len(v)>5 else '',
                'status': '正常', 'device_prog_bin': dpb, 'import_batch': 'MIGRATION',
            })
            cnt += 1
        print(f"  ✅ osat库存: {cnt} 条")

    # hold
    if 'hold' in wb.sheetnames:
        ws = wb['hold']
        cnt = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row): continue
            v = [safe_str(x) for x in row]
            if len(v) < 6: continue
            device = v[0]; mark = v[1]; b = v[3]; tp = v[4]
            dpb = f"{device}{tp}{b}" if device and tp and b else ''
            all_rows.append({
                'device': device, 'marking': mark, 'qty': safe_int(v[2]),
                'bin': b, 'test_program': tp, 'location_code': v[5] if len(v)>5 else '',
                'warehouse_type': 'hold', 'warehouse_name': v[5] if len(v)>5 else '',
                'status': 'hold', 'device_prog_bin': dpb, 'import_batch': 'MIGRATION',
            })
            cnt += 1
        print(f"  ✅ hold: {cnt} 条")

    wb.close()

    # 清空库存表
    delete_where('inventory')
    cnt = insert_many('inventory', all_rows)
    print(f"\n  ✅ 库存总计: {cnt} 条")

if __name__ == '__main__':
    migrate_inventory()