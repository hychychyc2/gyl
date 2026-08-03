"""
精简迁移 - 只读必需列，快速导入
"""
import os, sys, json, re
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))
from database import init_db, insert_many, delete_where, count as get_count
import openpyxl

WORKSPACE = os.path.join(os.path.dirname(__file__), "..", "..")
EXCEL = os.path.join(WORKSPACE, "芯片齐套表_lastest (78).xlsx")
DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

def safe_int(v, d=0):
    try: return int(float(str(v)))
    except: return d

def safe_float(v, d=0.0):
    try: return float(str(v))
    except: return d

def safe_str(v):
    if v is None: return ''
    return str(v).replace('\r','').replace('\n','').strip()

def fast_migrate():
    print("🦞 快速数据迁移")
    init_db()

    # 1. 机型对照表
    print("\n📋 机型对照表...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['机型对照表']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=35, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        device = v[2]; tp = v[3]; b = v[4]
        dpb = f"{device}{tp}{b}" if device and tp and b else ''
        rows.append({
            'device': device, 'test_program': tp, 'bin': b,
            'model1': v[5], 'model2': v[6], 'model3': v[7], 'model4': v[8], 'model5': v[9],
            'device_prog_bin': dpb, 'exclusive_bin': 0,
            'product': v[31] if len(v)>31 else '', 'osat_model': v[32] if len(v)>32 else '',
            'project': v[33] if len(v)>33 else '',
            'chip_total': safe_int(v[34]) if len(v)>34 else 0,
            'unit_count': safe_int(v[35]) if len(v)>35 else 0,
        })
    wb.close()
    delete_where('model_mapping')
    cnt = insert_many('model_mapping', rows)
    print(f"  ✅ {cnt} 条")

    # 2. 出货明细
    print("\n📋 出货明细...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['出货明细表']
    rows = []
    row_count = 0
    for row in ws.iter_rows(min_row=3, max_col=18, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if not v[0] and not v[2]: continue
        rows.append({
            'entity': v[0], 'ship_date': v[1], 'device_pn': v[2], 'wafer_lot_id': v[3],
            'marking': v[4], 'good_qty': safe_int(v[5]), 'bin': v[6],
            'invoice_no': v[7], 'test_program': v[8], 'osat': v[9], 'ship_to': v[10],
            'test_wo': v[11], 'date_code': v[12] if len(v)>12 else '',
            'po': v[13] if len(v)>13 else '暂无', 'so': v[14] if len(v)>14 else '',
            'model1': v[15] if len(v)>15 else '', 'line_no': v[16] if len(v)>16 else '',
            'material_code': v[17] if len(v)>17 else '', 'source': 'migration',
            'import_batch': 'MIGRATION',
        })
        row_count += 1
        if row_count % 5000 == 0: print(f"  已处理 {row_count}...")
    wb.close()
    delete_where('shipping_detail')
    cnt = insert_many('shipping_detail', rows)
    print(f"  ✅ {cnt} 条")

    # 3. 用量对照
    print("\n📋 用量对照...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['用量对照表']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if not v[0]: continue
        rows.append({'device': v[0], 'model_name': v[1], 'project': v[2], 'usage_qty': safe_int(v[3])})
    wb.close()
    delete_where('usage_mapping')
    cnt = insert_many('usage_mapping', rows)
    print(f"  ✅ {cnt} 条")

    # 4. 混BIN
    print("\n📋 混BIN组合...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['混BIN组合']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if not v[0]: continue
        rows.append({
            'device_prog_bin': v[0], 'material_code': v[1] if len(v)>1 else '',
            'device': v[2] if len(v)>2 else '', 'test_program': v[3] if len(v)>3 else '',
            'bin': v[4] if len(v)>4 else '', 'col': v[5] if len(v)>5 else '',
            'model_name': v[6] if len(v)>6 else '', 'mix_group': v[7] if len(v)>7 else '',
            'stock_qty': safe_int(v[8]) if len(v)>8 else 0,
            'chips_per_unit': safe_int(v[9]) if len(v)>9 else 0,
            'convertible_qty': safe_float(v[10]) if len(v)>10 else 0,
            'summary_actual': safe_int(v[12]) if len(v)>12 else 0,
        })
    wb.close()
    delete_where('mix_bin')
    cnt = insert_many('mix_bin', rows)
    print(f"  ✅ {cnt} 条")

    # 5. 外协代码
    print("\n📋 外协代码...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['外协代码对照表']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if len(v) < 3: continue
        rows.append({'type': v[0], 'short_name': v[1], 'internal_code': v[2],
                     'external_name': v[2], 'ship_to_code': v[7] if len(v)>7 else '',
                     'address': v[3] if len(v)>3 else '', 'contact': v[6] if len(v)>6 else ''})
    wb.close()
    delete_where('subcontractor_mapping')
    cnt = insert_many('subcontractor_mapping', rows)
    print(f"  ✅ {cnt} 条")

    # 6. 物流时间
    print("\n📋 物流时间...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['物流时间']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if not v[0]: continue
        days = int(re.findall(r'\d+', v[1])[0]) if re.findall(r'\d+', v[1]) else 0
        rows.append({'destination': v[0], 'transit_days': days, 'latest_ship_day': v[2] if len(v)>2 else ''})
    wb.close()
    delete_where('logistics_time')
    cnt = insert_many('logistics_time', rows)
    print(f"  ✅ {cnt} 条")

    # 7. 料号Device
    print("\n📋 料号Device...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['料号对应Device']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if len(v) < 3: continue
        rows.append({'erp_code': v[1], 'device': v[2], 'wafer_pn': v[3] if len(v)>3 else '',
                     'description': v[4] if len(v)>4 else '', 'package_desc': v[5] if len(v)>5 else ''})
    wb.close()
    delete_where('material_device')
    cnt = insert_many('material_device', rows)
    print(f"  ✅ {cnt} 条")

    print("\n" + "="*50)
    print("✅ 迁移完成!")
    for t in ['shipping_detail','inventory','model_mapping','usage_mapping','mix_bin','kit_completion','subcontractor_mapping','logistics_time','material_device']:
        print(f"  {t}: {get_count(t)} 条")
    print("="*50)

if __name__ == '__main__':
    fast_migrate()