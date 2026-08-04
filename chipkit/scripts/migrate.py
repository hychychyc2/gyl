"""
芯片齐套管理系统 - 数据迁移脚本
从 Excel 一次性导入所有数据到 SQLite
"""
import os, sys, re, json
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))
from database import init_db, insert_many, delete_where, count as get_count, insert
import openpyxl

WORKSPACE = os.path.join(os.path.dirname(__file__), "..", "..")
EXCEL = os.path.join(WORKSPACE, "芯片齐套表_lastest (78).xlsx")

def safe_int(v, d=0):
    try: return int(float(str(v)))
    except: return d

def safe_float(v, d=0.0):
    try: return float(str(v))
    except: return d

def safe_str(v):
    if v is None: return ''
    if isinstance(v, (int, float)):
        return str(int(v)) if v == int(v) else str(v)
    return str(v).replace('\r', '').replace('\n', '').replace('\t', '').strip()

def main():
    print("=" * 60)
    print("🦞 芯片齐套管理系统 - 数据迁移")
    print("=" * 60)

    if not os.path.exists(EXCEL):
        print(f"❌ 找不到 Excel 文件: {EXCEL}")
        print("   请将 芯片齐套表_lastest (78).xlsx 放在 chipkit/ 的同级目录")
        sys.exit(1)

    init_db()

    # ========== 1. 机型对照表 ==========
    print("\n📋 机型对照表...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['机型对照表']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=35, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if not v[2]: continue
        device = v[2]; tp = v[3]; b = v[4]
        dpb = f"{device}{tp}{b}" if device and tp and b else ''
        rows.append({
            'device': device, 'test_program': tp, 'bin': b,
            'model1': v[5], 'model2': v[6], 'model3': v[7], 'model4': v[8],
            'model5': v[9], 'model6': v[10], 'model7': v[11], 'model8': v[12],
            'device_prog_bin': dpb, 'exclusive_bin': 0,
            'product': v[31] if len(v) > 31 else '',
            'osat_model': v[32] if len(v) > 32 else '',
            'project': v[33] if len(v) > 33 else '',
        })
    wb.close()
    delete_where('model_mapping')
    cnt = insert_many('model_mapping', rows)
    print(f"  ✅ {cnt} 条")

    # ========== 2. 出货明细 ==========
    print("\n📋 出货明细...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['出货明细表']
    rows = []
    for row in ws.iter_rows(min_row=3, max_col=18, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if not v[0] and not v[2]: continue
        rows.append({
            'entity': v[0], 'ship_date': v[1], 'device_pn': v[2],
            'wafer_lot_id': v[3], 'marking': v[4], 'good_qty': safe_int(v[5]),
            'bin': v[6], 'invoice_no': v[7], 'test_program': v[8],
            'osat': v[9], 'ship_to': v[10], 'test_wo': v[11],
            'date_code': v[12] if len(v) > 12 else '',
            'po': v[13] if len(v) > 13 else '暂无',
            'so': v[14] if len(v) > 14 else '',
            'model1': v[15] if len(v) > 15 else '',
            'line_no': v[16] if len(v) > 16 else '',
            'material_code': v[17] if len(v) > 17 else '',
            'source': 'migration', 'import_batch': 'MIGRATION',
        })
    wb.close()
    delete_where('shipping_detail')
    cnt = insert_many('shipping_detail', rows)
    print(f"  ✅ {cnt} 条")

    # ========== 3. 用量对照 ==========
    print("\n📋 用量对照...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['用量对照表']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if not v[0]: continue
        rows.append({
            'device': v[0], 'model_name': v[1],
            'project': v[2], 'usage_qty': safe_int(v[3])
        })
    wb.close()
    delete_where('usage_mapping')
    cnt = insert_many('usage_mapping', rows)
    print(f"  ✅ {cnt} 条")

    # ========== 4. 混BIN组合 ==========
    print("\n📋 混BIN组合...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['混BIN组合']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if not v[0]: continue
        rows.append({
            'device_prog_bin': v[0],
            'material_code': v[1] if len(v) > 1 else '',
            'device': v[2] if len(v) > 2 else '',
            'test_program': v[3] if len(v) > 3 else '',
            'bin': v[4] if len(v) > 4 else '',
            'col': v[5] if len(v) > 5 else '',
            'model_name': v[6] if len(v) > 6 else '',
            'mix_group': v[7] if len(v) > 7 else '',
            'stock_qty': safe_int(v[8]) if len(v) > 8 else 0,
            'chips_per_unit': safe_int(v[9]) if len(v) > 9 else 0,
            'convertible_qty': safe_float(v[10]) if len(v) > 10 else 0,
            'summary_actual': safe_int(v[12]) if len(v) > 12 else 0,
        })
    wb.close()
    delete_where('mix_bin')
    cnt = insert_many('mix_bin', rows)
    print(f"  ✅ {cnt} 条")

    # ========== 5. 外协代码 ==========
    print("\n📋 外协代码...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['外协代码对照表']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if len(v) < 3: continue
        rows.append({
            'type': v[0], 'short_name': v[1],
            'internal_code': v[2], 'external_name': v[2],
            'ship_to_code': v[7] if len(v) > 7 else '',
            'address': v[3] if len(v) > 3 else '',
            'contact': v[6] if len(v) > 6 else '',
        })
    wb.close()
    delete_where('subcontractor_mapping')
    cnt = insert_many('subcontractor_mapping', rows)
    print(f"  ✅ {cnt} 条")

    # ========== 6. 物流时间 ==========
    print("\n📋 物流时间...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['物流时间']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if not v[0]: continue
        days = int(re.findall(r'\d+', v[1])[0]) if re.findall(r'\d+', v[1]) else 0
        rows.append({
            'destination': v[0], 'transit_days': days,
            'latest_ship_day': v[2] if len(v) > 2 else '',
        })
    wb.close()
    delete_where('logistics_time')
    cnt = insert_many('logistics_time', rows)
    print(f"  ✅ {cnt} 条")

    # ========== 7. 料号Device ==========
    print("\n📋 料号Device...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    ws = wb['料号对应Device']
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        if not row or not any(row): continue
        v = [safe_str(x) for x in row]
        if len(v) < 3: continue
        rows.append({
            'erp_code': v[1], 'device': v[2],
            'wafer_pn': v[3] if len(v) > 3 else '',
            'description': v[4] if len(v) > 4 else '',
            'package_desc': v[5] if len(v) > 5 else '',
        })
    wb.close()
    delete_where('material_device')
    cnt = insert_many('material_device', rows)
    print(f"  ✅ {cnt} 条")

    # ========== 8. 库存 ==========
    print("\n📋 库存数据...")
    wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True)
    all_inv = []

    # MES 格式库存
    mes_sheets = [
        ('SZKXYCL', 'SZKXYCL', 'other'),
        ('HSJXYCL', 'HSJXYCL', 'other'),
        ('SZKYCGL', 'SZKYCGL', 'other'),
        ('HSJYCGL', 'HSJYCGL', 'other'),
    ]
    for sn, wn, wt in mes_sheets:
        if sn not in wb.sheetnames:
            print(f"  ⚠️ {sn} 不存在，跳过")
            continue
        ws = wb[sn]
        cnt = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row): continue
            v = [safe_str(x) for x in row]
            if len(v) < 10: continue
            device = v[19] if len(v) > 19 else ''
            mark = v[11] if len(v) > 11 else ''  # 生产批次
            b = v[22] if len(v) > 22 else ''  # BIN
            tp = v[23] if len(v) > 23 else ''  # 程序
            dpb = f"{device}{tp}{b}" if device and tp and b else ''
            all_inv.append({
                'device': device, 'marking': mark,
                'qty': safe_int(v[9]),
                'bin': b, 'test_program': tp,
                'warehouse_type': wt, 'warehouse_name': wn,
                'material_code': v[1] if len(v) > 1 else '',
                'product_desc': v[2] if len(v) > 2 else '',
                'batch': v[10] if len(v) > 10 else '',
                'date_code': v[20] if len(v) > 20 else '',
                'status': v[12] if len(v) > 12 else '正常',
                'device_prog_bin': dpb, 'import_batch': 'MIGRATION',
            })
            cnt += 1
        print(f"  ✅ {sn}: {cnt} 条")

    # QHBS 保税仓
    if 'QHBS' in wb.sheetnames:
        ws = wb['QHBS']
        cnt = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not any(row): continue
            v = [safe_str(x) for x in row]
            if len(v) < 12: continue
            # QHBS列: 0=org, 1=物料编码, 8=子库存, 11=批次, 12=数量, 13=marking, 15=程序, 16=BIN, 17=device, 19=cooke
            device = v[17] if len(v) > 17 else ''
            b = v[16] if len(v) > 16 else ''
            tp = v[15] if len(v) > 15 else ''
            dpb = f"{device}{tp}{b}" if device and tp and b else ''
            all_inv.append({
                'device': device, 'marking': v[13] if len(v) > 13 else '',
                'qty': safe_int(v[12]) if len(v) > 12 else 0,
                'bin': b, 'test_program': tp,
                'warehouse_type': 'bonded', 'warehouse_name': 'QHBS',
                'material_code': v[1] if len(v) > 1 else '',
                'product_desc': v[2] if len(v) > 2 else '',
                'batch': v[11] if len(v) > 11 else '',
                'sub_inventory': v[8] if len(v) > 8 else '',
                'location': v[10] if len(v) > 10 else '',
                'org': v[0] if len(v) > 0 else '',
                'status': '正常',
                'remark': v[19] if len(v) > 19 else '',
                'device_prog_bin': dpb, 'import_batch': 'MIGRATION',
            })
            cnt += 1
        print(f"  ✅ QHBS: {cnt} 条")

    # osat库存
    if 'osat库存' in wb.sheetnames:
        ws = wb['osat库存']
        cnt = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not any(row): continue
            v = [safe_str(x) for x in row]
            if len(v) < 6: continue
            device = v[0]; mark = v[1]; b = v[3]; tp = v[4]
            dpb = f"{device}{tp}{b}" if device and tp and b else ''
            all_inv.append({
                'device': device, 'marking': mark,
                'qty': safe_int(v[2]),
                'bin': b, 'test_program': tp,
                'warehouse_type': 'osat',
                'warehouse_name': v[5] if len(v) > 5 else '',
                'status': '正常',
                'device_prog_bin': dpb, 'import_batch': 'MIGRATION',
            })
            cnt += 1
        print(f"  ✅ osat库存: {cnt} 条")

    # hold
    if 'hold' in wb.sheetnames:
        ws = wb['hold']
        cnt = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row): continue
            v = [safe_str(x) for x in row]
            if len(v) < 6: continue
            device = v[0]; mark = v[1]; b = v[3]; tp = v[4]
            dpb = f"{device}{tp}{b}" if device and tp and b else ''
            all_inv.append({
                'device': device, 'marking': mark,
                'qty': safe_int(v[2]),
                'bin': b, 'test_program': tp,
                'warehouse_type': 'hold',
                'warehouse_name': v[5] if len(v) > 5 else '',
                'status': 'hold',
                'device_prog_bin': dpb, 'import_batch': 'MIGRATION',
            })
            cnt += 1
        print(f"  ✅ hold: {cnt} 条")

    wb.close()
    delete_where('inventory')
    cnt = insert_many('inventory', all_inv)
    print(f"  ✅ 库存总计: {cnt} 条")

    # ========== 默认用户 ==========
    print("\n📋 创建默认用户...")
    existing = query('users', where='email=?', params=('yuchuan.he@casue.com',))
    if not existing:
        insert('users', {
            'email': 'yuchuan.he@casue.com',
            'name': '何宇川',
            'role': 'admin',
            'password_hash': 'chipkit2026',
            'active': 1,
        })
        print("  ✅ 管理员: yuchuan.he@casue.com / chipkit2026")
    else:
        print("  ⚠️ 用户已存在")

    # ========== 汇总 ==========
    print("\n" + "=" * 60)
    print("✅ 迁移完成！数据统计：")
    for t in ['shipping_detail', 'inventory', 'model_mapping', 'usage_mapping',
              'mix_bin', 'subcontractor_mapping', 'logistics_time', 'material_device']:
        print(f"  {t}: {get_count(t)} 条")
    print("=" * 60)

if __name__ == '__main__':
    main()