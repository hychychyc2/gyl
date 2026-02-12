import pandas as pd
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
ssc_file = "E:\\供应链\\总表\\191125\\SSC备料计划1112 V1.4.xlsx"
power_supply_file = "E:\\供应链\\总表\\201125\\电源欠料表20251120.xlsx"
overseas_bom_file = "E:\\供应链\\总表\\191125\\海外拼BOM版本--2025.11.17 vs 1112版大计划-v1.0.1.xlsx"
global_order_file = "E:\\供应链\\总表\\191125\\全球订单分配（1112版大计分配结果）20251117 V1.0.3.xlsx"
def read_ssc_xlsx(file_path):
    """读取ssc.xlsx文件，获取内部机型到电源的映射"""
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在")
        return {}
    
    try:
        df = pd.read_excel(file_path, header=1, sheet_name=1)  # 第二行是表头
        # 假设内部机型列和电源列分别为第一列和最后一列
        print(df)
        tot =0
        for i in df.iterrows():
            id=0
            for col in df.columns:
            
                if col == '内部机型':
                    break
                id+=1
            id2=0
            for col in df.columns:
            
                if col == '电源':
                   break
                id2+=1
            tot+=1
        #print(tot)
        for i in range(tot):
            #print(i)
            #print(df.iloc[i,id2])
            df.iloc[i,id2]= df.iloc[i-1,id2]if pd.isna(df.iloc[i,id2]) else df.iloc[i,id2]
            #print(df.iloc[i,5])
        power_dict = dict(zip(df.iloc[:, id].astype(str).str[:9], df.iloc[:, id2]))
        return power_dict
    except Exception as e:
        print(f"读取 {file_path} 时出错: {e}")
        return {}

    # 读取ssc.xlsx
power_dict = read_ssc_xlsx(ssc_file)
print(power_dict)
def unmerge_and_fill(input_file, output_file):
    # 加载工作簿
    wb = load_workbook(input_file)
    ws = wb.active
    
    # 遍历所有合并单元格
    merged_ranges = ws.merged_cells.ranges.copy()
    for merged_range in merged_ranges:
        # 获取合并区域的起始和结束坐标
        min_col, min_row, max_col, max_row = merged_range.bounds
        
        # 解除合并
        ws.unmerge_cells(start_row=min_row, start_column=min_col,
                         end_row=max_row, end_column=max_col)
        
        # 获取合并区域左上角单元格的值
        top_left_cell = ws.cell(row=min_row, column=min_col)
        fill_value = top_left_cell.value
        
        # 填充整个区域
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.value = fill_value
    
    # 保存修改后的工作簿
    wb.save(output_file)
    print(f"处理完成，已保存至 {output_file}")

# 使用示例


def read_power_supply_xlsx(file_path, sheet_name="电源成品供应"):
    """读取电源千聊表20250516.xlsx文件，获取表头信息"""
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在")
        return None, None, None
    
    try:
        # 读取第二行表头(海外外协、马来PIE等)
        df_header1 = pd.read_excel(file_path, sheet_name=sheet_name, header=1, nrows=1)
        # 读取第三行表头(内部机型代码、APW11等)
        df_header2 = pd.read_excel(file_path, sheet_name=sheet_name, header=2, nrows=1)
        
        # 从第四行开始加载数据(月份数据)
        df_data = pd.read_excel(file_path, sheet_name=sheet_name, header=3)
        
        return df_header1, df_header2, df_data
    except Exception as e:
        print(f"读取 {file_path} 时出错: {e}")
        return None, None, None

def read_overseas_bom_xlsx(file_path, months):
    """读取海外拼BOM版本.xlsx文件，按月份获取数据"""
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在")
        return {}
    
    monthly_data = {}
    input_excel = "input.xlsx"  # 替换为你的输入文件路径
    output_excel = "output.xlsx"  # 替换为你的输出文件路径
    # unmerge_and_fill(input_excel, output_excel)
    for month in months:
        sheet_name = f"2025.{str(month).zfill(2)}"
        try:
            # 第二行是表头
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
            # 解除合并单元格并填充
            # df = df.fillna(method='ffill')
            # print(df)
            monthly_data[month] = df
        except Exception as e:
            print(f"读取 {file_path} 的 {sheet_name} 工作表时出错: {e}")
        sheet_name = f"2026.{str(month).zfill(2)}"
        try:
            # 第二行是表头
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
            # 解除合并单元格并填充
            # df = df.fillna(method='ffill')
            # print(df)
            monthly_data[month] = df
        except Exception as e:
            print(f"读取 {file_path} 的 {sheet_name} 工作表时出错: {e}")
    
    return monthly_data

def read_global_order_xlsx(file_path, current_month, next_month):
    """读取全球订单分配.xlsx文件，获取当月和次月的订单数据"""
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在")
        return {}, {}
    
    current_sheet = f"国内整机{current_month}月-{next_month}月订单分配 "
    next_sheet = current_sheet  # 假设在同一个sheet中
    
    current_data = {}
    next_data = {}
    
    try:
        # 第三行是表头
        df = pd.read_excel(file_path, sheet_name=current_sheet, header=2)
        
        # 解除合并单元格并填充
        df.iloc[:, 2] = df.iloc[:, 2].fillna(method='ffill')        
        # 处理当月数据
        # print(df.iterrows())
        id=0
        # print(df.columns)
        for col in df.columns:
            
            if col == '内部代码':
                break
            id+=1
        id=2
        id2=0
        print(id,id2)
        for col in df.columns:
            
            if col == '分配外协':
               break
            id2+=1
        id3=0
        id4=0

        id2=4
        id_tmp=0
        for col in df.columns:
            # print(type(col))
            # print(col[:5])
            if type(col) is str and col[:6] == f"本次分配数量":
                id3=id4
                id4=id_tmp
            id_tmp+=1
        
        # print("PPPP")
        # print(df)
        print(id3,id4)
        for _, row in df.iterrows():
            # print(row)
            # # 假设第四列是外协，第三列是电源，第七列是当月数量
            # print(row.iloc[1])
            if(row.iloc[1]== '合计'):
                break
            supplier = row.iloc[id2]
            power = row.iloc[id][:9]
            print(supplier,power)
            if power in power_dict:
                power = power_dict[power]
            # power = power_dict[power]

            quantity = row.iloc[id3]  # 第七列
            quantity2 = row.iloc[id4]  # 第16列
            if(supplier == '奥海' and power == 'APW17'):
                print(supplier,power,quantity,quantity2)
            if pd.notna(quantity):
                if supplier not in current_data:
                    current_data[supplier] = {}
                if power not in current_data[supplier]:
                    current_data[supplier][power] = 0
                current_data[supplier][power] += int(quantity+0.5)
            if pd.notna(quantity2):
                if supplier not in next_data:
                    next_data[supplier] = {}
                if power not in next_data[supplier]:
                    next_data[supplier][power] = 0
                next_data[supplier][power] += int(quantity2+0.5)
       
    
    except Exception as e:
        print(f"读取 {file_path} 的 {current_sheet} 工作表时出错: {e}")
    
    return current_data, next_data
def find_latest_day_in_month(df: pd.DataFrame) -> str:
    """
    查找DataFrame表头中当月日期最晚的订单量列
    
    Args:
        df: 输入的DataFrame
    
    Returns:
        当月最晚日期对应的订单量列名，如果不存在则返回None
    """
    # 提取当前年份和月份
    current_year = datetime.now().year
    current_month = datetime.now().month
    #current_month = 8
    # 筛选以"订单量"结尾的列，并解析日期
    valid_columns = []
    print(df.columns)
    for col in df.columns:
        print(col)
        if col.endswith("订单量"):
            try:
                # 假设列名格式为"MMDD订单量"，如"0514订单量"
                month_str = col[:2]
                day_str = col[2:4]
                month = int(month_str)
                day = int(day_str)
                
                # 验证日期有效性和是否属于当月
                if month == current_month:
                    valid_columns.append((month, day, col))
            except (ValueError, IndexError):
                continue  # 跳过解析失败的列
    
    if not valid_columns:
        return None
    
    # 按日期排序，找出最晚的一天
    valid_columns.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return valid_columns[0][2]  # 返回最晚日期对应的列名
def process_data():
    """处理所有数据并生成最终字典"""
    # 获取当前月份
    today = datetime.now()
    current_month = datetime.now().month

    #current_month = 8  # 例如5
    next_month =12 if current_month==11  else (current_month + 1)%12  # 例如6
    months = list(range(1, 13))  # 1月到12月
   
    
    # 读取电源千聊表
    header1, header2, power_supply_data = read_power_supply_xlsx(power_supply_file)
    print("++++++++++++++")
    # 读取海外拼BOM版本.xlsx
    overseas_bom_data = read_overseas_bom_xlsx(overseas_bom_file, months)
    # print(overseas_bom_data)
    # 读取全球订单分配.xlsx
    current_order_data, next_order_data = read_global_order_xlsx(global_order_file, current_month, next_month)
    # import pdb;pdb.set_trace()
    print(current_order_data)
    print(next_order_data)
    # 构建最终结果字典
    final_result = {}
    
    # 处理全球订单分配数据
    final_result[f"{current_month}月"] = current_order_data
    final_result[f"{next_month}月"] = next_order_data
    
    # 处理海外拼BOM版本数据
    for month, df in overseas_bom_data.items():
        print(month)
        month_key = f"{month}月"
        if(month_key not in final_result.keys()):
            final_result[month_key] = {}
        print(df)
        # 假设最后一列表头以"订单量"结尾
        order_quantity_column =  find_latest_day_in_month(df)

        # for col in df.columns:
        #     if col.endswith("订单量"):
        #         order_quantity_column = col
        #         break
        
        if order_quantity_column is None:
            print(f"在 {month} 月的数据中找不到以'订单量'结尾的列")
            continue
        # 遍历每一行
        for _, row in df.iterrows():
            # 假设外协信息在某一列(例如第二列)，内部机型在某一列(例如第三列)
            id=0
            for col in df.columns:
            
                if col == '外协':
                    break
                id+=1
            id2=0
            for col in df.columns:
            
                if col == '成品料号':
                   break
                id2+=1
            supplier = row.iloc[id]  # 第二列
            print(row.iloc[0])
            model = row.iloc[0][:9]    # 第三列
            # print("______")
            # print(model)
            # print(model in power_dict)
            # print(supplier)
            
            # if(row.iloc[0][:9] == "A3HB70503"):
            #     print("++++++")
            #     print(row.iloc[7])
            print("__+_+")
            print(row.iloc[id2])
            print(model)
            #print(power_dict)
            #print(model in power_dict and not pd.isna(row.iloc[id2]))
            # 通过内部机型查找对应的电源
            if model in power_dict and not pd.isna(row.iloc[id2]):
                # if(row.iloc[0][:9] == "A3HB70503"):
                # print("+++++ii+")
                # print(row[order_quantity_column])
                # print(model)
                # print(supplier)
                # print(month)
                # if(supplier == "PIEM" and month=='6' and model in power_dict and power_dict[model]=="APW17"):
                #                 print("++++++")
                #                 print(row[order_quantity_column])
                #     print(row.iloc[7])
                #     print(row[order_quantity_column])
                power = power_dict[model]
                quantity = row[order_quantity_column]
                print(quantity)
                if pd.notna(quantity):
                    if supplier not in final_result[month_key]:
                        final_result[month_key][supplier] = {}
                    if power not in final_result[month_key][supplier]:
                        final_result[month_key][supplier][power] = 0
                    # if month_key=="6月" and supplier=="欧陆通/墨西哥" and power=="APW17":
                    #     print(row)
                    final_result[month_key][supplier][power] += int(quantity)
            print(month_key)
        # break
    print(final_result)
    return final_result

if __name__ == "__main__":
    data= process_data()


    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # 1. 加载工作簿并获取工作表
    wb = load_workbook(
        power_supply_file,
        data_only=False,
        keep_vba=True
    )
    ws = wb['电源成品供应']

    # 2. 解析多级表头，构建列索引映射
    header1 = [cell for cell in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    header2 = [cell for cell in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
    print(header1)
    print(header2)
    print("++++++++")
    column_mapping = {
        (supplier, model): col_idx 
        for col_idx, (supplier, model) in enumerate(zip(header1, header2))
    }

    # 3. 定义供应商和机型映射
    SUPPLIER_MAPPING = {
        '江元': ['江元制造', '江元智造'],
        '合权': ['合权', '东莞合权'],
        '欧陆通': ['欧陆通', '东莞欧陆通'],
        '欧陆通/墨西哥': ['墨西哥欧陆通'],
        '奥海': ['奥海', '东莞奥海'],
        'PIEM': ['马来PIE'],
        'NGSB': ['马来NSM'],
        'YNAH': ['印尼奥海'],
        'THQG': ['泰国群光'],
        'KPMI': ['马来KPMI'],
        '星创': ['宁夏星创'],
        'ONETEC': ['ONETEC'],
        '美国/Anzer':['美国ANZE'],
        '美国/Keytronic':['美国Keytronic'],
        '美国/Moonshot':['美国MNST'],
        '美国/Foxlink':['美国Foxlink'],
        '美国/A2Z':['美国AZEC'],
        '美国/Aspower':['美国ASPW'],
        '美国/CMS':['美国CMSI'],
        '美国/Meritronics':['美国MERI'],
        '海南':["比特方舟"],
        '星创（EMS出口）':['宁夏星创'],
        '欧陆通（保税）': ['欧陆通', '东莞欧陆通'],
        '合权(保税)':['合权', '东莞合权']
    }

    MODEL_MAPPING = {
        'APW17': 'W17',
        'APW17E0': 'W17E0',
        'APW11': 'APW11',
        'APW12': 'APW12',
        'APW12+': 'APW12+',
        'APW11H': 'APW11H',
        'APW11G0': 'APW11G0',
        'APW11G1': 'APW11G1',
        'APW11C0': 'APW11C0',
    }

    # 4. 示例数据（以5月为例）
    print(data)
    # 5. 遍历数据并写入Excel
    for month, suppliers in data.items():
        # 定位月份行（假设月份从第4行开始，第1列为“内部机型代码”）
        month_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_col=1, values_only=True), start=4):
            if row[0] == month:
                month_row = row_idx  # 获取行号（修正后）
                break
        if not month_row:
            print(f"警告：未找到月份 {month} 的行")
            continue

        for supplier, models in suppliers.items():
            # 处理供应商映射（可能对应多个列）
            for mapped_supplier in SUPPLIER_MAPPING.get(supplier, [supplier]):
                for model, qty in models.items():
                    # 处理机型映射
                    mapped_model = MODEL_MAPPING.get(model, model)
                    print(mapped_model)
                    # 查找列索引
                    col_idx = None
                    for key in column_mapping:
                        if key[0] == mapped_supplier and key[1] == mapped_model:
                            col_idx = column_mapping[key]
                            break
                    if col_idx is None:
                        print(f"警告：未找到列 {mapped_supplier} - {mapped_model}")
                        continue
                    # 写入数据
                    col_letter = get_column_letter(col_idx + 1)  # 列索引从0开始
                    cell = ws[f"{col_letter}{month_row}"]
                    if mapped_model == "APW11H" or mapped_model == "APW11G1":
                        cell.value = qty *2
                    else:

                        cell.value = qty

    # 6. 保存文件
    wb.save('E:\\供应链\\总表\\201125\\电源欠料表_更新1114大计划.xlsx')