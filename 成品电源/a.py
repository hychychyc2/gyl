import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import time
import re
from datetime import datetime, timedelta
def calculate_all_formulas(sheet):
    """计算工作表中的所有公式"""
    # 获取最大行和列
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # 遍历所有单元格
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # 如果单元格包含公式
            if cell.data_type == 'f':
                try:
                    # 获取公式
                    formula = cell.value
                    
                    # 计算并存储结果（如果公式有效）
                    if formula and formula.startswith('='):
                        # 这里只是标记公式已计算，实际计算由Excel完成
                        # openpyxl不直接计算公式，而是依赖Excel或其他软件
                        pass
                    
                except Exception as e:
                    print(f"计算单元格 {get_column_letter(col_idx)}{row_idx} 公式时出错: {str(e)}")
def process_excel_files(): #每天更新
    # 记录开始时间
    start_time = time.time()
    current_day = datetime.now().strftime('%Y%m%d')

    try:
        # 文件路径
        base_dir = os.path.dirname(os.path.abspath(__file__))
        file1_path = os.path.join(base_dir, "E:\\供应链\\总表\\081225\\电源欠料表20251206.xlsx")
        file2_path = os.path.join(base_dir, "E:\\供应链\\总表\\081225\\BHSC_库存收发存报表(不_081225.xls")
        file3_path = os.path.join(base_dir, "E:\\供应链\\总表\\081225\\BHSC_外协厂投料明细表(_081225.xls")
        file5_path = os.path.join(base_dir, "E:\\供应链\\总表\\081225\INTEORD_HN2512084818.xls")

        file4_path = os.path.join(base_dir, "E:\\供应链\\总表\\081225\\电源欠料表"+current_day+".xlsx")

        # 检查文件是否存在
        for file_path in [file1_path, file2_path, file3_path]:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
        
        print("开始处理文件...")
        
        # ===== 处理电子料总欠料表.xlsx =====
        print("处理电子料总欠料表.xlsx...")
        # 打开工作簿
        wb1 = openpyxl.load_workbook(file1_path)
        
        # 获取库存sheet
        inventory_sheet = wb1["库存"]
        # 删除库存sheet中的数据（保留表头）
        if inventory_sheet.max_row > 1:
            inventory_sheet.delete_rows(2, inventory_sheet.max_row)
        
        # 获取ICK智能投料明细表sheet
        ick_sheet = wb1["投料状态"]
        st_sheet = wb1["在途物料"]
        # 删除ICK智能投料明细表sheet中的数据（保留表头）
        if ick_sheet.max_row > 1:
            ick_sheet.delete_rows(2, ick_sheet.max_row)
        # tot_sheet = wb1["欠料汇总(查询替换) (2)"]
        print("已清空电子料总欠料表中的数据")
        
        # ===== 处理BHSC_库存收发存报表.xls（作为HTML文件处理） =====
        print("处理BHSC_库存收发存报表.xls...")
        # 读取HTML文件
        df2 = read_html_file(file2_path, skiprows=10)
        
        # 删除指定列（5,7,8,10,11,13,14,16,18列，注意pandas列索引从0开始）
        columns_to_delete = [1,2,4, 6, 7,8, 9, 10,11, 13, 14, 16, 18]
        df2 = df2.drop(df2.columns[columns_to_delete], axis=1)
        print(df2)
        # print(df2.values)
        # 确保数据框不为空
        if not df2.empty:
            # 将处理后的数据复制到电子料总欠料表的库存sheet
            for row_idx, row in enumerate(df2.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    inventory_sheet.cell(row=row_idx, column=col_idx).value = value
            print(f"已将{len(df2)}行数据从BHSC_库存收发存报表复制到库存sheet")
        else:
            print("BHSC_库存收发存报表处理后没有数据可复制")
        
        # ===== 处理BHSC_外协厂投料明细表_250425.xls（作为HTML文件处理） =====
        print("处理BHSC_外协厂投料明细表_250425.xls...")
        # 读取HTML文件
        df3 = read_html_file(file3_path, skiprows=1)
        # 删除指定列（3,11列，注意pandas列索引从0开始）
        columns_to_delete = [2,5,10]
        df3 = df3.drop(df3.columns[columns_to_delete], axis=1)
        # print(df3)
        df5 = read_html_file(file5_path, skiprows=1)
        # 确保数据框不为空
        if not df3.empty:
            # 提取2-3列的数据（任务号、工单类型）
            task_data = df3.iloc[:, 0:4]
            # 提取6-15列的数据（装配件到未投料数）
            tt_data = df3.iloc[:, 4:14]
            
            # 将数据写入ICK智能投料明细表sheet
            start_row = 2
            # 写入任务号、工单类型
            for row_idx, row in enumerate(task_data.values, start=start_row):
                ick_sheet.cell(row=row_idx, column=1).value = row[0]  # 任务号
                ick_sheet.cell(row=row_idx, column=2).value = row[1]  # 工单类型
                ick_sheet.cell(row=row_idx, column=3).value = row[2] 
                ick_sheet.cell(row=row_idx, column=4).value = row[3] 

                # 设置任务子库列的公式
                ick_sheet.cell(row=row_idx, column=5).value = f'=VLOOKUP(A{row_idx}&C{row_idx},Sheet1!C:D,2,0)'#VLOOKUP(A{row_idx}&C{row_idx},Sheet1!C:D,2,0)
            
            # 写入装配件到未投料数
            for row_idx, row in enumerate(tt_data.values, start=start_row):
                for col_idx, value in enumerate(row, start=6):  # 从第4列开始写入
                    ick_sheet.cell(row=row_idx, column=col_idx).value = value
                print(col_idx)
                # if(ick_sheet.cell(row=row_idx, column=1).value in ['DGQG20241201001-ZB','DGQG20250101006-ZB','DGQG20250101005-ZB', 'DGQG20250101004-ZB']):
                #     ick_sheet.cell(row=row_idx, column=col_idx).value = 0
                # ick_sheet.cell(row=row_idx, column=col_idx+1).value = -ick_sheet.cell(row=row_idx, column=col_idx).value
                # ick_sheet.cell(row=row_idx, column=col_idx+2).value = f'=VLOOKUP(H{row_idx},物料清单!B:F,5,0)'

            print(f"已将{len(df3)}行数据从BHSC_外协厂投料明细表复制到ICK智能投料明细表sheet")
        else:
            print("BHSC_外协厂投料明细表处理后没有数据可复制")


        if not df5.empty:
            # 提取2-3列的数据（任务号、工单类型）
            task_data = df5.iloc[:]
            print(task_data)
            start_row = 3
        
            # 写入装配件到未投料数
            for row_idx, row in enumerate(task_data.values, start=start_row):
                for col_idx, value in enumerate(row, start=1):  # 从第4列开始写入
                    st_sheet.cell(row=row_idx, column=col_idx).value = value
             
            print(f"已将{len(df5)}行数据从BHSC_外协厂投料明细表复制到ICK智能投料明细表sheet")
        else:
            print("BHSC_外协厂投料明细表处理后没有数据可复制")
        # calculate_all_formulas(tot_sheet)
        # 保存修改后的电子料总欠料表.xlsx
        wb1.save(file4_path)
        print(f"文件处理完成，已保存到: {file4_path}")
        
        # 计算处理时间
        end_time = time.time()
        print(f"处理完成，耗时: {end_time - start_time:.2f}秒")
        
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")


def process_excel_files_product(): #每天更新
    # 记录开始时间
    start_time = time.time()
    current_day = datetime.now().strftime('%Y%m%d')

    try:
        # 文件路径
        base_dir = os.path.dirname(os.path.abspath(__file__))
        file1_path = os.path.join(base_dir, "电子料总欠料表4-27.xlsx")
        file2_path = os.path.join(base_dir, "BHSC_库存收发存报表(不_300425.xls")
        file3_path = os.path.join(base_dir, "BHSC_外协厂投料明细表(_300425.xls")
        file4_path = os.path.join(base_dir, "电子料总欠料表"+current_day+".xlsx")

        # 检查文件是否存在
        for file_path in [file1_path, file2_path, file3_path]:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文件不存在: {file_path}")
        
        print("开始处理文件...")
        
        # ===== 处理电子料总欠料表.xlsx =====
        print("处理电子料总欠料表.xlsx...")
        # 打开工作簿
        wb1 = openpyxl.load_workbook(file1_path)
        
        # 获取库存sheet
        inventory_sheet = wb1["库存"]
        # 删除库存sheet中的数据（保留表头）
        if inventory_sheet.max_row > 1:
            inventory_sheet.delete_rows(2, inventory_sheet.max_row)
        
        # 获取ICK智能投料明细表sheet
        ick_sheet = wb1["ICK智能投料明细表"]
        # 删除ICK智能投料明细表sheet中的数据（保留表头）
        if ick_sheet.max_row > 1:
            ick_sheet.delete_rows(2, ick_sheet.max_row)
        tot_sheet = wb1["欠料汇总(查询替换) (2)"]
        print("已清空电子料总欠料表中的数据")
        
        # ===== 处理BHSC_库存收发存报表.xls（作为HTML文件处理） =====
        print("处理BHSC_库存收发存报表.xls...")
        # 读取HTML文件
        df2 = read_html_file(file2_path, skiprows=10)
        
        # 删除指定列（5,7,8,10,11,13,14,16,18列，注意pandas列索引从0开始）
        columns_to_delete = [0,1,2,4, 6, 7,8, 9, 10, 12, 13, 15, 17]
        df2 = df2.drop(df2.columns[columns_to_delete], axis=1)
        # print(df2.values)
        # 确保数据框不为空
        if not df2.empty:
            # 将处理后的数据复制到电子料总欠料表的库存sheet
            for row_idx, row in enumerate(df2.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    inventory_sheet.cell(row=row_idx, column=col_idx).value = value
            print(f"已将{len(df2)}行数据从BHSC_库存收发存报表复制到库存sheet")
        else:
            print("BHSC_库存收发存报表处理后没有数据可复制")
        
        # ===== 处理BHSC_外协厂投料明细表_250425.xls（作为HTML文件处理） =====
        print("处理BHSC_外协厂投料明细表_250425.xls...")
        # 读取HTML文件
        df3 = read_html_file(file3_path, skiprows=1)
        # 删除指定列（3,11列，注意pandas列索引从0开始）
        columns_to_delete = [0,2,4,5,10]
        df3 = df3.drop(df3.columns[columns_to_delete], axis=1)
        # print(df3)

        # 确保数据框不为空
        if not df3.empty:
            # 提取2-3列的数据（任务号、工单类型）
            task_data = df3.iloc[:, 0:2]
            # 提取6-15列的数据（装配件到未投料数）
            tt_data = df3.iloc[:, 2:12]
            
            # 将数据写入ICK智能投料明细表sheet
            start_row = 2
            # 写入任务号、工单类型
            for row_idx, row in enumerate(task_data.values, start=start_row):
                ick_sheet.cell(row=row_idx, column=1).value = row[0]  # 任务号
                ick_sheet.cell(row=row_idx, column=2).value = row[1]  # 工单类型
                # 设置任务子库列的公式
                ick_sheet.cell(row=row_idx, column=3).value = f'=VLOOKUP(B{row_idx},子库!F:G,2,0)'
            
            # 写入装配件到未投料数
            for row_idx, row in enumerate(tt_data.values, start=start_row):
                for col_idx, value in enumerate(row, start=4):  # 从第4列开始写入
                    ick_sheet.cell(row=row_idx, column=col_idx).value = value
                print(col_idx)
                if(ick_sheet.cell(row=row_idx, column=1).value in ['DGQG20241201001-ZB','DGQG20250101006-ZB','DGQG20250101005-ZB', 'DGQG20250101004-ZB']):
                    ick_sheet.cell(row=row_idx, column=col_idx).value = 0
                ick_sheet.cell(row=row_idx, column=col_idx+1).value = -ick_sheet.cell(row=row_idx, column=col_idx).value
                ick_sheet.cell(row=row_idx, column=col_idx+2).value = f'=VLOOKUP(H{row_idx},物料清单!B:F,5,0)'

            print(f"已将{len(df3)}行数据从BHSC_外协厂投料明细表复制到ICK智能投料明细表sheet")
        else:
            print("BHSC_外协厂投料明细表处理后没有数据可复制")
        # calculate_all_formulas(tot_sheet)
        # 保存修改后的电子料总欠料表.xlsx
        wb1.save(file4_path)
        print(f"文件处理完成，已保存到: {file4_path}")
        
        # 计算处理时间
        end_time = time.time()
        print(f"处理完成，耗时: {end_time - start_time:.2f}秒")
        
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")


def updatebig_plan(source_file, target_file): # 每月更新
    # 获取当前年月
    now = datetime.now()
    year = now.year
    month = now.month
    
    # 创建新文件名称
    new_file = f'电子料总欠料表{year}年{month}月.xlsx'
    
    # 读取源文件中的量产欠料sheet，跳过前四行，第五行作为表头
    source_df = pd.read_excel(source_file, sheet_name='量产欠料 ', header=14)
    
    # 读取目标文件中的欠料汇总(查询替换) (2) sheet，跳过前四行，第五行作为表头
    target_df = pd.read_excel(target_file, sheet_name='欠料汇总(查询替换) (2)', header=5)
    
    # 使用正则表达式查找匹配的列
    current_month_column = None
    pattern = f'{month+1}月欠料合并'
    
    for col in source_df.columns:
        print(col)
        if isinstance(col, str) and col == pattern:
            current_month_column = col
            print(f"找到匹配的列: {col}")
            break
    
    if current_month_column is None:
        print(f"警告：在源表中未找到匹配'%{month}月*合并欠料'模式的列")
    
    # 计算需要添加的行数
    rows_to_add = max(0, len(source_df) - len(target_df))
    
    # 复制目标文件到新文件
    import shutil
    shutil.copy2(target_file, new_file)
    
    # 打开新文件
    wb = openpyxl.load_workbook(new_file)
    ws = wb['欠料汇总(查询替换) (2)']
    
    # 定义需要更新的列映射（目标列索引 -> 源列索引）
    column_mapping = {
        0: 1,  # 目标第2列(index=1) <- 源第2列(index=1)
        1: 2,  # 目标第3列(index=2) <- 源第3列(index=2)
        2: 3,  # 目标第4列(index=3) <- 源第4列(index=3)
        3: 5,  # 目标第5列(index=4) <- 源第6列(index=5)
        7: 7,  # 目标第9列(index=8) <- 源第8列(index=7)
    }
    
    # 如果找到匹配的列，则添加第18列的映射
    if current_month_column is not None:
        source_col_idx = source_df.columns.get_loc(current_month_column)
        column_mapping[19] = source_col_idx  # 目标第18列(index=17)
    
    # 更新已有行的数据
    for row_idx in range(len(source_df)):
        excel_row = row_idx + 6  # Excel行号从1开始，第1-5行是标题和保留行
        
        for target_col_idx, source_col_idx in column_mapping.items():
            # Excel列号从1开始
            excel_col = target_col_idx + 1
            
            # 获取源数据的值
            value = source_df.iloc[row_idx, source_col_idx]
            # 更新单元格的值
            print(ws.cell(row=excel_row, column=excel_col).value)
            print(value)
            if(source_col_idx == 1):
                ws.cell(row=excel_row, column=5).value = value

            ws.cell(row=excel_row, column=excel_col).value = value
    
    # 扩展表格并填充公式
    if rows_to_add > 0:
        # 获取已有数据的最后一行（不包括表头）
        last_data_row = len(target_df) + 5  # Excel行号
        
        # 为每一行添加公式
        for new_row in range(last_data_row + 1, last_data_row + rows_to_add + 1):
            # 复制上一行的公式并调整引用
            for col in range(1, ws.max_column + 1):
                # 跳过我们更新数据的列
                if col - 1 in column_mapping:
                    continue
                
                # 获取上一行对应单元格
                prev_cell = ws.cell(row=new_row - 1, column=col)
                
                # 如果上一行单元格有公式，复制并调整
                if prev_cell.data_type == 'f':
                    formula = prev_cell.value
                    
                    # 简单处理相对引用（这是一个简化的实现，可能无法处理所有复杂公式）
                    # 基本思路：查找公式中的单元格引用并调整行号
                    new_formula = adjust_formula(formula, new_row - 1, new_row)
                    
                    # 设置新单元格的公式
                    ws.cell(row=new_row, column=col).value = new_formula
    
    # 保存新文件
    wb.save(new_file)
    
    print(f"数据已成功更新到新文件 {new_file} 中，其他列的公式已保留并扩展到新行。")


def update_schedue(source_file, target_file): # 每周更新
    # 获取当前年月
    now = datetime.now()
    year = now.year
    month = now.month
    
    # 创建新文件名称
    new_file = '交期_'+target_file
    
    # 读取源文件中的量产欠料sheet，跳过前四行，第五行作为表头
    result=process_material_data(source_file)
    
    # 读取目标文件中的欠料汇总(查询替换) (2) sheet，跳过前四行，第五行作为表头
    target_df = pd.read_excel(target_file, sheet_name='欠料汇总(查询替换) (2)', header=5)
    
    # 使用正则表达式查找匹配的列
  
 
    # 计算需要添加的行数
    
    # 复制目标文件到新文件
    import shutil
    shutil.copy2(target_file, new_file)
    
    # 打开新文件
    wb = openpyxl.load_workbook(new_file)
    ws = wb['欠料汇总(查询替换) (2)']
    
    # 定义需要更新的列映射（目标列索引 -> 源列索引）
    column_mapping = {
        0: 1,  # 目标第2列(index=1) <- 源第2列(index=1)
        1: 2,  # 目标第3列(index=2) <- 源第3列(index=2)
        2: 3,  # 目标第4列(index=3) <- 源第4列(index=3)
        3: 5,  # 目标第5列(index=4) <- 源第6列(index=5)
        7: 7,  # 目标第9列(index=8) <- 源第8列(index=7)
    }
    
 
    
    # 更新已有行的数据
    for row_idx in range(6,len(target_df)):
        
        
        if ws.cell(row=row_idx, column=2).value in result.keys():
    
            value_to_save=str(result[ws.cell(row=row_idx, column=2).value])
            ws.cell(row=row_idx, column=22).value=f"'{value_to_save}"
            ws.cell(row=row_idx, column=22).number_format = '@'
            print(ws.cell(row=row_idx, column=22).value)
    
    # 保存新文件
    wb.save(new_file)
    
    print(f"数据已成功更新到新文件 {new_file} 中，其他列的公式已保留并扩展到新行。")

# def process_material_data(file_path):
#     print(file_path)
#     df = pd.read_excel(file_path, sheet_name=-1,header=1)
#     result_dict = {}
#     print(df)
#     for index, row in df.iterrows():
#         print(row)
#         material_code = row['物料编码']
#         if material_code not in result_dict:
#             result_dict[material_code] = []
#         # 筛选出符合 '%m月%d日' 格式的列名
#         for col in df.columns :print(col)
#         date_columns = [col for col in df.columns if '/' not in col and '月' in col and '日' in col]
#         for date_col in date_columns:
#             value = row[date_col]
#             if pd.notnull(value):
#                 result_dict[material_code].append({date_col: value})
#     return result_dict
base_date = datetime(1899, 12, 30)
def process_material_data(file_path):
    # 读取 Excel 文件并获取所有表名
    wb = openpyxl.load_workbook(file_path, read_only=False, keep_vba=False)
    # 获取所有未隐藏的表名
    visible_sheet_names = [sheet.title for sheet in wb.worksheets if sheet.sheet_state == 'visible']
    # 获取最后一个未隐藏的表名
    last_sheet_name = visible_sheet_names[-1] if visible_sheet_names else None
    if not last_sheet_name:
        raise ValueError("Excel 文件中没有可见的工作表")
    # 读取指定工作表中的数据
    df = pd.read_excel(file_path, sheet_name=last_sheet_name, header=2)
    print(df)
    result_dict = {}
   
    date_columns = [col for col in df.columns if type(col) in [int, float]]
    print(date_columns)
    for index, row in df.iterrows():
        material_code = row['物料编码']
        if material_code not in result_dict:
            result_dict[material_code] = []
        # 筛选出符合 '%m月%d日' 格式的列名
        # print(row)
        for date_col in date_columns:
            value = row[date_col]
            if pd.notnull(value):
                result_dict[material_code].append({(base_date+timedelta(days=date_col)).strftime('%Y-%m-%d'): value})
    print(result_dict)
    return result_dict
def adjust_formula(formula, old_row, new_row):
    """调整公式中的相对引用"""
    # 这是一个简化的实现，处理A1样式的相对引用
    # 例如：=SUM(A1:B5) 会被调整为 =SUM(A2:B6)
    import re
    
    # 匹配Excel单元格引用（如A1, B2, AA10）
    cell_ref_pattern = r'([A-Z]+)(\d+)'
    
    def replace_ref(match):
        col = match.group(1)
        row = int(match.group(2))
        
        # 如果是绝对引用（如$A$1），不做修改
        if '$' in match.group(0):
            return match.group(0)
        
        # 计算新的行号
        new_row_num = row + (new_row - old_row)
        return f"{col}{new_row_num}"
    
    # 替换公式中的所有单元格引用
    return re.sub(cell_ref_pattern, replace_ref, formula)
def read_html_file(file_path, skiprows=0):
    """读取HTML文件并返回DataFrame"""
    try:
        # 读取HTML文件内容
        with open(file_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # 使用pandas的read_html解析表格
        dfs = pd.read_html(html_content, skiprows=skiprows)
        
        # 通常第一个表格是我们需要的，但也可能需要根据实际情况调整
        if not dfs:
            raise ValueError(f"在{file_path}中未找到表格")
        
        # 返回第一个表格
        return dfs[0]
    
    except Exception as e:
        print(f"读取HTML文件时发生错误: {str(e)}")
        raise
def merge_excel_data(source_file, target_file, source_sheet='Sheet1', target_sheet='电源欠料', 
                     material_code_col='物料编码', source_owe_col='合并欠料', target_owe_col='总欠料'):
    """
    合并两个Excel文件中特定工作表的数据
    
    参数:
    source_file: 源Excel文件路径
    target_file: 目标Excel文件路径
    source_sheet: 源工作表名称
    target_sheet: 目标工作表名称
    material_code_col: 物料编码列名
    source_owe_col: 源文件中欠料列名
    target_owe_col: 目标文件中欠料列名
    """
    # 检查文件是否存在
    if not os.path.exists(source_file):
        raise FileNotFoundError(f"源文件不存在: {source_file}")
    if not os.path.exists(target_file):
        raise FileNotFoundError(f"目标文件不存在: {target_file}")
    
    try:
        # 读取源Excel文件
        print(f"正在读取源文件: {source_file}")
        source_df = pd.read_excel(source_file, sheet_name=source_sheet,header=4)
        
        # 检查源DataFrame是否包含必要的列
        print(source_df.columns)
        if material_code_col not in source_df.columns:
            raise ValueError(f"源工作表 '{source_sheet}' 中不存在列 '{material_code_col}'")
        if source_owe_col not in source_df.columns:
            raise ValueError(f"源工作表 '{source_sheet}' 中不存在列 '{source_owe_col}'")
        
        # 创建物料编码到合并欠料的映射
        print(source_df)
        material_to_owe = dict(zip(source_df[material_code_col], source_df[source_owe_col]))
        print(material_to_owe)
        print(f"已从源工作表中提取 {len(material_to_owe)} 条物料数据")
        
        # 使用openpyxl加载目标Excel文件
        print(f"正在加载目标文件: {target_file}")
        wb = openpyxl.load_workbook(target_file)
        
        # 检查目标工作表是否存在
        if target_sheet not in wb.sheetnames:
            raise ValueError(f"目标文件中不存在工作表 '{target_sheet}'")
        
        ws = wb[target_sheet]
        
        # 找到物料编码列和总欠料列的索引
        header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        material_col_idx = None
        target_owe_col_idx = None
        
        for idx, cell_value in enumerate(header_row, 1):
            print(cell_value)
            if cell_value == material_code_col:
                material_col_idx = idx
            if cell_value == target_owe_col:
                target_owe_col_idx = idx
        
        if material_col_idx is None:
            raise ValueError(f"目标工作表 '{target_sheet}' 中不存在列 '{material_code_col}'")
        if target_owe_col_idx is None:
            raise ValueError(f"目标工作表 '{target_sheet}' 中不存在列 '{target_owe_col}'")
        
        # 确定要插入的列位置（总欠料列的下一列）
        insert_col_idx = target_owe_col_idx + 1
        
        # 插入新列并设置标题
        ws.insert_cols(insert_col_idx)
        ws.cell(row=2, column=insert_col_idx, value=f"{source_owe_col}（合并）")
        
        # 填充数据
        matched_count = 0
        total_rows = ws.max_row
        print("+++++++++++")
        for row_idx in range(3, total_rows + 1):
            material_code = ws.cell(row=row_idx, column=material_col_idx).value
            if material_code in material_to_owe:
                print(material_to_owe[material_code])
                ws.cell(row=row_idx, column=insert_col_idx).value=material_to_owe[material_code]
                matched_count += 1
            print(ws.cell(row=row_idx, column=insert_col_idx).value)
            # 每处理1000行显示进度
            if row_idx % 1000 == 0:
                print(f"已处理 {row_idx}/{total_rows} 行")
        
        print(f"共匹配到 {matched_count} 条物料数据")
        
        # 保存修改后的文件
        file_dir, file_name = os.path.split(target_file)
        new_file_name = os.path.splitext(file_name)[0] + '_合并版' + os.path.splitext(file_name)[1]
        new_file_path = os.path.join(file_dir, new_file_name)
        
        print(f"正在保存结果到: {new_file_path}")
        wb.save(new_file_path)
        print("数据合并完成!")
        
        return new_file_path
        
    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        raise

if __name__ == "__main__":
    process_excel_files()
    # 
    # updatebig_plan("3月需求数据--2025.04.26 vs 0413版大计划.xlsx","电子料总欠料表20250427.xlsx")    

    # from openpyxl import load_workbook

    # wb = load_workbook('交期_电子料总欠料表20250427.xlsx')
    # ws = wb['欠料汇总(查询替换) (2)']
    # print(ws)
    # for row_idx in range(6,10):
    #     print(ws.cell(row=row_idx, column=20).value)  # 应输出 =[{'2025-04-28': 600.0}]
    #     print(ws.cell(row=row_idx, column=20).number_format)  # 应输出 '@'
    # update_schedue("0425交期回复.xlsx","电子料总欠料表20250430.xlsx")    

    # merge_excel_data("交期_电子料总欠料表20250427.xlsx","群光-世纪云芯电源欠料明细表0428.xlsx","欠料汇总(查询替换) (2)","电源欠料表" ,'物料编码', '合并欠料','总欠料')