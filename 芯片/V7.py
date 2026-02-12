import imaplib
import email
from email.header import decode_header
import os
import json
import openpyxl
import xlrd
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Tuple, Optional, Any, Union
import base64
import re
import shutil
import json
import glob

def init_cache_dir(temp_dir: str) -> str:
    """初始化缓存目录"""
    cache_dir = os.path.join(temp_dir, "attach_cache")
    os.makedirs(cache_dir, exist_ok=True)
    return cache_dir

def get_latest_cache_file(cache_dir: str, match_key: str) -> Optional[str]:
    """获取缓存目录中匹配关键词的最新文件"""
    # 匹配所有包含match_key的文件
    cache_files = glob.glob(os.path.join(cache_dir, f"*{match_key}*"))
    if not cache_files:
        return None
    # 按修改时间排序，取最新的
    cache_files.sort(key=os.path.getmtime, reverse=True)
    latest_file = cache_files[0]
    print(f"\n📦 缓存中找到最新匹配文件：{latest_file}")
    return latest_file

def update_cache_file(cache_dir: str, new_file_path: str, match_key: str) -> None:
    """更新缓存文件（删除旧缓存，复制新文件到缓存）"""
    # 删除旧的匹配缓存文件
    old_cache_files = glob.glob(os.path.join(cache_dir, f"*{match_key}*"))
    for old_file in old_cache_files:
        try:
            if os.path.isfile(old_file):
                os.remove(old_file)
            elif os.path.isdir(old_file):
                shutil.rmtree(old_file)
            print(f"🗑️ 删除旧缓存文件：{old_file}")
        except Exception as e:
            print(f"⚠️ 删除旧缓存失败：{e}")
    
    # 复制新文件到缓存
    file_name = os.path.basename(new_file_path)
    cache_file_path = os.path.join(cache_dir, file_name)
    shutil.copy2(new_file_path, cache_file_path)
    
    # 记录缓存信息（可选：保存到json）
    cache_info = {
        "latest_file": cache_file_path,
        "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "match_key": match_key
    }
    cache_info_path = os.path.join(cache_dir, "cache_info.json")
    with open(cache_info_path, "w", encoding="utf-8") as f:
        json.dump(cache_info, f, ensure_ascii=False, indent=2)
    
    print(f"✅ 缓存更新成功：{cache_file_path}")
try:
    from tqdm import tqdm  # 进度条
except ImportError:
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "tqdm"])
    from tqdm import tqdm

# ===================== RAR解压相关依赖与函数（核心修复完整版） =====================
# 自动安装rarfile库
try:
    import rarfile
except ImportError:
    print("⚠️  未检测到rarfile库，正在自动安装...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "rarfile"])
    import rarfile

# 配置UnRAR工具路径（Windows需手动指定，Linux/macOS安装后可注释）
# 替换为你的WinRAR安装路径，示例：
rarfile.UNRAR_TOOL = r"C:\Program Files\WinRAR\UnRAR.exe"
# 32位系统默认路径：r"C:\Program Files (x86)\WinRAR\UnRAR.exe"
# Linux/macOS无需配置，安装unrar后自动识别

def clean_invalid_filename_chars(filename: str) -> str:
    """
    彻底清理Windows文件名中的所有非法字符（强化版，解决空格/控制字符问题）
    1. 移除Windows可见非法字符
    2. 移除所有不可见控制字符（\r\n\t等）
    3. 移除首尾空格，替换连续空格为单个下划线
    4. 兜底处理空文件名
    :param filename: 原始文件名
    :return: 合法、无冗余字符的文件名
    """
    if not filename:
        return "unnamed_file"
    
    # 1. 定义Windows禁止的可见非法字符
    invalid_chars = r'\/:*?"<>|'
    # 2. 替换可见非法字符为下划线
    for char in invalid_chars:
        filename = filename.replace(char, "_")
    
    # 3. 移除不可见控制字符（ASCII 0-31 和 127）
    control_chars = ''.join([chr(c) for c in range(0, 32)] + [chr(127)])
    filename = ''.join([c for c in filename if c not in control_chars])
    
    # 4. 移除首尾空格，合并连续空格为单个，再替换为下划线（彻底规避空格问题）
    filename = filename.strip()  # 移除首尾空格
    filename = ' '.join(filename.split())  # 合并连续空格为单个空格
    filename = filename.replace(" ", "_")  # 空格替换为下划线
    
    # 5. 兜底：如果文件名为空，赋予默认名称
    if not filename:
        filename = "unnamed_file"
    
    return filename

def extract_rar_file(rar_path: str, extract_dir: str) -> List[str]:
    """
    解压RAR文件到指定目录（核心修复：目录名/内部文件名清理，兼容所有rarfile版本）
    :param rar_path: RAR文件路径
    :param extract_dir: 解压目标目录
    :return: 解压后的所有文件路径列表
    """
    extracted_files = []
    try:
        # 创建解压目录（独立子目录，目录名严格清理，避免无效字符）
        rar_basename = os.path.basename(rar_path)
        rar_filename = os.path.splitext(rar_basename)[0]
        cleaned_rar_filename = clean_invalid_filename_chars(rar_filename)  # 清理解压目录名
        unique_extract_dir = os.path.join(extract_dir, f"extracted_{cleaned_rar_filename}")
        os.makedirs(unique_extract_dir, exist_ok=True)

        # 解压RAR文件
        with rarfile.RarFile(rar_path, 'r') as rf:
            rf.extractall(unique_extract_dir)
            # 获取所有解压后的文件路径，并清理内部文件名
            for file in rf.namelist():
                cleaned_file_name = file#clean_invalid_filename_chars(file)  # 清理压缩包内文件名
                file_path = os.path.join(unique_extract_dir, cleaned_file_name)
                print(file_path)
                # 仅保留文件路径，过滤目录
                if os.path.isfile(file_path):
                    extracted_files.append(file_path)
        
        print(f"✅ RAR解压成功：{rar_path} → {unique_extract_dir}")
        print(f"   解压出{len(extracted_files)}个文件")
        return extracted_files
    except Exception as e:
        # 兼容不同版本rarfile，通过错误信息分类提示
        error_msg = str(e).lower()
        if "cannot open" in error_msg or "not a rar file" in error_msg:
            raise Exception(f"❌ 无法打开RAR文件：{rar_path}（文件损坏或非RAR格式）")
        elif "password" in error_msg:
            raise Exception(f"❌ RAR文件{rar_path}受密码保护，无法解压")
        elif "cannot find working tool" in error_msg:
            raise Exception(f"❌ 未找到UnRAR工具，请安装并配置路径（参考代码注释）")
        else:
            raise Exception(f"❌ RAR解压失败：{str(e)}")

def find_xlsx_in_files(file_list: List[str]) -> Optional[str]:
    """
    从文件列表中查找xlsx文件（增加路径存在性校验，优先返回第一个有效xlsx）
    :param file_list: 文件路径列表
    :return: 第一个xlsx文件路径，无则返回None
    """
    for file_path in file_list:
        # 先校验路径是否存在，跳过无效路径
        if not os.path.exists(file_path):
            print(f"⚠️  跳过无效路径：{file_path}")
            continue
        
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext == ".xlsx":
            # 过滤临时文件和隐藏文件
            file_name = os.path.basename(file_path)
            if not file_name.startswith("~$") and not file_name.startswith("."):
                print(f"✅ 找到解压后的XLSX文件：{file_path}")
                return file_path
    print(f"❌ 解压后的文件中未找到有效.xlsx文件")
    return None

# ===================== 全局常量/兼容配置 =====================
XLRD_OFFSET = 1  # xlrd 0-based → 1-based

# 文件魔数（Magic Number）
FILE_MAGIC_NUMBERS = {
    b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1": "xls",
    b"\xd0\xcf\x11\xe0": "xls",
    b"PK\x03\x04": "xlsx",
    b"PK\x05\x06": "xlsx",
    b"PK\x07\x08": "xlsx",
    b"Rar!\x1a\x07\x00": "rar",
    b"Rar!\x1a\x07\x01\x00": "rar"
}

# 全局输出文件变量（唯一文件模式）
out_workbook: Optional[openpyxl.Workbook] = None
out_file_path: str = ""
use_global_out_file: bool = False  # 是否启用全局唯一文件模式

# ===================== 公式引用自动调整函数 =====================
def adjust_formula_row_reference(formula: str, template_row: int, target_row: int) -> str:
    if not formula or not str(formula).startswith("="):
        return formula

    row_offset = target_row - template_row
    if row_offset == 0:
        return formula

    # 正则匹配单元格引用（支持A1、$A1、A$1、$A$1格式）
    cell_pattern = re.compile(r'([$]?)([A-Za-z]+)([$]?)(\d+)')

    def replace_cell_ref(match):
        col_abs, col_letter, row_abs, row_num = match.groups()
        row_num = int(row_num)

        # 只有行不是绝对引用时，才调整行号
        if row_abs == "":
            new_row_num = row_num + row_offset
            return f"{col_abs}{col_letter}{row_abs}{new_row_num}"
        else:
            return f"{col_abs}{col_letter}{row_abs}{row_num}"

    # 替换所有符合条件的单元格引用
    adjusted_formula = cell_pattern.sub(replace_cell_ref, formula)
    return adjusted_formula

# ===================== 核心：文件格式检测与重命名 =====================
def detect_file_real_format(file_path: str) -> str:
    """检测真实格式并输出详细日志"""
    with open(file_path, "rb") as f:
        header = f.read(8)
    header_hex = header.hex()
    print(f"🔍 文件魔数检测：{file_path} → 前8字节魔数（16进制）：{header_hex}")
    
    # 匹配魔数
    for magic, fmt in FILE_MAGIC_NUMBERS.items():
        if header.startswith(magic):
            print(f"✅ 魔数匹配成功 → 真实格式：.{fmt}")
            return fmt
    
    # 兜底
    ext = os.path.splitext(file_path)[1].lower()
    print(f"⚠️  魔数未匹配！扩展名：{ext} → 兜底按扩展名判断")
    if ext in [".xls", ".xlsx", ".rar"]:
        return ext[1:]
    return "xlsx"

def force_rename_by_magic(file_path: str) -> str:
    """
    根据魔数强制重命名文件（修正扩展名）
    返回：重命名后的文件路径
    """
    real_fmt = detect_file_real_format(file_path)
    file_dir, file_name = os.path.split(file_path)
    name_without_ext, ext = os.path.splitext(file_name)
    
    # 如果扩展名和真实格式一致，直接返回原路径
    if ext.lower() == f".{real_fmt}":
        print(f"✅ 扩展名已匹配真实格式：{file_path}")
        return file_path
    
    # 强制重命名
    new_file_name = f"{name_without_ext}.{real_fmt}"
    new_file_path = os.path.join(file_dir, new_file_name)
    
    # 避免重名覆盖
    counter = 1
    while os.path.exists(new_file_path):
        new_file_name = f"{name_without_ext}_{counter}.{real_fmt}"
        new_file_path = os.path.join(file_dir, new_file_name)
        counter += 1
    
    # 复制并重命名（保留原文件，避免破坏）
    shutil.copy2(file_path, new_file_path)
    print(f"📝 强制重命名文件（修正扩展名）：")
    print(f"   原文件：{file_path}")
    print(f"   新文件：{new_file_path}")
    
    return new_file_path

def read_excel_final(file_path: str, sheet_name: str) -> Tuple[Any, str]:
    """
    终极读取函数：先重命名再读取
    """
    # 第一步：强制重命名修正扩展名
    renamed_path = force_rename_by_magic(file_path)
    real_fmt = os.path.splitext(renamed_path)[1].lower()[1:]
    
    # 非Excel格式直接抛出异常
    if real_fmt not in ["xls", "xlsx"]:
        raise Exception(f"❌ 文件{renamed_path}不是Excel格式（真实格式：.{real_fmt}）")
    
    # 第二步：按修正后的格式读取
    try:
        if real_fmt == "xlsx":
            wb = openpyxl.load_workbook(renamed_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"无Sheet：{sheet_name}，可用Sheet：{wb.sheetnames}")
            ws = wb[sheet_name]
            print(f"✅ 成功读取修正后的.xlsx文件：{renamed_path}")
            return ws, "xlsx"
        else:
            wb = xlrd.open_workbook(renamed_path, encoding_override="gbk")
            sheet_names = wb.sheet_names()
            if isinstance(sheet_name, str):
                if sheet_name not in sheet_names:
                    raise ValueError(f"无Sheet：{sheet_name}，可用Sheet：{sheet_names}")
                sheet_idx = sheet_names.index(sheet_name)
            else:
                sheet_idx = sheet_name
                if sheet_idx >= wb.nsheets:
                    raise ValueError(f"Sheet索引超出范围：{sheet_idx}，共{wb.nsheets}个Sheet")
            ws = wb.sheet_by_index(sheet_idx)
            print(f"✅ 成功读取修正后的.xls文件：{renamed_path}")
            return ws, "xls"
    except Exception as e:
        raise Exception(f"读取修正后的文件失败：{str(e)}")

# ===================== 核心：列解析函数（保留原始顺序，仅去重） =====================
def parse_col_range(range_str: str) -> List[str]:
    """
    解析单个列字符串/范围字符串为列字母列表
    :param range_str: 单列字符串（如"D"）或范围字符串（如"A-J"）
    :return: 列字母列表
    """
    # 去除空格，统一转大写
    range_str = range_str.strip().upper()
    if "-" not in range_str:
        # 无分隔符，视为单列
        return [range_str]
    
    # 分割起始和结束列（仅分割第一次出现的"-"）
    parts = range_str.split("-", 1)
    start_col_str = parts[0].strip()
    end_col_str = parts[1].strip()
    
    if not start_col_str or not end_col_str:
        raise ValueError(f"无效的列范围格式：{range_str}，正确格式为「单列」或「起始列-结束列」（如D、A-J）")
    
    # 转换为列索引（1-based）
    try:
        start_idx = column_index_from_string(start_col_str)
        end_idx = column_index_from_string(end_col_str)
    except Exception as e:
        raise ValueError(f"列范围解析失败：{range_str} → 无效列标识：{str(e)}")
    
    # 处理倒序（自动转为正序，保留整体列表顺序）
    if start_idx > end_idx:
        start_idx, end_idx = end_idx, start_idx
    
    # 生成连续列字母列表
    col_list = []
    for col_idx in range(start_idx, end_idx + 1):
        col_list.append(get_column_letter(col_idx))
    
    return col_list

def parse_col_list(col_input_list: list) -> List[str]:
    """
    批量解析列列表（保留原始顺序，仅去重）
    :param col_input_list: 输入列列表，如["A-J", "L", "P-R"]
    :return: 扁平化后的完整列字母列表
    """
    if not isinstance(col_input_list, list):
        raise ValueError(f"输入必须是列表，当前类型：{type(col_input_list)}")
    
    full_col_list = []
    for item in col_input_list:
        if not isinstance(item, str):
            raise ValueError(f"列表内每个项必须是字符串，当前无效项：{item}（类型：{type(item)}）")
        # 解析单个项
        parsed_cols = parse_col_range(item)
        full_col_list.extend(parsed_cols)
    
    # 按原始顺序去重
    seen = set()
    full_col_list_no_dup = []
    for col in full_col_list:
        if col not in seen:
            seen.add(col)
            full_col_list_no_dup.append(col)
    
    # 打印解析日志
    print(f"📌 列配置原始输入：{col_input_list}")
    print(f"📌 解析后（保留原始顺序+去重）：{full_col_list_no_dup}")
    
    return full_col_list_no_dup

# ===================== IMAP邮件相关工具函数 =====================
def imap_utf7_decode(s: str) -> str:
    if not s or "&" not in s:
        return s
    pattern = re.compile(r'&([^-]+)-')
    def decode_match(match):
        encoded_part = match.group(1).replace(",", "/")
        if not encoded_part:
            return "&"
        try:
            decoded = base64.b64decode(encoded_part + "==", altchars=b"+/").decode("utf-16be")
            return decoded
        except:
            return match.group(0)
    return pattern.sub(decode_match, s)

def imap_utf7_encode(s: str) -> str:
    if not s or all(ord(c) < 128 for c in s):
        return s
    result = []
    buffer = []
    for c in s:
        if ord(c) < 128:
            if buffer:
                b64 = base64.b64encode(''.join(buffer).encode("utf-16be")).decode("ascii").rstrip("=").replace("/", ",")
                result.append(f"&{b64}-")
                buffer = []
            result.append(c)
        else:
            buffer.append(c)
    if buffer:
        b64 = base64.b64encode(''.join(buffer).encode("utf-16be")).decode("ascii").rstrip("=").replace("/", ",")
        result.append(f"&{b64}-")
    return ''.join(result)

def get_imap_date_str(date_obj: datetime) -> str:
    month_abbr = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{date_obj.day:02d}-{month_abbr[date_obj.month-1]}-{date_obj.year}"

def get_today_imap_criteria() -> str:
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    today_str = get_imap_date_str(datetime(today.year, today.month, today.day))
    tomorrow_str = get_imap_date_str(datetime(tomorrow.year, tomorrow.month, tomorrow.day))
    return f'SINCE "{today_str}" BEFORE "{tomorrow_str}"'

def decode_email_header(header: Any) -> str:
    if not header:
        return ""
    decoded_parts = []
    for part, encoding in decode_header(header):
        if isinstance(part, bytes):
            try:
                decoded_part = part.decode(encoding or "gbk")
            except (UnicodeDecodeError, LookupError):
                try:
                    decoded_part = part.decode("utf-8")
                except UnicodeDecodeError:
                    decoded_part = part.decode("latin-1")
        else:
            decoded_part = str(part)
        decoded_parts.append(decoded_part)
    return "".join(decoded_parts).strip()

def encode_imap_folder(folder_name: str) -> bytes:
    if not folder_name:
        return b"INBOX"
    imap_encoded = imap_utf7_encode(folder_name)
    for encoding in ["utf-8", "gbk"]:
        try:
            return imap_encoded.encode(encoding)
        except:
            continue
    return imap_encoded.encode("latin-1")

def parse_imap_folders(mail: imaplib.IMAP4_SSL) -> List[Tuple[str, str]]:
    try:
        status, folders_raw = mail.list()
        if status != "OK":
            print("⚠️  获取文件夹列表失败")
            return [("INBOX", "INBOX")]
        
        folder_list = []
        for folder in folders_raw:
            if not folder:
                continue
            folder_str = folder.decode("utf-8", errors="ignore")
            parts = folder_str.split('"')
            if len(parts) < 3:
                continue
            encoded_name = parts[-2].strip()
            if not encoded_name or encoded_name in [f[0] for f in folder_list]:
                continue
            decoded_name = imap_utf7_decode(encoded_name)
            folder_list.append((encoded_name, decoded_name))
        
        folder_list = list(dict.fromkeys(folder_list))
        folder_list.sort(key=lambda x: (x[0] != "INBOX", x[0].lower()))
        print(f"\n📂 解析到邮箱所有文件夹（共{len(folder_list)}个）：")
        for i, (encoded, decoded) in enumerate(folder_list):
            print(f"   [{i+1}] 编码名：{encoded:<20} 中文名：{decoded}")
        return folder_list
    except Exception as e:
        print(f"⚠️  解析文件夹失败：{e}，仅使用收件箱")
        return [("INBOX", "INBOX")]

def get_root_folder_children(folder_list: List[Tuple[str, str]], root_folder: str) -> List[str]:
    target_encoded_root = None
    root_lower = root_folder.lower()
    
    for encoded, decoded in folder_list:
        if decoded.lower() == root_lower or encoded.lower() == root_lower:
            target_encoded_root = encoded
            root_decoded = decoded
            break
    
    if not target_encoded_root:
        print(f"⚠️  未找到根文件夹「{root_folder}」，默认使用INBOX")
        target_encoded_root = "INBOX"
        root_decoded = "INBOX"
    
    target_encoded_folders = []
    for encoded, decoded in folder_list:
        if encoded == target_encoded_root or encoded.startswith(f"{target_encoded_root}/"):
            target_encoded_folders.append((encoded, decoded))
    
    if not target_encoded_folders:
        print(f"⚠️  未找到根文件夹「{root_decoded}」及其子文件夹，默认使用INBOX")
        target_encoded_folders = [("INBOX", "INBOX")]
    else:
        target_encoded_folders.sort(key=lambda x: (x[0] != target_encoded_root, x[0].count("/"), x[0].lower()))
    
    print(f"\n🎯 待递归的文件夹（根：{root_decoded} / {target_encoded_root}）：")
    for i, (encoded, decoded) in enumerate(target_encoded_folders):
        level = encoded.count("/") - target_encoded_root.count("/")
        prefix = "  " * level + "└─ " if level > 0 else "┌─ "
        print(f"   {prefix}中文名：{decoded:<20} 编码名：{encoded}")
    
    return [encoded for encoded, decoded in target_encoded_folders]

def filter_folders(folder_list: List[str], exclude_folders: List[str], all_folders: List[Tuple[str, str]]) -> List[str]:
    if not exclude_folders:
        return folder_list
    
    filtered = []
    exclude_lower = [f.lower() for f in exclude_folders]
    
    for encoded in folder_list:
        decoded = next((d for e, d in all_folders if e == encoded), encoded)
        if any(ex in decoded.lower() or ex in encoded.lower() for ex in exclude_lower):
            print(f"🚫 排除子文件夹：{decoded} / {encoded}（匹配排除关键词）")
            continue
        filtered.append(encoded)
    
    print(f"\n🔍 过滤后最终递归的文件夹（共{len(filtered)}个）：")
    for encoded in filtered:
        decoded = next((d for e, d in all_folders if e == encoded), encoded)
        print(f"   中文名：{decoded:<20} 编码名：{encoded}")
    return filtered

def clean_text(text: Any) -> str:
    if isinstance(text, bytes):
        try:
            return text.decode("gbk")
        except:
            return text.decode("utf-8", errors="ignore")
    if isinstance(text, (int, float)):
        return str(text)
    return str(text).replace("\r", "").replace("\n", "").replace("\t", "").strip() if text else ""

def get_cell_value(ws, row: int, col: int, file_format: str) -> Any:
    if file_format == "xlsx":
        return ws.cell(row=row, column=col).value
    else:
        try:
            return ws.cell_value(row - XLRD_OFFSET, col - XLRD_OFFSET)
        except IndexError:
            return None

def get_excel_max_row(ws, file_format: str) -> int:
    if file_format == "xlsx":
        return ws.max_row
    else:
        return ws.nrows + XLRD_OFFSET - 1

def get_col_index_compatible(ws, col_identifier: str, header_row: int, file_format: str) -> int:
    col_identifier = clean_text(col_identifier)
    try:
        return column_index_from_string(col_identifier)
    except:
        pass
    
    if file_format == "xlsx":
        for col in ws.iter_cols(min_row=header_row, max_row=header_row):
            for cell in col:
                if clean_text(cell.value).upper() == col_identifier.upper():
                    return cell.column
    else:
        header_row_0 = header_row - XLRD_OFFSET
        for col_idx_0 in range(ws.ncols):
            cell_value = ws.cell_value(header_row_0, col_idx_0)
            if clean_text(cell_value).upper() == col_identifier.upper():
                return col_idx_0 + XLRD_OFFSET
    
    raise ValueError(f"未找到列：{col_identifier}（header行：{header_row}）")

def get_cols_range_compatible(ws, col_input_list: list, header_row: int, file_format: str) -> List[int]:
    """辅助函数：解析列列表为索引列表"""
    full_col_letter_list = parse_col_list(col_input_list)
    col_indexes = []
    for col_identifier in full_col_letter_list:
        try:
            if col_identifier.isdigit():
                col_idx = int(col_identifier)
            else:
                col_idx = column_index_from_string(col_identifier)
            if col_idx < 1:
                raise ValueError(f"无效的列索引：{col_idx}")
            col_indexes.append(col_idx)
        except Exception as e:
            try:
                col_idx = None
                if file_format == "xlsx":
                    for col in ws.iter_cols(min_row=header_row, max_row=header_row):
                        for cell in col:
                            if clean_text(cell.value).upper() == col_identifier.upper():
                                col_idx = cell.column
                                break
                        if col_idx:
                            break
                else:
                    header_row_0 = header_row - XLRD_OFFSET
                    for col_idx_0 in range(ws.ncols):
                        cell_value = clean_text(ws.cell_value(header_row_0, col_idx_0)).upper()
                        if cell_value == col_identifier.upper():
                            col_idx = col_idx_0 + XLRD_OFFSET
                            break
                if col_idx:
                    col_indexes.append(col_idx)
                else:
                    raise ValueError(f"未找到列：{col_identifier}")
            except Exception as inner_e:
                raise ValueError(f"列{col_identifier}转换失败：{str(inner_e)}")
    
    # 去重（保留原始顺序）
    seen = set()
    col_indexes_no_dup = []
    for idx in col_indexes:
        if idx not in seen:
            seen.add(idx)
            col_indexes_no_dup.append(idx)
    return col_indexes_no_dup

def parse_filter_conditions(filter_conditions: List[Dict[str, Any]], ws, header_row: int, file_format: str) -> List[Dict[str, Any]]:
    parsed_conditions = []
    if not filter_conditions:
        return parsed_conditions
    
    print(f"\n🔍 解析数据源筛选条件（共{len(filter_conditions)}条）：")
    for i, cond in enumerate(filter_conditions):
        required = ["col", "operator", "value"]
        for k in required:
            if k not in cond:
                raise ValueError(f"第{i+1}条筛选条件缺失{k}字段")
        
        col_idx = get_col_index_compatible(ws, cond["col"], header_row, file_format)
        col_letter = get_column_letter(col_idx)
        col_name = clean_text(get_cell_value(ws, header_row, col_idx, file_format))
        value = cond["value"]
        
        try:
            if isinstance(value, str) and value.replace(".", "").isdigit():
                value = float(value) if "." in value else int(value)
        except:
            pass
        
        if cond["operator"] in ["between", "in"] and isinstance(value, (str, list)):
            if isinstance(value, str):
                value = [v.strip() for v in value.split(",")]
            for j, v in enumerate(value):
                try:
                    value[j] = float(v) if "." in str(v) else int(v)
                except:
                    pass
        
        parsed_cond = {
            "col_idx": col_idx,
            "col_letter": col_letter,
            "col_name": col_name,
            "operator": cond["operator"].lower(),
            "value": value
        }
        parsed_conditions.append(parsed_cond)
        print(f"   [{i+1}] 列{col_letter}({col_name}) {parsed_cond['operator']} {parsed_cond['value']}")
    return parsed_conditions

def filter_data_rows(ws, header_row: int, filter_conditions: List[Dict[str, Any]], file_format: str) -> List[int]:
    if not filter_conditions:
        max_row = get_excel_max_row(ws, file_format)
        data_rows = list(range(header_row + 1, max_row + 1))
        print(f"\n📊 无筛选条件，保留所有{len(data_rows)}行数据")
        return data_rows
    
    operator_map = {
        "eq": lambda x, y: clean_text(x) == clean_text(y),
        "ne": lambda x, y: clean_text(x) != clean_text(y),
        "gt": lambda x, y: float(x) > float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "lt": lambda x, y: float(x) < float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "ge": lambda x, y: float(x) >= float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "le": lambda x, y: float(x) <= float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "contains": lambda x, y: clean_text(y) in clean_text(x),
        "not_contains": lambda x, y: clean_text(y) not in clean_text(x),
        "between": lambda x, y: float(y[0]) <= float(x) <= float(y[1]) if str(x).replace(".", "").isdigit() and all(str(v).replace(".", "").isdigit() for v in y) else False,
        "in": lambda x, y: clean_text(x) in [clean_text(v) for v in y],
        "not_in": lambda x, y: clean_text(x) not in [clean_text(v) for v in y]
    }
    
    max_row = get_excel_max_row(ws, file_format)
    data_start_row = header_row + 1
    valid_rows = []
    
    print(f"\n📝 开始筛选数据行（总行数：{max_row - header_row}）：")
    for row in range(data_start_row, max_row + 1):
        row_valid = True
        for cond in filter_conditions:
            cell_value = get_cell_value(ws, row, cond["col_idx"], file_format)
            operator_func = operator_map.get(cond["operator"])
            
            if not operator_func:
                raise ValueError(f"不支持的运算符：{cond['operator']}")
            
            if not operator_func(cell_value, cond["value"]):
                row_valid = False
                break
        
        if row_valid:
            valid_rows.append(row)
    
    print(f"✅ 筛选完成，保留{len(valid_rows)}行符合条件的数据（总行数：{max_row - header_row}）")
    return valid_rows

def deduplicate_data_rows(ws, header_row: int, deduplicate_cols: List[str], start_row: int) -> int:
    if not deduplicate_cols:
        return 0
    
    dedup_col_indexes = []
    print(f"\n🆔 开始数据去重（依据列：{deduplicate_cols}）：")
    for col_id in deduplicate_cols:
        col_idx = get_col_index_compatible(ws, col_id, header_row, "xlsx")
        col_letter = get_column_letter(col_idx)
        col_name = clean_text(ws.cell(row=header_row, column=col_idx).value)
        dedup_col_indexes.append(col_idx)
        print(f"   去重列：列{col_letter}({col_name})")
    
    seen = set()
    delete_rows = []
    data_row_count = 0
    
    for row in range(start_row, ws.max_row + 1):
        row_key = tuple(clean_text(ws.cell(row=row, column=col_idx).value) for col_idx in dedup_col_indexes)
        data_row_count += 1
        
        if row_key in seen:
            delete_rows.append(row)
        else:
            seen.add(row_key)
    
    delete_count = len(delete_rows)
    if delete_count > 0:
        for row in reversed(delete_rows):
            ws.delete_rows(row)
        print(f"🗑️ 删除{delete_count}行重复数据（剩余{data_row_count - delete_count}行唯一数据）")
    else:
        print(f"✅ 无重复数据，无需删除")
    
    return delete_count

# ===================== 全局输出文件初始化/保存 =====================
def init_global_out_file(out_config: Dict[str, Any]) -> bool:
    """初始化全局唯一输出文件（只加载一次）"""
    global out_workbook, out_file_path, use_global_out_file
    out_file_path = out_config["file_path"]
    backup = out_config.get("backup", True)
    backup_prefix = out_config.get("backup_prefix", "backup_")
    
    try:
        # 备份原文件
        if backup and os.path.exists(out_file_path):
            backup_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{os.path.splitext(out_file_path)[0]}_{backup_prefix}{backup_time}{os.path.splitext(out_file_path)[1]}"
            shutil.copy2(out_file_path, backup_path)
            print(f"📁 已备份全局输出文件：{backup_path}")
        
        # 加载或创建文件
        if os.path.exists(out_file_path):
            out_workbook = openpyxl.load_workbook(out_file_path, data_only=False)
            print(f"✅ 成功加载全局输出文件：{out_file_path}")
        else:
            out_workbook = openpyxl.Workbook()
            print(f"📄 全局输出文件不存在，创建新文件：{out_file_path}")
        
        use_global_out_file = True
        return True
    except Exception as e:
        print(f"❌ 加载全局输出文件失败：{str(e)}")
        use_global_out_file = False
        return False

def save_global_out_file() -> bool:
    """保存全局输出文件（所有规则处理完后调用）"""
    global out_workbook, out_file_path, use_global_out_file
    if not use_global_out_file or not out_workbook or not out_file_path:
        return False
    
    try:
        out_workbook.save(out_file_path)
        out_workbook.close()
        print(f"\n✅ 成功保存全局输出文件：{out_file_path}")
        # 重置全局变量
        out_workbook = None
        out_file_path = ""
        use_global_out_file = False
        return True
    except Exception as e:
        print(f"\n❌ 保存全局输出文件失败：{str(e)}")
        return False

# ===================== 配置加载 =====================
def load_json_config(config_path: str = "./config.json") -> Dict[str, Any]:
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"配置文件不存在：{config_path}")
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON格式错误：{str(e)}")

    # 核心配置检查（原有代码）
    required = ["email_config", "rules", "other"]
    for k in required:
        if k not in config:
            raise KeyError(f"缺失核心配置：{k}")
    if not config["rules"]:
        raise ValueError("规则列表不能为空")
    
    # 全局输出文件配置检查（原有代码）
    if "out_file" in config:
        print(f"\n🌐 检测到全局输出文件配置，启用批量处理模式")
        if "file_path" not in config["out_file"]:
            raise KeyError("全局out_file配置缺失file_path字段")
    
    # 规则校验（原有代码，新增处理规则解析）
    for i, rule in enumerate(config["rules"]):
        # 新增：初始化处理规则（默认空列表）
        rule.setdefault("data_process_rules", [])
        # 原有：本地文件路径配置
        rule.setdefault("local_file_path", "")
        rule["attach_rule"].setdefault("filter_conditions", [])
        local_rule = rule["local_rule"]
        local_rule.setdefault("write_mode", "overwrite")
        local_rule.setdefault("deduplicate", False)
        local_rule.setdefault("deduplicate_cols", [])
        local_rule.setdefault("force_formula_cols", [])
        
        # 校验cols_range（原有代码）
        if "cols_range" not in rule["attach_rule"] or "cols_range" not in rule["local_rule"]:
            raise KeyError(f"第{i+1}条规则缺失cols_range配置")
        
        # 校验attach_rule的cols_range（原有代码）
        attach_col_list = rule["attach_rule"]["cols_range"]
        if not isinstance(attach_col_list, list) or len(attach_col_list) == 0:
            raise ValueError(f"第{i+1}条规则attach_rule的cols_range必须是非空列表")
        try:
            parsed_attach_cols = parse_col_list(attach_col_list)
        except Exception as e:
            raise ValueError(f"第{i+1}条规则attach_rule的cols_range格式无效：{str(e)}")
        
        # 校验local_rule的cols_range（原有代码）
        local_col_list = local_rule["cols_range"]
        if not isinstance(local_col_list, list) or len(local_col_list) == 0:
            raise ValueError(f"第{i+1}条规则local_rule的cols_range必须是非空列表")
        try:
            parsed_local_cols = parse_col_list(local_col_list)
        except Exception as e:
            raise ValueError(f"第{i+1}条规则local_rule的cols_range格式无效：{str(e)}")
        
        # 校验列数一致（原有代码）
        if len(parsed_attach_cols) != len(parsed_local_cols):
            raise ValueError(f"第{i+1}条规则：附件列({len(parsed_attach_cols)}列)与本地列({len(parsed_local_cols)}列)数量不一致")
        
        # ========== 新增：打印处理规则配置 ==========
        process_rules = rule.get("data_process_rules", [])
        print(f"\n📋 第{i+1}条规则配置：")
        if rule["local_file_path"]:
            print(f"   本地文件路径：{rule['local_file_path']}")
        else:
            print(f"   模式：邮件下载附件")
        print(f"   筛选条件：{rule['attach_rule']['filter_conditions'] or '无'}")
        print(f"   附件提取列（原始）：{attach_col_list}")
        print(f"   附件提取列（解析后）：{parsed_attach_cols}")
        print(f"   本地写入列（原始）：{local_col_list}")
        print(f"   本地写入列（解析后）：{parsed_local_cols}")
        print(f"   写入模式：{local_rule['write_mode']}")
        print(f"   去重开关：{local_rule['deduplicate']}")
        print(f"   去重列：{local_rule['deduplicate_cols'] or '无'}")
        print(f"   强制公式列：{local_rule['force_formula_cols'] or '无'}")
        print(f"   数据处理规则：{process_rules or '无'}")
    
    # 邮箱配置默认值（原有代码）
    email_conf = config["email_config"]
    email_conf.setdefault("root_folder", "INBOX")
    email_conf.setdefault("exclude_subfolders", ["临时", "测试", "垃圾", "deleted"])
    email_conf.setdefault("temp_dir", "./temp_attachments")
    
    print(f"\n✅ 加载配置成功，共{len(config['rules'])}条规则")
    print(f"🔧 根文件夹配置：{email_conf['root_folder']}")
    print(f"🚫 排除子文件夹关键词：{email_conf['exclude_subfolders']}")
    return config

# ===================== 附件查找与下载 =====================
def search_attach_in_folder(mail: imaplib.IMAP4_SSL, folder_name: str, search_criteria: str, match_key: str, suffix: str, all_folders: List[Tuple[str, str]]) -> List[Tuple[int, str, bytes, str]]:
    attach_list = []
    folder_decoded = next((d for e, d in all_folders if e == folder_name), folder_name)
    
    try:
        encoded_folder = encode_imap_folder(folder_name)
        status, _ = mail.select(encoded_folder, readonly=True)
        if status != "OK":
            print(f"⚠️  无法访问子文件夹：{folder_decoded} / {folder_name}，跳过")
            return attach_list
        
        status, messages = mail.search(None, search_criteria)
        if status != "OK" or not messages[0]:
            return attach_list
        
        email_ids = messages[0].split()
        for idx, eid in enumerate(reversed(email_ids)):
            status, msg_data = mail.fetch(eid, "(RFC822)")
            if status != "OK":
                continue
            
            msg = email.message_from_bytes(msg_data[0][1])
            subject = decode_email_header(msg["Subject"])
            
            for part in msg.walk():
                if part.get_content_maintype() == "multipart" or not part.get("Content-Disposition"):
                    continue
                
                filename = decode_email_header(part.get_filename())
                if not filename or not filename.lower().endswith(suffix.lower()) or match_key not in filename:
                    continue
                
                attach_data = part.get_payload(decode=True)
                attach_list.append((idx, filename, attach_data, folder_decoded))
                print(f"   📦 找到匹配附件：[{folder_decoded}] {filename}（邮件主题：{subject}）")
    
    except Exception as e:
        print(f"⚠️  处理子文件夹{folder_decoded} / {folder_name}出错：{e}")
    return attach_list

def download_latest_attach_from_root(rule: Dict[str, Any], email_config: Dict[str, Any]) -> Optional[str]:
    match_key = rule["attach_rule"]["match_key"]
    suffix = rule["attach_rule"]["suffix"]
    temp_dir = email_config["temp_dir"]
    all_attach_candidates = []
    # 初始化缓存目录
    cache_dir = init_cache_dir(temp_dir)

    try:
        mail = imaplib.IMAP4_SSL(email_config["imap_server"])
        mail.login(email_config["account"], email_config["password"])

        # 配置搜索条件
        search_criteria = email_config.get("search_criteria", "")
        if search_criteria.upper() == "TODAY" or not search_criteria:
            search_criteria = get_today_imap_criteria()
            print(f"\n🔍 筛选当日邮件，搜索条件：{search_criteria}")

        # 解析文件夹
        all_folders = parse_imap_folders(mail)
        root_folder = email_config["root_folder"]
        target_encoded_folders = get_root_folder_children(all_folders, root_folder)
        filtered_encoded_folders = filter_folders(target_encoded_folders, email_config["exclude_subfolders"], all_folders)

        # 查找附件
        print(f"\n🚀 开始递归遍历{len(filtered_encoded_folders)}个子文件夹查找附件（匹配关键词：{match_key}，后缀：{suffix}）")
        for encoded_folder in filtered_encoded_folders:
            folder_attach = search_attach_in_folder(mail, encoded_folder, search_criteria, match_key, suffix, all_folders)
            all_attach_candidates.extend(folder_attach)

        if not all_attach_candidates:
            root_decoded = next((d for e, d in all_folders if e == root_folder), root_folder)
            print(f"\n❌ 根文件夹「{root_decoded}」及其子文件夹中未找到匹配【{match_key}】的附件")
            mail.close()
            mail.logout()
            return None
        
        # 排序取最新附件
        all_attach_candidates.sort(key=lambda x: x[0])
        _, filename, payload, folder_decoded = all_attach_candidates[0]
        
        # 清理附件文件名
        filename = clean_invalid_filename_chars(filename)
        save_path = os.path.join(temp_dir, filename)
        
        # 保存附件
        os.makedirs(temp_dir, exist_ok=True)
        with open(save_path, "wb") as f:
            f.write(payload)
        
        print(f"\n✅ 找到最新匹配附件：")
        print(f"   📂 所在子文件夹：{folder_decoded}")
        print(f"   📄 附件名称（已清理）：{filename}")
        print(f"   💾 保存路径：{save_path}")

        # 处理RAR文件
        file_ext = os.path.splitext(save_path)[1].lower()
        final_attach_path = None
        if file_ext == ".rar":
            try:
                extracted_files = extract_rar_file(save_path, temp_dir)
                final_attach_path = find_xlsx_in_files(extracted_files)
            except Exception as e:
                print(f"\n❌ RAR文件处理失败：{str(e)}")
                mail.close()
                mail.logout()
                return None
        else:
            final_attach_path = save_path
        
        # ========== 新增：更新缓存 ==========
        if final_attach_path and os.path.exists(final_attach_path):
            update_cache_file(cache_dir, final_attach_path, match_key)

        mail.close()
        mail.logout()
        return final_attach_path
    
    except Exception as e:
        print(f"\n❌ 附件下载失败：{str(e)}")
        try:
            mail.close()
            mail.logout()
        except:
            pass
        return None
# ===================== 新增：本地文件加载函数 =====================
def load_local_file(local_file_path: str, temp_dir: str) -> Optional[str]:
    """
    加载本地文件（支持Excel和RAR格式）
    :param local_file_path: 本地文件路径
    :param temp_dir: 临时目录（用于RAR解压）
    :return: 处理后的Excel文件路径
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(local_file_path):
            raise FileNotFoundError(f"本地文件不存在：{local_file_path}")
        
        print(f"\n📂 开始加载本地文件：")
        print(f"   📄 文件路径：{local_file_path}")
        print(f"   📁 文件大小：{os.path.getsize(local_file_path) / 1024:.2f} KB")
        
        # 获取文件扩展名
        file_ext = os.path.splitext(local_file_path)[1].lower()
        
        # 处理RAR文件
        if file_ext == ".rar":
            print(f"   🔍 检测到RAR压缩包，开始解压...")
            extracted_files = extract_rar_file(local_file_path, temp_dir)
            xlsx_file_path = find_xlsx_in_files(extracted_files)
            if not xlsx_file_path:
                raise Exception("RAR解压后未找到有效XLSX文件")
            return xlsx_file_path
        
        # 处理Excel文件（xls/xlsx）
        elif file_ext in [".xls", ".xlsx"]:
            print(f"   🔍 检测到Excel文件，直接使用")
            return local_file_path
        
        # 其他格式
        else:
            raise Exception(f"不支持的文件格式：{file_ext}（仅支持.xls/.xlsx/.rar）")
    
    except Exception as e:
        print(f"\n❌ 加载本地文件失败：{str(e)}")
        return None

# ===================== 数据提取与写入 =====================


def replace_local_cols(extract_data: Dict[str, List[Any]], local_rule: Dict[str, Any]) -> bool:
    """按列标识精准写入本地列，保留原始映射"""
    global out_workbook, use_global_out_file
    
    if not extract_data:
        print("❌ 无提取数据，跳过替换")
        return False

    def batch_write_column_by_letter(ws, col_letter, start_row, data, header_row):
        """按列标识批量写入数据"""
        col_idx = get_col_index_compatible(ws, col_letter, header_row, "xlsx")
        end_row = start_row + len(data) - 1
        if end_row < start_row:
            return
        for row_offset, value in enumerate(data):
            target_row = start_row + row_offset
            ws.cell(row=target_row, column=col_idx, value=value)
        print(f"   ✅ 列{col_letter}批量写入{len(data)}行数据（起始行：{start_row}）")

    def clear_config_cols_by_letter(ws, col_letter_list, header_row):
        """按列标识清空配置列旧数据"""
        if ws.max_row <= header_row:
            return
        for col_letter in col_letter_list:
            col_idx = get_col_index_compatible(ws, col_letter, header_row, "xlsx")
            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=row, column=col_idx, value=None)
        print(f"✅ 精准清空{len(col_letter_list)}个配置列的旧数据行")

    # 通用配置
    sheet_name = local_rule["sheet"]
    header_row = local_rule["header_row"]
    local_col_list = local_rule["cols_range"]
    write_mode = local_rule["write_mode"].lower()
    deduplicate = local_rule["deduplicate"]
    deduplicate_cols = local_rule["deduplicate_cols"]
    force_formula_cols = local_rule.get("force_formula_cols", [])

    try:
        # 解析列列表
        local_full_col_letter_list = parse_col_list(local_col_list)
        attach_col_letter_list = list(extract_data.keys())

        # 校验列数一致
        if len(attach_col_letter_list) != len(local_full_col_letter_list):
            raise ValueError(f"附件列数({len(attach_col_letter_list)})与本地列数({len(local_full_col_letter_list)})不匹配")

        # 建立列映射
        col_mapping = dict(zip(attach_col_letter_list, local_full_col_letter_list))
        print(f"\n🔗 建立附件列→本地列精准映射（保留原始顺序）：")
        for attach_col, local_col in col_mapping.items():
            print(f"   附件列{attach_col} → 本地列{local_col}")

        # 选择文件模式
        if use_global_out_file and out_workbook:
            print(f"\n🌐 使用全局输出文件模式，写入Sheet：{sheet_name}")
            if sheet_name in out_workbook.sheetnames:
                ws = out_workbook[sheet_name]
            else:
                ws = out_workbook.create_sheet(sheet_name)
                print(f"⚠️  Sheet「{sheet_name}」不存在，新建空白Sheet")
            output_path = out_file_path
        else:
            file_path = local_rule.get("file_path", "")
            output_path = local_rule.get("output_path", "")
            if not file_path or not os.path.exists(file_path):
                print(f"❌ 本地文件不存在：{file_path}")
                return False
            wb = openpyxl.load_workbook(file_path, data_only=False)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                print(f"⚠️  Sheet「{sheet_name}」不存在，新建空白Sheet")

        # 打印本地列信息
        print(f"\n📊 本地目标列（保留映射顺序）：")
        for local_col in local_full_col_letter_list:
            col_idx = get_col_index_compatible(ws, local_col, header_row, "xlsx")
            col_name = clean_text(ws.cell(row=header_row, column=col_idx).value)
            print(f"   列{local_col}：{col_name}")

        # 清空旧数据
        if write_mode == "overwrite":
            clear_config_cols_by_letter(ws, local_full_col_letter_list, header_row)

        # 确定写入起始行
        if write_mode == "overwrite":
            write_start_row = header_row + 1
        else:
            write_start_row = ws.max_row + 1

        # 获取数据行数
        data_rows_count = len(list(extract_data.values())[0]) if extract_data else 0
        print(f"\n🚀 开始批量写入{data_rows_count}行数据...")

        # 写入数据
        for attach_col, local_col in col_mapping.items():
            col_data = extract_data[attach_col]
            batch_write_column_by_letter(ws, local_col, write_start_row, col_data, header_row)

            # 公式下拉逻辑
            if (write_mode == "overwrite" or (write_mode == "append" and data_rows_count > 0)) and force_formula_cols:
                formula_cols = []
                for col in force_formula_cols:
                    try:
                        formula_cols.append(column_index_from_string(col) if isinstance(col, str) else col)
                    except:
                        print(f"⚠️  无效的强制公式列：{col}，跳过")
                formula_cols = list(set(formula_cols))
                print(f"[强制模式] 待处理公式列：{[get_column_letter(c) for c in formula_cols]}")

                if formula_cols:
                    formula_template_row = header_row + 1
                    if formula_template_row > ws.max_row:
                        formula_template_row = write_start_row

                    for f_col_idx in formula_cols:
                        f_col_letter = get_column_letter(f_col_idx)
                        formula_template = ws.cell(row=formula_template_row, column=f_col_idx).value
                        if not formula_template or not str(formula_template).startswith("="):
                            print(f"[调试] 列{f_col_letter} 无有效公式模板，跳过")
                            continue
                        print(f"[调试] 列{f_col_letter} 公式模板：{formula_template}")

                        # 批量下拉公式
                        for row_offset in range(len(col_data)):
                            target_row = write_start_row + row_offset
                            adjusted_formula = adjust_formula_row_reference(
                                formula=formula_template,
                                template_row=formula_template_row,
                                target_row=target_row
                            )
                            ws.cell(row=target_row, column=f_col_idx, value=adjusted_formula)

                        print(f"   📌 本地列{local_col}后续列{f_col_letter}公式下拉完成")

        # 数据去重
        if deduplicate and deduplicate_cols:
            print(f"\n🆔 开始数据去重...")
            deduplicate_data_rows(ws, header_row, deduplicate_cols, header_row + 1)

        # 保存文件
        if use_global_out_file and out_workbook:
            print(f"\n✅ 全局模式：Sheet「{sheet_name}」处理完成（暂存内存）")
            return True
        else:
            print(f"\n💾 保存文件到：{output_path}...")
            wb.save(output_path)
            wb.close()
            print(f"\n🎉 本地文件处理完成 → {output_path}")
            print(f"📋 处理总结：")
            print(f"   写入模式：{write_mode}")
            print(f"   写入行数：{data_rows_count}")
            print(f"   去重开关：{deduplicate}")
            print(f"   强制公式列：{force_formula_cols or '无'}")
            return True

    except Exception as e:
        print(f"\n❌ 替换失败：{str(e)}")
        if not use_global_out_file and 'wb' in locals():
            wb.close()
        return False
def process_data_by_rules(extract_data: Dict[str, List[Any]], process_rules: List[Dict[str, Any]]) -> Dict[str, List[Any]]:
    """
    按顺序执行数据处理规则：delete_str（删除指定字符串）、slice_combine（切片+拼接）
    :param extract_data: 原始提取数据 {列字母: [值列表]}
    :param process_rules: 处理规则列表，格式示例：
        [
            {"type": "delete_str", "col": "A", "target_str": "测试"},
            {"type": "slice_combine", "col": "B", "slice_rule": "[3:-6]", "combine_str": "BIN"}
        ]
    :return: 处理后的数据
    """
    if not process_rules or not extract_data:
        return extract_data
    
    print(f"\n🔧 开始按顺序执行{len(process_rules)}条数据处理规则：")
    processed_data = {col: values.copy() for col, values in extract_data.items()}
    
    for i, rule in enumerate(process_rules):
        proc_type = rule.get("type")
        col = rule.get("col")
        if col not in processed_data:
            print(f"   [{i+1}] 跳过 → 列{col}不存在于提取数据中")
            continue
        
        values = processed_data[col]
        if not values:
            print(f"   [{i+1}] 跳过 → 列{col}无数据")
            continue
        
        print(f"   [{i+1}] 处理列{col} → 规则类型：{proc_type}")
        
        # 规则1：删除指定字符串
        if proc_type == "delete_str":
            target_str = rule.get("target_str", "")
            new_values = []
            for val in values:
                if val is None:
                    new_values.append(val)
                    continue
                str_val = str(val)
                new_val = str_val.replace(target_str, "")
                new_values.append(new_val)
                # 打印前3条示例
                if len(new_values) <= 3:
                    print(f"      示例：{str_val} → {new_val}")
            processed_data[col] = new_values
            print(f"      ✅ 完成删除字符串'{target_str}'，共处理{len(new_values)}行")
        
        # 规则2：切片+拼接（支持Python str切片语法）
        elif proc_type == "slice_combine":
            slice_rule = rule.get("slice_rule", "")
            combine_str = rule.get("combine_str", "")
            new_values = []
            
            # 安全解析切片规则（支持[3:-6]、[:5]、[2:]、[:-3]等）
            slice_pattern = re.compile(r'^\[(-?\d*)?:(-?\d*)?\]$')
            match = slice_pattern.match(slice_rule)
            if not match:
                raise ValueError(f"无效的切片规则：{slice_rule}，正确格式如[3:-6]、[:5]、[2:]、[:-3]")
            
            start_str, end_str = match.groups()
            start = int(start_str) if start_str else None
            end = int(end_str) if end_str else None
            
            for val in values:
                if val is None:
                    new_values.append(val)
                    continue
                str_val = str(val)
                # 执行切片（异常时返回空字符串）
                try:
                    slice_val = str_val[start:end]
                except:
                    slice_val = ""
                # 拼接字符串
                new_val = combine_str + slice_val 
                new_values.append(new_val)
                # 打印前3条示例
                if len(new_values) <= 3:
                    print(f"      示例：{str_val} → {str_val}[{start or ''}:{end or ''}] + '{combine_str}' = {new_val}")
            
            processed_data[col] = new_values
            print(f"      ✅ 完成切片拼接，共处理{len(new_values)}行")
        
        else:
            raise ValueError(f"不支持的规则类型：{proc_type}（仅支持delete_str/slice_combine）")
    
    return processed_data
def extract_cols_from_attach(attach_path: str, attach_rule: Dict[str, Any], process_rules: List[Dict[str, Any]] = []) -> Dict[str, List[Any]]:
    """按列标识提取数据，保留原始顺序（新增：数据处理规则）"""
    sheet_name = attach_rule["sheet"]
    header_row = attach_rule["header_row"]
    col_list = attach_rule["cols_range"]
    filter_conditions = attach_rule["filter_conditions"]
    
    try:
        # 读取附件（原有代码）
        ws, real_format = read_excel_final(attach_path, sheet_name)
        
        # 解析列列表（原有代码）
        full_col_letter_list = parse_col_list(col_list)
        print(f"\n📊 提取附件列（共{len(full_col_letter_list)}列，保留原始顺序）：")
        for col_letter in full_col_letter_list:
            col_idx = get_col_index_compatible(ws, col_letter, header_row, real_format)
            col_name = clean_text(get_cell_value(ws, header_row, col_idx, real_format))
            print(f"   列{col_letter}：{col_name}")

        # 筛选数据行（原有代码）
        parsed_conditions = parse_filter_conditions(filter_conditions, ws, header_row, real_format)
        valid_rows = filter_data_rows(ws, header_row, parsed_conditions, real_format)
        if not valid_rows:
            print("❌ 无符合筛选条件的数据行")
            return {}

        # 提取数据（原有代码）
        extract_data = {}
        for col_letter in full_col_letter_list:
            col_idx = get_col_index_compatible(ws, col_letter, header_row, real_format)
            col_data = []
            for row in valid_rows:
                cell_value = get_cell_value(ws, row, col_idx, real_format)
                col_data.append(cell_value)
            extract_data[col_letter] = col_data
            print(f"   列{col_letter}提取{len(col_data)}行筛选后数据")

        # ========== 新增：执行数据处理规则 ==========
        if process_rules:
            extract_data = process_data_by_rules(extract_data, process_rules)

        return extract_data
    except Exception as e:
        print(f"\n❌ 提取附件数据失败：{str(e)}")
        return {}
# ===================== 主函数 =====================
def main(config_path: str = "./config.json"):
    try:
        # 加载配置（原有代码）
        print("📋 加载配置文件...")
        config = load_json_config(config_path)
        email_config = config["email_config"]
        rules = config["rules"]
        other = config["other"]
        temp_dir = other["temp_dir"]
        os.makedirs(temp_dir, exist_ok=True)
        # 初始化缓存目录
        cache_dir = init_cache_dir(temp_dir)

        # 初始化全局输出文件（原有代码）
        if "out_file" in config:
            if not init_global_out_file(config["out_file"]):
                print("⚠️  全局输出文件初始化失败，降级为原有模式")
        
        # 进度条（原有代码）
        success_count = 0
        pbar = tqdm(total=len(rules), desc="整体处理进度", unit="规则", ncols=100)
        
        for i, rule in enumerate(rules):
            pbar.set_description(f"处理规则 {i+1}/{len(rules)}")
            
            print(f"\n==================================================")
            print(f"📋 处理第{i+1}/{len(rules)}条规则：匹配【{rule['attach_rule']['match_key']}】")
            
            # 加载本地文件/下载附件（修改部分）
            local_file_path = rule.get("local_file_path", "")
            attach_path = None
            if local_file_path:
                pbar.set_postfix({"状态": "加载本地文件"})
                attach_path = load_local_file(local_file_path, temp_dir)
            else:
                pbar.set_postfix({"状态": "下载附件"})
                # 优先下载邮箱附件
                attach_path = download_latest_attach_from_root(rule, email_config)
                
                # ========== 新增：下载失败时读取缓存 ==========
                if not attach_path:
                    print(f"\n⚠️  邮箱下载失败，尝试读取缓存文件...")
                    attach_path = get_latest_cache_file(cache_dir, rule["attach_rule"]["match_key"])
                    if not attach_path:
                        print(f"❌ 第{i+1}条规则：文件获取失败（邮箱+缓存均无匹配文件）")
                        pbar.update(1)
                        pbar.set_postfix({"状态": "文件获取失败"})
                        continue
                    print(f"✅ 从缓存加载文件：{attach_path}")
            
            if not attach_path:
                print(f"❌ 第{i+1}条规则：文件获取失败（本地/邮件/缓存）")
                pbar.update(1)
                pbar.set_postfix({"状态": "文件获取失败"})
                continue
            
            # 获取处理规则
            process_rules = rule.get("data_process_rules", [])
            
            # 提取数据
            pbar.set_postfix({"状态": "提取数据"})
            extract_data = extract_cols_from_attach(attach_path, rule["attach_rule"], process_rules)
            if not extract_data:
                print(f"❌ 第{i+1}条规则：数据提取失败")
                pbar.update(1)
                pbar.set_postfix({"状态": "提取失败"})
                continue
            
            # 写入数据（原有代码）
            pbar.set_postfix({"状态": "写入数据"})
            if replace_local_cols(extract_data, rule["local_rule"]):
                success_count += 1
                print(f"✅ 第{i+1}条规则：处理成功")
                pbar.set_postfix({"状态": "处理成功"})
            else:
                print(f"❌ 第{i+1}条规则：替换失败")
                pbar.set_postfix({"状态": "替换失败"})
            
            pbar.update(1)
        
        # 保存全局输出文件（原有代码）
        if use_global_out_file:
            save_global_out_file()
        
        # 清理临时文件（原有代码）
        if other.get("clean_temp", True) and os.path.exists(temp_dir):
            # 保留缓存目录，只清理其他临时文件
            temp_subdirs = [d for d in glob.glob(os.path.join(temp_dir, "*")) if d != cache_dir]
            for d in temp_subdirs:
                if os.path.isdir(d):
                    shutil.rmtree(d)
                else:
                    os.remove(d)
            print(f"\n🗑️ 清理临时目录（保留缓存）：{temp_dir}")
        
        # 进度条结束（原有代码）
        pbar.close()
        
        # 输出结果（原有代码）
        print(f"\n==================================================")
        print(f"📊 总处理结果：成功{success_count}/{len(rules)}条规则")
        if success_count == len(rules):
            print(f"🎉 所有规则执行成功！")

    except Exception as e:
        print(f"\n❌ 程序执行失败：{str(e)}")
        # 异常时保存全局文件（原有代码）
        if use_global_out_file:
            save_global_out_file()
if __name__ == "__main__":
    # 请修改为你的配置文件实际路径
    main(config_path="D:\\供应链\\芯片\\configchv7.json")