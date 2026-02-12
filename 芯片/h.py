import imaplib
import email
from email.header import decode_header
import os
import json
import openpyxl
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Tuple, Optional, Any
import base64
import re

# ===================== IMAP UTF-7编解码核心函数（解决中文文件夹编码问题） =====================
def imap_utf7_decode(s: str) -> str:
    """
    解码IMAP Modified UTF-7编码的字符串（如&i6FSEg- → 中文）
    参考：https://datatracker.ietf.org/doc/html/rfc3501#section-5.1.3
    """
    if not s or "&" not in s:
        return s
    
    # 匹配IMAP UTF-7编码段：&xxxx-
    pattern = re.compile(r'&([^-]+)-')
    
    def decode_match(match):
        encoded_part = match.group(1).replace(",", "/")  # 替换分隔符
        if not encoded_part:
            return "&"  # 处理单独的&
        try:
            # base64解码 + UTF-16BE解码
            decoded = base64.b64decode(encoded_part + "==", altchars=b"+/").decode("utf-16be")
            return decoded
        except:
            return match.group(0)  # 解码失败则返回原字符串
    
    return pattern.sub(decode_match, s)

def imap_utf7_encode(s: str) -> str:
    """
    将中文编码为IMAP Modified UTF-7格式（如中文 → &i6FSEg-）
    """
    if not s or all(ord(c) < 128 for c in s):
        return s
    
    # 拆分ASCII和非ASCII部分
    result = []
    buffer = []
    
    for c in s:
        if ord(c) < 128:
            if buffer:
                # 编码非ASCII缓冲区
                b64 = base64.b64encode(''.join(buffer).encode("utf-16be")).decode("ascii").rstrip("=").replace("/", ",")
                result.append(f"&{b64}-")
                buffer = []
            result.append(c)
        else:
            buffer.append(c)
    
    # 处理剩余的非ASCII字符
    if buffer:
        b64 = base64.b64encode(''.join(buffer).encode("utf-16be")).decode("ascii").rstrip("=").replace("/", ",")
        result.append(f"&{b64}-")
    
    return ''.join(result)

# ===================== 日期工具函数 =====================
def get_imap_date_str(date_obj: datetime) -> str:
    """转换为IMAP协议要求的日期格式（DD-Mon-YYYY）"""
    month_abbr = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{date_obj.day:02d}-{month_abbr[date_obj.month-1]}-{date_obj.year}"

def get_today_imap_criteria() -> str:
    """生成当日邮件的IMAP搜索条件"""
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    today_str = get_imap_date_str(datetime(today.year, today.month, today.day))
    tomorrow_str = get_imap_date_str(datetime(tomorrow.year, tomorrow.month, tomorrow.day))
    return f'SINCE "{today_str}" BEFORE "{tomorrow_str}"'

# ===================== 核心工具函数（适配IMAP UTF-7编码） =====================
def decode_email_header(header: Any) -> str:
    """兼容多编码解析邮件头"""
    if not header:
        return ""
    decoded_parts = []
    for part, encoding in decode_header(header):
        if isinstance(part, bytes):
            try:
                decoded_part = part.decode(encoding or "gbk")
            except (UnicodeDecodeError, LookupError):
                try:
                    decoded_part = part.decode("utf-8")
                except UnicodeDecodeError:
                    decoded_part = part.decode("latin-1")
        else:
            decoded_part = str(part)
        decoded_parts.append(decoded_part)
    return "".join(decoded_parts).strip()

def encode_imap_folder(folder_name: str) -> bytes:
    """
    编码IMAP文件夹名（先转IMAP UTF-7，再转字节串）
    支持输入中文或编码后的字符串，自动处理
    """
    if not folder_name:
        return b"INBOX"
    
    # 先将中文编码为IMAP UTF-7格式
    imap_encoded = imap_utf7_encode(folder_name)
    
    # 再转字节串（兼容不同服务器）
    for encoding in ["utf-8", "gbk"]:
        try:
            return imap_encoded.encode(encoding)
        except:
            continue
    return imap_encoded.encode("latin-1")

def parse_imap_folders(mail: imaplib.IMAP4_SSL) -> List[Tuple[str, str]]:
    """
    解析所有邮箱文件夹，返回(编码名, 解码后中文名)的列表
    例如：("&i6FSEg-", "芯片数据"), ("&i6FSEg-/PC", "芯片数据/PC")
    """
    try:
        status, folders_raw = mail.list()
        if status != "OK":
            print("⚠️  获取文件夹列表失败")
            return [("INBOX", "INBOX")]
        
        folder_list = []
        for folder in folders_raw:
            if not folder:
                continue
            
            # 解码原始文件夹数据
            folder_str = folder.decode("utf-8", errors="ignore")
            # 解析IMAP LIST返回格式：(\HasNoChildren) "/" "INBOX/子文件夹"
            parts = folder_str.split('"')
            if len(parts) < 3:
                continue
            
            # 提取编码后的文件夹名
            encoded_name = parts[-2].strip()
            if not encoded_name or encoded_name in [f[0] for f in folder_list]:
                continue
            
            # 解码为中文
            decoded_name = imap_utf7_decode(encoded_name)
            folder_list.append((encoded_name, decoded_name))
        
        # 去重并排序，确保INBOX在最前
        folder_list = list(dict.fromkeys(folder_list))
        folder_list.sort(key=lambda x: (x[0] != "INBOX", x[0].lower()))
        
        # 打印解析结果（编码名 + 中文）
        print(f"\n📂 解析到邮箱所有文件夹（共{len(folder_list)}个）：")
        for i, (encoded, decoded) in enumerate(folder_list):
            print(f"   [{i+1}] 编码名：{encoded:<20} 中文名：{decoded}")
        return folder_list
    except Exception as e:
        print(f"⚠️  解析文件夹失败：{e}，仅使用收件箱")
        return [("INBOX", "INBOX")]

def get_root_folder_children(folder_list: List[Tuple[str, str]], root_folder: str) -> List[str]:
    """
    根据配置的根文件夹名（中文/编码名），筛选其下所有子文件夹的编码名
    支持输入：中文（如"芯片数据"）或编码名（如"&i6FSEg-"）
    """
    # 先统一处理根文件夹名（匹配中文或编码名）
    target_encoded_root = None
    root_lower = root_folder.lower()
    
    # 第一步：查找根文件夹对应的编码名
    for encoded, decoded in folder_list:
        if decoded.lower() == root_lower or encoded.lower() == root_lower:
            target_encoded_root = encoded
            root_decoded = decoded
            break
    
    # 根文件夹不存在则默认INBOX
    if not target_encoded_root:
        print(f"⚠️  未找到根文件夹「{root_folder}」，默认使用INBOX")
        target_encoded_root = "INBOX"
        root_decoded = "INBOX"
    
    # 第二步：筛选根文件夹下的所有子文件夹（编码名）
    target_encoded_folders = []
    for encoded, decoded in folder_list:
        # 匹配条件：等于根文件夹编码名 OR 以根文件夹编码名+分隔符为前缀
        if encoded == target_encoded_root or encoded.startswith(f"{target_encoded_root}/"):
            target_encoded_folders.append((encoded, decoded))
    
    if not target_encoded_folders:
        print(f"⚠️  未找到根文件夹「{root_decoded}」及其子文件夹，默认使用INBOX")
        target_encoded_folders = [("INBOX", "INBOX")]
    else:
        # 排序：根文件夹在前，子文件夹按层级排序
        target_encoded_folders.sort(key=lambda x: (x[0] != target_encoded_root, x[0].count("/"), x[0].lower()))
    
    # 打印递归范围（树形展示）
    print(f"\n🎯 待递归的文件夹（根：{root_decoded} / {target_encoded_root}）：")
    for i, (encoded, decoded) in enumerate(target_encoded_folders):
        level = encoded.count("/") - target_encoded_root.count("/")  # 计算子文件夹层级
        prefix = "  " * level + "└─ " if level > 0 else "┌─ "
        print(f"   {prefix}中文名：{decoded:<20} 编码名：{encoded}")
    
    # 只返回编码名（用于后续访问）
    return [encoded for encoded, decoded in target_encoded_folders]

def filter_folders(folder_list: List[str], exclude_folders: List[str], all_folders: List[Tuple[str, str]]) -> List[str]:
    """
    过滤根文件夹子集中不需要的子文件夹
    exclude_folders：排除的关键词（中文/编码名）
    """
    if not exclude_folders:
        return folder_list
    
    filtered = []
    exclude_lower = [f.lower() for f in exclude_folders]
    
    for encoded in folder_list:
        # 获取该文件夹的中文名
        decoded = next((d for e, d in all_folders if e == encoded), encoded)
        # 匹配排除关键词（中文/编码名都匹配）
        if any(ex in decoded.lower() or ex in encoded.lower() for ex in exclude_lower):
            print(f"🚫 排除子文件夹：中文名={decoded} / 编码名={encoded}（匹配排除关键词）")
            continue
        filtered.append(encoded)
    
    print(f"\n🔍 过滤后最终递归的文件夹（共{len(filtered)}个）：")
    for encoded in filtered:
        decoded = next((d for e, d in all_folders if e == encoded), encoded)
        print(f"   中文名：{decoded:<20} 编码名：{encoded}")
    return filtered

def clean_text(text: Any) -> str:
    """文本清理"""
    if isinstance(text, bytes):
        try:
            return text.decode("gbk")
        except:
            return text.decode("utf-8", errors="ignore")
    return str(text).replace("\r", "").replace("\n", "").replace("\t", "").strip()

def get_col_index(ws, col_identifier: str, header_row: int) -> int:
    """解析列标识为列索引"""
    col_identifier = clean_text(col_identifier)
    try:
        return column_index_from_string(col_identifier)
    except:
        pass
    for col in ws.iter_cols(min_row=header_row, max_row=header_row):
        for cell in col:
            if clean_text(cell.value) == col_identifier:
                return cell.column
    raise ValueError(f"未找到列：{col_identifier}")

def get_cols_range(ws, start_col: str, end_col: str, header_row: int) -> List[int]:
    """获取连续列的索引列表"""
    start_idx = get_col_index(ws, start_col, header_row)
    end_idx = get_col_index(ws, end_col, header_row)
    if start_idx > end_idx:
        start_idx, end_idx = end_idx, start_idx
    return list(range(start_idx, end_idx + 1))

def load_json_config(config_path: str = "./config.json") -> Dict[str, Any]:
    """读取JSON配置（适配中文根文件夹）"""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"配置文件不存在：{config_path}")
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON格式错误：{str(e)}")

    # 核心校验
    required = ["email_config", "rules", "other"]
    for k in required:
        if k not in config:
            raise KeyError(f"缺失核心配置：{k}")
    if not config["rules"]:
        raise ValueError("规则列表不能为空")
    
    # 列范围配置校验
    for i, rule in enumerate(config["rules"]):
        if "cols_range" not in rule["attach_rule"] or "cols_range" not in rule["local_rule"]:
            raise KeyError(f"第{i+1}条规则缺失cols_range配置")
        for side in ["attach_rule", "local_rule"]:
            cr = rule[side]["cols_range"]
            if "start_col" not in cr or "end_col" not in cr:
                raise KeyError(f"第{i+1}条规则{side}缺失start_col/end_col")
    
    # 根文件夹配置
    email_conf = config["email_config"]
    email_conf.setdefault("root_folder", "INBOX")  # 支持中文/编码名
    email_conf.setdefault("exclude_subfolders", ["临时", "测试", "垃圾", "deleted"])  # 排除关键词（中文）
    email_conf.setdefault("temp_dir", "./temp_attachments")
    
    print(f"✅ 加载配置成功，共{len(config['rules'])}条规则")
    print(f"🔧 根文件夹配置：{email_conf['root_folder']}")
    print(f"🚫 排除子文件夹关键词：{email_conf['exclude_subfolders']}")
    return config

# ===================== 递归指定根文件夹的子文件夹下载附件 =====================
def search_attach_in_folder(mail: imaplib.IMAP4_SSL, folder_name: str, search_criteria: str, match_key: str, suffix: str, all_folders: List[Tuple[str, str]]) -> List[Tuple[int, str, bytes, str]]:
    """
    在单个子文件夹中搜索匹配的附件
    folder_name：文件夹编码名
    返回：[(邮件相对新旧度, 附件名, 附件数据, 文件夹中文名), ...]
    """
    attach_list = []
    # 获取文件夹中文名（用于日志）
    folder_decoded = next((d for e, d in all_folders if e == folder_name), folder_name)
    
    try:
        # 选择文件夹（readonly=True 防止修改邮件状态）
        encoded_folder = encode_imap_folder(folder_name)
        status, _ = mail.select(encoded_folder, readonly=True)
        if status != "OK":
            print(f"⚠️  无法访问子文件夹：{folder_decoded} / {folder_name}，跳过")
            return attach_list
        
        # 执行邮件搜索
        status, messages = mail.search(None, search_criteria)
        if status != "OK" or not messages[0]:
            return attach_list
        
        # 遍历该子文件夹下的邮件（最新到最旧）
        email_ids = messages[0].split()
        for idx, eid in enumerate(reversed(email_ids)):
            status, msg_data = mail.fetch(eid, "(RFC822)")
            if status != "OK":
                continue
            
            # 解析邮件
            msg = email.message_from_bytes(msg_data[0][1])
            subject = decode_email_header(msg["Subject"])
            
            # 遍历附件
            for part in msg.walk():
                if part.get_content_maintype() == "multipart" or not part.get("Content-Disposition"):
                    continue
                
                # 解析附件名并匹配条件
                filename = decode_email_header(part.get_filename())
                if not filename or not filename.endswith(suffix) or match_key not in filename:
                    continue
                
                # 获取附件数据
                attach_data = part.get_payload(decode=True)
                # idx越小表示邮件越新（reversed遍历）
                attach_list.append((idx, filename, attach_data, folder_decoded))
                print(f"   📦 找到匹配附件：[{folder_decoded}] {filename}（邮件主题：{subject}）")
    
    except Exception as e:
        print(f"⚠️  处理子文件夹{folder_decoded} / {folder_name}出错：{e}")
    return attach_list

def download_latest_attach_from_root(rule: Dict[str, Any], email_config: Dict[str, Any]) -> Optional[str]:
    """
    递归遍历指定根文件夹下的所有子文件夹，下载匹配的最新附件
    """
    match_key = rule["attach_rule"]["match_key"]
    suffix = rule["attach_rule"]["suffix"]
    temp_dir = email_config["temp_dir"]
    all_attach_candidates = []

    try:
        # 邮箱连接
        mail = imaplib.IMAP4_SSL(email_config["imap_server"])
        mail.login(email_config["account"], email_config["password"])

        # 1. 构造搜索条件
        search_criteria = email_config.get("search_criteria", "")
        if search_criteria.upper() == "TODAY" or not search_criteria:
            search_criteria = get_today_imap_criteria()
            print(f"\n🔍 筛选当日邮件，搜索条件：{search_criteria}")

        # 2. 获取所有文件夹（编码名+中文名）
        all_folders = parse_imap_folders(mail)
        
        # 3. 筛选根文件夹下的子文件夹（编码名）
        root_folder = email_config["root_folder"]
        target_encoded_folders = get_root_folder_children(all_folders, root_folder)
        
        # 4. 过滤不需要的子文件夹
        filtered_encoded_folders = filter_folders(target_encoded_folders, email_config["exclude_subfolders"], all_folders)

        # 5. 遍历根文件夹的子文件夹搜索附件
        print(f"\n🚀 开始递归遍历{len(filtered_encoded_folders)}个子文件夹查找附件（匹配关键词：{match_key}，后缀：{suffix}）")
        for encoded_folder in filtered_encoded_folders:
            folder_attach = search_attach_in_folder(mail, encoded_folder, search_criteria, match_key, suffix, all_folders)
            all_attach_candidates.extend(folder_attach)

        # 6. 处理搜索结果
        if not all_attach_candidates:
            root_decoded = next((d for e, d in all_folders if e == root_folder), root_folder)
            print(f"\n❌ 根文件夹「{root_decoded}」及其子文件夹中未找到匹配【{match_key}】的附件")
            mail.close()
            mail.logout()
            return None
        
        # 按邮件新旧排序（idx越小越新），取最新的附件
        all_attach_candidates.sort(key=lambda x: x[0])
        _, filename, payload, folder_decoded = all_attach_candidates[0]
        
        # 清理附件名非法字符
        filename = filename.replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_")
        save_path = os.path.join(temp_dir, filename)
        
        # 保存附件
        os.makedirs(temp_dir, exist_ok=True)
        with open(save_path, "wb") as f:
            f.write(payload)
        
        print(f"\n✅ 找到最新匹配附件：")
        print(f"   📂 所在子文件夹：{folder_decoded}")
        print(f"   📄 附件名称：{filename}")
        print(f"   💾 保存路径：{save_path}")

        mail.close()
        mail.logout()
        return save_path
    
    except Exception as e:
        print(f"\n❌ 附件下载失败：{str(e)}")
        try:
            mail.close()
            mail.logout()
        except:
            pass
        return None

# ===================== 数据提取/替换（保留原有逻辑） =====================
def extract_cols_from_attach(attach_path: str, attach_rule: Dict[str, Any]) -> Dict[int, List[Any]]:
    """提取附件中连续列的数据"""
    sheet_name = attach_rule["sheet"]
    header_row = attach_rule["header_row"]
    cols_range = attach_rule["cols_range"]
    try:
        wb = openpyxl.load_workbook(attach_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"附件无Sheet：{sheet_name}")
        ws = wb[sheet_name]

        col_indexes = get_cols_range(ws, cols_range["start_col"], cols_range["end_col"], header_row)
        print(f"\n📊 提取附件列（共{len(col_indexes)}列）：")
        for idx in col_indexes:
            col_name = ws.cell(row=header_row, column=idx).value
            print(f"   列{get_column_letter(idx)}：{clean_text(col_name)}")

        extract_data = {}
        data_start_row = header_row + 1
        for col_idx in col_indexes:
            col_data = []
            for row in range(data_start_row, ws.max_row + 1):
                col_data.append(ws.cell(row=row, column=col_idx).value)
            extract_data[col_idx] = col_data
            print(f"   列{get_column_letter(col_idx)}提取{len(col_data)}行数据")

        wb.close()
        return extract_data
    except Exception as e:
        print(f"\n❌ 提取附件数据失败：{str(e)}")
        return {}

def replace_local_cols(extract_data: Dict[int, List[Any]], local_rule: Dict[str, Any]) -> bool:
    """批量替换本地文件的连续列"""
    if not extract_data:
        print("❌ 无提取数据，跳过替换")
        return False
    file_path = local_rule["file_path"]
    sheet_name = local_rule["sheet"]
    header_row = local_rule["header_row"]
    cols_range = local_rule["cols_range"]
    output_path = local_rule["output_path"]

    if not os.path.exists(file_path):
        print(f"❌ 本地文件不存在：{file_path}")
        return False

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"本地文件无Sheet：{sheet_name}")
        ws = wb[sheet_name]

        local_col_indexes = get_cols_range(ws, cols_range["start_col"], cols_range["end_col"], header_row)
        if len(local_col_indexes) != len(extract_data):
            raise ValueError(f"附件列数({len(extract_data)})与本地列数({len(local_col_indexes)})不匹配")
        print(f"\n📝 替换本地列（共{len(local_col_indexes)}列）：")
        for idx in local_col_indexes:
            col_name = ws.cell(row=header_row, column=idx).value
            print(f"   列{get_column_letter(idx)}：{clean_text(col_name)}")

        data_start_row = header_row + 1
        for i, (attach_col_idx, local_col_idx) in enumerate(zip(extract_data.keys(), local_col_indexes)):
            col_data = extract_data[attach_col_idx]
            col_letter = get_column_letter(local_col_idx)
            for row_idx, value in enumerate(col_data):
                ws[f"{col_letter}{data_start_row + row_idx}"].value = value
            print(f"   ✅ 列{col_letter}替换{len(col_data)}行数据")

            # 公式下拉
            formula_cols = []
            for col in ws.iter_cols(min_row=data_start_row, max_row=data_start_row):
                cell = col[0]
                if cell.column > local_col_idx and cell.value and str(cell.value).startswith("="):
                    formula_cols.append(cell.column)
            if formula_cols:
                for f_col_idx in formula_cols:
                    f_col_letter = get_column_letter(f_col_idx)
                    formula = ws[f"{f_col_letter}{data_start_row}"].value
                    for row in range(data_start_row + 1, data_start_row + len(col_data)):
                        ws[f"{f_col_letter}{row}"].value = formula
                print(f"   📌 列{col_letter}后续{len(formula_cols)}列公式下拉完成")

        # 删除多余行
        max_data_row = data_start_row + len(list(extract_data.values())[0]) - 1
        if ws.max_row > max_data_row:
            delete_count = ws.max_row - max_data_row
            for row in range(ws.max_row, max_data_row, -1):
                ws.delete_rows(row)
            print(f"\n🗑️ 删除{delete_count}行多余数据")

        wb.save(output_path)
        wb.close()
        print(f"\n🎉 本地文件处理完成 → {output_path}")
        return True
    except Exception as e:
        print(f"\n❌ 替换本地文件失败：{str(e)}")
        return False

# ===================== 主流程 =====================
def main(config_path: str = "./config.json"):
    try:
        # 1. 加载配置
        config = load_json_config(config_path)
        email_config = config["email_config"]
        rules = config["rules"]
        other = config["other"]
        temp_dir = other["temp_dir"]
        os.makedirs(temp_dir, exist_ok=True)

        # 2. 遍历规则处理
        success_count = 0
        for i, rule in enumerate(rules):
            print(f"\n==================================================")
            print(f"📋 处理第{i+1}/{len(rules)}条规则：匹配【{rule['attach_rule']['match_key']}】")
            
            # 步骤1：递归根文件夹的子文件夹下载最新附件
            attach_path = download_latest_attach_from_root(rule, email_config)
            if not attach_path:
                print(f"❌ 第{i+1}条规则：附件下载失败")
                continue
            
            # 步骤2：提取多列数据
            extract_data = extract_cols_from_attach(attach_path, rule["attach_rule"])
            if not extract_data:
                print(f"❌ 第{i+1}条规则：数据提取失败")
                continue
            
            # 步骤3：替换本地多列
            if replace_local_cols(extract_data, rule["local_rule"]):
                success_count += 1
                print(f"✅ 第{i+1}条规则：处理成功")
            else:
                print(f"❌ 第{i+1}条规则：本地替换失败")

        # 3. 清理临时目录
        if other.get("clean_temp", True) and os.path.exists(temp_dir):
            import shutil
            shutil.rmtree(temp_dir)
            print(f"\n🗑️ 清理临时目录：{temp_dir}")

        # 4. 结果汇总
        print(f"\n==================================================")
        print(f"📊 总处理结果：成功{success_count}/{len(rules)}条规则")
        if success_count == len(rules):
            print(f"🎉 所有规则执行成功！")

    except Exception as e:
        print(f"\n❌ 程序执行失败：{str(e)}")

if __name__ == "__main__":
    main(config_path="D:\\供应链\\芯片\\config.json")