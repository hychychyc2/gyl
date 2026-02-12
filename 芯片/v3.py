import imaplib
import email
from email.header import decode_header
import os
import json
import openpyxl
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Tuple, Optional, Any, Union
import base64
import re

# ===================== IMAP UTF-7编解码核心函数 =====================
def imap_utf7_decode(s: str) -> str:
    if not s or "&" not in s:
        return s
    pattern = re.compile(r'&([^-]+)-')
    def decode_match(match):
        encoded_part = match.group(1).replace(",", "/")
        if not encoded_part:
            return "&"
        try:
            decoded = base64.b64decode(encoded_part + "==", altchars=b"+/").decode("utf-16be")
            return decoded
        except:
            return match.group(0)
    return pattern.sub(decode_match, s)

def imap_utf7_encode(s: str) -> str:
    if not s or all(ord(c) < 128 for c in s):
        return s
    result = []
    buffer = []
    for c in s:
        if ord(c) < 128:
            if buffer:
                b64 = base64.b64encode(''.join(buffer).encode("utf-16be")).decode("ascii").rstrip("=").replace("/", ",")
                result.append(f"&{b64}-")
                buffer = []
            result.append(c)
        else:
            buffer.append(c)
    if buffer:
        b64 = base64.b64encode(''.join(buffer).encode("utf-16be")).decode("ascii").rstrip("=").replace("/", ",")
        result.append(f"&{b64}-")
    return ''.join(result)

# ===================== 日期工具函数 =====================
def get_imap_date_str(date_obj: datetime) -> str:
    month_abbr = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{date_obj.day:02d}-{month_abbr[date_obj.month-1]}-{date_obj.year}"

def get_today_imap_criteria() -> str:
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    today_str = get_imap_date_str(datetime(today.year, today.month, today.day))
    tomorrow_str = get_imap_date_str(datetime(tomorrow.year, tomorrow.month, tomorrow.day))
    return f'SINCE "{today_str}" BEFORE "{tomorrow_str}"'

# ===================== 核心工具函数 =====================
def decode_email_header(header: Any) -> str:
    if not header:
        return ""
    decoded_parts = []
    for part, encoding in decode_header(header):
        if isinstance(part, bytes):
            try:
                decoded_part = part.decode(encoding or "gbk")
            except (UnicodeDecodeError, LookupError):
                try:
                    decoded_part = part.decode("utf-8")
                except UnicodeDecodeError:
                    decoded_part = part.decode("latin-1")
        else:
            decoded_part = str(part)
        decoded_parts.append(decoded_part)
    return "".join(decoded_parts).strip()

def encode_imap_folder(folder_name: str) -> bytes:
    if not folder_name:
        return b"INBOX"
    imap_encoded = imap_utf7_encode(folder_name)
    for encoding in ["utf-8", "gbk"]:
        try:
            return imap_encoded.encode(encoding)
        except:
            continue
    return imap_encoded.encode("latin-1")

def parse_imap_folders(mail: imaplib.IMAP4_SSL) -> List[Tuple[str, str]]:
    try:
        status, folders_raw = mail.list()
        if status != "OK":
            print("⚠️  获取文件夹列表失败")
            return [("INBOX", "INBOX")]
        
        folder_list = []
        for folder in folders_raw:
            if not folder:
                continue
            folder_str = folder.decode("utf-8", errors="ignore")
            parts = folder_str.split('"')
            if len(parts) < 3:
                continue
            encoded_name = parts[-2].strip()
            if not encoded_name or encoded_name in [f[0] for f in folder_list]:
                continue
            decoded_name = imap_utf7_decode(encoded_name)
            folder_list.append((encoded_name, decoded_name))
        
        folder_list = list(dict.fromkeys(folder_list))
        folder_list.sort(key=lambda x: (x[0] != "INBOX", x[0].lower()))
        print(f"\n📂 解析到邮箱所有文件夹（共{len(folder_list)}个）：")
        for i, (encoded, decoded) in enumerate(folder_list):
            print(f"   [{i+1}] 编码名：{encoded:<20} 中文名：{decoded}")
        return folder_list
    except Exception as e:
        print(f"⚠️  解析文件夹失败：{e}，仅使用收件箱")
        return [("INBOX", "INBOX")]

def get_root_folder_children(folder_list: List[Tuple[str, str]], root_folder: str) -> List[str]:
    target_encoded_root = None
    root_lower = root_folder.lower()
    
    for encoded, decoded in folder_list:
        if decoded.lower() == root_lower or encoded.lower() == root_lower:
            target_encoded_root = encoded
            root_decoded = decoded
            break
    
    if not target_encoded_root:
        print(f"⚠️  未找到根文件夹「{root_folder}」，默认使用INBOX")
        target_encoded_root = "INBOX"
        root_decoded = "INBOX"
    
    target_encoded_folders = []
    for encoded, decoded in folder_list:
        if encoded == target_encoded_root or encoded.startswith(f"{target_encoded_root}/"):
            target_encoded_folders.append((encoded, decoded))
    
    if not target_encoded_folders:
        print(f"⚠️  未找到根文件夹「{root_decoded}」及其子文件夹，默认使用INBOX")
        target_encoded_folders = [("INBOX", "INBOX")]
    else:
        target_encoded_folders.sort(key=lambda x: (x[0] != target_encoded_root, x[0].count("/"), x[0].lower()))
    
    print(f"\n🎯 待递归的文件夹（根：{root_decoded} / {target_encoded_root}）：")
    for i, (encoded, decoded) in enumerate(target_encoded_folders):
        level = encoded.count("/") - target_encoded_root.count("/")
        prefix = "  " * level + "└─ " if level > 0 else "┌─ "
        print(f"   {prefix}中文名：{decoded:<20} 编码名：{encoded}")
    
    return [encoded for encoded, decoded in target_encoded_folders]

def filter_folders(folder_list: List[str], exclude_folders: List[str], all_folders: List[Tuple[str, str]]) -> List[str]:
    if not exclude_folders:
        return folder_list
    
    filtered = []
    exclude_lower = [f.lower() for f in exclude_folders]
    
    for encoded in folder_list:
        decoded = next((d for e, d in all_folders if e == encoded), encoded)
        if any(ex in decoded.lower() or ex in encoded.lower() for ex in exclude_lower):
            print(f"🚫 排除子文件夹：{decoded} / {encoded}（匹配排除关键词）")
            continue
        filtered.append(encoded)
    
    print(f"\n🔍 过滤后最终递归的文件夹（共{len(filtered)}个）：")
    for encoded in filtered:
        decoded = next((d for e, d in all_folders if e == encoded), encoded)
        print(f"   中文名：{decoded:<20} 编码名：{encoded}")
    return filtered

def clean_text(text: Any) -> str:
    if isinstance(text, bytes):
        try:
            return text.decode("gbk")
        except:
            return text.decode("utf-8", errors="ignore")
    return str(text).replace("\r", "").replace("\n", "").replace("\t", "").strip()

def get_col_index(ws, col_identifier: str, header_row: int) -> int:
    col_identifier = clean_text(col_identifier)
    try:
        return column_index_from_string(col_identifier)
    except:
        pass
    for col in ws.iter_cols(min_row=header_row, max_row=header_row):
        for cell in col:
            if clean_text(cell.value) == col_identifier:
                return cell.column
    raise ValueError(f"未找到列：{col_identifier}")

def get_cols_range(ws, start_col: str, end_col: str, header_row: int) -> List[int]:
    start_idx = get_col_index(ws, start_col, header_row)
    end_idx = get_col_index(ws, end_col, header_row)
    if start_idx > end_idx:
        start_idx, end_idx = end_idx, start_idx
    return list(range(start_idx, end_idx + 1))

# ===================== 新增：数据筛选/去重核心函数 =====================
def parse_filter_conditions(filter_conditions: List[Dict[str, Any]], ws, header_row: int) -> List[Dict[str, Any]]:
    """解析筛选条件，转换列标识为列索引"""
    parsed_conditions = []
    if not filter_conditions:
        return parsed_conditions
    
    print(f"\n🔍 解析数据源筛选条件（共{len(filter_conditions)}条）：")
    for i, cond in enumerate(filter_conditions):
        # 校验必填项
        required = ["col", "operator", "value"]
        for k in required:
            if k not in cond:
                raise ValueError(f"第{i+1}条筛选条件缺失{k}字段")
        
        # 转换列标识为列索引
        col_idx = get_col_index(ws, cond["col"], header_row)
        col_letter = get_column_letter(col_idx)
        col_name = clean_text(ws.cell(row=header_row, column=col_idx).value)
        
        # 转换值类型（数字/字符串）
        value = cond["value"]
        try:
            if isinstance(value, str) and value.replace(".", "").isdigit():
                value = float(value) if "." in value else int(value)
        except:
            pass
        
        # 支持区间值（如 [10,20] 或 "10,20"）
        if cond["operator"] in ["between", "in"] and isinstance(value, (str, list)):
            if isinstance(value, str):
                value = [v.strip() for v in value.split(",")]
            # 转换区间值类型
            for j, v in enumerate(value):
                try:
                    value[j] = float(v) if "." in str(v) else int(v)
                except:
                    pass
        
        parsed_cond = {
            "col_idx": col_idx,
            "col_letter": col_letter,
            "col_name": col_name,
            "operator": cond["operator"].lower(),
            "value": value
        }
        parsed_conditions.append(parsed_cond)
        print(f"   [{i+1}] 列{col_letter}({col_name}) {parsed_cond['operator']} {parsed_cond['value']}")
    
    return parsed_conditions

def filter_data_rows(ws, header_row: int, filter_conditions: List[Dict[str, Any]]) -> List[int]:
    """根据筛选条件过滤数据行，返回符合条件的行索引"""
    if not filter_conditions:
        # 无筛选条件，返回所有数据行
        data_rows = list(range(header_row + 1, ws.max_row + 1))
        print(f"\n📊 无筛选条件，保留所有{len(data_rows)}行数据")
        return data_rows
    
    # 逐行校验筛选条件
    valid_rows = []
    data_start_row = header_row + 1
    operator_map = {
        "eq": lambda x, y: clean_text(x) == clean_text(y),
        "ne": lambda x, y: clean_text(x) != clean_text(y),
        "gt": lambda x, y: float(x) > float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "lt": lambda x, y: float(x) < float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "ge": lambda x, y: float(x) >= float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "le": lambda x, y: float(x) <= float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "contains": lambda x, y: clean_text(y) in clean_text(x),
        "not_contains": lambda x, y: clean_text(y) not in clean_text(x),
        "between": lambda x, y: float(y[0]) <= float(x) <= float(y[1]) if str(x).replace(".", "").isdigit() and all(str(v).replace(".", "").isdigit() for v in y) else False,
        "in": lambda x, y: clean_text(x) in [clean_text(v) for v in y],
        "not_in": lambda x, y: clean_text(x) not in [clean_text(v) for v in y]
    }
    
    print(f"\n📝 开始筛选数据行（总行数：{ws.max_row - header_row}）：")
    for row in range(data_start_row, ws.max_row + 1):
        row_valid = True
        for cond in filter_conditions:
            cell_value = ws.cell(row=row, column=cond["col_idx"]).value
            operator_func = operator_map.get(cond["operator"])
            
            if not operator_func:
                raise ValueError(f"不支持的运算符：{cond['operator']}")
            
            if not operator_func(cell_value, cond["value"]):
                row_valid = False
                break
        
        if row_valid:
            valid_rows.append(row)
    
    print(f"✅ 筛选完成，保留{len(valid_rows)}行符合条件的数据（总行数：{ws.max_row - header_row}）")
    return valid_rows

def deduplicate_data_rows(ws, header_row: int, deduplicate_cols: List[str], start_row: int) -> int:
    """根据指定列去重数据行，返回删除的行数"""
    if not deduplicate_cols:
        return 0
    
    # 转换去重列标识为列索引
    dedup_col_indexes = []
    print(f"\n🆔 开始数据去重（依据列：{deduplicate_cols}）：")
    for col_id in deduplicate_cols:
        col_idx = get_col_index(ws, col_id, header_row)
        col_letter = get_column_letter(col_idx)
        col_name = clean_text(ws.cell(row=header_row, column=col_idx).value)
        dedup_col_indexes.append(col_idx)
        print(f"   去重列：列{col_letter}({col_name})")
    
    # 记录已存在的行数据，去重
    seen = set()
    delete_rows = []
    data_row_count = 0
    
    for row in range(start_row, ws.max_row + 1):
        # 拼接去重列的值作为唯一标识
        row_key = tuple(clean_text(ws.cell(row=row, column=col_idx).value) for col_idx in dedup_col_indexes)
        data_row_count += 1
        
        if row_key in seen:
            delete_rows.append(row)
        else:
            seen.add(row_key)
    
    # 倒序删除重复行（避免行索引错乱）
    delete_count = len(delete_rows)
    if delete_count > 0:
        for row in reversed(delete_rows):
            ws.delete_rows(row)
        print(f"🗑️ 删除{delete_count}行重复数据（剩余{data_row_count - delete_count}行唯一数据）")
    else:
        print(f"✅ 无重复数据，无需删除")
    
    return delete_count

def load_json_config(config_path: str = "./config.json") -> Dict[str, Any]:
    """加载配置（新增筛选/写入/去重配置校验）"""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"配置文件不存在：{config_path}")
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON格式错误：{str(e)}")

    # 核心校验
    required = ["email_config", "rules", "other"]
    for k in required:
        if k not in config:
            raise KeyError(f"缺失核心配置：{k}")
    if not config["rules"]:
        raise ValueError("规则列表不能为空")
    
    # 遍历规则，补充新增配置默认值并校验
    for i, rule in enumerate(config["rules"]):
        # 1. 附件筛选条件默认值
        rule["attach_rule"].setdefault("filter_conditions", [])
        
        # 2. 本地写入配置默认值
        local_rule = rule["local_rule"]
        local_rule.setdefault("write_mode", "overwrite")  # overwrite/append
        local_rule.setdefault("deduplicate", False)       # 是否去重
        local_rule.setdefault("deduplicate_cols", [])     # 去重依据列
        
        # 原有列范围校验
        if "cols_range" not in rule["attach_rule"] or "cols_range" not in rule["local_rule"]:
            raise KeyError(f"第{i+1}条规则缺失cols_range配置")
        for side in ["attach_rule", "local_rule"]:
            cr = rule[side]["cols_range"]
            if "start_col" not in cr or "end_col" not in cr:
                raise KeyError(f"第{i+1}条规则{side}缺失start_col/end_col")
        
        # 新增配置日志
        print(f"\n📋 第{i+1}条规则新增配置：")
        print(f"   筛选条件：{rule['attach_rule']['filter_conditions'] or '无'}")
        print(f"   写入模式：{local_rule['write_mode']}")
        print(f"   去重开关：{local_rule['deduplicate']}")
        print(f"   去重列：{local_rule['deduplicate_cols'] or '无'}")
    
    # 邮箱配置默认值
    email_conf = config["email_config"]
    email_conf.setdefault("root_folder", "INBOX")
    email_conf.setdefault("exclude_subfolders", ["临时", "测试", "垃圾", "deleted"])
    email_conf.setdefault("temp_dir", "./temp_attachments")
    
    print(f"\n✅ 加载配置成功，共{len(config['rules'])}条规则")
    print(f"🔧 根文件夹配置：{email_conf['root_folder']}")
    print(f"🚫 排除子文件夹关键词：{email_conf['exclude_subfolders']}")
    return config

# ===================== 附件下载（原有逻辑） =====================
def search_attach_in_folder(mail: imaplib.IMAP4_SSL, folder_name: str, search_criteria: str, match_key: str, suffix: str, all_folders: List[Tuple[str, str]]) -> List[Tuple[int, str, bytes, str]]:
    attach_list = []
    folder_decoded = next((d for e, d in all_folders if e == folder_name), folder_name)
    
    try:
        encoded_folder = encode_imap_folder(folder_name)
        status, _ = mail.select(encoded_folder, readonly=True)
        if status != "OK":
            print(f"⚠️  无法访问子文件夹：{folder_decoded} / {folder_name}，跳过")
            return attach_list
        
        status, messages = mail.search(None, search_criteria)
        if status != "OK" or not messages[0]:
            return attach_list
        
        email_ids = messages[0].split()
        for idx, eid in enumerate(reversed(email_ids)):
            status, msg_data = mail.fetch(eid, "(RFC822)")
            if status != "OK":
                continue
            
            msg = email.message_from_bytes(msg_data[0][1])
            subject = decode_email_header(msg["Subject"])
            
            for part in msg.walk():
                if part.get_content_maintype() == "multipart" or not part.get("Content-Disposition"):
                    continue
                
                filename = decode_email_header(part.get_filename())
                if not filename or not filename.endswith(suffix) or match_key not in filename:
                    continue
                
                attach_data = part.get_payload(decode=True)
                attach_list.append((idx, filename, attach_data, folder_decoded))
                print(f"   📦 找到匹配附件：[{folder_decoded}] {filename}（邮件主题：{subject}）")
    
    except Exception as e:
        print(f"⚠️  处理子文件夹{folder_decoded} / {folder_name}出错：{e}")
    return attach_list

def download_latest_attach_from_root(rule: Dict[str, Any], email_config: Dict[str, Any]) -> Optional[str]:
    match_key = rule["attach_rule"]["match_key"]
    suffix = rule["attach_rule"]["suffix"]
    temp_dir = email_config["temp_dir"]
    all_attach_candidates = []

    try:
        mail = imaplib.IMAP4_SSL(email_config["imap_server"])
        mail.login(email_config["account"], email_config["password"])

        search_criteria = email_config.get("search_criteria", "")
        if search_criteria.upper() == "TODAY" or not search_criteria:
            search_criteria = get_today_imap_criteria()
            print(f"\n🔍 筛选当日邮件，搜索条件：{search_criteria}")

        all_folders = parse_imap_folders(mail)
        root_folder = email_config["root_folder"]
        target_encoded_folders = get_root_folder_children(all_folders, root_folder)
        filtered_encoded_folders = filter_folders(target_encoded_folders, email_config["exclude_subfolders"], all_folders)

        print(f"\n🚀 开始递归遍历{len(filtered_encoded_folders)}个子文件夹查找附件（匹配关键词：{match_key}，后缀：{suffix}）")
        for encoded_folder in filtered_encoded_folders:
            folder_attach = search_attach_in_folder(mail, encoded_folder, search_criteria, match_key, suffix, all_folders)
            all_attach_candidates.extend(folder_attach)

        if not all_attach_candidates:
            root_decoded = next((d for e, d in all_folders if e == root_folder), root_folder)
            print(f"\n❌ 根文件夹「{root_decoded}」及其子文件夹中未找到匹配【{match_key}】的附件")
            mail.close()
            mail.logout()
            return None
        
        all_attach_candidates.sort(key=lambda x: x[0])
        _, filename, payload, folder_decoded = all_attach_candidates[0]
        
        filename = filename.replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_")
        save_path = os.path.join(temp_dir, filename)
        
        os.makedirs(temp_dir, exist_ok=True)
        with open(save_path, "wb") as f:
            f.write(payload)
        
        print(f"\n✅ 找到最新匹配附件：")
        print(f"   📂 所在子文件夹：{folder_decoded}")
        print(f"   📄 附件名称：{filename}")
        print(f"   💾 保存路径：{save_path}")

        mail.close()
        mail.logout()
        return save_path
    
    except Exception as e:
        print(f"\n❌ 附件下载失败：{str(e)}")
        try:
            mail.close()
            mail.logout()
        except:
            pass
        return None

# ===================== 数据提取（新增筛选逻辑） =====================
def extract_cols_from_attach(attach_path: str, attach_rule: Dict[str, Any]) -> Dict[int, List[Any]]:
    """提取附件数据（新增筛选逻辑）"""
    sheet_name = attach_rule["sheet"]
    header_row = attach_rule["header_row"]
    cols_range = attach_rule["cols_range"]
    filter_conditions = attach_rule["filter_conditions"]
    
    try:
        wb = openpyxl.load_workbook(attach_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"附件无Sheet：{sheet_name}")
        ws = wb[sheet_name]

        # 1. 解析列范围
        col_indexes = get_cols_range(ws, cols_range["start_col"], cols_range["end_col"], header_row)
        print(f"\n📊 提取附件列（共{len(col_indexes)}列）：")
        for idx in col_indexes:
            col_name = ws.cell(row=header_row, column=idx).value
            print(f"   列{get_column_letter(idx)}：{clean_text(col_name)}")

        # 2. 解析并执行筛选条件
        parsed_conditions = parse_filter_conditions(filter_conditions, ws, header_row)
        valid_rows = filter_data_rows(ws, header_row, parsed_conditions)
        if not valid_rows:
            print("❌ 无符合筛选条件的数据行")
            wb.close()
            return {}

        # 3. 提取筛选后的数据
        extract_data = {}
        for col_idx in col_indexes:
            col_data = []
            for row in valid_rows:
                col_data.append(ws.cell(row=row, column=col_idx).value)
            extract_data[col_idx] = col_data
            print(f"   列{get_column_letter(col_idx)}提取{len(col_data)}行筛选后数据")

        wb.close()
        return extract_data
    except Exception as e:
        print(f"\n❌ 提取附件数据失败：{str(e)}")
        return {}

# ===================== 数据替换（新增追加/覆盖/去重逻辑） =====================
def replace_local_cols(extract_data: Dict[int, List[Any]], local_rule: Dict[str, Any]) -> bool:
    """替换本地数据（新增追加/覆盖/去重逻辑）"""
    if not extract_data:
        print("❌ 无提取数据，跳过替换")
        return False
    
    file_path = local_rule["file_path"]
    sheet_name = local_rule["sheet"]
    header_row = local_rule["header_row"]
    cols_range = local_rule["cols_range"]
    output_path = local_rule["output_path"]
    write_mode = local_rule["write_mode"].lower()  # overwrite/append
    deduplicate = local_rule["deduplicate"]
    deduplicate_cols = local_rule["deduplicate_cols"]

    if not os.path.exists(file_path):
        print(f"❌ 本地文件不存在：{file_path}")
        return False

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"本地文件无Sheet：{sheet_name}")
        ws = wb[sheet_name]

        # 解析目标列范围
        local_col_indexes = get_cols_range(ws, cols_range["start_col"], cols_range["end_col"], header_row)
        if len(local_col_indexes) != len(extract_data):
            raise ValueError(f"附件列数({len(extract_data)})与本地列数({len(local_col_indexes)})不匹配")
        print(f"\n📝 目标列配置（共{len(local_col_indexes)}列）：")
        for idx in local_col_indexes:
            col_name = ws.cell(row=header_row, column=idx).value
            print(f"   列{get_column_letter(idx)}：{clean_text(col_name)}")

        # 1. 确定数据写入起始行
        if write_mode == "overwrite":
            # 覆盖模式：从header_row+1开始写入，先清空原有数据
            write_start_row = header_row + 1
            # 删除原有数据行
            if ws.max_row > header_row:
                delete_count = ws.max_row - header_row
                for row in range(ws.max_row, header_row, -1):
                    ws.delete_rows(row)
                print(f"🗑️ 覆盖模式：删除原有{delete_count}行数据")
        else:
            # 追加模式：从现有数据最后一行+1开始写入
            write_start_row = ws.max_row + 1
            print(f"📥 追加模式：从第{write_start_row}行开始写入数据")

        # 2. 写入数据
        data_rows_count = len(list(extract_data.values())[0])
        for i, (attach_col_idx, local_col_idx) in enumerate(zip(extract_data.keys(), local_col_indexes)):
            col_data = extract_data[attach_col_idx]
            col_letter = get_column_letter(local_col_idx)
            for row_idx, value in enumerate(col_data):
                target_row = write_start_row + row_idx
                ws[f"{col_letter}{target_row}"].value = value
            print(f"   ✅ 列{col_letter}写入{len(col_data)}行数据（起始行：{write_start_row}）")

            # 公式下拉（仅覆盖模式生效）
            if write_mode == "overwrite":
                formula_cols = []
                for col in ws.iter_cols(min_row=write_start_row, max_row=write_start_row):
                    cell = col[0]
                    if cell.column > local_col_idx and cell.value and str(cell.value).startswith("="):
                        formula_cols.append(cell.column)
                if formula_cols:
                    for f_col_idx in formula_cols:
                        f_col_letter = get_column_letter(f_col_idx)
                        formula = ws[f"{f_col_letter}{write_start_row}"].value
                        for row in range(write_start_row + 1, write_start_row + len(col_data)):
                            ws[f"{f_col_letter}{row}"].value = formula
                    print(f"   📌 列{col_letter}后续{len(formula_cols)}列公式下拉完成")

        # 3. 数据去重（最终数据行）
        if deduplicate and deduplicate_cols:
            deduplicate_data_rows(ws, header_row, deduplicate_cols, header_row + 1)

        # 4. 保存文件
        wb.save(output_path)
        wb.close()
        print(f"\n🎉 本地文件处理完成 → {output_path}")
        print(f"📋 处理总结：")
        print(f"   写入模式：{write_mode}")
        print(f"   写入行数：{data_rows_count}")
        print(f"   去重开关：{deduplicate}（依据列：{deduplicate_cols or '无'}）")
        return True
    except Exception as e:
        print(f"\n❌ 替换本地文件失败：{str(e)}")
        return False

# ===================== 主流程 =====================
def main(config_path: str = "./config.json"):
    try:
        config = load_json_config(config_path)
        email_config = config["email_config"]
        rules = config["rules"]
        other = config["other"]
        temp_dir = other["temp_dir"]
        os.makedirs(temp_dir, exist_ok=True)

        success_count = 0
        for i, rule in enumerate(rules):
            print(f"\n==================================================")
            print(f"📋 处理第{i+1}/{len(rules)}条规则：匹配【{rule['attach_rule']['match_key']}】")
            
            attach_path = download_latest_attach_from_root(rule, email_config)
            if not attach_path:
                print(f"❌ 第{i+1}条规则：附件下载失败")
                continue
            
            extract_data = extract_cols_from_attach(attach_path, rule["attach_rule"])
            if not extract_data:
                print(f"❌ 第{i+1}条规则：数据提取失败")
                continue
            
            if replace_local_cols(extract_data, rule["local_rule"]):
                success_count += 1
                print(f"✅ 第{i+1}条规则：处理成功")
            else:
                print(f"❌ 第{i+1}条规则：本地替换失败")

        if other.get("clean_temp", True) and os.path.exists(temp_dir):
            import shutil
            shutil.rmtree(temp_dir)
            print(f"\n🗑️ 清理临时目录：{temp_dir}")

        print(f"\n==================================================")
        print(f"📊 总处理结果：成功{success_count}/{len(rules)}条规则")
        if success_count == len(rules):
            print(f"🎉 所有规则执行成功！")

    except Exception as e:
        print(f"\n❌ 程序执行失败：{str(e)}")

if __name__ == "__main__":
    main(config_path="D:\\供应链\\芯片\\config.json")