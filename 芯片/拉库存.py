import imaplib
import email
from email.header import decode_header
import os
import json
import openpyxl
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Tuple, Optional, Any

# ===================== 日期工具函数（新增当日筛选） =====================
def get_imap_date_str(date_obj: datetime) -> str:
    """
    转换为IMAP协议要求的日期格式（DD-Mon-YYYY，如22-Dec-2025）
    月份为3个字母的英文缩写，全部大写/首字母大写均可
    """
    # 月份缩写映射（IMAP兼容）
    month_abbr = [
        "Jan", "Feb", "Mar", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
    ]
    return f"{date_obj.day:02d}-{month_abbr[date_obj.month-1]}-{date_obj.year}"

def get_today_imap_criteria() -> str:
    """生成当日邮件的IMAP搜索条件（SINCE 今日 BEFORE 明日）"""
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    
    # 转换为IMAP日期格式
    today_str = get_imap_date_str(datetime(today.year, today.month, today.day))
    tomorrow_str = get_imap_date_str(datetime(tomorrow.year, tomorrow.month, tomorrow.day))
    
    # 构造搜索条件：当日（包含今日，不包含明日）
    return f'SINCE "{today_str}" BEFORE "{tomorrow_str}"'

# ===================== 核心工具函数（保留编码修复） =====================
def decode_email_header(header: Any) -> str:
    """兼容多编码解析邮件头（附件名/邮件主题）"""
    if not header:
        return ""
    decoded_parts = []
    for part, encoding in decode_header(header):
        if isinstance(part, bytes):
            # 尝试多种编码解码，优先级：GBK(中文) > UTF-8 > Latin-1(兜底)
            try:
                decoded_part = part.decode(encoding or "gbk")
            except (UnicodeDecodeError, LookupError):
                try:
                    decoded_part = part.decode("utf-8")
                except UnicodeDecodeError:
                    decoded_part = part.decode("latin-1")  # 兜底，不会报错
        else:
            decoded_part = str(part)
        decoded_parts.append(decoded_part)
    return "".join(decoded_parts).strip()

def load_json_config(config_path: str = "./config.json") -> Dict[str, Any]:
    """读取JSON配置"""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"配置文件不存在：{config_path}")
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON格式错误：{str(e)}")

    # 核心校验
    required = ["email_config", "rules", "other"]
    for k in required:
        if k not in config:
            raise KeyError(f"缺失核心配置：{k}")
    if not config["rules"]:
        raise ValueError("规则列表不能为空")
    
    # 列范围配置校验
    for i, rule in enumerate(config["rules"]):
        if "cols_range" not in rule["attach_rule"] or "cols_range" not in rule["local_rule"]:
            raise KeyError(f"第{i+1}条规则缺失cols_range配置")
        for side in ["attach_rule", "local_rule"]:
            cr = rule[side]["cols_range"]
            if "start_col" not in cr or "end_col" not in cr:
                raise KeyError(f"第{i+1}条规则{side}缺失start_col/end_col")
    print(f"✅ 加载配置成功，共{len(config['rules'])}条规则")
    return config

def clean_text(text: Any) -> str:
    """文本清理"""
    if isinstance(text, bytes):
        try:
            return text.decode("gbk")
        except:
            return text.decode("utf-8", errors="ignore")
    return str(text).replace("\r", "").replace("\n", "").replace("\t", "").strip()

def get_col_index(ws, col_identifier: str, header_row: int) -> int:
    """解析列标识为列索引"""
    col_identifier = clean_text(col_identifier)
    try:
        return column_index_from_string(col_identifier)
    except:
        pass
    for col in ws.iter_cols(min_row=header_row, max_row=header_row):
        for cell in col:
            if clean_text(cell.value) == col_identifier:
                return cell.column
    raise ValueError(f"未找到列：{col_identifier}")

def get_cols_range(ws, start_col: str, end_col: str, header_row: int) -> List[int]:
    """获取连续列的索引列表"""
    start_idx = get_col_index(ws, start_col, header_row)
    end_idx = get_col_index(ws, end_col, header_row)
    if start_idx > end_idx:
        start_idx, end_idx = end_idx, start_idx
    return list(range(start_idx, end_idx + 1))

# ===================== 附件下载（新增当日筛选） =====================
def download_latest_attach(rule: Dict[str, Any], email_config: Dict[str, Any]) -> Optional[str]:
    """下载当日匹配的最新附件"""
    match_key = rule["attach_rule"]["match_key"]
    suffix = rule["attach_rule"]["suffix"]
    temp_dir = email_config.get("temp_dir", "./temp_attachments")
    attach_candidates = []

    try:
        # 邮箱连接
        mail = imaplib.IMAP4_SSL(email_config["imap_server"])
        mail.login(email_config["account"], email_config["password"])
        mail.select("INBOX")

        # 🔥 核心修改：优先使用当日筛选条件，兼容配置中的自定义条件
        search_criteria = email_config.get("search_criteria", "")
        if search_criteria.upper() == "TODAY":
            # 配置为"TODAY"则自动筛选当日邮件
            search_criteria = get_today_imap_criteria()
            print(f"\n🔍 筛选当日邮件，搜索条件：{search_criteria}")
        elif not search_criteria:
            # 无配置则默认筛选当日
            search_criteria = get_today_imap_criteria()
            print(f"\n🔍 未配置搜索条件，默认筛选当日邮件：{search_criteria}")

        # 执行搜索
        status, messages = mail.search(None, search_criteria)
        if status != "OK":
            print("❌ 邮件搜索失败（可能日期格式错误）")
            return None

        # 无当日邮件直接返回
        if not messages[0]:
            print(f"\nℹ️  未找到当日（{get_imap_date_str(datetime.now())}）的邮件")
            return None

        # 遍历当日邮件（最新到最旧）
        email_ids = messages[0].split()
        print(f"\n📮 找到{len(email_ids)}封当日邮件，开始遍历...")
        
        for idx, eid in enumerate(reversed(email_ids)):
            status, msg_data = mail.fetch(eid, "(RFC822)")
            if status != "OK":
                continue
            
            # 解析邮件主题
            msg = email.message_from_bytes(msg_data[0][1])
            subject = decode_email_header(msg["Subject"])
            print(f"\n📧 检查当日邮件 {idx+1}/{len(email_ids)}：{subject}")

            # 遍历附件
            for part in msg.walk():
                if part.get_content_maintype() == "multipart" or not part.get("Content-Disposition"):
                    continue
                
                # 解析附件名（修复编码）
                filename = decode_email_header(part.get_filename())
                if not filename or not filename.endswith(suffix) or match_key not in filename:
                    continue
                
                # 获取附件二进制数据
                attach_data = part.get_payload(decode=True)
                attach_candidates.append((idx, filename, attach_data))
                print(f"   ✅ 匹配附件：{filename}")

        # 筛选最新附件
        if not attach_candidates:
            print(f"\n❌ 当日邮件中未找到匹配【{match_key}】的附件")
            return None
        
        # 按邮件顺序排序（idx越小越新）
        attach_candidates.sort(key=lambda x: x[0])
        _, filename, payload = attach_candidates[0]
        
        # 清理附件名非法字符
        filename = filename.replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_")
        save_path = os.path.join(temp_dir, filename)
        
        # 保存附件
        with open(save_path, "wb") as f:
            f.write(payload)
        
        print(f"\n✅ 下载当日最新附件：{filename} → {save_path}")
        mail.close()
        mail.logout()
        return save_path
    
    except Exception as e:
        print(f"\n❌ 附件下载失败：{str(e)}")
        try:
            mail.close()
            mail.logout()
        except:
            pass
        return None

# ===================== 数据提取/替换（保留原有逻辑） =====================
def extract_cols_from_attach(attach_path: str, attach_rule: Dict[str, Any]) -> Dict[int, List[Any]]:
    """提取附件中连续列的数据"""
    sheet_name = attach_rule["sheet"]
    header_row = attach_rule["header_row"]
    cols_range = attach_rule["cols_range"]
    try:
        wb = openpyxl.load_workbook(attach_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"附件无Sheet：{sheet_name}")
        ws = wb[sheet_name]

        col_indexes = get_cols_range(ws, cols_range["start_col"], cols_range["end_col"], header_row)
        print(f"\n📊 提取附件列（共{len(col_indexes)}列）：")
        for idx in col_indexes:
            col_name = ws.cell(row=header_row, column=idx).value
            print(f"   列{get_column_letter(idx)}：{clean_text(col_name)}")

        extract_data = {}
        data_start_row = header_row + 1
        for col_idx in col_indexes:
            col_data = []
            for row in range(data_start_row, ws.max_row + 1):
                col_data.append(ws.cell(row=row, column=col_idx).value)
            extract_data[col_idx] = col_data
            print(f"   列{get_column_letter(col_idx)}提取{len(col_data)}行数据")

        wb.close()
        return extract_data
    except Exception as e:
        print(f"\n❌ 提取附件数据失败：{str(e)}")
        return {}

def replace_local_cols(extract_data: Dict[int, List[Any]], local_rule: Dict[str, Any]) -> bool:
    """批量替换本地文件的连续列"""
    if not extract_data:
        print("❌ 无提取数据，跳过替换")
        return False
    file_path = local_rule["file_path"]
    sheet_name = local_rule["sheet"]
    header_row = local_rule["header_row"]
    cols_range = local_rule["cols_range"]
    output_path = local_rule["output_path"]

    if not os.path.exists(file_path):
        print(f"❌ 本地文件不存在：{file_path}")
        return False

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"本地文件无Sheet：{sheet_name}")
        ws = wb[sheet_name]

        local_col_indexes = get_cols_range(ws, cols_range["start_col"], cols_range["end_col"], header_row)
        if len(local_col_indexes) != len(extract_data):
            raise ValueError(f"附件列数({len(extract_data)})与本地列数({len(local_col_indexes)})不匹配")
        print(f"\n📝 替换本地列（共{len(local_col_indexes)}列）：")
        for idx in local_col_indexes:
            col_name = ws.cell(row=header_row, column=idx).value
            print(f"   列{get_column_letter(idx)}：{clean_text(col_name)}")

        data_start_row = header_row + 1
        for i, (attach_col_idx, local_col_idx) in enumerate(zip(extract_data.keys(), local_col_indexes)):
            col_data = extract_data[attach_col_idx]
            col_letter = get_column_letter(local_col_idx)
            for row_idx, value in enumerate(col_data):
                ws[f"{col_letter}{data_start_row + row_idx}"].value = value
            print(f"   ✅ 列{col_letter}替换{len(col_data)}行数据")

            # 公式下拉
            formula_cols = []
            for col in ws.iter_cols(min_row=data_start_row, max_row=data_start_row):
                cell = col[0]
                if cell.column > local_col_idx and cell.value and str(cell.value).startswith("="):
                    formula_cols.append(cell.column)
            if formula_cols:
                for f_col_idx in formula_cols:
                    f_col_letter = get_column_letter(f_col_idx)
                    formula = ws[f"{f_col_letter}{data_start_row}"].value
                    for row in range(data_start_row + 1, data_start_row + len(col_data)):
                        ws[f"{f_col_letter}{row}"].value = formula
                print(f"   📌 列{col_letter}后续{len(formula_cols)}列公式下拉完成")

        # 删除多余行
        max_data_row = data_start_row + len(list(extract_data.values())[0]) - 1
        if ws.max_row > max_data_row:
            delete_count = ws.max_row - max_data_row
            for row in range(ws.max_row, max_data_row, -1):
                ws.delete_rows(row)
            print(f"\n🗑️ 删除{delete_count}行多余数据")

        wb.save(output_path)
        wb.close()
        print(f"\n🎉 本地文件处理完成 → {output_path}")
        return True
    except Exception as e:
        print(f"\n❌ 替换本地文件失败：{str(e)}")
        return False

# ===================== 主流程 =====================
def main(config_path: str = "./config.json"):
    try:
        # 1. 加载配置
        config = load_json_config(config_path)
        email_config = config["email_config"]
        rules = config["rules"]
        other = config["other"]
        temp_dir = other["temp_dir"]
        os.makedirs(temp_dir, exist_ok=True)

        # 2. 遍历规则处理
        success_count = 0
        for i, rule in enumerate(rules):
            print(f"\n==================================================")
            print(f"📋 处理第{i+1}/{len(rules)}条规则：匹配【{rule['attach_rule']['match_key']}】")
            
            # 步骤1：下载当日附件
            attach_path = download_latest_attach(rule, email_config)
            if not attach_path:
                print(f"❌ 第{i+1}条规则：附件下载失败")
                continue
            
            # 步骤2：提取多列数据
            extract_data = extract_cols_from_attach(attach_path, rule["attach_rule"])
            if not extract_data:
                print(f"❌ 第{i+1}条规则：数据提取失败")
                continue
            
            # 步骤3：替换本地多列
            if replace_local_cols(extract_data, rule["local_rule"]):
                success_count += 1
                print(f"✅ 第{i+1}条规则：处理成功")
            else:
                print(f"❌ 第{i+1}条规则：本地替换失败")

        # 3. 清理临时目录
        if other.get("clean_temp", True) and os.path.exists(temp_dir):
            import shutil
            shutil.rmtree(temp_dir)
            print(f"\n🗑️ 清理临时目录：{temp_dir}")

        # 4. 结果汇总
        print(f"\n==================================================")
        print(f"📊 总处理结果：成功{success_count}/{len(rules)}条规则")
        if success_count == len(rules):
            print(f"🎉 所有规则执行成功！")

    except Exception as e:
        print(f"\n❌ 程序执行失败：{str(e)}")

if __name__ == "__main__":
    main(config_path="D:\\供应链\\芯片\\config.json")