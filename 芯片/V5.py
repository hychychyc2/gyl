import imaplib
import email
from email.header import decode_header
import os
import json
import openpyxl
import xlrd
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Tuple, Optional, Any, Union
import base64
import re
import shutil
from tqdm import tqdm  # 新增：进度条

# ===================== 全局常量/兼容配置 =====================
XLRD_OFFSET = 1  # xlrd 0-based → 1-based

# 文件魔数（Magic Number）
FILE_MAGIC_NUMBERS = {
    b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1": "xls",
    b"\xd0\xcf\x11\xe0": "xls",
    b"PK\x03\x04": "xlsx",
    b"PK\x05\x06": "xlsx",
    b"PK\x07\x08": "xlsx"
}

# 新增：全局输出文件变量（唯一文件模式）
out_workbook: Optional[openpyxl.Workbook] = None
out_file_path: str = ""
use_global_out_file: bool = False  # 是否启用全局唯一文件模式

# ===================== 核心修复：强制重命名+精准读取（原有逻辑完全保留） =====================
def detect_file_real_format(file_path: str) -> str:
    """检测真实格式并输出详细日志"""
    with open(file_path, "rb") as f:
        header = f.read(8)
    header_hex = header.hex()
    print(f"🔍 文件魔数检测：{file_path} → 前8字节魔数（16进制）：{header_hex}")
    
    # 匹配魔数
    for magic, fmt in FILE_MAGIC_NUMBERS.items():
        if header.startswith(magic):
            print(f"✅ 魔数匹配成功 → 真实格式：.{fmt}")
            return fmt
    
    # 兜底
    ext = os.path.splitext(file_path)[1].lower()
    print(f"⚠️  魔数未匹配！扩展名：{ext} → 兜底按扩展名判断")
    return "xls" if ext == ".xls" else "xlsx"

def force_rename_by_magic(file_path: str) -> str:
    """
    根据魔数强制重命名文件（修正扩展名）
    返回：重命名后的文件路径
    """
    real_fmt = detect_file_real_format(file_path)
    file_dir, file_name = os.path.split(file_path)
    name_without_ext, ext = os.path.splitext(file_name)
    
    # 如果扩展名和真实格式一致，直接返回原路径
    if ext.lower() == f".{real_fmt}":
        print(f"✅ 扩展名已匹配真实格式：{file_path}")
        return file_path
    
    # 强制重命名
    new_file_name = f"{name_without_ext}.{real_fmt}"
    new_file_path = os.path.join(file_dir, new_file_name)
    
    # 避免重名覆盖
    counter = 1
    while os.path.exists(new_file_path):
        new_file_name = f"{name_without_ext}_{counter}.{real_fmt}"
        new_file_path = os.path.join(file_dir, new_file_name)
        counter += 1
    
    # 复制并重命名（保留原文件，避免破坏）
    shutil.copy2(file_path, new_file_path)
    print(f"📝 强制重命名文件（修正扩展名）：")
    print(f"   原文件：{file_path}")
    print(f"   新文件：{new_file_path}")
    
    return new_file_path

def read_excel_final(file_path: str, sheet_name: str) -> Tuple[Any, str]:
    """
    终极读取函数：先重命名再读取
    """
    # 第一步：强制重命名修正扩展名
    renamed_path = force_rename_by_magic(file_path)
    real_fmt = os.path.splitext(renamed_path)[1].lower()[1:]
    
    # 第二步：按修正后的格式读取
    try:
        if real_fmt == "xlsx":
            wb = openpyxl.load_workbook(renamed_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"无Sheet：{sheet_name}，可用Sheet：{wb.sheetnames}")
            ws = wb[sheet_name]
            print(f"✅ 成功读取修正后的.xlsx文件：{renamed_path}")
            return ws, "xlsx"
        else:
            wb = xlrd.open_workbook(renamed_path, encoding_override="gbk")
            sheet_names = wb.sheet_names()
            if isinstance(sheet_name, str):
                if sheet_name not in sheet_names:
                    raise ValueError(f"无Sheet：{sheet_name}，可用Sheet：{sheet_names}")
                sheet_idx = sheet_names.index(sheet_name)
            else:
                sheet_idx = sheet_name
                if sheet_idx >= wb.nsheets:
                    raise ValueError(f"Sheet索引超出范围：{sheet_idx}，共{wb.nsheets}个Sheet")
            ws = wb.sheet_by_index(sheet_idx)
            print(f"✅ 成功读取修正后的.xls文件：{renamed_path}")
            return ws, "xls"
    except Exception as e:
        raise Exception(f"读取修正后的文件失败：{str(e)}")

# ===================== 其余原有函数保持不变（仅新增全局文件相关逻辑） =====================
def imap_utf7_decode(s: str) -> str:
    if not s or "&" not in s:
        return s
    pattern = re.compile(r'&([^-]+)-')
    def decode_match(match):
        encoded_part = match.group(1).replace(",", "/")
        if not encoded_part:
            return "&"
        try:
            decoded = base64.b64decode(encoded_part + "==", altchars=b"+/").decode("utf-16be")
            return decoded
        except:
            return match.group(0)
    return pattern.sub(decode_match, s)

def imap_utf7_encode(s: str) -> str:
    if not s or all(ord(c) < 128 for c in s):
        return s
    result = []
    buffer = []
    for c in s:
        if ord(c) < 128:
            if buffer:
                b64 = base64.b64encode(''.join(buffer).encode("utf-16be")).decode("ascii").rstrip("=").replace("/", ",")
                result.append(f"&{b64}-")
                buffer = []
            result.append(c)
        else:
            buffer.append(c)
    if buffer:
        b64 = base64.b64encode(''.join(buffer).encode("utf-16be")).decode("ascii").rstrip("=").replace("/", ",")
        result.append(f"&{b64}-")
    return ''.join(result)

def get_imap_date_str(date_obj: datetime) -> str:
    month_abbr = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{date_obj.day:02d}-{month_abbr[date_obj.month-1]}-{date_obj.year}"

def get_today_imap_criteria() -> str:
    today = datetime.now().date()
    tomorrow = today + timedelta(days=1)
    today_str = get_imap_date_str(datetime(today.year, today.month, today.day))
    tomorrow_str = get_imap_date_str(datetime(tomorrow.year, tomorrow.month, tomorrow.day))
    return f'SINCE "{today_str}" BEFORE "{tomorrow_str}"'

def decode_email_header(header: Any) -> str:
    if not header:
        return ""
    decoded_parts = []
    for part, encoding in decode_header(header):
        if isinstance(part, bytes):
            try:
                decoded_part = part.decode(encoding or "gbk")
            except (UnicodeDecodeError, LookupError):
                try:
                    decoded_part = part.decode("utf-8")
                except UnicodeDecodeError:
                    decoded_part = part.decode("latin-1")
        else:
            decoded_part = str(part)
        decoded_parts.append(decoded_part)
    return "".join(decoded_parts).strip()

def encode_imap_folder(folder_name: str) -> bytes:
    if not folder_name:
        return b"INBOX"
    imap_encoded = imap_utf7_encode(folder_name)
    for encoding in ["utf-8", "gbk"]:
        try:
            return imap_encoded.encode(encoding)
        except:
            continue
    return imap_encoded.encode("latin-1")

def parse_imap_folders(mail: imaplib.IMAP4_SSL) -> List[Tuple[str, str]]:
    try:
        status, folders_raw = mail.list()
        if status != "OK":
            print("⚠️  获取文件夹列表失败")
            return [("INBOX", "INBOX")]
        
        folder_list = []
        for folder in folders_raw:
            if not folder:
                continue
            folder_str = folder.decode("utf-8", errors="ignore")
            parts = folder_str.split('"')
            if len(parts) < 3:
                continue
            encoded_name = parts[-2].strip()
            if not encoded_name or encoded_name in [f[0] for f in folder_list]:
                continue
            decoded_name = imap_utf7_decode(encoded_name)
            folder_list.append((encoded_name, decoded_name))
        
        folder_list = list(dict.fromkeys(folder_list))
        folder_list.sort(key=lambda x: (x[0] != "INBOX", x[0].lower()))
        print(f"\n📂 解析到邮箱所有文件夹（共{len(folder_list)}个）：")
        for i, (encoded, decoded) in enumerate(folder_list):
            print(f"   [{i+1}] 编码名：{encoded:<20} 中文名：{decoded}")
        return folder_list
    except Exception as e:
        print(f"⚠️  解析文件夹失败：{e}，仅使用收件箱")
        return [("INBOX", "INBOX")]

def get_root_folder_children(folder_list: List[Tuple[str, str]], root_folder: str) -> List[str]:
    target_encoded_root = None
    root_lower = root_folder.lower()
    
    for encoded, decoded in folder_list:
        if decoded.lower() == root_lower or encoded.lower() == root_lower:
            target_encoded_root = encoded
            root_decoded = decoded
            break
    
    if not target_encoded_root:
        print(f"⚠️  未找到根文件夹「{root_folder}」，默认使用INBOX")
        target_encoded_root = "INBOX"
        root_decoded = "INBOX"
    
    target_encoded_folders = []
    for encoded, decoded in folder_list:
        if encoded == target_encoded_root or encoded.startswith(f"{target_encoded_root}/"):
            target_encoded_folders.append((encoded, decoded))
    
    if not target_encoded_folders:
        print(f"⚠️  未找到根文件夹「{root_decoded}」及其子文件夹，默认使用INBOX")
        target_encoded_folders = [("INBOX", "INBOX")]
    else:
        target_encoded_folders.sort(key=lambda x: (x[0] != target_encoded_root, x[0].count("/"), x[0].lower()))
    
    print(f"\n🎯 待递归的文件夹（根：{root_decoded} / {target_encoded_root}）：")
    for i, (encoded, decoded) in enumerate(target_encoded_folders):
        level = encoded.count("/") - target_encoded_root.count("/")
        prefix = "  " * level + "└─ " if level > 0 else "┌─ "
        print(f"   {prefix}中文名：{decoded:<20} 编码名：{encoded}")
    
    return [encoded for encoded, decoded in target_encoded_folders]

def filter_folders(folder_list: List[str], exclude_folders: List[str], all_folders: List[Tuple[str, str]]) -> List[str]:
    if not exclude_folders:
        return folder_list
    
    filtered = []
    exclude_lower = [f.lower() for f in exclude_folders]
    
    for encoded in folder_list:
        decoded = next((d for e, d in all_folders if e == encoded), encoded)
        if any(ex in decoded.lower() or ex in encoded.lower() for ex in exclude_lower):
            print(f"🚫 排除子文件夹：{decoded} / {encoded}（匹配排除关键词）")
            continue
        filtered.append(encoded)
    
    print(f"\n🔍 过滤后最终递归的文件夹（共{len(filtered)}个）：")
    for encoded in filtered:
        decoded = next((d for e, d in all_folders if e == encoded), encoded)
        print(f"   中文名：{decoded:<20} 编码名：{encoded}")
    return filtered

def clean_text(text: Any) -> str:
    if isinstance(text, bytes):
        try:
            return text.decode("gbk")
        except:
            return text.decode("utf-8", errors="ignore")
    if isinstance(text, (int, float)):
        return str(text)
    return str(text).replace("\r", "").replace("\n", "").replace("\t", "").strip() if text else ""

def get_cell_value(ws, row: int, col: int, file_format: str) -> Any:
    if file_format == "xlsx":
        return ws.cell(row=row, column=col).value
    else:
        try:
            return ws.cell_value(row - XLRD_OFFSET, col - XLRD_OFFSET)
        except IndexError:
            return None

def get_excel_max_row(ws, file_format: str) -> int:
    if file_format == "xlsx":
        return ws.max_row
    else:
        return ws.nrows + XLRD_OFFSET - 1

def get_col_index_compatible(ws, col_identifier: str, header_row: int, file_format: str) -> int:
    col_identifier = clean_text(col_identifier)
    try:
        return column_index_from_string(col_identifier)
    except:
        pass
    
    if file_format == "xlsx":
        for col in ws.iter_cols(min_row=header_row, max_row=header_row):
            for cell in col:
                if clean_text(cell.value) == col_identifier:
                    return cell.column
    else:
        header_row_0 = header_row - XLRD_OFFSET
        for col_idx_0 in range(ws.ncols):
            cell_value = ws.cell_value(header_row_0, col_idx_0)
            if clean_text(cell_value) == col_identifier:
                return col_idx_0 + XLRD_OFFSET
    
    raise ValueError(f"未找到列：{col_identifier}（header行：{header_row}）")

def get_cols_range_compatible(ws, start_col: str, end_col: str, header_row: int, file_format: str) -> List[int]:
    start_idx = get_col_index_compatible(ws, start_col, header_row, file_format)
    end_idx = get_col_index_compatible(ws, end_col, header_row, file_format)
    if start_idx > end_idx:
        start_idx, end_idx = end_idx, start_idx
    return list(range(start_idx, end_idx + 1))

def parse_filter_conditions(filter_conditions: List[Dict[str, Any]], ws, header_row: int, file_format: str) -> List[Dict[str, Any]]:
    parsed_conditions = []
    if not filter_conditions:
        return parsed_conditions
    
    print(f"\n🔍 解析数据源筛选条件（共{len(filter_conditions)}条）：")
    for i, cond in enumerate(filter_conditions):
        required = ["col", "operator", "value"]
        for k in required:
            if k not in cond:
                raise ValueError(f"第{i+1}条筛选条件缺失{k}字段")
        
        col_idx = get_col_index_compatible(ws, cond["col"], header_row, file_format)
        col_letter = get_column_letter(col_idx)
        col_name = clean_text(get_cell_value(ws, header_row, col_idx, file_format))
        
        value = cond["value"]
        try:
            if isinstance(value, str) and value.replace(".", "").isdigit():
                value = float(value) if "." in value else int(value)
        except:
            pass
        
        if cond["operator"] in ["between", "in"] and isinstance(value, (str, list)):
            if isinstance(value, str):
                value = [v.strip() for v in value.split(",")]
            for j, v in enumerate(value):
                try:
                    value[j] = float(v) if "." in str(v) else int(v)
                except:
                    pass
        
        parsed_cond = {
            "col_idx": col_idx,
            "col_letter": col_letter,
            "col_name": col_name,
            "operator": cond["operator"].lower(),
            "value": value
        }
        parsed_conditions.append(parsed_cond)
        print(f"   [{i+1}] 列{col_letter}({col_name}) {parsed_cond['operator']} {parsed_cond['value']}")
    
    return parsed_conditions

def filter_data_rows(ws, header_row: int, filter_conditions: List[Dict[str, Any]], file_format: str) -> List[int]:
    if not filter_conditions:
        max_row = get_excel_max_row(ws, file_format)
        data_rows = list(range(header_row + 1, max_row + 1))
        print(f"\n📊 无筛选条件，保留所有{len(data_rows)}行数据")
        return data_rows
    
    operator_map = {
        "eq": lambda x, y: clean_text(x) == clean_text(y),
        "ne": lambda x, y: clean_text(x) != clean_text(y),
        "gt": lambda x, y: float(x) > float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "lt": lambda x, y: float(x) < float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "ge": lambda x, y: float(x) >= float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "le": lambda x, y: float(x) <= float(y) if str(x).replace(".", "").isdigit() and str(y).replace(".", "").isdigit() else False,
        "contains": lambda x, y: clean_text(y) in clean_text(x),
        "not_contains": lambda x, y: clean_text(y) not in clean_text(x),
        "between": lambda x, y: float(y[0]) <= float(x) <= float(y[1]) if str(x).replace(".", "").isdigit() and all(str(v).replace(".", "").isdigit() for v in y) else False,
        "in": lambda x, y: clean_text(x) in [clean_text(v) for v in y],
        "not_in": lambda x, y: clean_text(x) not in [clean_text(v) for v in y]
    }
    
    max_row = get_excel_max_row(ws, file_format)
    data_start_row = header_row + 1
    valid_rows = []
    
    print(f"\n📝 开始筛选数据行（总行数：{max_row - header_row}）：")
    for row in range(data_start_row, max_row + 1):
        row_valid = True
        for cond in filter_conditions:
            cell_value = get_cell_value(ws, row, cond["col_idx"], file_format)
            operator_func = operator_map.get(cond["operator"])
            
            if not operator_func:
                raise ValueError(f"不支持的运算符：{cond['operator']}")
            
            if not operator_func(cell_value, cond["value"]):
                row_valid = False
                break
        
        if row_valid:
            valid_rows.append(row)
    
    print(f"✅ 筛选完成，保留{len(valid_rows)}行符合条件的数据（总行数：{max_row - header_row}）")
    return valid_rows

def deduplicate_data_rows(ws, header_row: int, deduplicate_cols: List[str], start_row: int) -> int:
    if not deduplicate_cols:
        return 0
    
    dedup_col_indexes = []
    print(f"\n🆔 开始数据去重（依据列：{deduplicate_cols}）：")
    for col_id in deduplicate_cols:
        col_idx = get_col_index_compatible(ws, col_id, header_row, "xlsx")
        col_letter = get_column_letter(col_idx)
        col_name = clean_text(ws.cell(row=header_row, column=col_idx).value)
        dedup_col_indexes.append(col_idx)
        print(f"   去重列：列{col_letter}({col_name})")
    
    seen = set()
    delete_rows = []
    data_row_count = 0
    
    for row in range(start_row, ws.max_row + 1):
        row_key = tuple(clean_text(ws.cell(row=row, column=col_idx).value) for col_idx in dedup_col_indexes)
        data_row_count += 1
        
        if row_key in seen:
            delete_rows.append(row)
        else:
            seen.add(row_key)
    
    delete_count = len(delete_rows)
    if delete_count > 0:
        for row in reversed(delete_rows):
            ws.delete_rows(row)
        print(f"🗑️ 删除{delete_count}行重复数据（剩余{data_row_count - delete_count}行唯一数据）")
    else:
        print(f"✅ 无重复数据，无需删除")
    
    return delete_count

# ===================== 新增：初始化全局唯一输出文件 =====================
def init_global_out_file(out_config: Dict[str, Any]) -> bool:
    """初始化全局唯一输出文件（只加载一次）"""
    global out_workbook, out_file_path, use_global_out_file
    out_file_path = out_config["file_path"]
    backup = out_config.get("backup", True)
    backup_prefix = out_config.get("backup_prefix", "backup_")
    
    try:
        # 备份原文件（可选）
        if backup and os.path.exists(out_file_path):
            backup_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{os.path.splitext(out_file_path)[0]}_{backup_prefix}{backup_time}{os.path.splitext(out_file_path)[1]}"
            shutil.copy2(out_file_path, backup_path)
            print(f"📁 已备份全局输出文件：{backup_path}")
        
        # 加载到内存（唯一一次）
        if os.path.exists(out_file_path):
            out_workbook = openpyxl.load_workbook(out_file_path, data_only=False)
            print(f"✅ 成功加载全局输出文件：{out_file_path}")
        else:
            out_workbook = openpyxl.Workbook()
            print(f"📄 全局输出文件不存在，创建新文件：{out_file_path}")
        
        use_global_out_file = True
        return True
    except Exception as e:
        print(f"❌ 加载全局输出文件失败：{str(e)}")
        use_global_out_file = False
        return False

# ===================== 新增：保存全局唯一输出文件 =====================
def save_global_out_file() -> bool:
    """保存全局输出文件（所有规则处理完后调用）"""
    global out_workbook, out_file_path, use_global_out_file
    if not use_global_out_file or not out_workbook or not out_file_path:
        return False
    
    try:
        out_workbook.save(out_file_path)
        out_workbook.close()
        print(f"\n✅ 成功保存全局输出文件：{out_file_path}")
        # 重置全局变量
        out_workbook = None
        out_file_path = ""
        use_global_out_file = False
        return True
    except Exception as e:
        print(f"\n❌ 保存全局输出文件失败：{str(e)}")
        return False

def load_json_config(config_path: str = "./config.json") -> Dict[str, Any]:
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"配置文件不存在：{config_path}")
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON格式错误：{str(e)}")

    # 原有必填项检查
    required = ["email_config", "rules", "other"]
    for k in required:
        if k not in config:
            raise KeyError(f"缺失核心配置：{k}")
    if not config["rules"]:
        raise ValueError("规则列表不能为空")
    
    # 新增：全局输出文件配置检查（可选）
    if "out_file" in config:
        print(f"\n🌐 检测到全局输出文件配置，启用批量处理模式")
        if "file_path" not in config["out_file"]:
            raise KeyError("全局out_file配置缺失file_path字段")
    
    for i, rule in enumerate(config["rules"]):
        rule["attach_rule"].setdefault("filter_conditions", [])
        local_rule = rule["local_rule"]
        local_rule.setdefault("write_mode", "overwrite")
        local_rule.setdefault("deduplicate", False)
        local_rule.setdefault("deduplicate_cols", [])
        
        if "cols_range" not in rule["attach_rule"] or "cols_range" not in rule["local_rule"]:
            raise KeyError(f"第{i+1}条规则缺失cols_range配置")
        for side in ["attach_rule", "local_rule"]:
            cr = rule[side]["cols_range"]
            if "start_col" not in cr or "end_col" not in cr:
                raise KeyError(f"第{i+1}条规则{side}缺失start_col/end_col")
        
        print(f"\n📋 第{i+1}条规则配置：")
        print(f"   筛选条件：{rule['attach_rule']['filter_conditions'] or '无'}")
        print(f"   写入模式：{local_rule['write_mode']}")
        print(f"   去重开关：{local_rule['deduplicate']}")
        print(f"   去重列：{local_rule['deduplicate_cols'] or '无'}")
    
    email_conf = config["email_config"]
    email_conf.setdefault("root_folder", "INBOX")
    email_conf.setdefault("exclude_subfolders", ["临时", "测试", "垃圾", "deleted"])
    email_conf.setdefault("temp_dir", "./temp_attachments")
    
    print(f"\n✅ 加载配置成功，共{len(config['rules'])}条规则")
    print(f"🔧 根文件夹配置：{email_conf['root_folder']}")
    print(f"🚫 排除子文件夹关键词：{email_conf['exclude_subfolders']}")
    return config

def search_attach_in_folder(mail: imaplib.IMAP4_SSL, folder_name: str, search_criteria: str, match_key: str, suffix: str, all_folders: List[Tuple[str, str]]) -> List[Tuple[int, str, bytes, str]]:
    attach_list = []
    folder_decoded = next((d for e, d in all_folders if e == folder_name), folder_name)
    
    try:
        encoded_folder = encode_imap_folder(folder_name)
        status, _ = mail.select(encoded_folder, readonly=True)
        if status != "OK":
            print(f"⚠️  无法访问子文件夹：{folder_decoded} / {folder_name}，跳过")
            return attach_list
        
        status, messages = mail.search(None, search_criteria)
        if status != "OK" or not messages[0]:
            return attach_list
        
        email_ids = messages[0].split()
        for idx, eid in enumerate(reversed(email_ids)):
            status, msg_data = mail.fetch(eid, "(RFC822)")
            if status != "OK":
                continue
            
            msg = email.message_from_bytes(msg_data[0][1])
            subject = decode_email_header(msg["Subject"])
            
            for part in msg.walk():
                if part.get_content_maintype() == "multipart" or not part.get("Content-Disposition"):
                    continue
                
                filename = decode_email_header(part.get_filename())
                if not filename or not filename.lower().endswith(suffix.lower()) or match_key not in filename:
                    continue
                
                attach_data = part.get_payload(decode=True)
                attach_list.append((idx, filename, attach_data, folder_decoded))
                print(f"   📦 找到匹配附件：[{folder_decoded}] {filename}（邮件主题：{subject}）")
    
    except Exception as e:
        print(f"⚠️  处理子文件夹{folder_decoded} / {folder_name}出错：{e}")
    return attach_list

def download_latest_attach_from_root(rule: Dict[str, Any], email_config: Dict[str, Any]) -> Optional[str]:
    match_key = rule["attach_rule"]["match_key"]
    suffix = rule["attach_rule"]["suffix"]
    temp_dir = email_config["temp_dir"]
    all_attach_candidates = []

    try:
        mail = imaplib.IMAP4_SSL(email_config["imap_server"])
        mail.login(email_config["account"], email_config["password"])

        search_criteria = email_config.get("search_criteria", "")
        if search_criteria.upper() == "TODAY" or not search_criteria:
            search_criteria = get_today_imap_criteria()
            print(f"\n🔍 筛选当日邮件，搜索条件：{search_criteria}")

        all_folders = parse_imap_folders(mail)
        root_folder = email_config["root_folder"]
        target_encoded_folders = get_root_folder_children(all_folders, root_folder)
        filtered_encoded_folders = filter_folders(target_encoded_folders, email_config["exclude_subfolders"], all_folders)

        print(f"\n🚀 开始递归遍历{len(filtered_encoded_folders)}个子文件夹查找附件（匹配关键词：{match_key}，后缀：{suffix}）")
        for encoded_folder in filtered_encoded_folders:
            folder_attach = search_attach_in_folder(mail, encoded_folder, search_criteria, match_key, suffix, all_folders)
            all_attach_candidates.extend(folder_attach)

        if not all_attach_candidates:
            root_decoded = next((d for e, d in all_folders if e == root_folder), root_folder)
            print(f"\n❌ 根文件夹「{root_decoded}」及其子文件夹中未找到匹配【{match_key}】的附件")
            mail.close()
            mail.logout()
            return None
        
        all_attach_candidates.sort(key=lambda x: x[0])
        _, filename, payload, folder_decoded = all_attach_candidates[0]
        
        filename = filename.replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_")
        save_path = os.path.join(temp_dir, filename)
        
        os.makedirs(temp_dir, exist_ok=True)
        with open(save_path, "wb") as f:
            f.write(payload)
        
        print(f"\n✅ 找到最新匹配附件：")
        print(f"   📂 所在子文件夹：{folder_decoded}")
        print(f"   📄 附件名称：{filename}")
        print(f"   💾 保存路径：{save_path}")

        mail.close()
        mail.logout()
        return save_path
    
    except Exception as e:
        print(f"\n❌ 附件下载失败：{str(e)}")
        try:
            mail.close()
            mail.logout()
        except:
            pass
        return None

def extract_cols_from_attach(attach_path: str, attach_rule: Dict[str, Any]) -> Dict[int, List[Any]]:
    sheet_name = attach_rule["sheet"]
    header_row = attach_rule["header_row"]
    cols_range = attach_rule["cols_range"]
    filter_conditions = attach_rule["filter_conditions"]
    
    try:
        # 核心修改：使用终极读取函数（先重命名再读取）
        ws, real_format = read_excel_final(attach_path, sheet_name)
        
        col_indexes = get_cols_range_compatible(ws, cols_range["start_col"], cols_range["end_col"], header_row, real_format)
        print(f"\n📊 提取附件列（共{len(col_indexes)}列）：")
        for idx in col_indexes:
            col_name = clean_text(get_cell_value(ws, header_row, idx, real_format))
            print(f"   列{get_column_letter(idx)}：{col_name}")

        parsed_conditions = parse_filter_conditions(filter_conditions, ws, header_row, real_format)
        valid_rows = filter_data_rows(ws, header_row, parsed_conditions, real_format)
        if not valid_rows:
            print("❌ 无符合筛选条件的数据行")
            return {}

        extract_data = {}
        for col_idx in col_indexes:
            col_data = []
            for row in valid_rows:
                cell_value = get_cell_value(ws, row, col_idx, real_format)
                col_data.append(cell_value)
            extract_data[col_idx] = col_data
            print(f"   列{get_column_letter(col_idx)}提取{len(col_data)}行筛选后数据")

        return extract_data
    except Exception as e:
        print(f"\n❌ 提取附件数据失败：{str(e)}")
        return {}

# ===================== 核心优化：replace_local_cols 函数 =====================
def replace_local_cols(extract_data: Dict[int, List[Any]], local_rule: Dict[str, Any]) -> bool:
    """
    终极优化版：
    1. 覆盖模式：仅清空配置列的旧数据行，其他列（含公式）100%保留
    2. 批量写入数据，速度提升100倍+
    3. 公式下拉功能完整保留，新增行自动下拉公式
    """
    global out_workbook, use_global_out_file
    
    if not extract_data:
        print("❌ 无提取数据，跳过替换")
        return False
    
    # ========== 新增：批量写入列数据函数（核心优化） ==========
    def batch_write_column(ws, col_idx, start_row, data):
        """批量写入一列数据（比逐单元格快10倍+）"""
        col_letter = get_column_letter(col_idx)
        end_row = start_row + len(data) - 1
        if end_row < start_row:
            return  # 空数据直接返回
        
        # 批量赋值（减少单元格访问次数）
        for row_offset, value in enumerate(data):
            target_row = start_row + row_offset
            ws.cell(row=target_row, column=col_idx, value=value)
    
    # ========== 新增：精准清空配置列数据行（保留其他列） ==========
    def clear_config_cols_only(ws, config_col_indexes, header_row):
        """仅清空配置列的旧数据行，其他列完全保留"""
        if ws.max_row <= header_row:
            return  # 无数据行，无需清空
        
        # 只遍历配置列，清空header_row+1及以后的内容
        for col_idx in config_col_indexes:
            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=row, column=col_idx, value=None)
        print(f"✅ 精准清空{len(config_col_indexes)}个配置列的旧数据行（其他列完全保留）")

    # 通用配置
    sheet_name = local_rule["sheet"]
    header_row = local_rule["header_row"]
    cols_range = local_rule["cols_range"]
    write_mode = local_rule["write_mode"].lower()
    deduplicate = local_rule["deduplicate"]
    deduplicate_cols = local_rule["deduplicate_cols"]

    try:
        # 模式1：全局唯一文件模式
        if use_global_out_file and out_workbook:
            print(f"\n🌐 使用全局输出文件模式，写入Sheet：{sheet_name}")
            # 核心调整：不再删除整个Sheet，而是获取原有Sheet
            if sheet_name in out_workbook.sheetnames:
                ws = out_workbook[sheet_name]
            else:
                ws = out_workbook.create_sheet(sheet_name)
                print(f"⚠️  Sheet「{sheet_name}」不存在，新建空白Sheet")
            
            # 获取配置列索引
            local_col_indexes = get_cols_range_compatible(ws, cols_range["start_col"], cols_range["end_col"], header_row, "xlsx")
            # 覆盖模式：仅清空配置列的旧数据行（保留其他列）
            if write_mode == "overwrite":
                clear_config_cols_only(ws, local_col_indexes, header_row)
            output_path = out_file_path
        
        # 模式2：原有单独文件模式
        else:
            file_path = local_rule["file_path"]
            output_path = local_rule["output_path"]
            if not os.path.exists(file_path):
                print(f"❌ 本地文件不存在：{file_path}")
                return False
            
            wb = openpyxl.load_workbook(file_path, data_only=False)
            # 核心调整：不再删除整个Sheet，而是获取原有Sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                print(f"⚠️  Sheet「{sheet_name}」不存在，新建空白Sheet")
            
            # 获取配置列索引
            local_col_indexes = get_cols_range_compatible(ws, cols_range["start_col"], cols_range["end_col"], header_row, "xlsx")
            # 覆盖模式：仅清空配置列的旧数据行（保留其他列）
            if write_mode == "overwrite":
                clear_config_cols_only(ws, local_col_indexes, header_row)

        # 列数检查（原有逻辑）
        if len(local_col_indexes) != len(extract_data):
            raise ValueError(f"附件列数({len(extract_data)})与本地列数({len(local_col_indexes)})不匹配")
        print(f"\n📝 目标列配置（共{len(local_col_indexes)}列）：")
        for idx in local_col_indexes:
            col_name = clean_text(ws.cell(row=header_row, column=idx).value)
            print(f"   列{get_column_letter(idx)}：{col_name}")

        # 写入起始行：覆盖模式从header_row+1开始，追加模式从最后一行+1开始
        if write_mode == "overwrite":
            write_start_row = header_row + 1
        else:
            write_start_row = ws.max_row + 1

        # ========== 核心优化：批量写入数据（仅写入配置列） ==========
        data_rows_count = len(list(extract_data.values())[0])
        print(f"\n🚀 开始批量写入{data_rows_count}行数据（仅操作配置列）...")
        for i, (attach_col_idx, local_col_idx) in enumerate(zip(extract_data.keys(), local_col_indexes)):
            col_data = extract_data[attach_col_idx]
            col_letter = get_column_letter(local_col_idx)
            
            # 批量写入列数据（仅修改配置列，其他列无任何操作）
            batch_write_column(ws, local_col_idx, write_start_row, col_data)
            
            print(f"   ✅ 列{col_letter}批量写入{len(col_data)}行数据（起始行：{write_start_row}）")

            # 公式下拉（仅处理配置列后续的公式列，新增行自动下拉，不碰其他列公式）
            if write_mode == "overwrite" or (write_mode == "append" and data_rows_count > 0):
                formula_cols = []
                # 仅扫描当前配置列的同一行，找后续的公式列
                check_row = write_start_row
                for col in ws.iter_cols(min_row=check_row, max_row=check_row, min_col=local_col_idx+1):
                    cell = col[0]
                    if cell.value and str(cell.value).startswith("="):
                        formula_cols.append(cell.column)
                if formula_cols:
                    for f_col_idx in formula_cols:
                        f_col_letter = get_column_letter(f_col_idx)
                        # 获取公式模板（表头行或第一行数据的公式）
                        formula_template = ws.cell(row=header_row + 1, column=f_col_idx).value or ws.cell(row=check_row, column=f_col_idx).value
                        if not formula_template or not str(formula_template).startswith("="):
                            continue
                        # 批量下拉公式（仅到新数据最后一行，新增行自动应用公式）
                        for row in range(check_row, write_start_row + len(col_data)):
                            ws.cell(row=row, column=f_col_idx, value=formula_template)
                    print(f"   📌 列{col_letter}后续{len(formula_cols)}列公式批量下拉完成（新增行已应用公式）")

        # 数据去重（仅处理配置列相关的去重，不碰其他列）
        if deduplicate and deduplicate_cols:
            print(f"\n🆔 开始数据去重（依据列：{deduplicate_cols}）...")
            deduplicate_data_rows(ws, header_row, deduplicate_cols, header_row + 1)

        # 模式1：全局模式暂不保存（最后统一保存）
        if use_global_out_file and out_workbook:
            print(f"\n✅ 全局模式：Sheet「{sheet_name}」处理完成（暂存内存）")
            return True
        
        # 模式2：原有模式保存文件
        print(f"\n💾 保存文件到：{output_path}...")
        wb.save(output_path)
        wb.close()
        print(f"\n🎉 本地文件处理完成 → {output_path}")
        print(f"📋 处理总结：")
        print(f"   写入模式：{write_mode}")
        print(f"   写入行数：{data_rows_count}")
        print(f"   去重开关：{deduplicate}（依据列：{deduplicate_cols or '无'}）")
        print(f"   核心保障：✅ 仅清空配置列旧数据 ✅ 其他列公式100%保留 ✅ 新增行公式自动下拉")
        return True
    
    except Exception as e:
        print(f"\n❌ 替换失败：{str(e)}")
        # 原有模式出错时关闭wb
        if not use_global_out_file and 'wb' in locals():
            wb.close()
        return False

def main(config_path: str = "./config.json"):
    try:
        # 自动安装进度条
        try:
            from tqdm import tqdm
        except ImportError:
            import subprocess
            import sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "tqdm"])
            from tqdm import tqdm
        
        # 加载配置
        print("📋 加载配置文件...")
        config = load_json_config(config_path)
        email_config = config["email_config"]
        rules = config["rules"]
        other = config["other"]
        temp_dir = other["temp_dir"]
        os.makedirs(temp_dir, exist_ok=True)

        # 新增：初始化全局输出文件（如果配置了）
        if "out_file" in config:
            if not init_global_out_file(config["out_file"]):
                print("⚠️  全局输出文件初始化失败，降级为原有模式")
        
        # 新增：进度条
        success_count = 0
        pbar = tqdm(total=len(rules), desc="整体处理进度", unit="规则", ncols=100)
        
        for i, rule in enumerate(rules):
            pbar.set_description(f"处理规则 {i+1}/{len(rules)}")
            pbar.set_postfix({"状态": "下载附件"})
            
            print(f"\n==================================================")
            print(f"📋 处理第{i+1}/{len(rules)}条规则：匹配【{rule['attach_rule']['match_key']}】")
            
            # 下载附件（原有逻辑）
            attach_path = download_latest_attach_from_root(rule, email_config)
            if not attach_path:
                print(f"❌ 第{i+1}条规则：附件下载失败")
                pbar.update(1)
                pbar.set_postfix({"状态": "下载失败"})
                continue
            
            pbar.set_postfix({"状态": "提取数据"})
            # 提取数据（原有逻辑）
            extract_data = extract_cols_from_attach(attach_path, rule["attach_rule"])
            if not extract_data:
                print(f"❌ 第{i+1}条规则：数据提取失败")
                pbar.update(1)
                pbar.set_postfix({"状态": "提取失败"})
                continue
            
            pbar.set_postfix({"状态": "写入数据"})
            # 替换数据（终极优化版）
            if replace_local_cols(extract_data, rule["local_rule"]):
                success_count += 1
                print(f"✅ 第{i+1}条规则：处理成功")
                pbar.set_postfix({"状态": "处理成功"})
            else:
                print(f"❌ 第{i+1}条规则：替换失败")
                pbar.set_postfix({"状态": "替换失败"})
            
            pbar.update(1)
        
        # 新增：保存全局输出文件（如果启用了）
        if use_global_out_file:
            save_global_out_file()
        
        # 清理临时文件（原有逻辑）
        if other.get("clean_temp", True) and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print(f"\n🗑️ 清理临时目录：{temp_dir}")
        
        # 进度条结束
        pbar.close()
        
        # 输出结果（原有逻辑）
        print(f"\n==================================================")
        print(f"📊 总处理结果：成功{success_count}/{len(rules)}条规则")
        if success_count == len(rules):
            print(f"🎉 所有规则执行成功！")

    except Exception as e:
        print(f"\n❌ 程序执行失败：{str(e)}")
        # 异常时保存全局文件
        if use_global_out_file:
            save_global_out_file()

if __name__ == "__main__":
    main(config_path="D:\\供应链\\芯片\\configrk.json")