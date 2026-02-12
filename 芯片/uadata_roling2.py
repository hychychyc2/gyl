import pandas as pd
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# ssc_file = "E:\\供应链\\总表\\191125\\SSC备料计划1112 V1.4.xlsx"
power_supply_file = "E:\\供应链\\芯片\\总表\\芯片齐套表_更新0114.xlsx"
overseas_bom_file = "E:\\供应链\\芯片\\总表\\海外拼BOM版本--2026.01.09 vs 12月版大计划-v1.0.1.xlsx"
global_order_file = "E:\\供应链\\芯片\\总表\\全球订单分配（260109版大计分配结果）20260109 V1.0.0.xlsx"

def read_ssc_xlsx(file_path):
    """读取ssc.xlsx文件，获取内部机型到电源的映射"""
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在")
        return {}
    
    try:
        df = pd.read_excel(file_path, header=1, sheet_name=1)  # 第二行是表头
        # 假设内部机型列和电源列分别为第一列和最后一列
        print(df)
        tot =0
        for i in df.iterrows():
            id=0
            for col in df.columns:
            
                if col == '内部机型':
                    break
                id+=1
            id2=0
            for col in df.columns:
            
                if col == '电源':
                   break
                id2+=1
            tot+=1
        #print(tot)
        for i in range(tot):
            #print(i)
            #print(df.iloc[i,id2])
            df.iloc[i,id2]= df.iloc[i-1,id2]if pd.isna(df.iloc[i,id2]) else df.iloc[i,id2]
            #print(df.iloc[i,5])
        power_dict = dict(zip(df.iloc[:, id].astype(str).str[:9], df.iloc[:, id2]))
        return power_dict
    except Exception as e:
        print(f"读取 {file_path} 时出错: {e}")
        return {}

    # 读取ssc.xlsx
# power_dict = read_ssc_xlsx(ssc_file)
# print(power_dictpower_dict)
def unmerge_and_fill(input_file, output_file):
    # 加载工作簿
    wb = load_workbook(input_file)
    ws = wb.active
    
    # 遍历所有合并单元格
    merged_ranges = ws.merged_cells.ranges.copy()
    for merged_range in merged_ranges:
        # 获取合并区域的起始和结束坐标
        min_col, min_row, max_col, max_row = merged_range.bounds
        
        # 解除合并
        ws.unmerge_cells(start_row=min_row, start_column=min_col,
                         end_row=max_row, end_column=max_col)
        
        # 获取合并区域左上角单元格的值
        top_left_cell = ws.cell(row=min_row, column=min_col)
        fill_value = top_left_cell.value
        
        # 填充整个区域
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.value = fill_value
    
    # 保存修改后的工作簿
    wb.save(output_file)
    print(f"处理完成，已保存至 {output_file}")

# 使用示例


def read_power_supply_xlsx(file_path, sheet_name="电源成品供应"):
    """读取电源千聊表20250516.xlsx文件，获取表头信息"""
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在")
        return None, None, None
    
    try:
        # 读取第二行表头(海外外协、马来PIE等)
        df_header1 = pd.read_excel(file_path, sheet_name=sheet_name, header=1, nrows=1)
        # 读取第三行表头(内部机型代码、APW11等)
        df_header2 = pd.read_excel(file_path, sheet_name=sheet_name, header=2, nrows=1)
        
        # 从第四行开始加载数据(月份数据)
        df_data = pd.read_excel(file_path, sheet_name=sheet_name, header=3)
        
        return df_header1, df_header2, df_data
    except Exception as e:
        print(f"读取 {file_path} 时出错: {e}")
        return None, None, None

def read_overseas_bom_xlsx(file_path, months):
    """读取海外拼BOM版本.xlsx文件，按月份获取数据"""
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在")
        return {}
    
    monthly_data = {}
    input_excel = "input.xlsx"  # 替换为你的输入文件路径
    output_excel = "output.xlsx"  # 替换为你的输出文件路径
    # unmerge_and_fill(input_excel, output_excel)
    for month in months:
        sheet_name = f"2025.{str(month).zfill(2)}"
        try:
            # 第二行是表头
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
            # 解除合并单元格并填充
            # df = df.fillna(method='ffill')
            # print(df)
            monthly_data[month] = df
        except Exception as e:
            print(f"读取 {file_path} 的 {sheet_name} 工作表时出错: {e}")
        sheet_name = f"2026.{str(month).zfill(2)}"
        try:
            # 第二行是表头
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
            # 解除合并单元格并填充
            # df = df.fillna(method='ffill')
            # print(df)
            monthly_data[month] = df
        except Exception as e:
            print(f"读取 {file_path} 的 {sheet_name} 工作表时出错: {e}")
    
    return monthly_data

def read_global_order_xlsx(file_path):
    """读取全球订单分配.xlsx文件，获取所有月份的订单数据（直到找不到本次分配数量列）"""
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在")
        return {}
    
    # 存储所有月份的订单数据 {月份: {外协: {项目: 数量}}}
    all_month_data = {}
    
    try:
        # 自动查找包含"国内整机"和"订单分配"的工作表
        xl = pd.ExcelFile(file_path)
        target_sheet = None
        for sheet_name in xl.sheet_names:
            if "国内整机" in sheet_name and "订单分配" in sheet_name:
                target_sheet = sheet_name
                break
        
        if not target_sheet:
            print(f"错误: 在 {file_path} 中未找到包含'国内整机'和'订单分配'的工作表")
            return all_month_data
       
        # 第三行是表头
        df = pd.read_excel(file_path, sheet_name=target_sheet, header=2)
        
        # 解除合并单元格并填充
        df.iloc[:, 2] = df.iloc[:, 2].fillna(method='ffill')        
                
        # 查找关键列索引
        id = 2  # 内部代码列
        id2 = 4  # 分配外协列
        
        # 找到所有以"本次分配数量"开头的列（核心修改：不再局限2列）
        quantity_cols = []  # 格式: [(列索引, 月份名), ...]
        id_tmp = 0
        for col in df.columns:
            if isinstance(col, str) and col.startswith("本次分配数量"):
                # 提取月份（支持列名格式："本次分配数量11月"、"本次分配数量12月"等）
                moth=(len(quantity_cols)+5)%12
                moth= 12 if moth==0 else moth
                month_str = f"{moth}月"
                quantity_cols.append((id_tmp, month_str))
            id_tmp += 1
        
        if not quantity_cols:
            print(f"警告: 在 {target_sheet} 工作表中未找到'本次分配数量'列")
            return all_month_data
        
        # 初始化每个月份的数据字典
        for _, month_str in quantity_cols:
            all_month_data[month_str] = {}
        
        # 遍历每一行数据，填充所有月份的订单量
        for _, row in df.iterrows():
            # 跳过合计行
            if row.iloc[1] == '合计':
                break
            
            # 获取外协和项目信息
            supplier = row.iloc[id2]
            print(supplier)
            if pd.isna(supplier):
                continue  # 跳过外协为空的行
            power = str(row.iloc[id])[:9]
            if not power:
                continue  # 跳过项目为空的行
            
            # 遍历所有"本次分配数量"列，填充对应月份数据
            for col_idx, month_str in quantity_cols:
                quantity = row.iloc[col_idx]
                if pd.notna(quantity):
                    qty_int = int(quantity + 0.5)
                    # 初始化层级字典
                    if supplier not in all_month_data[month_str]:
                        all_month_data[month_str][supplier] = {}
                    if power not in all_month_data[month_str][supplier]:
                        all_month_data[month_str][supplier][power] = 0
                    all_month_data[month_str][supplier][power] += qty_int
                    
                    # 调试输出（可选保留）
                    if supplier == '奥海' and power == 'APW17':
                        print(f"{month_str} - {supplier} - {power}: {qty_int}")
    
    except Exception as e:
        print(f"读取 {file_path} 时出错: {e}")
    
    return all_month_data

def find_latest_day_in_month(df: pd.DataFrame) -> str:
    """
    查找DataFrame表头中当月日期最晚的订单量列
    
    Args:
        df: 输入的DataFrame
    
    Returns:
        当月最晚日期对应的订单量列名，如果不存在则返回None
    """
    # 提取当前年份和月份
    current_year = datetime.now().year
    current_month = datetime.now().month
    #current_month = 12
    # 筛选以"订单量"结尾的列，并解析日期
    valid_columns = []
    print(df.columns)
    for col in df.columns:
        print(col)
        if col.endswith("订单量"):
            try:
                # 假设列名格式为"MMDD订单量"，如"0514订单量"
                month_str = col[:2]
                day_str = col[2:4]
                month = int(month_str)
                day = int(day_str)
                
                # 验证日期有效性和是否属于当月
                if month == current_month:
                    valid_columns.append((month, day, col))
            except (ValueError, IndexError):
                continue  # 跳过解析失败的列
    
    if not valid_columns:
        return None
    
    # 按日期排序，找出最晚的一天
    valid_columns.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return valid_columns[0][2]  # 返回最晚日期对应的列名
def process_data():
    """处理所有数据并生成最终字典"""
    # 获取当前月份
    today = datetime.now()
    current_month = datetime.now().month

    #current_month = 1  # 例如5
    next_month =12 if current_month==11  else (current_month + 1)%12  # 例如6
    months = list(range(1, 13))  # 1月到12月
   
    
    # 读取电源千聊表
    header1, header2, power_supply_data = read_power_supply_xlsx(power_supply_file)
    print("++++++++++++++")
    # 读取海外拼BOM版本.xlsx
    overseas_bom_data = read_overseas_bom_xlsx(overseas_bom_file, months)
    # print(overseas_bom_data)
    # import pdb;pdb.set_trace()
    # 读取全球订单分配.xlsx（核心修改：调用新的多月份版本）
    global_order_data = read_global_order_xlsx(global_order_file)
    print("全球订单数据（多月份）：", global_order_data)
    # import pdb;pdb.set_trace()
    # 构建最终结果字典
    final_result = {}
    
    # 处理全球订单分配数据（所有月份）
    for month_key, month_data in global_order_data.items():
        final_result[month_key] = month_data
    
    # 处理海外拼BOM版本数据
    for month, df in overseas_bom_data.items():
        print(month)
        month_key = f"{month}月"
        if(month_key not in final_result.keys()):
            final_result[month_key] = {}
        print(df)
        # 假设最后一列表头以"订单量"结尾
        order_quantity_column =  find_latest_day_in_month(df)

        # for col in df.columns:
        #     if col.endswith("订单量"):
        #         order_quantity_column = col
        #         break
        
        if order_quantity_column is None:
            print(f"在 {month} 月的数据中找不到以'订单量'结尾的列")
            continue
        # 遍历每一行
        for _, row in df.iterrows():
            # 假设外协信息在某一列(例如第二列)，内部机型在某一列(例如第三列)
            id=0
            for col in df.columns:
            
                if col == '外协':
                    break
                id+=1
            id2=0
            for col in df.columns:
            
                if col == '成品料号':
                   break
                id2+=1
            supplier = row.iloc[id]  # 第二列
            print(row.iloc[0])
            model = row.iloc[0][:9]    # 第三列
            # print("______")
            # print(model)
            # print(model in power_dict)
            # print(supplier)
            
            # if(row.iloc[0][:9] == "A3HB70503"):
            #     print("++++++")
            #     print(row.iloc[7])
            print("__+_+")
            print(row.iloc[id2])
            print(model)
            #print(power_dict)
            #print(model in power_dict and not pd.isna(row.iloc[id2]))
            # 通过内部机型查找对应的电源
            if  not pd.isna(row.iloc[id2]):
                # if(row.iloc[0][:9] == "A3HB70503"):
                # print("+++++ii+")
                # print(row[order_quantity_column])
                # print(model)
                # print(supplier)
                # print(month)
                # if(supplier == "PIEM" and month=='6' and model in power_dict and power_dict[model]=="APW17"):
                #                 print("++++++")
                #                 print(row[order_quantity_column])
                #     print(row.iloc[7])
                #     print(row[order_quantity_column])
                power =model
                quantity = row[order_quantity_column]
                print(quantity)
                if pd.notna(quantity):
                    if supplier not in final_result[month_key]:
                        final_result[month_key][supplier] = {}
                    if power not in final_result[month_key][supplier]:
                        final_result[month_key][supplier][power] = 0
                    # if month_key=="6月" and supplier=="欧陆通/墨西哥" and power=="APW17":
                    #     print(row)
                    final_result[month_key][supplier][power] += int(quantity)
            print(month_key)
        # break
    print(final_result)
    return final_result

if __name__ == "__main__":
    data= process_data()
    # print(data["5月"]["奥海"]["A3HB70702"])
    # 1. 加载工作簿并获取工作表
    wb = load_workbook(
        power_supply_file,
        data_only=False,
        keep_vba=True
    )
    ws = wb['各外协齐套达成情况']

    # ===================== 关键配置：列内合并规则（每3行合并） =====================
    MERGE_ROW_COUNT = 3  # 列内每3行合并成1个单元格
    # 记录原有合并区域（用于后续恢复格式）
    original_merged_ranges = []
    for rng in ws.merged_cells.ranges:
        # 只记录垂直合并（列数=1，行数=3）的区域（列内三行合并）
        if rng.max_col - rng.min_col + 1 == 1 and rng.max_row - rng.min_row + 1 == MERGE_ROW_COUNT:
            original_merged_ranges.append({
                "col": rng.min_col,
                "start_row": rng.min_row,
                "end_row": rng.max_row
            })

    # ===================== 步骤1：解除所有列内垂直合并（避免只读错误） =====================
    print("开始解除列内三行合并的单元格并填充值...")
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        # 只处理列内垂直合并（列数=1，行数=3）
        if merged_range.max_col - merged_range.min_col + 1 != 1 or merged_range.max_row - merged_range.min_row + 1 != MERGE_ROW_COUNT:
            continue
        
        # 获取合并区域主单元格值（左上角）
        master_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        # 解除合并
        ws.unmerge_cells(str(merged_range))
        # 填充该列下3行的所有单元格（保证数据不丢失）
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            ws.cell(row=row, column=merged_range.min_col).value = master_val
    print("列内合并单元格已解除并填充值")

    # 3. 定义供应商和机型映射（完全保留）
    SUPPLIER_MAPPING = {
        '江元': ['江元'],
        '合权': ['合权'],
        '欧陆通': ['欧陆通'],
        '欧陆通/墨西哥': ['墨西哥欧陆通'],
        '奥海': ['奥海'],
        'PIEM': ['PIE'],
        'NGSB': ['Nationgate'],
        'YNAH': ['印尼奥海'],
        'THQG': ['泰国群光'],
        'KPMI': ['马来KPMI'],
        '星创': ['星创'],
        'ONETEC': ['泰国ONETEC'],
        '美国/Anzer':['美国ANZE'],
        '美国/Keytronic':['Keytronic'],
        '美国/Moonshot':['美国MNST'],
        '美国/Foxlink':['美国Foxlink'],
        '美国/A2Z':['美国AZEC'],
        '美国/Aspower':['美国ASPW'],
        '美国/CMS':['美国CMSI'],
        '美国/Meritronics':['Meritronics'],
        '海南':["HKBTFZ"],
        '星创（EMS出口）':['星创'],
        '欧陆通（保税）': ['欧陆通（保税）'],
        '合权（保税）':['合权（保税）'],
        'HUT8':["HUT8"]
    }

    # ===================== 步骤2：定位关键列 =====================
    header_row = 1  # 表头行
    project_col = None    # 项目列索引
    supplier_col = None   # 外协分配列索引
    month_cols = {}       # 月份列映射 {月份名: 列索引}

    # 遍历表头找到关键列
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=col).value
        if cell_val is None:
            continue
        cell_text = str(cell_val).strip()
        if cell_text == "项目":
            project_col = col
        elif cell_text == "外协分配":
            supplier_col = col
        elif cell_text.split("-")[-1].endswith("月") and len(cell_text.split("-")[-1]) in (2,3):
            month_cols[cell_text.split("-")[-1]] = col
    print(f"项目列：{project_col}，外协分配列：{supplier_col}，月份列：{month_cols}")

    # 校验关键列
    if not project_col or not supplier_col:
        print(f"错误：未找到【项目】列或【外协分配】列")
    elif not month_cols:
        print("错误：未找到任何月份列")
    else:
        # ===================== 步骤3：计算实际数据行（按3行一组对齐） =====================
        def get_actual_last_row(worksheet, check_cols):
            """获取实际有数据的最后一行（按3行一组向上取整）"""
            last_row = header_row
            for row in range(header_row + 1, min(100000, worksheet.max_row + 1)):
                has_data = False
                for col in check_cols:
                    if worksheet.cell(row=row, column=col).value is not None:
                        has_data = True
                        break
                if has_data:
                    last_row = row
                else:
                    if row - last_row > 10:
                        break
            # 按3行一组对齐（比如最后行是10 → 调整为12，保证新增行从13开始，也是3行一组）
            if (last_row - header_row) % MERGE_ROW_COUNT != 0:
                last_row = last_row + (MERGE_ROW_COUNT - (last_row - header_row) % MERGE_ROW_COUNT)
            return last_row

        actual_last_row = get_actual_last_row(ws, [project_col, supplier_col])
        print(f"实际有数据的最后一行（3行对齐）：{actual_last_row}")

        # 步骤4：构建已有行索引（按3行一组的起始行索引）
        existing_rows = {}
        # 按3行一组遍历原有数据
        for row_group_start in range(header_row + 1, actual_last_row + 1, MERGE_ROW_COUNT):
            # 取每组第一行的数值作为索引（因为3行值相同）
            proj_val = ws.cell(row=row_group_start, column=project_col).value
            supp_val = ws.cell(row=row_group_start, column=supplier_col).value
            if proj_val and supp_val:
                key = (str(proj_val).strip(), str(supp_val).strip())
                existing_rows[key] = row_group_start  # 索引指向组起始行

        # Excel最大行数限制
        EXCEL_MAX_ROW = 1048576
        # 记录需要合并的新增行组（列+行范围）
        new_merge_groups = []
        print(existing_rows)
        # ===================== 步骤5：填充数据（新增行按3行一组） =====================
        for month_key, supplier_dict in data.items():
            if month_key not in month_cols:
                print(f"提示：Excel中无{month_key}列，跳过该月份数据")
                continue
            target_col = month_cols[month_key]

            for supp_orig, product_dict in supplier_dict.items():
                target_suppliers = SUPPLIER_MAPPING.get(supp_orig, [supp_orig])
                
                for supp_target in target_suppliers:
                    for product, qty in product_dict.items():
                        if not product or not supp_target:
                            continue
                        
                        lookup_key = (str(product).strip(), str(supp_target).strip())
                        print(lookup_key)
                        if lookup_key in existing_rows:
                            # 更新已有行（3行都填充，保证合并后值一致）
                            row_group_start = existing_rows[lookup_key]
                            for row in range(row_group_start, row_group_start + MERGE_ROW_COUNT):
                                ws.cell(row=row, column=target_col, value=qty)
                        elif qty != 0:
                            # 新增3行一组（列内合并的基础）
                            new_group_start = actual_last_row + 1
                            new_group_end = new_group_start + MERGE_ROW_COUNT - 1
                            
                            # 检查行数限制
                            if new_group_end > EXCEL_MAX_ROW:
                                print(f"警告：Excel行数超限，无法新增3行组：{lookup_key}")
                                continue
                            
                            # 填充该组3行的所有单元格（项目、外协、数量）
                            for row in range(new_group_start, new_group_end + 1):
                                ws.cell(row=row, column=project_col, value=product)
                                ws.cell(row=row, column=supplier_col, value=supp_target)
                                ws.cell(row=row, column=target_col, value=qty)
                            
                            # 记录新增组的合并信息（后续恢复列内合并）
                            for col in [project_col, supplier_col, target_col] + list(month_cols.values()):
                                new_merge_groups.append({
                                    "col": col,
                                    "start_row": new_group_start,
                                    "end_row": new_group_end
                                })
                            
                            # 更新索引和实际最后行
                            existing_rows[lookup_key] = new_group_start
                            actual_last_row = new_group_end  # 指向组最后一行

        # ===================== 步骤6：恢复列内三行合并（原有+新增） =====================
        print("开始恢复列内三行合并格式...")
        # 恢复原有合并区域
        for rng in original_merged_ranges:
            ws.merge_cells(
                start_row=rng["start_row"],
                start_column=rng["col"],
                end_row=rng["end_row"],
                end_column=rng["col"]
            )
        # 合并新增行组（列内三行合并）
        for merge_info in new_merge_groups:
            ws.merge_cells(
                start_row=merge_info["start_row"],
                start_column=merge_info["col"],
                end_row=merge_info["end_row"],
                end_column=merge_info["col"]
            )
        print("列内三行合并格式已恢复完成")
    
    # 6. 保存文件
    output_path = 'E:\\供应链\\芯片\\总表\\芯片齐套表_更新0111大计划.xlsx'
    wb.save(output_path)
    print(f"\n文件已保存至：{output_path}")