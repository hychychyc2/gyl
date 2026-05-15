from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from webdriver_manager.microsoft import EdgeChromiumDriverManager
import openpyxl
from datetime import datetime
import requests
from time import sleep
import os

# -------------------------- 配置参数 --------------------------
XLSX_PATH = "10月FT问题汇总.xlsx"  # 替换为你的XLSX文件路径
WAIT_TIME = 20  # 延长等待时间（内网加载可能慢）
BROWSER_TYPE = "Edge"  # 保持Edge不变
TARGET_URL = "https://icop.bitmain.vip/home"
RETRY_TIMES = 3
RETRY_INTERVAL = 5
DEBUG_MODE = True  # 调试模式：启动后暂停，让你手动验证浏览器网络

# 网页元素定位表达式（需替换为实际XPATH）
INTERNAL_USER_BTN = "//button[contains(text(),'内部用户')]"
FT_WO_TAB = "//div[contains(text(),'FT WO')]"
PO_NO_INPUT = "//input[@name='PO_NO']"
SEARCH_BTN = "//button[contains(text(),'搜索')]"
UNSYNC_ERP_CHECKBOX = "//input[@value='未同步erp']"
ERP_FT_CREATE_BTN = "//button[contains(text(),'ERP-FT工单创建')]"
FT_OUT_LIST_TAB = "//div[contains(text(),'FT Out List')]"
FG_LIST_TAB = "//div[contains(text(),'FG List')]"
STAGE_SELECT = "//select[@name='Stage']"
WIP_STATUS_SELECT = "//select[@name='WIP状态']"
ERP_FT_STOCK_IN_BTN = "//button[contains(text(),'ERP-FT完工入库')]"
ALERT_MSG = "//div[@class='alert-message']"
EMPTY_TABLE_MSG = "//div[contains(text(),'无数据')]"

# -------------------------- 工具函数 --------------------------
def check_network(url):
    """检测系统网络是否能访问目标网页（仅参考，以浏览器实际访问为准）"""
    try:
        response = requests.head(url, timeout=10, proxies={"http": "", "https": ""})  # 不使用requests代理，仅检测系统网络
        return response.status_code in [200, 302, 301]
    except requests.exceptions.RequestException as e:
        print(f"❌ 系统网络检测提示：{str(e)}（忽略，以浏览器实际访问为准）")
        return False

def init_browser():
    """初始化浏览器（关键：启用系统代理，继承VPN/内网配置）"""
    options = webdriver.EdgeOptions()
    # 核心配置：启用系统代理，让Selenium浏览器用和手动打开时相同的网络
    options.add_argument("--proxy-server=system")  # 关键：继承系统代理
    options.add_argument("--no-sandbox")  # 解决部分内网环境权限问题
    options.add_argument("--disable-dev-shm-usage")  # 避免内存不足导致的崩溃
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    
    # 调试模式：保留浏览器缓存（避免每次启动都是新环境）
    if DEBUG_MODE:
        options.add_experimental_option("detach", True)  # 程序执行后不关闭浏览器
    
    service = Service(EdgeChromiumDriverManager().install())
    driver = webdriver.Edge(service=service, options=options)
    driver.maximize_window()
    return driver

def get_work_orders(xlsx_path):
    """读取XLSX工单号（不变）"""
    wb = openpyxl.load_workbook(xlsx_path)
    work_orders = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        work_order_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "完工":
                work_order_col = col
                break
        if not work_order_col:
            print(f"⚠️  sheet[{sheet_name}]未找到'完工'列，跳过")
            continue
        if "备注" not in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
            ws.cell(row=1, column=ws.max_column + 1, value="备注")
        wo_list = []
        for row in range(2, ws.max_row + 1):
            wo = ws.cell(row=row, column=work_order_col).value
            if wo and str(wo).strip():
                wo_list.append((row, str(wo).strip()))
        work_orders[sheet_name] = wo_list
    wb.save(xlsx_path)
    return work_orders, wb

def open_url_with_retry(driver, url):
    """带重试打开网页（不变）"""
    for i in range(RETRY_TIMES):
        try:
            driver.get(url)
            if DEBUG_MODE:
                print(f"⏳ 调试模式：请手动在浏览器中验证是否能访问 {url}，5秒后继续...")
                sleep(5)  # 暂停5秒，让你手动测试
            WebDriverWait(driver, WAIT_TIME).until(
                lambda d: d.title != ""
            )
            print(f"✅ 第{i+1}次尝试访问网页成功")
            return True
        except (TimeoutException, WebDriverException) as e:
            print(f"❌ 第{i+1}次访问网页失败：{str(e)}")
            if i < RETRY_TIMES - 1:
                print(f"⏳ {RETRY_INTERVAL}秒后重试...")
                sleep(RETRY_INTERVAL)
    return False

def handle_single_wo(driver, wo):
    """处理单个工单号（不变）"""
    try:
        if not open_url_with_retry(driver, TARGET_URL):
            return "网络异常：Selenium浏览器未继承内网/VPN配置"
        WebDriverWait(driver, WAIT_TIME).until(
            EC.element_to_be_clickable((By.XPATH, INTERNAL_USER_BTN))
        ).click()
        WebDriverWait(driver, WAIT_TIME).until(
            EC.element_to_be_clickable((By.XPATH, FT_WO_TAB))
        ).click()
        po_input = WebDriverWait(driver, WAIT_TIME).until(
            EC.visibility_of_element_located((By.XPATH, PO_NO_INPUT))
        )
        po_input.clear()
        po_input.send_keys(wo)
        driver.find_element(By.XPATH, SEARCH_BTN).click()
        try:
            unsync_checkbox = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, UNSYNC_ERP_CHECKBOX))
            )
            if not unsync_checkbox.is_selected():
                unsync_checkbox.click()
            driver.find_element(By.XPATH, ERP_FT_CREATE_BTN).click()
            alert = WebDriverWait(driver, WAIT_TIME).until(
                EC.visibility_of_element_located((By.XPATH, ALERT_MSG))
            ).text
            if "失败" in alert or "不通过" in alert:
                return f"工单创建不通过：{alert}"
        except TimeoutException:
            return "未找到'未同步erp'选项"
        WebDriverWait(driver, WAIT_TIME).until(
            EC.element_to_be_clickable((By.XPATH, FT_OUT_LIST_TAB))
        ).click()
        po_input = WebDriverWait(driver, WAIT_TIME).until(
            EC.visibility_of_element_located((By.XPATH, PO_NO_INPUT))
        )
        po_input.clear()
        po_input.send_keys(wo)
        driver.find_element(By.XPATH, SEARCH_BTN).click()
        try:
            WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, EMPTY_TABLE_MSG))
            )
            return "没有Ft outlist"
        except TimeoutException:
            pass
        WebDriverWait(driver, WAIT_TIME).until(
            EC.element_to_be_clickable((By.XPATH, FG_LIST_TAB))
        ).click()
        po_input = WebDriverWait(driver, WAIT_TIME).until(
            EC.visibility_of_element_located((By.XPATH, PO_NO_INPUT))
        )
        po_input.clear()
        po_input.send_keys(wo)
        Select(WebDriverWait(driver, WAIT_TIME).until(
            EC.visibility_of_element_located((By.XPATH, STAGE_SELECT))
        )).select_by_visible_text("FT/SLT")
        Select(WebDriverWait(driver, WAIT_TIME).until(
            EC.visibility_of_element_located((By.XPATH, WIP_STATUS_SELECT))
        )).select_by_visible_text("OK")
        driver.find_element(By.XPATH, SEARCH_BTN).click()
        try:
            WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, EMPTY_TABLE_MSG))
            )
            current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return f"{current_date} 无问题"
        except TimeoutException:
            driver.find_element(By.XPATH, ERP_FT_STOCK_IN_BTN).click()
            alert = WebDriverWait(driver, WAIT_TIME).until(
                EC.visibility_of_element_located((By.XPATH, ALERT_MSG))
            ).text
            current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if "成功" in alert:
                return f"{current_date} 已同步"
            else:
                return f"{current_date} 入库异常：{alert}"
    except Exception as e:
        return f"流程异常：{str(e)}"

# -------------------------- 主程序 --------------------------
if __name__ == "__main__":
    print("🔍 注意：当前为调试模式，浏览器不会自动关闭，可手动验证网络")
    print("1. 若Selenium浏览器无法访问目标网页，检查VPN是否放行Edge进程")
    print("2. 若手动能访问，说明网络配置生效，按回车继续执行...")
    input()  # 暂停，让你查看提示
    
    # 初始化浏览器和XLSX（跳过系统网络检测，以浏览器实际访问为准）
    driver = init_browser()
    work_orders, wb = get_work_orders(XLSX_PATH)
    
    try:
        for sheet_name, wo_list in work_orders.items():
            ws = wb[sheet_name]
            print(f"\n📊 开始处理sheet：{sheet_name}（共{len(wo_list)}个工单号）")
            for row, wo in wo_list:
                print(f"🔍 处理工单号：{wo}")
                remark = handle_single_wo(driver, wo)
                ws.cell(row=row, column=ws.max_column, value=remark)
                print(f"✅ 备注：{remark}\n")
                if DEBUG_MODE:
                    print("⏳ 调试模式：按回车继续处理下一个工单号...")
                    input()  # 逐个工单号暂停，方便排查
            wb.save(XLSX_PATH)
        print(f"🎉 所有工单号处理完成，文件已保存至：{XLSX_PATH}")
    except Exception as e:
        wb.save(XLSX_PATH)
        print(f"❌ 程序异常中断：{str(e)}，已保存当前进度")
    finally:
        if not DEBUG_MODE:
            driver.quit()
        wb.close()