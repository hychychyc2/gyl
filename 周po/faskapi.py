from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse
import shutil
import tempfile
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# 初始化FastAPI应用
app = FastAPI(title="Excel Processor", description="处理Excel文件的富余清单数据")

# 复用原有的处理逻辑（稍作修改以适应文件路径）
def process_excel(file_path,output_path):
    try:
        # 加载工作簿（保留公式）
        wb = load_workbook(file_path, data_only=False)
        
        # 选择工作表
        ws = wb['结存清单-汇总']
        
        # 获取表头行和数据起始行
        header_row = 3
        data_start_row = 4
        
        # 提取表头
        headers = [cell.value for cell in ws[header_row]]
        
        # 查找需要处理的列的索引
        column_indices = {
            '替代': headers.index('替代') if '替代' in headers else None,
            '在途PO/pcs': headers.index('在途PO/pcs') if '在途PO/pcs' in headers else None,
            '结存合计/pcs': headers.index('结存合计/pcs') if '结存合计/pcs' in headers else None,
            '结存替代合并/pcs': headers.index('结存替代合并/pcs') if '结存替代合并/pcs' in headers else None,
            '库存结存/pcs': headers.index('库存结存/pcs') if '库存结存/pcs' in headers else None,
            'PO结存/pcs': headers.index('PO结存/pcs') if 'PO结存/pcs' in headers else None,

        }
        
        # 验证所有需要的列都存在
        if None in column_indices.values():
            missing = [col for col, idx in column_indices.items() if idx is None]
            print(f"错误：找不到以下列：{', '.join(missing)}")
            return
        
        # 创建数据字典，用于分组处理
        data = []
        for row_idx in range(data_start_row, ws.max_row + 1):
            row_data = {
                '替代': ws.cell(row=row_idx, column=column_indices['替代'] + 1).value,
                '在途PO/pcs': ws.cell(row=row_idx, column=column_indices['在途PO/pcs'] + 1).value,
                '结存合计/pcs': ws.cell(row=row_idx, column=column_indices['结存合计/pcs'] + 1).value,
                '结存替代合并/pcs': ws.cell(row=row_idx, column=column_indices['结存替代合并/pcs'] + 1).value,
                # 'c': ws.cell(row=row_idx, column=column_indices['库存结存/pcs'] + 1).value,
                # 'PO结存/pcs': ws.cell(row=row_idx, column=column_indices['PO结存/pcs'] + 1).value,
                'row_idx': row_idx
            }
            # 跳过空行
            if all(value is None for value in row_data.values() if value != 'row_idx'):
                continue
            data.append(row_data)
        
        # 转换为DataFrame进行处理
        df = pd.DataFrame(data)
        
        # 确保数据类型正确
        df['在途PO/pcs'] = pd.to_numeric(df['在途PO/pcs'], errors='coerce')
        df['结存合计/pcs'] = pd.to_numeric(df['结存合计/pcs'], errors='coerce')
        df['结存替代合并/pcs'] = pd.to_numeric(df['结存替代合并/pcs'], errors='coerce')
        # df['库存结存/pcs'] = pd.to_numeric(df['库存结存/pcs'], errors='coerce')
        # df['PO结存/pcs'] = pd.to_numeric(df['PO结存/pcs'], errors='coerce')
        # 删除包含NaN的行
        df['在途PO/pcs'] = df['在途PO/pcs'].fillna(0)
        df['结存合计/pcs'] = df['结存合计/pcs'].fillna(0)

        
        df['结存替代合并/pcs'] = df['结存替代合并/pcs'].fillna(0)
        processed_df = df.dropna(subset=['替代'])
        print(processed_df['在途PO/pcs'].astype(int))
        # processed_df = processed_df[processed_df['在途PO/pcs'].astype(int) != 0]
        # 将替代和在途PO/pcs列中的NaN值填充为0
        
        # 按替代列进行分组处理
        grouped = processed_df.groupby('替代')
        result_dfs = []
        rows_to_deletes = []

        for name, group in grouped:
            # 计算在途PO/pcs的和与结存替代合并/pcs
            sum_po = group['在途PO/pcs'].sum()
            sum_merge = group['结存替代合并/pcs'].max()
            print(group.iterrows())
            # import pdb;pdb.set_trace()
            for idx,_ in group.iterrows():
                    print()
                    group.loc[idx, '结存替代合并/pcs']=sum_merge
            print("+++++++++++")
            print(sum_po)
            print(sum_merge)
            if sum_po<=sum_merge:


                temp_group = group.copy()
                # 按结存合计/pcs升序排序
                temp_group = temp_group.sort_values(by=['在途PO/pcs', '结存合计/pcs'], ascending=[False, True])

                # temp_group = temp_group.sort_values('结存合计/pcs')
                print(temp_group)
                print("---------")
                rows_to_delete = []
                sum_total=0
                # 从小的开始处理，将结存合计/pcs小于在途PO/pcs的改为在途PO/pcs的值
                for idx, row in temp_group.iterrows():
                    # if row['结存合计/pcs'] < row['在途PO/pcs']:

                    if sum_total >= sum_merge:
                        temp_group.loc[idx, '结存合计/pcs'] = 0
                        rows_to_delete.append(idx)
                        rows_to_deletes.append(row['row_idx'])

                        continue

                    sum_total += row['在途PO/pcs']

                    # temp_group.loc[idx, '结存合计/pcs'] = row['在途PO/pcs']
                    if(row['在途PO/pcs']<=0.001 and row['结存合计/pcs']<500 and len(rows_to_delete)+1<len(temp_group)):
                        rows_to_delete.append(idx)
                        rows_to_deletes.append(row['row_idx'])
                        temp_group.loc[idx, '结存合计/pcs'] = row['在途PO/pcs']

                    elif row['在途PO/pcs']>=0.001:
                        temp_group.loc[idx, '结存合计/pcs'] = row['在途PO/pcs']

                   
                    sum_total += temp_group.loc[idx, '结存合计/pcs']

                
                # 计算修改后的结存合计/pcs的总和
                temp_group = temp_group.drop(rows_to_delete)
                sum_total = temp_group['结存合计/pcs'].sum()
                print(sum_total)
                print(sum_merge)
                # 如果总和小于结存替代合并/pcs，调整最后一行
                if 1 and len(temp_group.index):
                    # 获取最后一行的索引（结存合计/pcs最大的行）
                    last_idx = temp_group.index[-1]
                    # 调整最后一行的值，使总和等于结存替代合并/pcs
                    temp_group.loc[last_idx, '结存合计/pcs'] = sum_merge - (sum_total - temp_group.loc[last_idx, '结存合计/pcs'])
                
                # 处理后的组添加到结果列表
                temp_group = temp_group.sort_values('row_idx')

                print(temp_group)
                
                result_dfs.append(temp_group)
                # # 情况1：在途PO/pcs的和大于结存替代合并/pcs
                # temp_group = group.copy()
                # # 按结存合计/pcs升序排序
                # temp_group = temp_group.sort_values('结存合计/pcs')
                
                # print(temp_group)
                # print("---------")
                # sum_total = temp_group['结存合计/pcs'].sum()
                # rows_to_delete = []
                
                # # 从结存合计/pcs最小的行开始删除，直到总和大于等于结存替代合并/pcs
                # for idx, row in temp_group.iterrows():
                #     if row['结存合计/pcs'] < 0:
                #         sum_total -= row['结存合计/pcs']
                #         rows_to_delete.append(idx)
                #         rows_to_deletes.append(row['row_idx'])
                #         continue
                #     if sum_total - row['结存合计/pcs']< sum_merge:
                #         break
                #     if row['结存合计/pcs'] > 0:
                #         sum_total -= row['结存合计/pcs']
                #         rows_to_delete.append(idx)
                #         rows_to_deletes.append(row['row_idx'])

                
                # # 移除需要删除的行
                # print(rows_to_delete)
                # temp_group = temp_group.drop(rows_to_delete)
                
                # # 如果总和仍然大于结存替代合并/pcs，调整最后一行
                # sum_total = temp_group['结存合计/pcs'].sum()
                # print(sum_total)
                # if sum_total > sum_merge:
                #     # 获取最后一行的索引
                #     last_idx = temp_group.index[-1]
                #     # 调整最后一行的值，使总和等于结存替代合并/pcs
                #     temp_group.loc[last_idx, '结存合计/pcs'] = sum_merge - (sum_total - temp_group.loc[last_idx, '结存合计/pcs'])
                
                # # 处理后的组添加到结果列表
                # temp_group = temp_group.sort_values('row_idx')

                # result_dfs.append(temp_group)
                
            else:
                # 情况2：在途PO/pcs的和小于等于结存替代合并/pcs
                temp_group = group.copy()
                # 按结存合计/pcs升序排序

                temp_group = temp_group.sort_values(by=['在途PO/pcs', '结存合计/pcs'], ascending=[True, True])
                print(temp_group)
                print("---------")
                rows_to_delete = []
                sum_total=0
                # 从小的开始处理，将结存合计/pcs小于在途PO/pcs的改为在途PO/pcs的值
                for idx, row in temp_group.iterrows():
                    # if row['结存合计/pcs'] < row['在途PO/pcs']:

                    if sum_total >= sum_merge:
                        temp_group.loc[idx, '结存合计/pcs'] = 0
                        rows_to_delete.append(idx)
                        rows_to_deletes.append(row['row_idx'])

                        continue
                    if(row['在途PO/pcs']==0 and len(rows_to_delete)+1<len(temp_group)):
                        rows_to_delete.append(idx)
                        rows_to_deletes.append(row['row_idx'])
                        temp_group.loc[idx, '结存合计/pcs'] = row['在途PO/pcs']

                    else:
                        sum_total += row['在途PO/pcs']

                        temp_group.loc[idx, '结存合计/pcs'] = row['在途PO/pcs']
                    
                   
                
                # 计算修改后的结存合计/pcs的总和
                temp_group = temp_group.drop(rows_to_delete)
                sum_total = temp_group['结存合计/pcs'].sum()
                print(sum_total)
                print(sum_merge)
                # 如果总和小于结存替代合并/pcs，调整最后一行
                if 1 and len(temp_group.index):
                    # 获取最后一行的索引（结存合计/pcs最大的行）
                    last_idx = temp_group.index[-1]
                    # 调整最后一行的值，使总和等于结存替代合并/pcs
                    temp_group.loc[last_idx, '结存合计/pcs'] = sum_merge - (sum_total - temp_group.loc[last_idx, '结存合计/pcs'])
                
                # 处理后的组添加到结果列表
                temp_group = temp_group.sort_values('row_idx')

                print(temp_group)
                
                result_dfs.append(temp_group)
        print(result_dfs)
        # 合并所有处理后的组
        if result_dfs:
            processed_result = pd.concat(result_dfs)
            
            # 创建一个字典，将原始行索引映射到处理后的值
            row_to_values = {}
            for _, row in processed_result.iterrows():
                original_row = row['row_idx']
                row_to_values[original_row] = {
                    '在途PO/pcs': row['在途PO/pcs'],
                    '结存合计/pcs': row['结存合计/pcs'],
                    '结存替代合并/pcs': row['结存替代合并/pcs']
                }
            
            # 只保留每组第一行的结存替代合并/pcs，其他行设为空
            # 首先，找出每个组的第一行
            first_rows = {}
            for alt, group_df in processed_result.groupby('替代'):
                first_idx = group_df.index[0]
                print(first_idx)
                first_rows[group_df.loc[first_idx, 'row_idx']] = True
            # 更新处理结果
            for row_idx in row_to_values:
                if row_idx not in first_rows:
                    row_to_values[row_idx]['结存替代合并/pcs'] = ''
            
            # 获取列字母表示
            total_col_letter = get_column_letter(column_indices['结存合计/pcs'] + 1)
            merge_col_letter = get_column_letter(column_indices['结存替代合并/pcs'] + 1)
            kucun_col_letter = get_column_letter(column_indices['库存结存/pcs'] + 1)
            po_col_letter = get_column_letter(column_indices['PO结存/pcs'] + 1)

            # 更新工作表中的数据
            for row_idx, values in row_to_values.items():
                # 更新结存合计/pcs列（覆盖可能存在的公式）
                ws[f"{total_col_letter}{row_idx}"] = values['结存合计/pcs']
                
                # 更新结存替代合并/pcs列（覆盖可能存在的公式）
                ws[f"{merge_col_letter}{row_idx}"] = values['结存替代合并/pcs']
                # print(values['结存替代合并/pcs'],values['在途PO/pcs'])
                ws[f"{po_col_letter}{row_idx}"].value = min(values['结存合计/pcs'],values['在途PO/pcs'])
                ws[f"{po_col_letter}{row_idx}"].number_format = '0'
                ws[f"{kucun_col_letter}{row_idx}"].value = values['结存合计/pcs']-ws[f"{po_col_letter}{row_idx}"].value
                ws[f"{kucun_col_letter}{row_idx}"].number_format = '0'
            # 保存工作簿



            header = [cell.value for cell in ws[3]]
    
            # 查找"在投PO"和"替代"列的索引
            try:
                po_col_idx = header.index('在途PO/pcs')
                replace_col_idx = header.index('替代')
                tot_col_idx = header.index('结存合计/pcs')

            except ValueError as e:
                print(f"错误: 找不到列 - {e}")
                return
            
            # 转换为字母索引
            po_col_letter = get_column_letter(po_col_idx + 1)
            replace_col_letter = get_column_letter(replace_col_idx + 1)
            tot_col_idx = get_column_letter(tot_col_idx + 1)

            # 从第2行开始检查(跳过表头)
            rows_to_delete = []
            # for row_idx in range(4, ws.max_row + 1):
            #     po_cell = ws[f"{po_col_letter}{row_idx}"]
            #     replace_cell = ws[f"{replace_col_letter}{row_idx}"]
            #     tot_cell = ws[f"{tot_col_idx}{row_idx}"]

            #     # 检查在投PO是否为0或NaN
            #     po_value = po_cell.value
            #     if po_value is None:
            #         rows_to_delete.append(row_idx)
            #     elif isinstance(po_value, (int, float)) and po_value == 0 and tot_cell.value <= 0:
            #         rows_to_delete.append(row_idx)
            #     elif isinstance(po_value, str) and po_value.strip().lower() in ['nan', 'nan%']:
            #         rows_to_delete.append(row_idx)
                
            #     # 检查替代是否为NaN
            #     replace_value = replace_cell.value
            #     if replace_value is None or (isinstance(replace_value, str) and replace_value.strip().lower() == 'nan'):
            #         if row_idx not in rows_to_delete:
            #             rows_to_delete.append(row_idx)
            
            # # 从底部向上删除行，避免索引变化问题
            for row_idx in sorted(rows_to_deletes, reverse=True):
                ws.delete_rows(row_idx)
            
            print(f"已删除 {len(rows_to_deletes)} 行数据")
            # output_file = 'processed_' + file_path
            wb.save(output_path)
            
            print(f"处理完成，结果已保存至 {output_path}")
        else:
            print("没有需要处理的数据")
        return 1
    except Exception as e:
        print(f"处理过程中发生错误: {e}")


@app.post("/process-excel/")
async def process_excel_file(upload_file: UploadFile = File(...)):
    # 创建临时目录
    temp_dir = tempfile.mkdtemp()
    input_path = os.path.join(temp_dir, upload_file.filename)
    output_path = f"processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        # 保存上传的文件
        with open(input_path, "wb") as f:
            shutil.copyfileobj(upload_file.file, f)
        
        # 调用处理函数
        success = process_excel(input_path, output_path)
        
        if not success:
            raise HTTPException(status_code=500, detail="文件处理失败")
        
        # 返回处理后的文件
        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"processed_{upload_file.filename}"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")
    finally:
        # 清理临时文件（生产环境建议使用更安全的清理策略）
        shutil.rmtree(temp_dir, ignore_errors=True)

# 可选：添加一个简单的上传页面
from fastapi.responses import HTMLResponse

@app.get("/", response_class=HTMLResponse)
async def get_upload_page():
    return """
    <html>
        <head>
            <title>Excel处理工具</title>
        </head>
        <body>
            <h1>上传Excel文件进行处理</h1>
            <form action="/process-excel/" method="post" enctype="multipart/form-data">
                <input type="file" name="upload_file" accept=".xlsx, .xls">
                <button type="submit">处理并下载</button>
            </form>
        </body>
    </html>
    """

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)