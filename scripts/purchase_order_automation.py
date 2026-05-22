#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
采购订单自动化脚本
每天晚上10点运行，自动完成：
1. 读取杨娜邮件获取订单信息
2. 匹配价格表
3. 生成采购订单号
4. 填充WebADI模板
5. 操作Oracle导入
6. 更新统计表格
7. 发送结果邮件

使用方法：
    python purchase_order_automation.py

依赖安装：
    pip install imapclient openpyxl pyautogui pyperclip python-dateutil
"""

import os
import sys
import json
import imaplib
import email
from email.header import decode_header
from datetime import datetime, date
import re
import openpyxl
from openpyxl import Workbook, load_workbook
import time
import subprocess
from pathlib import Path

# ============== 配置区 ==============

# 凭证配置（从加密文件读取，或直接配置）
CONFIG = {
    "email": {
        "account": "yuchuan.he@casue.com",
        "password": "-DxpOD5kkN)(RuPgAK-p",
        "imap_server": "imap.appia.vip",
        "source_email": "na.yang_w@casue.com"
    },
    "erp": {
        "username": "607693",
        "password": "hyc010815"
    },
    "report_emails": ["yuchuan.he@casue.com", "haixia.lu@casue.com"],  # 报告发送目标（多个收件人）
}

# 文件路径配置
BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data"

TEMPLATES_DIR = DATA_DIR / "templates"
STATISTICS_DIR = DATA_DIR / "statistics"
PRICES_DIR = DATA_DIR / "prices"
OUTPUT_DIR = DATA_DIR / "output"

# 主体配置
ENTITY_CONFIG = {
    "SZK": {"name": "世纪云芯", "currency": "CNY", "prefix": "SZK"},
    "ICK": {"name": "智能云芯", "currency": "CNY", "prefix": "ICK"},
    "HSJ": {"name": "海南世纪", "currency": "CNY", "prefix": "HSJ"},
    "DPT": {"name": "Bitmain Development PTE. LTD.", "currency": "USD", "prefix": "DPT"},
    "BJK": {"name": "Bitmain Beijing", "currency": "CNY", "prefix": "BJK"},
}

# ============== 编码和价格查询模块 ==============

# 物料编码对照表（启动时从国内统计表中的编码对账表动态加载）
MODEL_CODE_MAP = {}

def load_model_code_map():
    """从国内统计表的编码对账表动态加载编码对照"""
    global MODEL_CODE_MAP
    stats_file = STATISTICS_DIR / "domestic_statistics.xlsx"
    if not stats_file.exists():
        print(f"统计文件不存在，无法加载编码对照表: {stats_file}")
        return
    
    try:
        wb = load_workbook(stats_file, data_only=True)
        if '编码对账表' not in wb.sheetnames:
            print("警告: 统计表中没有'编码对账表'sheet")
            wb.close()
            return
        
        ws = wb['编码对账表']
        codes = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            chip_name = row[1]  # 列B=芯片简称
            material_code = row[3]  # 列D=物料编码
            if chip_name and material_code and str(material_code).startswith('Y'):
                codes[str(chip_name).upper().strip()] = str(material_code)
        
        wb.close()
        MODEL_CODE_MAP = codes
        print(f"已加载 {len(MODEL_CODE_MAP)} 条物料编码")
    except Exception as e:
        print(f"加载编码对照表失败: {e}")

# 价格对照表（从价格表提取）
MODEL_PRICE_MAP = {
    "BM1370": 8.4438,
    "BM1373CC": 31.53,
    "BM1373AA": 31.53,
    "BM1746": 57.64,
    "BM1489": 5.21,
    "BM1491": 12.75,
}

def get_model_code(model):
    """根据型号获取物料编码"""
    model_upper = model.upper().strip()
    
    # 先查完整型号
    if model_upper in MODEL_CODE_MAP:
        return MODEL_CODE_MAP[model_upper]
    
    # 再查基础型号（去掉后缀）
    for base_model in MODEL_CODE_MAP:
        if model_upper.startswith(base_model):
            return MODEL_CODE_MAP[base_model]
    
    # 如果找不到，返回None（写入时留空，后续手动填写）
    return None

def get_model_price(model):
    """根据型号获取价格"""
    model_upper = model.upper().strip()
    
    for base_model in MODEL_PRICE_MAP:
        if model_upper.startswith(base_model):
            return MODEL_PRICE_MAP[base_model]
    
    return None  # 如果找不到，需要手动填写

def classify_email(subject, has_attachment):
    """
    判断邮件应该写入哪个表格
    
    规则（用户确认）：
    - 有附件的"进口产品统计表"邮件 → 国内表格（domestic）
    - 正文直接写的出货通知 → 海外表格（international）
    
    返回: "domestic" 或 "international"
    """
    # 有附件且包含"进口产品统计表" → 国内
    if has_attachment and "进口产品统计表" in subject:
        return "domestic"
    
    # 其他情况 → 海外（正文直接写的出货通知、出口资料等）
    return "international"

def get_price_source(table_type, item_data):
    """
    根据表格类型获取价格来源
    
    规则（用户确认）：
    - 国内表格：价格从附件获得（邮件解析时已提取）
    - 海外表格：价格从价格表查询
    
    参数：
        table_type: "domestic" 或 "international"
        item_data: 包含model和可能的price字段
    
    返回：价格（float）或 None
    """
    if table_type == "domestic":
        # 国内：从附件数据获取价格（item_data中应该有price字段）
        return item_data.get("price")
    
    elif table_type == "international":
        # 海外：从价格表查询
        model = item_data.get("model", "")
        return get_model_price(model)
    
    return None

def decode_email_header(header):
    """解码邮件头部"""
    if header is None:
        return ""
    decoded_parts = decode_header(header)
    result = []
    for part, charset in decoded_parts:
        if isinstance(part, bytes):
            if charset:
                try:
                    result.append(part.decode(charset))
                except:
                    result.append(part.decode('utf-8', errors='ignore'))
            else:
                result.append(part.decode('utf-8', errors='ignore'))
        else:
            result.append(str(part))
    return ''.join(result)

def get_today_emails_from_source():
    """获取今日来自杨娜的邮件"""
    print("正在连接邮件服务器...")
    
    imap = imaplib.IMAP4_SSL(CONFIG["email"]["imap_server"])
    imap.login(CONFIG["email"]["account"], CONFIG["email"]["password"])
    
    # 邮件在 MC/po 子文件夹，不在主收件箱
    target_folder = "MC/po"
    print(f"搜索文件夹: {target_folder}")
    
    try:
        status = imap.select(target_folder)
        if status[0] != "OK":
            print(f"无法访问文件夹 {target_folder}，尝试搜索收件箱")
            imap.select("INBOX")
    except Exception as e:
        print(f"文件夹访问失败: {e}，使用收件箱")
        imap.select("INBOX")
    
    # 搜索今日邮件
    today = date.today().strftime("%d-%b-%Y")
    search_criteria = f'(FROM "{CONFIG["email"]["source_email"]}" ON "{today}")'
    
    status, messages = imap.search(None, search_criteria)
    
    if status != "OK":
        print("搜索邮件失败")
        imap.logout()
        return []
    
    email_ids = messages[0].split()
    print(f"找到 {len(email_ids)} 封今日邮件")
    
    emails_data = []
    for email_id in email_ids:
        status, msg_data = imap.fetch(email_id, "(RFC822)")
        if status == "OK":
            msg = email.message_from_bytes(msg_data[0][1])
            
            subject = decode_email_header(msg.get("Subject"))
            from_addr = decode_email_header(msg.get("From"))
            date_str = decode_email_header(msg.get("Date"))
            
            # 提取邮件正文
            body = ""
            if msg.is_multipart():
                for part in msg.walk():
                    content_type = part.get_content_type()
                    if content_type == "text/plain" or content_type == "text/html":
                        try:
                            charset = part.get_content_charset() or 'utf-8'
                            payload = part.get_payload(decode=True)
                            if payload:
                                body += payload.decode(charset, errors='ignore')
                        except:
                            pass
            else:
                try:
                    charset = msg.get_content_charset() or 'utf-8'
                    payload = msg.get_payload(decode=True)
                    if payload:
                        body = payload.decode(charset, errors='ignore')
                except:
                    pass
            
            emails_data.append({
                "id": email_id.decode(),
                "subject": subject,
                "from": from_addr,
                "date": date_str,
                "body": body
            })
    
    imap.logout()
    return emails_data

def parse_order_info_from_email(email_body, subject):
    """从邮件内容解析订单信息"""
    print(f"解析邮件: {subject}")
    
    order_info = {
        "items": [],  # 芯片型号和数量列表
        "entity": None,  # 主体
        "address": None,  # 地址
        "test_factory": None,  # 测试厂
        "notes": None  # 备注
    }
    
    # 尝试解析芯片型号和数量
    # 常见格式: BM1362 10000pcs 或 BM1362AC: 5000等
    
    # 模式1: 型号 + 数量
    pattern1 = r'(BM\d{4}[A-Z]{0,3}[\+\w]*)\s*[:\s]\s*(\d+)\s*(pcs|个|片)?'
    matches1 = re.findall(pattern1, email_body, re.IGNORECASE)
    
    # 模式2: 型号 + 数量（更宽松）
    pattern2 = r'(BM\d{4}[A-Z]{0,3}[\+\w]*)\s+(\d+)'
    matches2 = re.findall(pattern2, email_body, re.IGNORECASE)
    
    # 模式3: 从表格格式解析（如果有）
    pattern3 = r'(\d+)\s*(pcs|个|片)?\s*(BM\d{4}[A-Z]{0,3}[\+\w]*)'
    matches3 = re.findall(pattern3, email_body, re.IGNORECASE)
    
    all_matches = matches1 + matches2 + [(m[2], m[0]) for m in matches3]
    
    # 去重：同一型号只保留一条
    seen_models = {}
    for match in all_matches:
        model = match[0].upper().strip()
        quantity = int(match[1])
        if model not in seen_models or quantity > seen_models[model]["quantity"]:
            seen_models[model] = {"model": model, "quantity": quantity}
    
    # 提取价格（仅国内邮件 - 进口产品统计表中的报关单价）
    # 只有主题包含"进口产品统计表"的邮件才从正文提取价格
    price_map = {}
    if "进口产品统计表" in subject:
        # 格式：BM1374CC 780 PCS 38.37 或 报关单价列
        price_pattern = r'(BM\d{4}[A-Z]{0,3}[\+\w]*)\s+(?:\d+)\s*(?:pcs|PCS|个|片)?\s*(\d+\.?\d*)'
        price_matches = re.findall(price_pattern, email_body, re.IGNORECASE)
        for pm in price_matches:
            price_map[pm[0].upper().strip()] = float(pm[1])
        
        # 也可以从HTML表格提取价格（报关单价列）
        html_price_pattern = r'(BM\d{4}[A-Z]{0,3}[\+\w]*)\s*</td>\s*<td[^>]*>\s*(\d+)\s*</td>\s*<td[^>]*>\s*PCS\s*</td>\s*<td[^>]*>\s*(\d+\.?\d*)'
        html_price_matches = re.findall(html_price_pattern, email_body, re.IGNORECASE)
        for hpm in html_price_matches:
            model = hpm[0].upper().strip()
            qty = int(hpm[1])
            price = float(hpm[2])
            price_map[model] = price
            # 更正数量
            if model in seen_models:
                seen_models[model]["quantity"] = qty
    
    for model, data in seen_models.items():
        item = {"model": data["model"], "quantity": data["quantity"]}
        # 国内：如果有邮件中的报关单价，使用它
        if model in price_map:
            item["price"] = price_map[model]
        order_info["items"].append(item)
    
    # 解析主体（从邮件内容或主题）
    for entity_code in ENTITY_CONFIG.keys():
        if entity_code in email_body.upper() or entity_code in subject.upper():
            order_info["entity"] = entity_code
            break
    
    # 如果没找到主体，根据内容推断（同时检查主题）
    combined = email_body + subject
    if not order_info["entity"]:
        if "深圳" in combined or "SZ" in combined:
            order_info["entity"] = "SZK"
        elif "新加坡" in combined or "SG" in combined or "ONETEC" in combined or "出口" in combined:
            order_info["entity"] = "DPT"
        elif "北京" in combined or "BJ" in combined:
            order_info["entity"] = "BJK"
        elif "海南" in combined:
            order_info["entity"] = "HSJ"
    
    # 解析地址
    address_patterns = [
        r'地址[:\s：]+([^\n]+)',
        r'收货地址[:\s：]+([^\n]+)',
        r'发货至[:\s：]+([^\n]+)',
        r'送至[:\s：]+([^\n]+)',
    ]
    for pattern in address_patterns:
        match = re.search(pattern, email_body)
        if match:
            order_info["address"] = match.group(1).strip()
            break
    
    # 解析测试厂
    test_factory_patterns = ["XJ", "捷策创", "SCK", "朗华", "Vtest", "确安", "HN"]
    for tf in test_factory_patterns:
        if tf in email_body:
            order_info["test_factory"] = tf
            break
    
    # 备注
    if "备注" in email_body or "注意" in email_body:
        match = re.search(r'(备注|注意)[:\s：]+([^\n]+)', email_body)
        if match:
            order_info["notes"] = match.group(2).strip()
    
    return order_info

# ============== 价格匹配模块 ==============

def load_price_table():
    """加载价格表"""
    price_file = PRICES_DIR / "current_prices.xlsx"
    if not price_file.exists():
        print(f"价格表文件不存在: {price_file}")
        return {}
    
    wb = load_workbook(price_file, data_only=True)
    ws = wb["PO"]
    
    prices = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[3]:  # 型号和价格
            model = str(row[0]).upper().strip()
            price = float(row[3])
            prices[model] = price
    
    wb.close()
    print(f"已加载 {len(prices)} 个型号的价格")
    return prices

def match_price(model, prices):
    """匹配芯片价格"""
    model_upper = model.upper().strip()
    
    # 直接匹配
    if model_upper in prices:
        return prices[model_upper]
    
    # 去掉后缀字母匹配（如BM1362AA -> BM1362）
    base_model = re.sub(r'[A-Z]{2,3}$', '', model_upper)
    if base_model in prices:
        return prices[base_model]
    
    # BM1368+ 特殊处理
    if "1368+" in model_upper or "1368PA" in model_upper or "1368PB" in model_upper:
        return prices.get("BM1368+", prices.get("BM1368", 0))
    
    print(f"警告: 未找到型号 {model} 的价格")
    return None

# ============== 订单号生成模块 ==============

def generate_order_number(entity, existing_numbers=None):
    """生成采购订单号"""
    today = date.today()
    date_str = today.strftime("%Y%m%d")
    prefix = ENTITY_CONFIG[entity]["prefix"]
    
    # 查找今日最大序号
    if existing_numbers:
        today_prefix = f"{prefix}{date_str}"
        max_seq = 0
        for num in existing_numbers:
            if str(num).startswith(today_prefix):
                try:
                    seq = int(str(num)[-4:])
                    max_seq = max(max_seq, seq)
                except:
                    pass
        next_seq = max_seq + 1
    else:
        next_seq = 1
    
    return f"{prefix}{date_str}{next_seq:04d}"

def get_existing_order_numbers(statistics_file):
    """从统计表获取现有订单号"""
    if not statistics_file.exists():
        return []
    
    wb = load_workbook(statistics_file, data_only=True)
    ws = wb.active
    
    numbers = []
    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        if row[8]:  # 采购订单号列
            numbers.append(row[8])
    
    wb.close()
    return numbers

# ============== WebADI模板填充模块 ==============

def fill_webadi_template(all_orders, output_file):
    """
    填充WebADI模板（.xlsm格式，保留宏和格式）
    
    所有订单汇总到一个文件，只写入数据行，不改变模板格式。
    使用copy复制模板文件，然后用openpyxl写入数据。
    """
    import shutil
    
    # 复制模板文件（保留宏和格式）
    template_file = TEMPLATES_DIR / "webadi_template.xlsm"
    if not template_file.exists():
        print(f"模板文件不存在: {template_file}")
        return False
    
    shutil.copy2(template_file, output_file)
    
    # 加载复制的模板（保留VBA）
    wb = load_workbook(output_file, keep_vba=True)
    ws = wb['WebADI']
    
    # 找到第一个空数据行（从行5开始，行3-4是表头和格式说明）
    row_num = 5
    while ws.cell(row=row_num, column=3).value is not None:
        row_num += 1
    
    print(f"从行{row_num}开始写入数据")
    
    line_num = 1
    for order_data in all_orders:
        entity = order_data["entity"]
        entity_config = ENTITY_CONFIG[entity]
        
        # 根据主体确定供应商和地点
        if entity in ["SZK", "ICK", "HSJ"]:
            supplier = "BITMAIN DEVELOPMENT PTE. LTD."
            supplier_site = "SG"
            source_subinv = "SZKXYCL"
            receiver = "1004.Bitmain Shenzhen"
            dest_subinv = "SZKXYCL"
            bill_to = "1004.Bitmain Shenzhen"
        elif entity == "DPT":
            supplier = "Chanhua Pte. Ltd."
            supplier_site = "SG"
            source_subinv = "DPTXYCL"
            receiver = "1004.Bitmain Singapore"
            dest_subinv = "DPTXYCL"
            bill_to = "1004.Bitmain Singapore"
        elif entity == "BJK":
            supplier = "Bitmain  Technologies Limited"
            supplier_site = "HK"
            source_subinv = "XAP"
            receiver = "1001.Bitmain Beijing"
            dest_subinv = "BJKDFC"
            bill_to = "1001.Bitmain Beijing"
        else:
            supplier = "BITMAIN DEVELOPMENT PTE. LTD."
            supplier_site = "SG"
            source_subinv = "SZKXYCL"
            receiver = "1004.Bitmain Shenzhen"
            dest_subinv = "SZKXYCL"
            bill_to = "1004.Bitmain Shenzhen"
        
        for item in order_data["items"]:
            model = item["model"]
            quantity = item["quantity"]
            price = item.get("price")
            
            if price is None:
                print(f"跳过无价格的型号: {model}")
                continue
            
            # 获取物料编码
            model_code = get_model_code(model)
            if model_code is None:
                print(f"警告: 未找到 {model} 的物料编码，留空")
                model_code = ""
            
            # 头部信息（按模板列顺序）
            ws.cell(row=row_num, column=3, value=entity)  # 业务实体
            ws.cell(row=row_num, column=4, value="标准采购订单")  # 类型
            ws.cell(row=row_num, column=5, value=order_data["order_number"])  # 采购订单号
            ws.cell(row=row_num, column=6, value=entity_config["currency"])  # 币种
            ws.cell(row=row_num, column=7, value="何宇川,")  # 采购员
            ws.cell(row=row_num, column=8, value=supplier)  # 供应商
            ws.cell(row=row_num, column=9, value=supplier_site)  # 供应商地点
            ws.cell(row=row_num, column=10, value=source_subinv)  # 来源子库存
            ws.cell(row=row_num, column=11, value=receiver)  # 收货方
            ws.cell(row=row_num, column=12, value=dest_subinv)  # 目的子库存
            ws.cell(row=row_num, column=13, value=bill_to)  # 收单方
            ws.cell(row=row_num, column=14, value="付款方式一")  # 付款方式
            ws.cell(row=row_num, column=15, value="生产用料销售")  # 内部申请类型
            ws.cell(row=row_num, column=17, value="Y")  # 是否报关
            
            # 行信息
            ws.cell(row=row_num, column=20, value="手工录入")  # 业务模式
            ws.cell(row=row_num, column=21, value=line_num)  # 行号
            ws.cell(row=row_num, column=22, value="BM系列")  # 行类型
            ws.cell(row=row_num, column=23, value=model_code)  # 物料编码
            ws.cell(row=row_num, column=25, value="个")  # 单位
            ws.cell(row=row_num, column=26, value=quantity)  # 数量
            ws.cell(row=row_num, column=27, value=date.today())  # 创建日期
            ws.cell(row=row_num, column=28, value=date.today())  # 承诺日期
            ws.cell(row=row_num, column=29, value=date.today())  # 需求日期
            ws.cell(row=row_num, column=30, value=price)  # 不含税单价
            ws.cell(row=row_num, column=31, value=price)  # 含税单价
            ws.cell(row=row_num, column=32, value=0)  # 税率
            ws.cell(row=row_num, column=33, value="ANTMINER")  # 品牌/厂商
            
            row_num += 1
            line_num += 1
    
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # 保存（兼容xlsm中的图片等资源）
    try:
        wb.save(output_file)
    except KeyError as e:
        # openpyxl保存xlsm时可能遇到mimetype问题
        # 直接修补openpyxl的manifest模块
        print(f"保存xlsm时遇到问题({e})，尝试兼容保存...")
        import mimetypes as _mt
        # 注册缺失的mimetype
        _mt.add_type('image/jpeg', '.JPG')
        _mt.add_type('image/jpeg', '.jpg')
        _mt.add_type('image/png', '.PNG')
        _mt.add_type('image/png', '.png')
        _mt.add_type('image/gif', '.GIF')
        _mt.add_type('image/gif', '.gif')
        _mt.add_type('image/bmp', '.BMP')
        _mt.add_type('image/bmp', '.bmp')
        _mt.add_type('application/x-msmetafile', '.EMF')
        _mt.add_type('application/x-msmetafile', '.emf')
        _mt.add_type('application/x-msmetafile', '.WMF')
        _mt.add_type('application/x-msmetafile', '.wmf')
        try:
            wb.save(output_file)
        except KeyError:
            # 最后手段：修补openpyxl的manifest
            from openpyxl.packaging import manifest
            original_register = manifest.Manifest._register_mimetypes
            def patched_register(self, filenames=None):
                try:
                    original_register(self, filenames)
                except KeyError:
                    # 忽略未知的mimetype扩展名
                    pass
            manifest.Manifest._register_mimetypes = patched_register
            wb.save(output_file)
    wb.close()
    
    print(f"WebADI模板已填充: {output_file}")
    return True
    wb.close()
    
    print(f"数据文件已生成: {xlsx_file}")
    print("注意: 请打开WebADI模板，将此文件数据复制粘贴到模板中")
    return True

# ============== 统计表格更新模块 ==============

def update_statistics_table(order_data, entity):
    """更新统计表格
    
    文件对应关系（用户确认）：
    - 国内（SZK/ICK/HSJ/BJK）→ international_statistics_new.xlsx
    - 海外（DPT等）→ domestic_statistics.xlsx
    
    注意：文件名和实际内容是反的，这是历史命名原因
    """
    if entity in ["SZK", "ICK", "HSJ", "BJK"]:
        # 国内表格（文件名是international但内容是国内数据）
        stats_file = STATISTICS_DIR / "international_statistics_new.xlsx"
    else:
        # 海外表格（文件名是domestic但内容是海外数据）
        stats_file = STATISTICS_DIR / "domestic_statistics.xlsx"
    
    if not stats_file.exists():
        print(f"统计文件不存在: {stats_file}")
        return False
    
    wb = load_workbook(stats_file)
    
    # 选择正确的sheet（当月）
    today = date.today()
    month_name = today.strftime("%m月")
    
    if month_name in wb.sheetnames:
        ws = wb[month_name]
    else:
        # 使用第一个sheet或创建新sheet
        ws = wb.active
    
    # 找到最后一行
    last_row = ws.max_row + 1
    
    # 添加订单记录 - 按 Excel 格式：序号, 抬头, 出货日期, 主体, 测试厂, 收货地址, 型号, 物料编码, 数量, 新PO, SO, 单价, 是否已出, 对比, 标记
    for i, item in enumerate(order_data["items"]):
        row_num = last_row + i
        
        # 序号（数字格式）
        cell = ws.cell(row=row_num, column=1, value=int(row_num - 1))
        cell.number_format = '0'
        
        ws.cell(row=row_num, column=2, value="chanhua")  # 抬头
        
        # 出货日期（日期格式）
        cell = ws.cell(row=row_num, column=3, value=date.today())
        cell.number_format = 'YYYY-MM-DD'
        
        ws.cell(row=row_num, column=4, value=ENTITY_CONFIG.get(entity, {}).get("name", entity))  # 主体
        ws.cell(row=row_num, column=5, value=order_data.get("test_factory", "XJ"))  # 测试厂
        ws.cell(row=row_num, column=6, value=order_data.get("address", ""))  # 收货地址
        ws.cell(row=row_num, column=7, value=item["model"])  # 型号
        
        # 物料编码（从编码对照表查询，找不到则留空）
        model_code = get_model_code(item["model"])
        if model_code:
            ws.cell(row=row_num, column=8, value=model_code)
        else:
            # 找不到编码，用VLOOKUP公式
            ws.cell(row=row_num, column=8, value=f"=VLOOKUP(G{row_num},编码对账表!C:D,2,0)")
        
        # 数量（数字格式）
        cell = ws.cell(row=row_num, column=9, value=int(item["quantity"]))
        cell.number_format = '0'
        
        ws.cell(row=row_num, column=10, value=order_data["order_number"])  # 新PO
        ws.cell(row=row_num, column=11, value=order_data.get("sales_order_number", ""))  # SO
        
        # 单价：根据表格类型获取
        # 国内（SZK,ICK,HSJ,BJK）→ 从附件数据获取
        # 海外（DPT等）→ 从价格表查询
        if entity in ["SZK", "ICK", "HSJ", "BJK"]:
            # 国内：价格从附件获得（item_data中的price字段）
            price = item.get("price")
        else:
            # 海外：价格从价格表查询
            price = get_model_price(item["model"])
        
        if price is not None:
            cell = ws.cell(row=row_num, column=12, value=float(price))
            cell.number_format = '0.00'
        else:
            ws.cell(row=row_num, column=12, value="")  # 留空，手动填写
        
        ws.cell(row=row_num, column=13, value="")  # 是否已出（留空，后续手动填写）
        ws.cell(row=row_num, column=14, value="")  # 对比（留空）
        ws.cell(row=row_num, column=15, value="")  # 标记（留空）
    
    wb.save(stats_file)
    wb.close()
    
    print(f"统计表已更新: {stats_file}")
    return True

# ============== Oracle自动化操作模块 ==============

def automate_oracle_import(template_file, coords=None):
    """
    使用pyautogui自动化Oracle导入操作
    注意：这需要Oracle客户端已经打开
    
    参数：
        template_file: 模板文件路径
        coords: 界面坐标字典，从config.py加载
    """
    try:
        import pyautogui
        import pyperclip
    except ImportError:
        print("请安装pyautogui和pyperclip: pip install pyautogui pyperclip")
        return None
    
    # 加载坐标配置
    try:
        from config import ORACLE_COORDS
        if coords is None:
            coords = ORACLE_COORDS
    except ImportError:
        print("警告: 未找到坐标配置，使用默认值")
        coords = {
            "import_button": (100, 100),
            "file_input": (150, 150),
            "confirm_button": (200, 200),
            "export_button": (250, 250),
        }
    
    print("开始Oracle自动化导入...")
    print("请确保Oracle客户端已打开并显示在屏幕上")
    print("等待10秒，请切换到Oracle窗口...")
    time.sleep(10)
    
    sales_order_number = None
    
    try:
        # 1. 激活Oracle窗口（按Alt+Tab切换）
        pyautogui.hotkey('alt', 'tab')
        time.sleep(1)
        
        # 2. 点击导入按钮
        print(f"点击导入按钮: {coords['import_button']}")
        pyautogui.click(coords['import_button'][0], coords['import_button'][1])
        time.sleep(2)
        
        # 3. 输入模板文件路径
        print(f"输入文件路径...")
        pyautogui.click(coords['file_input'][0], coords['file_input'][1])
        time.sleep(0.5)
        pyperclip.copy(str(template_file))
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(1)
        
        # 4. 点击确认导入
        print(f"点击确认按钮: {coords['confirm_button']}")
        pyautogui.click(coords['confirm_button'][0], coords['confirm_button'][1])
        time.sleep(15)  # 等待导入处理
        
        # 5. 导出销售订单号
        print(f"点击导出按钮: {coords['export_button']}")
        pyautogui.click(coords['export_button'][0], coords['export_button'][1])
        time.sleep(3)
        
        # 6. 复制销售订单号（假设在某位置显示）
        # 这里需要根据实际情况调整
        # 尝试选中并复制
        pyautogui.click(coords.get('sales_order_field', (300, 300))[0],
                       coords.get('sales_order_field', (300, 300))[1])
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.5)
        sales_order_number = pyperclip.paste()
        
        print(f"获取销售订单号: {sales_order_number}")
        
    except Exception as e:
        print(f"Oracle自动化出错: {e}")
        return None
    
    print("Oracle自动化操作完成")
    return sales_order_number

# ============== 邮件发送模块 ==============

def send_result_email(result_data):
    """发送结果报告邮件（含附件：统计表格+采购订单模板）"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders
    
    print("正在发送结果邮件...")
    
    msg = MIMEMultipart()
    msg['From'] = CONFIG["email"]["account"]
    msg['To'] = ", ".join(CONFIG["report_emails"])
    msg['Subject'] = f"采购订单自动化报告 - {date.today().strftime('%Y-%m-%d')}"
    
    # 构建邮件正文
    new_items_text = ""
    for item in result_data.get("new_items", []):
        new_items_text += f"  - {item['model']}: {item['quantity']} PCS, 单价: {item.get('price', 'N/A')} USD, PO: {item.get('po', 'N/A')}\n"
    
    if not new_items_text:
        new_items_text = "  今日无新增数据\n"
    
    body = f"""
采购订单自动化执行报告
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
执行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
执行日期: {date.today().strftime('%Y-%m-%d')}

📊 今日新增数据:
{new_items_text}
处理结果:
- 处理邮件数: {result_data['emails_processed']}
- 生成订单数: {result_data['orders_created']}
- 状态: {result_data['status']}

📎 附件说明:
1. 国内统计表（international_statistics_new.xlsx）- 今日国内进出口数据
2. 海外统计表（domestic_statistics.xlsx）- 今日海外出口数据
3. 采购订单模板 - 今日生成的采购订单（如有）

如有问题请及时处理。
"""
    
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    # 添加附件
    today_str = date.today().strftime("%Y%m%d")
    attachments = []
    
    # 国内统计表
    domestic_file = STATISTICS_DIR / "international_statistics_new.xlsx"
    if domestic_file.exists():
        attachments.append((f"国内统计表_{today_str}.xlsx", domestic_file))
    
    # 海外统计表
    international_file = STATISTICS_DIR / "domestic_statistics.xlsx"
    if international_file.exists():
        attachments.append((f"海外统计表_{today_str}.xlsx", international_file))
    
    # 采购订单模板（所有订单汇总，.xlsm格式）
    template_file = OUTPUT_DIR / f"采购订单_{today_str}.xlsm"
    if template_file.exists():
        attachments.append((f"采购订单_{today_str}.xlsm", template_file))
    
    for filename, filepath in attachments:
        try:
            with open(filepath, 'rb') as f:
                part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                msg.attach(part)
            print(f"  附加: {filename}")
        except Exception as e:
            print(f"  附加失败 {filename}: {e}")
    
    try:
        smtp_server = CONFIG["email"]["imap_server"].replace("imap", "smtp")
        
        with smtplib.SMTP(smtp_server, 587) as server:
            server.starttls()
            server.login(CONFIG["email"]["account"], CONFIG["email"]["password"])
            server.send_message(msg)
        
        print("结果邮件已发送")
        return True
    except Exception as e:
        print(f"发送邮件失败: {e}")
        return False

# ============== 主流程 ==============

def main():
    """主执行流程"""
    print("="*60)
    print(f"采购订单自动化脚本启动 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*60)
    
    result_data = {
        "emails_processed": 0,
        "orders_created": 0,
        "order_details": "",
        "status": "成功",
        "sales_order_numbers": [],
        "new_items": [],
        "order_number": None
    }
    
    try:
        # 1. 读取今日邮件
        print("\n[1] 读取今日邮件...")
        emails = get_today_emails_from_source()
        result_data["emails_processed"] = len(emails)
        
        if not emails:
            print("今日无新邮件，退出")
            result_data["status"] = "无新邮件"
            send_result_email(result_data)
            return
        
        # 2. 加载价格表
        print("\n[2] 加载价格表...")
        prices = load_price_table()
        
        # 2.5 加载编码对照表
        print("\n[2.5] 加载编码对照表...")
        load_model_code_map()
        
        # 3. 解析邮件并创建订单
        print("\n[3] 解析邮件内容...")
        
        all_orders = []
        for email_data in emails:
            order_info = parse_order_info_from_email(email_data["body"], email_data["subject"])
            
            if not order_info["items"]:
                print(f"邮件 '{email_data['subject']}' 未解析出订单信息，跳过")
                continue
            
            if not order_info["entity"]:
                print("警告: 未检测到主体，默认使用SZK")
                order_info["entity"] = "SZK"
            
            # 匹配价格：国内从邮件内容提取，海外从价格表查询
            entity = order_info["entity"]
            for item in order_info["items"]:
                if entity in ["SZK", "ICK", "HSJ", "BJK"]:
                    # 国内：价格从邮件内容获取（item可能已有price字段）
                    if item.get("price") is None:
                        # 如果邮件没提取到价格，从价格表查
                        item["price"] = match_price(item["model"], prices)
                else:
                    # 海外：价格从价格表查询
                    item["price"] = match_price(item["model"], prices)
            
            all_orders.append(order_info)
        
        if not all_orders:
            print("无有效订单信息，退出")
            result_data["status"] = "无有效订单"
            send_result_email(result_data)
            return
        
        # 4. 生成订单号
        print("\n[4] 生成采购订单...")
        
        for order_info in all_orders:
            entity = order_info["entity"]
            
            # 获取现有订单号避免重复
            existing_numbers = get_existing_order_numbers(
                STATISTICS_DIR / "domestic_statistics.xlsx"
            ) + get_existing_order_numbers(
                STATISTICS_DIR / "international_statistics_new.xlsx"
            )
            
            order_number = generate_order_number(entity, existing_numbers)
            order_info["order_number"] = order_number
            
            print(f"订单号: {order_number}")
            print(f"主体: {ENTITY_CONFIG[entity]['name']}")
            print(f"币种: {ENTITY_CONFIG[entity]['currency']}")
            print(f"型号数量:")
            for item in order_info["items"]:
                print(f"  - {item['model']}: {item['quantity']}pcs @ {item.get('price')}")
        
        # 5. 所有订单汇总到一个WebADI模板
        print("\n[5] 汇总写入WebADI模板...")
        today_str = date.today().strftime("%Y%m%d")
        template_output = OUTPUT_DIR / f"采购订单_{today_str}.xlsm"
        fill_webadi_template(all_orders, template_output)
        result_data["order_number"] = f"采购订单_{today_str}"
        
        # 6. Oracle导入（需要Oracle客户端已打开）
        print("\n[6] Oracle导入...")
        sales_order_number = automate_oracle_import(template_output)
        if sales_order_number:
            for order_info in all_orders:
                order_info["sales_order_number"] = sales_order_number
            print(f"销售订单号: {sales_order_number}")
        
        # 7. 更新统计表
        for order_info in all_orders:
            entity = order_info["entity"]
            print(f"\n[7] 更新统计表 - {ENTITY_CONFIG[entity]['name']}...")
            update_statistics_table(order_info, entity)
            
            result_data["orders_created"] += 1
            result_data["order_details"] += f"\n订单号: {order_number}\n"
            result_data["order_number"] = order_number
            for item in order_info["items"]:
                result_data["order_details"] += f"  {item['model']}: {item['quantity']}pcs\n"
                result_data["new_items"].append({
                    "model": item["model"],
                    "quantity": item["quantity"],
                    "price": item.get("price"),
                    "po": order_number
                })
        
        # 7. 发送结果邮件
        print("\n[7] 发送结果邮件...")
        send_result_email(result_data)
        
        print("\n" + "="*60)
        print("采购订单自动化完成")
        print("="*60)
        
    except Exception as e:
        print(f"\n执行出错: {e}")
        import traceback
        traceback.print_exc()
        result_data["status"] = f"失败: {str(e)}"
        send_result_email(result_data)

if __name__ == "__main__":
    main()