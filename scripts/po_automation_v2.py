#!/usr/bin/env python3
"""
采购订单自动化脚本 v2
- 从杨娜邮件获取订单数据
- 生成WebADI xlsm (XML字符串操作)
- 更新国内采购订单模板 (XML字符串操作, 7MB)
- 更新海外统计表 (openpyxl)
- 发送结果邮件
"""

import zipfile, re, os, sys, json, shutil, smtplib, email as email_lib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header, decode_header
from email.utils import formataddr
from datetime import datetime, date, timedelta
from urllib.parse import quote
import imaplib
from openpyxl import load_workbook

# ==================== 配置 ====================
WORKSPACE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
IMAP_SERVER = "imap.appia.vip"
SMTP_SERVER = "smtp.appia.vip"
SMTP_PORT = 587
EMAIL_ACCOUNT = "yuchuan.he@casue.com"
EMAIL_PASSWORD = "-DxpOD5kkN)(RuPgAK-p"
SOURCE_EMAIL = "na.yang_w@casue.com"
REPORT_EMAILS = ["yuchuan.he@casue.com", "haixia.lu@casue.com", "yunrui.chen@casue.com", "na.yang_w@casue.com", "yujia.cheng@casue.com"]
DOMESTIC_REPORT_EMAIL = "LH-SJXPC@cbscs.com"

# ==================== 数据 (动态从邮件获取) ====================
DOMESTIC_ITEMS = []
DOMESTIC_MERGED = []
OVERSEAS_ITEMS = []
TODAY = date.today().strftime("%Y%m%d")
TODAY_DATE = date.today()
# Excel serial date: days from 1900-01-01 (Excel bug: treats 1900 as leap year)
EXCEL_DATE = (TODAY_DATE - date(1899, 12, 30)).days

# 物料编码对照表（从国内统计表的编码对账表加载）
MODEL_CODE_MAP = {}

# 价格对照表（从价格表加载）
MODEL_PRICE_MAP = {}
# 物料编码→价格对照表（从价目表0601 sheet加载，精确匹配）
CODE_PRICE_MAP = {}

# ==================== 邮件读取 ====================
def decode_email_header(header):
    """解码邮件头部"""
    if header is None:
        return ""
    decoded_parts = decode_header(header)
    result = []
    for part, charset in decoded_parts:
        if isinstance(part, bytes):
            try:
                result.append(part.decode(charset or 'utf-8', errors='ignore'))
            except:
                result.append(part.decode('utf-8', errors='ignore'))
        else:
            result.append(str(part))
    return ''.join(result)

def load_model_code_map():
    """从国内统计表的编码对账表加载物料编码对照"""
    global MODEL_CODE_MAP
    stats_file = os.path.join(WORKSPACE, "data/statistics/domestic_statistics.xlsx")
    if not os.path.exists(stats_file):
        print(f"  编码对照表不存在: {stats_file}")
        return
    try:
        wb = load_workbook(stats_file, data_only=True)
        if '编码对账表' not in wb.sheetnames:
            print("  警告: 统计表中没有'编码对账表'sheet")
            wb.close()
            return
        ws = wb['编码对账表']
        codes = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            chip_name = row[1] or row[2]   # B列=简称 or C列=名称
            material_code = row[3]  # D列=物料编码
            if chip_name and material_code and str(material_code).startswith('Y'):
                codes[str(chip_name).upper().strip()] = str(material_code)
        wb.close()
        MODEL_CODE_MAP = codes
        print(f"  已加载 {len(MODEL_CODE_MAP)} 条物料编码")
    except Exception as e:
        print(f"  加载编码对照表失败: {e}")

def load_price_map():
    """从价格表加载价格对照（价目表0601 sheet 按物料编码精确匹配）"""
    global MODEL_PRICE_MAP, CODE_PRICE_MAP
    prices_file = os.path.join(WORKSPACE, "data/prices/current_prices.xlsx")
    if not os.path.exists(prices_file):
        print(f"  价格表不存在: {prices_file}")
        return
    try:
        wb = load_workbook(prices_file, data_only=True)
        
        # 1. 从 PO sheet 加载型号→价格（fallback用）
        ws_po = wb.active
        prices = {}
        for row in ws_po.iter_rows(min_row=2, values_only=True):
            model = row[0]
            price = row[3] if len(row) > 3 else None
            if model and price:
                prices[str(model).upper().strip()] = float(price)
        MODEL_PRICE_MAP = prices
        
        # 2. 从 价目表0601 sheet 加载物料编码→价格（精确匹配）
        code_prices = {}
        if '价目表0601' in wb.sheetnames:
            ws_price = wb['价目表0601']
            for row in ws_price.iter_rows(min_row=2, values_only=True):
                code = row[2]  # C列=项目编号（物料编码）
                price = row[9]  # J列=值（价格）
                if code and price:
                    code_prices[str(code).strip()] = float(price)
        CODE_PRICE_MAP = code_prices
        
        wb.close()
        print(f"  已加载 {len(MODEL_PRICE_MAP)} 条型号价格, {len(CODE_PRICE_MAP)} 条编码价格")
    except Exception as e:
        print(f"  加载价格表失败: {e}")

def get_model_code(model):
    """根据型号获取物料编码"""
    model_upper = model.upper().strip()
    if model_upper in MODEL_CODE_MAP:
        return MODEL_CODE_MAP[model_upper]
    for base in MODEL_CODE_MAP:
        if model_upper.startswith(base):
            return MODEL_CODE_MAP[base]
    return None

def get_model_price(model, material_code=None):
    """根据型号获取价格。优先按物料编码从CODE_PRICE_MAP查，否则按型号前缀匹配。
    例如 BM1368PB(编码Y31010515) → CODE_PRICE_MAP → 4.0901(BM1368+)
    而不是前缀匹配 BM1368 → 5.0536"""
    # 1. 优先按物料编码精确查找（价目表0601 sheet）
    if material_code and material_code in CODE_PRICE_MAP:
        return CODE_PRICE_MAP[material_code]
    
    # 2. Fallback: 按型号前缀匹配
    model_upper = model.upper().strip()
    if model_upper in MODEL_PRICE_MAP:
        return MODEL_PRICE_MAP[model_upper]
    for length in range(len(model_upper), 5, -1):
        prefix = model_upper[:length]
        if prefix in MODEL_PRICE_MAP:
            return MODEL_PRICE_MAP[prefix]
    return None

def fetch_today_emails():
    """用yuchuan.he@casue.com登录IMAP，搜索杨娜今日邮件"""
    print("=== 从邮件获取今日订单 ===")
    print(f"  登录: {EMAIL_ACCOUNT}@{IMAP_SERVER}")
    print(f"  搜索: {SOURCE_EMAIL} 的今日邮件")

    imap = imaplib.IMAP4_SSL(IMAP_SERVER)
    imap.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)

    # 选择MC/po文件夹
    folder = "MC/po"
    try:
        status = imap.select(folder)
        if status[0] != "OK":
            print(f"  无法访问 {folder}，使用INBOX")
            imap.select("INBOX")
        else:
            print(f"  文件夹: {folder}")
    except Exception as e:
        print(f"  文件夹访问失败: {e}，使用INBOX")
        imap.select("INBOX")

    # 搜索今日邮件
    today_str = date.today().strftime("%d-%b-%Y")
    criteria = f'(FROM "{SOURCE_EMAIL}" ON "{today_str}")'
    status, messages = imap.search(None, criteria)
    if status != "OK":
        print("  搜索邮件失败")
        imap.logout()
        return []

    email_ids = messages[0].split()
    print(f"  找到 {len(email_ids)} 封今日邮件")

    emails_data = []
    for eid in email_ids:
        status, msg_data = imap.fetch(eid, "(RFC822)")
        if status != "OK":
            continue
        msg = email_lib.message_from_bytes(msg_data[0][1])
        subject = decode_email_header(msg.get("Subject"))
        from_addr = decode_email_header(msg.get("From"))
        date_str = decode_email_header(msg.get("Date"))

        body = ""
        attachments = []
        if msg.is_multipart():
            for part in msg.walk():
                ct = part.get_content_type()
                filename = part.get_filename()
                if filename:
                    decoded_fn = decode_email_header(filename)
                    payload = part.get_payload(decode=True)
                    if payload:
                        att_dir = os.path.join(WORKSPACE, "data/attachments")
                        os.makedirs(att_dir, exist_ok=True)
                        att_path = os.path.join(att_dir, decoded_fn)
                        with open(att_path, 'wb') as f:
                            f.write(payload)
                        attachments.append(att_path)
                        print(f"  下载附件: {decoded_fn} ({len(payload)} bytes)")
                if ct in ("text/plain", "text/html"):
                    try:
                        charset = part.get_content_charset() or 'utf-8'
                        payload = part.get_payload(decode=True)
                        if payload:
                            body += payload.decode(charset, errors='ignore')
                    except:
                        pass
        else:
            try:
                charset = msg.get_content_charset() or 'utf-8'
                payload = msg.get_payload(decode=True)
                if payload:
                    body = payload.decode(charset, errors='ignore')
            except:
                pass

        emails_data.append({
            "subject": subject,
            "from": from_addr,
            "date": date_str,
            "body": body,
            "attachments": attachments,
        })

    imap.logout()
    return emails_data

def parse_domestic_from_attachment(attachment_path):
    """从邮件附件Excel提取国内订单（进口产品统计表）"""
    print(f"  解析国内订单附件: {os.path.basename(attachment_path)}")
    today_int = int(TODAY)
    wb = None
    try:
        wb = load_workbook(attachment_path, data_only=True)
    except Exception:
        # Fallback: try xlrd for .xls files
        if attachment_path.endswith('.xls'):
            try:
                import xlrd
                print(f"    尝试用 xlrd 读取 .xls 文件...")
                return _parse_domestic_xls(attachment_path)
            except Exception as e2:
                print(f"    加载附件失败 (xlrd also failed): {e2}")
                return []
        else:
            print(f"    加载附件失败")
            return []

    # 各实体sheet的列配置
    sheet_configs = {
        "世纪云芯": {"entity": "SZK", "date_col": 4, "model_col": 8, "qty_col": 9, "price_col": 11, "supplier_col": 7, "tax_agent_col": 16, "po_col": 17, "so_col": 18, "code_col": 19},
        "智能云芯": {"entity": "ICK", "date_col": 3, "model_col": 6, "qty_col": 7, "price_col": 9, "supplier_col": 5, "tax_agent_col": 16, "po_col": 16, "so_col": 0, "code_col": 15},
        "比特方舟": {"entity": "HSJ", "date_col": 4, "model_col": 8, "qty_col": 9, "price_col": 11, "supplier_col": 7, "tax_agent_col": 16, "po_col": 17, "so_col": 18, "code_col": 19},
        "海口世纪": {"entity": "HSJ", "date_col": 4, "model_col": 8, "qty_col": 9, "price_col": 11, "supplier_col": 7, "tax_agent_col": 16, "po_col": 17, "so_col": 18, "code_col": 19},
    }

    all_items = []
    for sheet_name, cfg in sheet_configs.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        entity = cfg["entity"]

        for row in range(2, ws.max_row + 1):
            row_date = ws.cell(row=row, column=cfg["date_col"]).value
            if row_date is None:
                continue
            try:
                d = int(row_date)
                if d != today_int:
                    continue
            except:
                continue

            model = ws.cell(row=row, column=cfg["model_col"]).value
            qty = ws.cell(row=row, column=cfg["qty_col"]).value
            price = ws.cell(row=row, column=cfg["price_col"]).value
            supplier = ws.cell(row=row, column=cfg["supplier_col"]).value
            row_date = ws.cell(row=row, column=cfg["date_col"]).value

            if not model or not qty:
                continue

            # 日期直接从附件取
            item_date = str(int(row_date)) if row_date else TODAY

            # 税代直接从附件取
            tax_agent_from_att = ""
            if cfg.get("tax_agent_col", 0) > 0:
                ta = ws.cell(row=row, column=cfg["tax_agent_col"]).value
                if ta:
                    tax_agent_from_att = str(ta).strip()
            if not tax_agent_from_att:
                tax_agent_from_att = "世纪通"  # fallback

            item = {
                "model": str(model).upper().strip(),
                "qty": int(qty) if qty else 0,
                "price": float(price) if price else get_model_price(str(model), None),  # 附件优先用自带价格
                "supplier": str(supplier).strip() if supplier else "",
                "material_code": "",
                "tax_agent": tax_agent_from_att,
                "po": "",
                "so": "",
                "date": item_date,
            }

            # 物料编码 — 只接受有效编码（以Y开头），否则从编码表查询
            if cfg.get("code_col", 0) > 0:
                code = ws.cell(row=row, column=cfg["code_col"]).value
                if code and str(code).upper().startswith('Y'):
                    item["material_code"] = str(code)
            if not item["material_code"]:
                item["material_code"] = get_model_code(item["model"]) or ""

            # PO号
            if cfg.get("po_col", 0) > 0:
                po = ws.cell(row=row, column=cfg["po_col"]).value
                if po:
                    item["po"] = str(po)

            # SO号
            if cfg.get("so_col", 0) > 0:
                so = ws.cell(row=row, column=cfg["so_col"]).value
                if so:
                    item["so"] = str(so)

            all_items.append(item)
            print(f"    {sheet_name}: {item['model']} {item['qty']} PCS @ {item['price']} USD, 供应商={item['supplier']}")

    wb.close()
    return all_items

def _parse_domestic_xls(attachment_path):
    """xlr fallback for .xls files - 动态解析列位置"""
    import xlrd
    today_int = int(TODAY)
    wb = xlrd.open_workbook(attachment_path)
    all_items = []
    for sn in wb.sheet_names():
        ws = wb.sheet_by_name(sn)
        if ws.nrows < 2:
            continue
        # 从表头动态查找列索引
        headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
        def find_col(*keywords):
            for c, h in enumerate(headers):
                for kw in keywords:
                    if kw in h:
                        return c
            return -1
        col_date = find_col('下单日期', '进口日期')
        col_model = find_col('型号')
        col_qty = find_col('数量')
        col_price = find_col('报关单价', '委托报关单价')
        col_supplier = find_col('供应商')
        col_code = find_col('物料编码', 'ERPl料号')
        col_po = find_col('PO')
        col_so = find_col('SO')
        if col_model < 0 or col_qty < 0:
            print(f"    {sn}: 找不到型号/数量列，跳过")
            continue
        for r in range(1, ws.nrows):
            if col_date >= 0:
                row_date = ws.cell_value(r, col_date)
                if not row_date:
                    continue
                try:
                    d = int(row_date)
                    if d != today_int:
                        continue
                except:
                    # Maybe it's an Excel date serial number
                    try:
                        from datetime import datetime
                        dt = xlrd.xldate_as_datetime(row_date, wb.datemode)
                        if dt.strftime('%Y%m%d') != TODAY:
                            continue
                    except:
                        continue
            model = str(ws.cell_value(r, col_model)).strip()
            qty = ws.cell_value(r, col_qty)
            if not model or not qty:
                continue
            # Skip non-BM models (like 接插件, 辅料 etc)
            if not model.upper().startswith('BM'):
                continue
            price = ws.cell_value(r, col_price) if col_price >= 0 else None
            supplier = str(ws.cell_value(r, col_supplier)).strip() if col_supplier >= 0 else ""
            code = str(ws.cell_value(r, col_code)).strip() if col_code >= 0 else ""
            po = str(ws.cell_value(r, col_po)).strip() if col_po >= 0 else ""
            so = str(ws.cell_value(r, col_so)).strip() if col_so >= 0 else ""
            item = {
                "model": model.upper(),
                "qty": int(qty) if qty else 0,
                "price": float(price) if price else get_model_price(model, None),
                "supplier": supplier,
                "material_code": code if code and str(code).upper().startswith('Y') else "",
                "tax_agent": "",
                "po": po,
                "so": so,
                "date": str(int(d)) if col_date >= 0 else TODAY,
            }
            if not item["material_code"] or not item["material_code"].startswith('Y'):
                item["material_code"] = get_model_code(item["model"]) or ""
            all_items.append(item)
            print(f"    {sn}: {item['model']} {item['qty']} PCS @ {item['price']} USD, 供应商={item['supplier']}")
    return all_items

def parse_qianhai_from_email(body, subject, attachments=None):
    """从邮件正文解析前海保税区结转订单（区间结转），归入海外类"""
    print(f"  解析前海区间结转邮件: {subject}")
    
    # 供应商识别优先级：标题 > 附件文件名 > 正文
    supplier = ""
    all_kw = ["NJVT", "XJ", "SPILSZ", "ASECL", "ASE", "HN"]
    # 1. 标题
    for kw in all_kw:
        if kw in subject:
            supplier = kw
            break
    # 2. 附件文件名
    if not supplier and attachments:
        for att in attachments:
            att_name = os.path.basename(att).upper()
            for kw in all_kw:
                if kw in att_name:
                    supplier = kw
                    break
            if supplier:
                break
    # 3. 正文（同时检查"海纳"映射到HN）
    if not supplier:
        for kw in all_kw + ["海纳"]:
            if kw in body:
                if kw == "海纳":
                    supplier = "HN"
                else:
                    supplier = kw
                break
    
    items = []
    
    # 表格格式: 序号 ERP料号 型号 ... 数量 PCS 报关单价 ...
    # 1 Y31010540 BM1373AA 集成电路 8542399000 ... 124,937 PCS 27.7858 ...
    pattern = r'(\d+)\s+(Y\d+)\s+(BM\d{4}\S*)\s+集成电路\s+\d+\s+.+?([\d,]+)\s+PCS\s+([\d.]+)\s+([\d,.]+)'
    
    for m in re.findall(pattern, body):
        code = m[1]  # ERP料号
        model = m[2].upper().strip()  # 型号
        qty = int(m[3].replace(',', ''))  # 数量
        
        # 价格从价格表获取，优先按物料编码精确匹配
        price = get_model_price(model, code)
        if not price:
            # fallback to email body
            price = float(m[4])
            print(f"    ⚠ 前海 {model} 未找到价格，使用邮件中价格 {price}")
        
        print(f"    前海: {model} {qty} PCS @ {price} USD, 编码={code}, 供应商={supplier}")
        items.append({
            "model": model,
            "qty": qty,
            "price": price,
            "material_code": code,
            "supplier": supplier or "未知",
            "destination": "前海保税区",
            "po": "",
        })
    
    return items

def parse_overseas_from_email(body, subject, attachments=None):
    """从邮件正文或附件解析海外订单"""
    print(f"  解析海外订单邮件: {subject}")

    items = []
    seen = {}
    
    # 1. 从正文提取 BM 型号+数量
    patterns = [
        r'(BM\d{4}[A-Z]{0,3}[\+\w]*)\s*[:\s]\s*(\d+)\s*(?:pcs|PCS|个|片)?',
        r'(BM\d{4}[A-Z]{0,3}[\+\w]*)\s+(\d{3,})',
    ]
    for pat in patterns:
        for m in re.findall(pat, body, re.IGNORECASE):
            model = m[0].upper().strip()
            qty = int(m[1])
            if model not in seen or qty > seen[model]:
                seen[model] = qty

    # 解析目的地
    destination = ""
    if "群光" in body or "群光" in subject:
        destination = "泰国群光"
    elif "ONETEC" in body or "ONETEC" in subject:
        destination = "泰国ONETEC"
    elif "新加坡" in body:
        destination = "新加坡比特"
    elif "墨西哥" in body or "墨西哥" in subject:
        destination = "墨西哥欧陆通"
    elif "泰国" in body:
        destination = "泰国"

    # 解析供应商 - 优先标题，其次附件文件名，最后正文
    supplier = ""
    # 1. 标题
    for kw in ["XJ", "NJVT", "SPILSZ", "ASECL", "ASE", "HN"]:
        if kw in subject:
            supplier = kw
            break
    # 2. 附件文件名
    if not supplier and attachments:
        for att in attachments:
            att_name = os.path.basename(att).upper()
            for kw in ["XJ", "NJVT", "SPILSZ", "ASECL", "ASE", "HN"]:
                if kw in att_name:
                    supplier = kw
                    break
            if supplier:
                break
    # 3. 正文
    if not supplier:
        for kw in ["XJ", "信佳", "NJVT", "南京", "SPILSZ", "HN", "海纳"]:
            if kw in body:
                if kw == "信佳":
                    supplier = "XJ"
                elif kw == "南京":
                    supplier = "NJVT"
                elif kw == "海纳":
                    supplier = "HN"
                else:
                    supplier = kw
                break
    
    # 4. 如果正文没有型号数据，从附件提取
    if not seen and attachments:
        print(f"    正文无型号数据，尝试从附件提取...")
        for att in attachments:
            if not att.endswith(('.xlsx', '.xls')):
                continue
            try:
                att_wb = load_workbook(att, data_only=True)
                for sn in att_wb.sheetnames:
                    if '发票' not in sn and '箱单' not in sn:
                        continue
                    ws = att_wb[sn]
                    for r in range(1, ws.max_row + 1):
                        for c in range(1, min(ws.max_column + 1, 12)):
                            val = str(ws.cell(row=r, column=c).value or '')
                            bm = re.search(r'(BM\d{4}[A-Z]{0,3})', val)
                            if bm:
                                model = bm.group(1).upper().strip()
                                # 找同一行的数量
                                for c2 in range(c + 1, min(ws.max_column + 1, c + 6)):
                                    qty_val = ws.cell(row=r, column=c2).value
                                    if qty_val and isinstance(qty_val, (int, float)) and qty_val > 100:
                                        # 确认是数量（不是价格）
                                        if model not in seen or int(qty_val) > seen[model]:
                                            seen[model] = int(qty_val)
                            # Also try TOTAL row
                            if 'TOTAL' in val.upper() and not seen:
                                for c3 in range(1, c):
                                    prev = str(ws.cell(row=r-2, column=c3).value or '')
                                    if not seen:
                                        # Look at rows above for model
                                        pass
                att_wb.close()
                if seen:
                    break
            except Exception as e:
                print(f"    附件解析失败: {e}")

    for model, qty in seen.items():
        code = get_model_code(model)
        price = get_model_price(model, code)
        if not price:
            print(f"    ⚠ 未找到 {model} 的价格，需手动填写")
            price = 0
        item = {
            "model": model,
            "qty": qty,
            "price": price,
            "supplier": supplier,
            "material_code": get_model_code(model) or "",
            "destination": destination,
            "po": "",
            "so": "",
        }
        items.append(item)
        print(f"    海外: {model} {qty} PCS @ {price} USD, 目的地={destination}")

    return items

def generate_po_number(prefix, existing_items):
    """生成PO号: prefix + YYYYMMDD + 序号 (统一从6001起)"""
    today_str = TODAY
    start_seq = 6000
    max_seq = start_seq
    for item in existing_items:
        po = item.get("po", "")
        if po.startswith(prefix + today_str):
            try:
                seq = int(po[len(prefix + today_str):])
                max_seq = max(max_seq, seq)
            except:
                pass
    return f"{prefix}{today_str}{max_seq + 1:04d}"

def fetch_and_parse_orders():
    """主函数：从邮件获取并解析订单，设置DOMESTIC_ITEMS和OVERSEAS_ITEMS"""
    global DOMESTIC_ITEMS, DOMESTIC_MERGED, OVERSEAS_ITEMS

    # 加载编码和价格对照表
    load_model_code_map()
    load_price_map()

    # 获取今日邮件
    emails = fetch_today_emails()
    if not emails:
        print("  今日无新邮件，无需处理")
        return False

    domestic_items = []
    overseas_items = []
    processed_attachments = set()  # 去重：同一附件只处理一次
    domestic_emails = set()  # 国内邮件的id，不再重复解析海外

    # Detect "此份为准" emails — they supersede the original with same base subject
    zhunque_bases = set()
    for em in emails:
        subj = em["subject"]
        if "此份为准" in subj:
            base = subj.replace("Re: ", "").replace("(此份为准）", "").replace("（此份为准）", "").replace("此份为准", "").strip().rstrip(")").strip()
            zhunque_bases.add(base)
    if zhunque_bases:
        print(f"  检测到 {len(zhunque_bases)} 封'此份为准'邮件，将跳过原始版本")

    for idx, em in enumerate(emails):
        subject = em["subject"]
        body = em["body"]
        attachments = em["attachments"]
        print(f"\n  处理邮件[{idx}]: {subject}")

        # Skip if a "此份为准" version exists for this subject
        if "此份为准" not in subject:
            base_subj = subject.replace("Re: ", "").strip()
            superseded = False
            for zb in zhunque_bases:
                if base_subj.startswith(zb[:30]) or zb.startswith(base_subj[:30]):
                    print(f"    跳过（已有'此份为准'版本）")
                    superseded = True
                    break
            if superseded:
                continue

        # 有附件且主题含"进口产品统计表" → 国内订单
        is_domestic = attachments and ("进口产品统计表" in subject)
        # 主题含"进口产品统计表"但无附件 → 跳过（附件已在其他邮件中）
        is_domestic_no_att = (not attachments) and ("进口产品统计表" in subject)
        # 前海保税区结转 → 单独的解析路径（不识别为海外或国内）
        is_qianhai = "区间结转" in subject or "前海" in subject
        # 海外关键词：出口/出货通知/海外/DPT/清关资料/墨西哥
        overseas_keywords = ["出口", "出货通知", "海外", "DPT", "清关资料", "墨西哥"]
        is_overseas = any(kw in subject for kw in overseas_keywords) and not is_domestic and not is_qianhai
        # 有出口相关附件但没命中海外关键字 → 也尝试解析
        has_overseas_attachment = attachments and any(
            kw in att for att in attachments for kw in ["出口", "墨西哥", "ONETEC", "群光"]
        )

        # 国内订单：从附件解析
        if is_domestic:
            for att in attachments:
                if att in processed_attachments:
                    print(f"    跳过重复附件: {os.path.basename(att)}")
                    continue
                if att.endswith(('.xlsx', '.xls')):
                    processed_attachments.add(att)
                    items = parse_domestic_from_attachment(att)
                    domestic_items.extend(items)
            domestic_emails.add(idx)
            continue  # 国内邮件不再解析海外

        # 国内主题但无附件 → 跳过
        if is_domestic_no_att:
            print(f"    主题含'进口产品统计表'但无附件，跳过")
            continue

        # 前海保税区结转 → 区间结转邮件，按邮件DATE过滤日期避免跨天重复
        if is_qianhai:
            # 提取邮件日期，只处理当天的前海邮件
            em_date_str = em.get("date", "")
            # 尝试从邮件Date头提取日期
            from email.utils import parsedate_to_datetime as pdt
            try:
                em_dt = pdt(em_date_str).date()
                if em_dt != TODAY_DATE:
                    print(f"    前海邮件日期={em_dt}≠今天={TODAY_DATE}，跳过")
                    continue
            except:
                pass  # 无法解析日期则继续处理
            items = parse_qianhai_from_email(body, subject, attachments)
            if items:
                overseas_items.extend(items)
            continue

        # 海外订单：从正文解析 OR 有墨西哥/出口附件
        if is_overseas or has_overseas_attachment or (not is_domestic and not is_qianhai and not attachments):
            items = parse_overseas_from_email(body, subject, attachments)
            if items:
                overseas_items.extend(items)

    # 国内不汇总；海外相同料号+目的地汇总
    def merge_same_model(items, key_fields=('model',)):
        merged = {}
        for item in items:
            key = tuple(item.get(f, '') for f in key_fields)
            if key in merged:
                merged[key]['qty'] += item['qty']
            else:
                merged[key] = item.copy()
        return list(merged.values())
    
    # 国内不合并（多条记录同一PO），海外按型号+目的地合并
    # domestic_items 保持不变
    # 海外：相同料号+目的地+供应商汇总
    overseas_items = merge_same_model(overseas_items, ('model', 'destination', 'supplier'))

    # 海外：合并后各自生成PO号
    for item in overseas_items:
        if not item.get("po"):
            item["po"] = generate_po_number("DPT", overseas_items)

    # 国内：按型号+供应商分组生成PO（同测试厂同料号共享PO），统计表保留所有行
    model_po = {}
    for item in domestic_items:
        key = (item['model'], item.get('supplier', ''))
        if key not in model_po:
            model_po[key] = generate_po_number("SZK", domestic_items)
        item['po'] = model_po[key]

    # 国内按PO合并数量（统计表保留明细，模板合并）——用于WebADI和summary
    domestic_merged = {}
    for item in domestic_items:
        po = item['po']
        if po in domestic_merged:
            domestic_merged[po]['qty'] += item['qty']
        else:
            domestic_merged[po] = item.copy()
    DOMESTIC_MERGED = list(domestic_merged.values())
    
    DOMESTIC_ITEMS = domestic_items
    OVERSEAS_ITEMS = overseas_items

    print(f"\n  国内订单: {len(DOMESTIC_ITEMS)} 条")
    for item in DOMESTIC_ITEMS:
        print(f"    {item['po']}: {item['model']} {item['qty']} PCS")
    print(f"  海外订单: {len(OVERSEAS_ITEMS)} 条")
    for item in OVERSEAS_ITEMS:
        print(f"    {item['po']}: {item['model']} {item['qty']} PCS")

    return len(DOMESTIC_ITEMS) > 0 or len(OVERSEAS_ITEMS) > 0

# ==================== 工具函数 ====================
def get_ss_map(ss_xml):
    """Build index→text map from sharedStrings XML using <si> tag counting"""
    idx_map = {}
    positions = list(re.finditer(r'<si\b', ss_xml))
    for i, pos in enumerate(positions):
        s = pos.start()
        e = ss_xml.find('</si>', s) + len('</si>')
        block = ss_xml[s:e]
        texts = re.findall(r'<t[^>]*>(.*?)</t>', block)
        idx_map[i] = ''.join(texts)
    return idx_map, len(positions)

def find_or_add_ss(ss_xml, target, total_count):
    """Find string in sharedStrings or return (index, new_total_count, new_ss_xml)"""
    idx_map, count = get_ss_map(ss_xml)
    for idx, text in idx_map.items():
        if text == target:
            return idx, count, ss_xml
    # Not found, add new
    new_entry = f'<si><t>{target}</t></si>'
    # Insert before </sst>
    insert_pos = ss_xml.rfind('</sst>')
    new_ss = ss_xml[:insert_pos] + new_entry + ss_xml[insert_pos:]
    return count, count + 1, new_ss

def add_multiple_ss(ss_xml, strings_to_add):
    """Add multiple strings to sharedStrings, return (idx_map, new_ss_xml)"""
    idx_map, count = get_ss_map(ss_xml)
    new_entries = []
    result_map = {}
    
    for s in strings_to_add:
        found = False
        for idx, text in idx_map.items():
            if text == s:
                result_map[s] = idx
                found = True
                break
        if not found:
            result_map[s] = count + len(new_entries)
            new_entries.append(f'<si><t>{s}</t></si>')
    
    if new_entries:
        insert_pos = ss_xml.rfind('</sst>')
        ss_xml = ss_xml[:insert_pos] + ''.join(new_entries) + ss_xml[insert_pos:]
        # Update count attribute
        new_count = count + len(new_entries)
        ss_xml = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{new_count}"', ss_xml, count=1)
    
    return result_map, ss_xml

# ==================== 1. 生成WebADI xlsm ====================
def generate_xlsm():
    print("=== 生成WebADI xlsm ===")
    template_path = os.path.join(WORKSPACE, "data/templates/webadi_template.xlsm")
    output_path = os.path.join(WORKSPACE, f"data/output/采购订单_{TODAY}.xlsm")
    
    z = zipfile.ZipFile(template_path, 'r')
    sheet2_raw = z.read('xl/worksheets/sheet2.xml').decode('utf-8')
    ss_raw = z.read('xl/sharedStrings.xml').decode('utf-8')
    
    # Read all other files
    other_files = {}
    for name in z.namelist():
        if name not in ['xl/worksheets/sheet2.xml', 'xl/sharedStrings.xml']:
            other_files[name] = z.read(name)
    z.close()
    
    # Template sharedStrings key indices (verified)
    IDX = {
        'SZK': 87, '标准采购订单': 88, 'USD': 52, '何宇川': 89,
        'BITMAIN': 90, 'SG': 91, 'XAP': 92, '1004': 93, 'SZKXYCL': 94,
        '付款方式一': 59, '生产用料销售': 60, 'Y': 61, '手工录入': 62,
        'BM系列': 63, '个': 65, '0': 66, 'ANTMINER': 67, 'J': 109,
        'CNY': 128,
    }
    
    # Strings to add to sharedStrings
    new_strings = []
    for item in DOMESTIC_ITEMS:
        new_strings.extend([item['po'], item['material_code'], item['model']])
    for item in OVERSEAS_ITEMS:
        new_strings.extend([item['po'], item['material_code'], item['model']])
    # Add model description strings
    for item in DOMESTIC_ITEMS:
        new_strings.append(f"{item['model']}芯片")
    for item in OVERSEAS_ITEMS:
        new_strings.append(f"{item['model']}芯片")
    # DPT海外专用字符串
    new_strings.extend([
        'DPT',                        # 业务实体(海外)
        'CHANHUA PTE. LTD.',          # 供应商(海外)
        '费用',                       # 供应商地点(海外)
        '1155.BITMAIN DEVELOPMENT PTE. LTD.',  # 收货方(海外)
    ])
    # 海外目的子库存 - 根据收货地址
    for item in OVERSEAS_ITEMS:
        dest_subinv_map = {
            '泰国群光': 'DPTHQGCP',
            '泰国ONETEC': 'DPTONETYCL',
            'PIE': 'DPTPIECL',
            '前海保税区': 'DPTQHBSC',
            '墨西哥欧陆通': 'DPTMOLTYCL',
        }
        dest_subinv = dest_subinv_map.get(item.get('destination', ''), 'DPTHQGCP')
        new_strings.append(dest_subinv)
        item['dest_subinv'] = dest_subinv
    
    new_strings = list(set(new_strings))  # deduplicate
    
    ss_map, ss_new = add_multiple_ss(ss_raw, new_strings)
    
    # Merge with existing known indices
    for k, v in IDX.items():
        if k not in ss_map:
            ss_map[k] = v
    
    # Generate row XML for each order
    def make_xlsm_row(row_num, item, is_domestic=True):
        row = f'<row r="{row_num}" spans="2:39" ht="14.25" outlineLevel="1">'
        
        if is_domestic:
            # 国内SZK订单
            entity_idx = ss_map["SZK"]
            supplier_idx = ss_map["BITMAIN"]
            supplier_loc_idx = ss_map["SG"]
            source_idx = ss_map["XAP"]
            recv_idx = ss_map["1004"]
            dest_idx = ss_map["SZKXYCL"]
            recv2_idx = ss_map["1004"]
        else:
            # 海外DPT订单
            entity_idx = ss_map["DPT"]
            supplier_idx = ss_map["CHANHUA PTE. LTD."]
            supplier_loc_idx = ss_map["费用"]
            source_idx = ss_map["XAP"]  # 来源子库存统一=XAP
            recv_idx = ss_map["1155.BITMAIN DEVELOPMENT PTE. LTD."]
            dest_idx = ss_map[item['dest_subinv']]  # 目的子库存根据收货地址
            recv2_idx = ss_map["1155.BITMAIN DEVELOPMENT PTE. LTD."]
        
        # B: (empty style)
        row += f'<c r="B{row_num}" s="7"/>'
        # C: 业务实体
        row += f'<c r="C{row_num}" s="4" t="s"><v>{entity_idx}</v></c>'
        # D: 类型 = 标准采购订单
        row += f'<c r="D{row_num}" s="5" t="s"><v>{ss_map["标准采购订单"]}</v></c>'
        # E: 采购订单号
        row += f'<c r="E{row_num}" s="5" t="s"><v>{ss_map[item["po"]]}</v></c>'
        # F: 币种 = USD
        row += f'<c r="F{row_num}" s="18" t="s"><v>{ss_map["USD"]}</v></c>'
        # G: 采购员 = 何宇川
        row += f'<c r="G{row_num}" s="4" t="s"><v>{ss_map["何宇川"]}</v></c>'
        # H: 供应商
        row += f'<c r="H{row_num}" s="4" t="s"><v>{supplier_idx}</v></c>'
        # I: 供应商地点
        row += f'<c r="I{row_num}" s="4" t="s"><v>{supplier_loc_idx}</v></c>'
        # J: 来源子库存
        row += f'<c r="J{row_num}" s="5" t="s"><v>{source_idx}</v></c>'
        # K: 收货方
        row += f'<c r="K{row_num}" s="4" t="s"><v>{recv_idx}</v></c>'
        # L: 目的子库存
        row += f'<c r="L{row_num}" s="5" t="s"><v>{dest_idx}</v></c>'
        # M: 收单方
        row += f'<c r="M{row_num}" s="4" t="s"><v>{recv2_idx}</v></c>'
        # N: 付款方式
        row += f'<c r="N{row_num}" s="5" t="s"><v>{ss_map["付款方式一"]}</v></c>'
        # O: 内部申请类型
        row += f'<c r="O{row_num}" s="5" t="s"><v>{ss_map["生产用料销售"]}</v></c>'
        # P: (empty style)
        row += f'<c r="P{row_num}" s="5"/>'
        # Q: 货贷 = Y
        row += f'<c r="Q{row_num}" s="5" t="s"><v>{ss_map["Y"]}</v></c>'
        # R: (empty)
        row += f'<c r="R{row_num}" s="5"/>'
        # S: (empty)
        row += f'<c r="S{row_num}" s="5"/>'
        # T: 摘要 = 手工录入
        row += f'<c r="T{row_num}" s="5" t="s"><v>{ss_map["手工录入"]}</v></c>'
        # U: 行号
        row += f'<c r="U{row_num}" s="4"><v>{item.get("line_no", 1)}</v></c>'
        # V: 行类型 = BM系列
        row += f'<c r="V{row_num}" s="4" t="s"><v>{ss_map["BM系列"]}</v></c>'
        # W: 物料编码
        row += f'<c r="W{row_num}" s="4" t="s"><v>{ss_map[item["material_code"]]}</v></c>'
        # X: 物料说明 = 不填(空)
        row += f'<c r="X{row_num}" s="5"/>'
        # Y: 单位 = 个
        row += f'<c r="Y{row_num}" s="5" t="s"><v>{ss_map["个"]}</v></c>'
        # Z: 数量
        row += f'<c r="Z{row_num}" s="4"><v>{item["qty"]}</v></c>'
        # AA: 创建日期
        row += f'<c r="AA{row_num}" s="6"><v>{EXCEL_DATE}</v></c>'
        # AB: 承诺日期
        row += f'<c r="AB{row_num}" s="6"><v>{EXCEL_DATE}</v></c>'
        # AC: 需求日期
        row += f'<c r="AC{row_num}" s="6"><v>{EXCEL_DATE}</v></c>'
        # AD: 不含税单价
        row += f'<c r="AD{row_num}" s="15"><v>{item["price"]}</v></c>'
        # AE: 含税单价
        row += f'<c r="AE{row_num}" s="15"><v>{item["price"]}</v></c>'
        # AF: 税率 = 0
        row += f'<c r="AF{row_num}" s="5" t="s"><v>{ss_map["0"]}</v></c>'
        # AG: 品牌 = ANTMINER
        row += f'<c r="AG{row_num}" s="5" t="s"><v>{ss_map["ANTMINER"]}</v></c>'
        # AH: 最小包装 (空)
        row += f'<c r="AH{row_num}" s="8"/>'
        # AI-AK: (empty)
        row += f'<c r="AI{row_num}" s="5"/>'
        row += f'<c r="AJ{row_num}" s="5"/>'
        row += f'<c r="AK{row_num}" s="8"/>'
        
        row += '</row>'
        return row
    
    # Build ALL rows from row 10 to template_max_row
    # - Rows 1-9: kept from template (header area)
    # - Rows 10+: regenerated — data rows + clean empty padding (no shifting of stale template rows)
    all_items = list(DOMESTIC_MERGED) + list(OVERSEAS_ITEMS)
    
    # Figure out how many rows the template has
    template_max = 1005
    row_matches = re.findall(r'<row r="(\d+)"', sheet2_raw)
    if row_matches:
        template_max = max(int(r) for r in row_matches)
    
    new_rows = []
    for i in range(template_max - 9):  # rows 10 through template_max inclusive
        rn = 10 + i
        if i < len(all_items):
            item = all_items[i]
            is_dom = i < len(DOMESTIC_MERGED)
            new_rows.append(make_xlsm_row(rn, item, is_domestic=is_dom))
        else:
            # Clean empty padding row — styles only, NO shared string content
            # Same 38-cell (B:AM) structure as template, no <v> tags to avoid corrupting shared strings
            empty_cells = (
                f'<c r="B{rn}" s="7"/>'
                f'<c r="C{rn}" s="4"/>'
                f'<c r="D{rn}" s="5"/>'
                f'<c r="E{rn}" s="5"/>'
                f'<c r="F{rn}" s="5"/>'
                f'<c r="G{rn}" s="4"/>'
                f'<c r="H{rn}" s="4"/>'
                f'<c r="I{rn}" s="4"/>'
                f'<c r="J{rn}" s="5"/>'
                f'<c r="K{rn}" s="4"/>'
                f'<c r="L{rn}" s="5"/>'
                f'<c r="M{rn}" s="4"/>'
                f'<c r="N{rn}" s="5"/>'
                f'<c r="O{rn}" s="5"/>'
                f'<c r="P{rn}" s="5"/>'
                f'<c r="Q{rn}" s="5"/>'
                f'<c r="R{rn}" s="5"/>'
                f'<c r="S{rn}" s="5"/>'
                f'<c r="T{rn}" s="5"/>'
                f'<c r="U{rn}" s="4"/>'
                f'<c r="V{rn}" s="4"/>'
                f'<c r="W{rn}" s="4"/>'
                f'<c r="X{rn}" s="5"/>'
                f'<c r="Y{rn}" s="5"/>'
                f'<c r="Z{rn}" s="4"/>'
                f'<c r="AA{rn}" s="6"/>'
                f'<c r="AB{rn}" s="6"/>'
                f'<c r="AC{rn}" s="6"/>'
                f'<c r="AD{rn}" s="4"/>'
                f'<c r="AE{rn}" s="5"/>'
                f'<c r="AF{rn}" s="5"/>'
                f'<c r="AG{rn}" s="5"/>'
                f'<c r="AH{rn}" s="5"/>'
                f'<c r="AI{rn}" s="5"/>'
                f'<c r="AJ{rn}" s="5"/>'
                f'<c r="AK{rn}" s="8"/>'
                f'<c r="AL{rn}" s="13"/>'
                f'<c r="AM{rn}" s="9"/>'
            )
            new_rows.append(f'<row r="{rn}" spans="2:39" ht="14.25" outlineLevel="1">{empty_cells}</row>')
    
    # Keep template rows 1-9, replace rows 10+ with generated rows, preserve template tail
    r9_end = sheet2_raw.find('</row>', sheet2_raw.find('<row r="9"')) + len('</row>')
    sd_pos = sheet2_raw.find('</sheetData>')
    tail = sheet2_raw[sd_pos:]  # keeps </sheetData> + mergeCells + conditionalFormatting + </worksheet> etc.
    sheet2_new = sheet2_raw[:r9_end] + ''.join(new_rows) + tail
    
    # Write output xlsm
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in other_files.items():
            zout.writestr(name, data)
        zout.writestr('xl/worksheets/sheet2.xml', sheet2_new.encode('utf-8'))
        zout.writestr('xl/sharedStrings.xml', ss_new.encode('utf-8'))
    
    print(f"  生成: {output_path} ({os.path.getsize(output_path)} bytes)")
    return output_path

# ==================== 2. 生成采购订单数据xlsx ====================
def generate_summary_xlsx():
    print("=== 生成采购订单数据xlsx ===")
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    
    # 用第一个国内PO号命名，没有则用TODAY
    first_po = DOMESTIC_ITEMS[0]['po'] if DOMESTIC_ITEMS else f"PO_{TODAY}"
    output_path = os.path.join(WORKSPACE, f"data/output/{first_po}.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "采购订单数据"
    
    headers = ['加载', '业务实体', '类型', '采购订单号', '币种', '采购员', '供应商',
               '供应商地点', '来源子库存', '收货方', '目的子库存', '收单方',
               '付款方式', '内部申请类型', '货贷', '是否报关', '加工费报价OA单据号',
               '摘要', '业务模式', '行号', '行类型', '物料', '物料说明', '单位',
               '数量', '创建日期', '承诺日期', '需求日期', '不含税单价', '含税单价',
               '税率', '品牌/厂商']
    
    ws.append(headers)
    
    all_items = []
    domestic_data = [
        ('Y', 'SZK', '标准采购订单', item['po'], 'USD', '何宇川,', 'BITMAIN DEVELOPMENT PTE.  LTD.',
         'SG', 'XAP', '1004.Bitmain Shenzhen', 'SZKXYCL', '1004.Bitmain Shenzhen',
         '付款方式一', '生产用料销售', '', 'Y', '', '手工录入', '', i+1, 'BM系列',
         item['material_code'], f"{item['model']}芯片", '个', item['qty'],
         TODAY_DATE, TODAY_DATE, TODAY_DATE, item['price'], item['price'], '0', 'ANTMINER')
        for i, item in enumerate(DOMESTIC_MERGED)
    ]
    
    dest_subinv_map = {
        '泰国群光': 'DPTHQGCP', '泰国ONETEC': 'DPTONETYCL', 'PIE': 'DPTPIECL',
        '前海保税区': 'DPTQHBSC', '墨西哥欧陆通': 'DPTMOLTYCL',
    }
    overseas_data = [
        ('Y', 'DPT', '标准采购订单', item['po'], 'USD', '何宇川,', 'CHANHUA PTE. LTD.',
         '费用', 'XAP', '1155.BITMAIN DEVELOPMENT PTE. LTD.', dest_subinv_map.get(item.get('destination',''), 'DPTHQGCP'), '1155.BITMAIN DEVELOPMENT PTE. LTD.',
         '付款方式一', '生产用料销售', '', 'Y', '', '手工录入', '', 1, 'BM系列',
         item['material_code'], f"{item['model']}芯片", '个', item['qty'],
         TODAY_DATE, TODAY_DATE, TODAY_DATE, item['price'], item['price'], '0', 'ANTMINER')
        for item in OVERSEAS_ITEMS
    ]
    
    all_rows = domestic_data + overseas_data
    for row_data in all_rows:
        ws.append(row_data)
    
    wb.save(output_path)
    print(f"  生成: {output_path} ({os.path.getsize(output_path)} bytes)")
    return output_path

# ==================== 3. 更新国内采购订单模板 ====================
def update_international_template():
    print("=== 更新国内采购订单模板 ===")
    output_path = os.path.join(WORKSPACE, "data/output/international_statistics_new.xlsx")
    
    # 找前一天的输出文件作为base，没有则用模板
    prev_output = os.path.join(WORKSPACE, "data/output/国内进口产品统计表.xlsx")
    template_path = os.path.join(WORKSPACE, "data/templates/国内采购订单模板.xlsx")
    
    base_path = prev_output if os.path.exists(prev_output) else template_path
    print(f"  使用base: {base_path} ({'前一天输出' if base_path == prev_output else '原始模板'})")
    
    z = zipfile.ZipFile(base_path, 'r')
    sheet13_raw = z.read('xl/worksheets/sheet13.xml').decode('utf-8')
    ss_raw = z.read('xl/sharedStrings.xml').decode('utf-8')
    
    other_files = {}
    for name in z.namelist():
        if name not in ['xl/worksheets/sheet13.xml', 'xl/sharedStrings.xml']:
            other_files[name] = z.read(name)
    z.close()
    
    # Key existing indices in国内模板 sharedStrings
    # 深圳世纪云芯=4468, 新加坡比特=828, 集成电路=5224, SPILSZ=58, XJ=13
    # BM1373AA=3388, PCS=5634, Y31010540=3901, 世纪通=44
    
    # Add new strings: PO numbers, SO numbers, supplier, model, material_code, tax_agent
    new_strings = []
    for item in DOMESTIC_ITEMS:
        new_strings.append(item['po'])
        new_strings.append(item['date'])
        if item.get('so'):
            new_strings.append(item['so'])
        if item.get('supplier'):
            new_strings.append(item['supplier'])
        if item.get('model'):
            new_strings.append(item['model'])
        if item.get('material_code'):
            new_strings.append(item['material_code'])
        if item.get('tax_agent'):
            new_strings.append(item['tax_agent'])
    
    ss_map, ss_new = add_multiple_ss(ss_raw, new_strings)
    
    # Add known indices
    KNOWN = {
        '深圳世纪云芯': 4468, '新加坡比特': 828, '集成电路': 5224,
        'SPILSZ': 58, 'XJ': 13, 'ASE': 7671,
        'BM1373AA': 3388, 'BM1374CC': 5562, 'BM1493AA': 3389,
        'BM1746AA': 3390, 'BM1374AA': 3387,
        'PCS': 5634, 'Y31010540': 3901, 'Y31010551': 5635,
        'Y31030503': 3903, 'Y09BM1746010': 3910,
        '世纪通': 44, '朗华': 41, '富森': 50,
    }
    for k, v in KNOWN.items():
        if k not in ss_map:
            ss_map[k] = v
    
    # Find last row (162) and add new rows after it
    # Row style from row162: B=260, C=261, D=262, E=261, F=261, G=261, H=261, I=261, J=263, K=101, L=264, N=261
    # But row160 style: B=87, C=36, D=35, E=36, F=36, G=36, H=36, I=36, J=41, K=151, L=259, M=720, N=36
    
    # Use row160 style (more standard)
    def make_domestic_row(row_num, item):
        supplier_idx = ss_map.get(item['supplier'], KNOWN.get(item['supplier']))
        model_idx = ss_map.get(item['model'], KNOWN.get(item['model']))
        material_idx = ss_map.get(item['material_code'], KNOWN.get(item['material_code']))
        po_idx = ss_map[item['po']]
        tax_agent_idx = ss_map.get(item['tax_agent'], KNOWN.get(item['tax_agent']))
        
        row = f'<row r="{row_num}" spans="1:17">'
        # B: 采购主体 = 深圳世纪云芯
        row += f'<c r="B{row_num}" s="87" t="s"><v>{ss_map["深圳世纪云芯"]}</v></c>'
        # C: 出货日期 = from attachment
        row += f'<c r="C{row_num}" s="36"><v>{item["date"]}</v></c>'
        # D: 销售主体 = 新加坡比特
        row += f'<c r="D{row_num}" s="35" t="s"><v>{ss_map["新加坡比特"]}</v></c>'
        # E: 协议 = 集成电路
        row += f'<c r="E{row_num}" s="36" t="s"><v>{ss_map["集成电路"]}</v></c>'
        # F: 供应商 = from attachment
        row += f'<c r="F{row_num}" s="36" t="s"><v>{supplier_idx}</v></c>'
        # G: 型号 = from attachment
        row += f'<c r="G{row_num}" s="36" t="s"><v>{model_idx}</v></c>'
        # H: 数量
        row += f'<c r="H{row_num}" s="36"><v>{item["qty"]}</v></c>'
        # I: 单位 = PCS
        row += f'<c r="I{row_num}" s="36" t="s"><v>{ss_map["PCS"]}</v></c>'
        # J: 报关单价
        row += f'<c r="J{row_num}" s="41"><v>{item["price"]}</v></c>'
        # K: 物料编码
        row += f'<c r="K{row_num}" s="151" t="s"><v>{material_idx}</v></c>'
        # L: PO号
        row += f'<c r="L{row_num}" s="259" t="s"><v>{po_idx}</v></c>'
        # M: SO号
        if item.get('so'):
            so_idx = ss_map[item['so']]
            row += f'<c r="M{row_num}" s="259" t="s"><v>{so_idx}</v></c>'
        # N: 税代 = from attachment
        row += f'<c r="N{row_num}" s="36" t="s"><v>{tax_agent_idx}</v></c>'
        row += '</row>'
        return row
    
    # Deduplicate: check existing PO numbers in sheet13, skip items already in the table
    # PO numbers are in column L, which are shared strings indices - we need to find actual PO text
    existing_po = set()
    for cell_match in re.finditer(r'<c r="L\d+"[^>]*t="s"[^>]*><v>(\d+)</v></c>', sheet13_raw):
        ss_idx = int(cell_match.group(1))
        # Extract the PO string from sharedStrings at this index
        si_elements = re.findall(r'<si>(.*?)</si>', ss_raw, re.DOTALL)
        if ss_idx < len(si_elements):
            po_text = re.sub(r'<[^>]+>', '', si_elements[ss_idx]).strip()
            if po_text.startswith('SZK') or po_text.startswith('BTD'):
                existing_po.add(po_text)
    
    domestic_new = [item for item in DOMESTIC_ITEMS if item['po'] not in existing_po]
    skipped = len(DOMESTIC_ITEMS) - len(domestic_new)
    if skipped > 0:
        print(f"  去重: 跳过 {skipped} 条已存在的记录 (PO号: {[item['po'] for item in DOMESTIC_ITEMS if item not in domestic_new]})")
    
    if not domestic_new:
        print("  无新增数据，不追加行")
        # Still need to write output
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name, data in other_files.items():
                zout.writestr(name, data)
            zout.writestr('xl/worksheets/sheet13.xml', sheet13_raw.encode('utf-8'))
            zout.writestr('xl/sharedStrings.xml', ss_new.encode('utf-8'))
        print(f"  生成: {output_path} ({os.path.getsize(output_path)} bytes)")
        return output_path

    # 动态找到最后一个数据行
    all_row_nums = [int(m.group(1)) for m in re.finditer(r'<row r="(\d+)"', sheet13_raw)]
    last_row_num = max(all_row_nums)
    print(f"  sheet13 last row: {last_row_num}")
    
    last_row_match = re.search(f'<row r="{last_row_num}".*?</row>', sheet13_raw, re.DOTALL)
    if not last_row_match:
        print("  ERROR: Last row not found!")
        return None
    
    # Insert new rows after last data row
    new_rows = []
    row_num = last_row_num + 1
    for item in domestic_new:
        new_rows.append(make_domestic_row(row_num, item))
        row_num += 1
    
    insert_pos = last_row_match.end()
    sheet13_new = sheet13_raw[:insert_pos] + ''.join(new_rows) + sheet13_raw[insert_pos:]
    
    # Write output
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in other_files.items():
            zout.writestr(name, data)
        zout.writestr('xl/worksheets/sheet13.xml', sheet13_new.encode('utf-8'))
        zout.writestr('xl/sharedStrings.xml', ss_new.encode('utf-8'))
    
    print(f"  生成: {output_path} ({os.path.getsize(output_path)} bytes)")
    return output_path

# ==================== 4. 更新海外统计表 ====================
def update_domestic_statistics():
    print("=== 更新海外统计表 ===")
    import openpyxl
    
    # 找前一天的输出文件作为base，没有则用模板
    prev_output = os.path.join(WORKSPACE, "data/output/海外统计表.xlsx")
    template_path = os.path.join(WORKSPACE, "data/statistics/domestic_statistics.xlsx")
    
    base_path = prev_output if os.path.exists(prev_output) else template_path
    print(f"  使用base: {base_path} ({'前一天输出' if base_path == prev_output else '原始模板'})")
    
    output_path = os.path.join(WORKSPACE, "data/output/domestic_statistics.xlsx")
    
    # Copy base to output (not template)
    shutil.copy2(base_path, output_path)
    
    wb = openpyxl.load_workbook(output_path)
    ws = wb['2025-2026']
    
    last_row = ws.max_row
    print(f"  Last row: {last_row}")
    
    # Deduplicate: check existing PO numbers, skip items already in the table
    existing_po = set()
    for r in range(1, last_row + 1):
        po_val = ws.cell(row=r, column=10).value
        if po_val:
            existing_po.add(po_val)
    
    new_items = [item for item in OVERSEAS_ITEMS if item['po'] not in existing_po]
    skipped = len(OVERSEAS_ITEMS) - len(new_items)
    if skipped > 0:
        print(f"  去重: 跳过 {skipped} 条已存在的记录 (PO号: {[item['po'] for item in OVERSEAS_ITEMS if item not in new_items]})")
    
    if not new_items:
        print("  无新增数据，不追加行")
        wb.save(output_path)
        print(f"  生成: {output_path} ({os.path.getsize(output_path)} bytes)")
        return output_path
    
    # Get last sequence number
    last_seq = ws.cell(row=last_row, column=1).value or (last_row - 1)
    
    for i, item in enumerate(new_items):
        row_num = last_row + 1 + i
        seq = int(last_seq) + 1 + i if isinstance(last_seq, (int, float)) else row_num - 1
        
        ws.cell(row=row_num, column=1, value=seq).number_format = '0'  # A: 序号
        ws.cell(row=row_num, column=2, value="chanhua")  # B: 抬头
        ws.cell(row=row_num, column=3, value=TODAY_DATE).number_format = 'YYYY-MM-DD'  # C: 出货日期
        ws.cell(row=row_num, column=4, value="Bitmain Development PTE. LTD.")  # D: 主体
        ws.cell(row=row_num, column=5, value=item.get('supplier', 'XJ'))  # E: 供应商
        ws.cell(row=row_num, column=6, value=item['destination'])  # F: 收货地址
        ws.cell(row=row_num, column=7, value=item['model'])  # G: 型号
        ws.cell(row=row_num, column=8, value=item['material_code'])  # H: 物料编码
        ws.cell(row=row_num, column=9, value=item['qty']).number_format = '0'  # I: 数量
        ws.cell(row=row_num, column=10, value=item['po'])  # J: PO号
        ws.cell(row=row_num, column=11, value=item.get('so', ''))  # K: SO号
        ws.cell(row=row_num, column=12, value=item['price']).number_format = '0.0000'  # L: 单价
        
        print(f"  Row {row_num}: {item['model']} {item['qty']} PCS → {item['po']}")
    
    wb.save(output_path)
    print(f"  生成: {output_path} ({os.path.getsize(output_path)} bytes)")
    return output_path

# ==================== 5. 发送邮件 ====================
def send_email(xlsm_path, summary_path, intl_path, overseas_path):
    print("=== 发送结果邮件 ===")
    
    msg = MIMEMultipart()
    msg['From'] = formataddr(("采购PO自动化", EMAIL_ACCOUNT))
    msg['To'] = ", ".join(REPORT_EMAILS)
    msg['Date'] = email_lib.utils.formatdate(localtime=True)
    
    subject = f"采购订单数据_{TODAY}"
    msg['Subject'] = Header(subject, 'utf-8')
    
    # Email body
    body_lines = [
        "各位好，",
        "",
        f"附件为{TODAY}采购订单数据，请查收：",
        "",
        "**国内进口订单：**",
    ]
    for item in DOMESTIC_ITEMS:
        body_lines.append(f"  {item['po']}: {item['model']} {item['qty']} PCS, 报关单价 {item['price']} USD, 供应商={item['supplier']}, 税代={item['tax_agent']}")
    
    body_lines.extend(["", "**海外出口订单：**"])
    for item in OVERSEAS_ITEMS:
        body_lines.append(f"  {item['po']}: {item['model']} {item['qty']} PCS, 单价 {item['price']} USD, 收货地址={item['destination']}")
    
    body_lines.extend([
        "",
        "附件包含：",
        "1. WebADI采购订单模板(xlsm) - 用于Oracle导入",
        "2. 采购订单数据(xlsx) - 订单明细",
        "3. 国内进口产品统计表(xlsx) - 已更新",
        "4. 海外出口统计表(xlsx) - 已更新",
        "",
        "谢谢！",
    ])
    
    body = "\n".join(body_lines)
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    # Attach files - dynamically named
    files = {}
    if xlsm_path:
        files[f"采购订单_{TODAY}.xlsm"] = xlsm_path
    if summary_path:
        fn = os.path.basename(summary_path)
        files[fn] = summary_path
    if intl_path:
        files[f"国际进口产品统计表-{TODAY}号更新.xlsx"] = intl_path
    if overseas_path:
        files[f"海外出口统计表-{TODAY}号更新.xlsx"] = overseas_path
    
    for display_name, filepath in files.items():
        if not filepath or not os.path.exists(filepath):
            print(f"  SKIP: {display_name} - file not found")
            continue
        with open(filepath, 'rb') as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            # Use RFC 2231 for Chinese filenames
            encoded_name = quote(display_name)
            part.add_header('Content-Disposition', 
                           f"attachment; filename*=UTF-8''{encoded_name}")
            msg.attach(part)
        print(f"  Attached: {display_name} ({os.path.getsize(filepath)} bytes)")
    
    # Send via SMTP
    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
    server.starttls()
    server.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)
    server.sendmail(EMAIL_ACCOUNT, REPORT_EMAILS, msg.as_string())
    server.quit()
    
    print("  邮件发送成功!")

def send_domestic_report():
    """单独发送国内进口报告给 LH-SJXPC@cbscs.com，HTML表格形式（带边框），隐藏价格"""
    if not DOMESTIC_ITEMS:
        return
    
    print(f"=== 发送国内进口报告给 {DOMESTIC_REPORT_EMAIL} ===")
    
    msg = MIMEMultipart('alternative')
    msg['From'] = formataddr(("采购PO自动化", EMAIL_ACCOUNT))
    msg['To'] = DOMESTIC_REPORT_EMAIL
    msg['Date'] = email_lib.utils.formatdate(localtime=True)
    msg['Subject'] = Header(f"国内进口采购订单_{TODAY}", 'utf-8')
    
    # --- 构造 HTML 表格（带边框，模仿国内进口附件表头）---
    html = f"""\
<html>
<head><meta charset="utf-8"></head>
<body>
<p>您好，</p>
<p>以下为{TODAY}国内进口采购订单：</p>
<table border="1" cellpadding="5" cellspacing="0" style="border-collapse:collapse; font-size:12px; font-family:Arial, sans-serif;">
  <tr style="background-color:#4472C4; color:#ffffff; text-align:center;">
    <th>序号</th>
    <th>采购主体</th>
    <th>出货日期</th>
    <th>销售主体</th>
    <th>协议</th>
    <th>供应商</th>
    <th>型号</th>
    <th>数量</th>
    <th>单位</th>
    <th>物料编码</th>
    <th>PO号</th>
    <th>SO号</th>
    <th>税代</th>
  </tr>
"""
    for i, item in enumerate(DOMESTIC_ITEMS):
        seq = str(i + 1)
        entity = "深圳世纪云芯"
        date = item.get('date', TODAY)
        dest = "新加坡比特"
        agreement = "集成电路"
        supplier = item.get('supplier', '')
        model = item['model']
        qty = str(item['qty'])
        unit = "PCS"
        code = item.get('material_code', '')
        po = item['po']
        so = item.get('so', '')
        tax = item.get('tax_agent', '')
        html += f"""\
  <tr style="text-align:center;">
    <td>{seq}</td>
    <td>{entity}</td>
    <td>{date}</td>
    <td>{dest}</td>
    <td>{agreement}</td>
    <td>{supplier}</td>
    <td>{model}</td>
    <td style="text-align:right;">{qty}</td>
    <td>{unit}</td>
    <td>{code}</td>
    <td>{po}</td>
    <td>{so}</td>
    <td>{tax}</td>
  </tr>
"""
    html += """\
</table>
<br/>
<p>谢谢！</p>
</body>
</html>"""
    
    msg.attach(MIMEText(html, 'html', 'utf-8'))
    
    server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
    server.starttls()
    server.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)
    server.sendmail(EMAIL_ACCOUNT, [DOMESTIC_REPORT_EMAIL], msg.as_string())
    server.quit()
    
    print(f"  报告发送成功!")

# ==================== 主流程 ====================
if __name__ == "__main__":
    print(f"采购订单自动化 - {TODAY}")
    print("=" * 50)
    
    # 从邮件获取订单数据
    has_orders = fetch_and_parse_orders()
    if not has_orders:
        print("今日无新订单，流程结束")
        sys.exit(0)
    
    xlsm_path = generate_xlsm()
    summary_path = generate_summary_xlsx()
    intl_path = update_international_template()
    overseas_path = update_domestic_statistics()
    
    if xlsm_path and summary_path and intl_path and overseas_path:
        send_email(xlsm_path, summary_path, intl_path, overseas_path)
        send_domestic_report()
        
        # 保存累计文件供第二天使用
        shutil.copy2(intl_path, os.path.join(WORKSPACE, "data/output/国内进口产品统计表.xlsx"))
        shutil.copy2(overseas_path, os.path.join(WORKSPACE, "data/output/海外统计表.xlsx"))
        print(f"\n  已保存累计文件供下次使用")
        print("\n全部完成!")
    else:
        print("\n部分文件生成失败，请检查!")
        sys.exit(1)