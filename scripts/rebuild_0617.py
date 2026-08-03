#!/usr/bin/env python3
"""Regenerate 6/17 xlsm with ALL items (3 domestic + 5 overseas)"""
import json, os, re, zipfile, openpyxl, smtplib
from datetime import date
from urllib.parse import quote
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
from email.utils import formataddr
from collections import defaultdict

WORKSPACE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TODAY = "20260617"
TODAY_DATE = date(2026, 6, 17)
EXCEL_DATE = (TODAY_DATE - date(1899, 12, 30)).days

# Domestic items from po_automation output
DOMESTIC_MERGED = [
    {'model': 'BM1376AA', 'qty': 2922, 'price': 29.89, 'supplier': 'SPILSZ',
     'po': 'SZK202606176001', 'material_code': 'Y09BM1376010', 'line_no': 1},
    {'model': 'BM1370PA', 'qty': 42056, 'price': 8.61, 'supplier': 'XJ',
     'po': 'SZK202606176002', 'material_code': 'Y31010530', 'line_no': 1},
    {'model': 'BM1746AA', 'qty': 376, 'price': 58.63, 'supplier': 'XJ',
     'po': 'SZK202606176003', 'material_code': 'Y09BM1746010', 'line_no': 1},
]

# Overseas items from stats table (source of truth)
from openpyxl import load_workbook
wb = load_workbook(os.path.join(WORKSPACE, 'data/output/domestic_statistics.xlsx'))
ws = wb['2025-2026']
OVERSEAS_ITEMS = []
for r in range(2, ws.max_row + 1):
    dt = str(ws.cell(row=r, column=3).value or '')[:10]
    if '2026-06-17' in dt:
        OVERSEAS_ITEMS.append({
            'model': str(ws.cell(row=r, column=7).value or ''),
            'qty': int(ws.cell(row=r, column=9).value or 0),
            'price': float(ws.cell(row=r, column=12).value or 0),
            'supplier': str(ws.cell(row=r, column=5).value or ''),
            'destination': str(ws.cell(row=r, column=6).value or ''),
            'po': str(ws.cell(row=r, column=10).value or ''),
            'material_code': str(ws.cell(row=r, column=8).value or ''),
            'line_no': 1,
        })
wb.close()

# Set dest_subinv
dest_map = {'泰国群光': 'DPTHQGCP', '泰国ONETEC': 'DPTONETYCL', 'PIE': 'DPTPIECL',
            '前海保税区': 'DPTQHBSC', '墨西哥欧陆通': 'DPTMOLTYCL'}
for item in OVERSEAS_ITEMS:
    item['dest_subinv'] = dest_map.get(item.get('destination', ''), 'DPTQHBSC')

print("Domestic: {} items".format(len(DOMESTIC_MERGED)))
for it in DOMESTIC_MERGED:
    print("  {}: {} {}PCS @{}".format(it['po'], it['model'], it['qty'], it['price']))
print("Overseas: {} items".format(len(OVERSEAS_ITEMS)))
for it in OVERSEAS_ITEMS:
    print("  {}: {}({}) {}PCS @{} {}->{}".format(
        it['po'], it['model'], it['material_code'], it['qty'], it['price'],
        it['supplier'], it['destination']))

# ====== Generate XLSM ======
print("\n=== Generate WebADI xlsm ===")
template_path = os.path.join(WORKSPACE, "data/templates/webadi_template.xlsm")
output_xlsm = os.path.join(WORKSPACE, "data/output/采购订单_{}.xlsm".format(TODAY))

z = zipfile.ZipFile(template_path, 'r')
sheet2_raw = z.read('xl/worksheets/sheet2.xml').decode('utf-8')
ss_raw = z.read('xl/sharedStrings.xml').decode('utf-8')
other_files = {}
for n in z.namelist():
    if n not in ['xl/worksheets/sheet2.xml', 'xl/sharedStrings.xml']:
        other_files[n] = z.read(n)
z.close()

# Known indices
IDX = {
    'SZK': 87, '标准采购订单': 88, 'USD': 52, '何宇川': 89,
    'BITMAIN': 90, 'SG': 91, 'XAP': 92, '1004': 93, 'SZKXYCL': 94,
    '付款方式一': 59, '生产用料销售': 60, 'Y': 61, '手工录入': 62,
    'BM系列': 63, '个': 65, '0': 66, 'ANTMINER': 67,
}

new_strings = []
for item in DOMESTIC_MERGED + OVERSEAS_ITEMS:
    new_strings.extend([item['po'], item['material_code'], item['model']])
for item in DOMESTIC_MERGED + OVERSEAS_ITEMS:
    new_strings.append("{}芯片".format(item['model']))
new_strings.extend([
    'DPT', 'CHANHUA PTE. LTD.', '费用', '1155.BITMAIN DEVELOPMENT PTE. LTD.',
    'DPTMOLTYCL',
])
for item in OVERSEAS_ITEMS:
    new_strings.append(item['dest_subinv'])
new_strings = list(set(new_strings))

def add_multiple_ss(ss_xml, strings_to_add):
    positions = list(re.finditer(r'<si\b', ss_xml))
    idx_map = {}
    for i, pos in enumerate(positions):
        s = pos.start()
        e = ss_xml.find('</si>', s) + len('</si>')
        block = ss_xml[s:e]
        texts = re.findall(r'<t[^>]*>(.*?)</t>', block)
        idx_map[i] = ''.join(texts)
    count = len(positions)
    new_entries = []
    result_map = {}
    for s in strings_to_add:
        found = False
        for idx, text in idx_map.items():
            if text == s:
                result_map[s] = idx
                found = True
                break
        if not found:
            result_map[s] = count + len(new_entries)
            new_entries.append('<si><t>{}</t></si>'.format(s))
    if new_entries:
        insert_pos = ss_xml.rfind('</sst>')
        ss_xml = ss_xml[:insert_pos] + ''.join(new_entries) + ss_xml[insert_pos:]
        new_count = count + len(new_entries)
        ss_xml = re.sub(r'uniqueCount="\d+"', 'uniqueCount="{}"'.format(new_count), ss_xml, count=1)
    return result_map, ss_xml

ss_map, ss_new = add_multiple_ss(ss_raw, new_strings)
for k, v in IDX.items():
    if k not in ss_map:
        ss_map[k] = v

def make_xlsm_row(row_num, item, is_domestic=True):
    row = '<row r="{}" spans="2:39" ht="14.25" outlineLevel="1">'.format(row_num)
    if is_domestic:
        entity_idx = ss_map["SZK"]
        supplier_idx = ss_map["BITMAIN"]
        supplier_loc_idx = ss_map["SG"]
        source_idx = ss_map["XAP"]
        recv_idx = ss_map["1004"]
        dest_idx = ss_map["SZKXYCL"]
        recv2_idx = ss_map["1004"]
    else:
        entity_idx = ss_map["DPT"]
        supplier_idx = ss_map["CHANHUA PTE. LTD."]
        supplier_loc_idx = ss_map["费用"]
        source_idx = ss_map["XAP"]
        recv_idx = ss_map["1155.BITMAIN DEVELOPMENT PTE. LTD."]
        dest_idx = ss_map[item['dest_subinv']]
        recv2_idx = ss_map["1155.BITMAIN DEVELOPMENT PTE. LTD."]
    
    for col, val in [
        ('B', None), ('C', entity_idx), ('D', ss_map["标准采购订单"]),
        ('E', ss_map[item["po"]]), ('F', ss_map["USD"]), ('G', ss_map["何宇川"]),
        ('H', supplier_idx), ('I', supplier_loc_idx), ('J', source_idx),
        ('K', recv_idx), ('L', dest_idx), ('M', recv2_idx),
        ('N', ss_map["付款方式一"]), ('O', ss_map["生产用料销售"]),
        ('Q', ss_map["Y"]), ('T', ss_map["手工录入"]),
        ('V', ss_map["BM系列"]), ('W', ss_map[item["material_code"]]),
        ('Y', ss_map["个"]), ('AF', ss_map["0"]), ('AG', ss_map["ANTMINER"]),
    ]:
        if val is None:
            row += '<c r="{}{}" s="7"/>'.format(col, row_num)
        else:
            row += '<c r="{}{}" s="5" t="s"><v>{}</v></c>'.format(col, row_num, val)
    
    row += '<c r="P{}" s="5"/>'.format(row_num)
    row += '<c r="R{}" s="5"/><c r="S{}" s="5"/>'.format(row_num, row_num)
    row += '<c r="U{}" s="4"><v>{}</v></c>'.format(row_num, item.get("line_no", 1))
    row += '<c r="X{}" s="5"/>'.format(row_num)
    row += '<c r="Z{}" s="4"><v>{}</v></c>'.format(row_num, item["qty"])
    row += '<c r="AA{}" s="6"><v>{}</v></c>'.format(row_num, EXCEL_DATE)
    row += '<c r="AB{}" s="6"><v>{}</v></c>'.format(row_num, EXCEL_DATE)
    row += '<c r="AC{}" s="6"><v>{}</v></c>'.format(row_num, EXCEL_DATE)
    row += '<c r="AD{}" s="15"><v>{}</v></c>'.format(row_num, item["price"])
    row += '<c r="AE{}" s="15"><v>{}</v></c>'.format(row_num, item["price"])
    row += '<c r="AH{}" s="8"/><c r="AI{}" s="5"/>'.format(row_num, row_num)
    row += '<c r="AJ{}" s="5"/><c r="AK{}" s="8"/>'.format(row_num, row_num)
    row += '</row>'
    return row

all_items = list(DOMESTIC_MERGED) + list(OVERSEAS_ITEMS)

row_matches = re.findall(r'<row r="(\d+)"', sheet2_raw)
template_max = max(int(r) for r in row_matches) if row_matches else 1005

new_rows = []
for i in range(template_max - 9):
    rn = 10 + i
    if i < len(all_items):
        is_dom = i < len(DOMESTIC_MERGED)
        new_rows.append(make_xlsm_row(rn, all_items[i], is_domestic=is_dom))
    else:
        empty = ''
        for col, s in [('B',7),('C',4),('D',5),('E',5),('F',5),('G',4),('H',4),
                        ('I',4),('J',5),('K',4),('L',5),('M',4),('N',5),('O',5),
                        ('P',5),('Q',5),('R',5),('S',5),('T',5),('U',4),('V',4),
                        ('W',4),('X',5),('Y',5),('Z',4),('AA',6),('AB',6),('AC',6),
                        ('AD',4),('AE',5),('AF',5),('AG',5),('AH',8),('AI',5),
                        ('AJ',5),('AK',8),('AL',13),('AM',9)]:
            empty += '<c r="{}{}" s="{}"/>'.format(col, rn, s)
        new_rows.append('<row r="{}" spans="2:39" ht="14.25" outlineLevel="1">{}</row>'.format(rn, empty))

r9_end = sheet2_raw.find('</row>', sheet2_raw.find('<row r="9"')) + len('</row>')
sd_pos = sheet2_raw.find('</sheetData>')
tail = sheet2_raw[sd_pos:]
sheet2_new = sheet2_raw[:r9_end] + ''.join(new_rows) + tail

with zipfile.ZipFile(output_xlsm, 'w', zipfile.ZIP_DEFLATED) as zout:
    for n, d in other_files.items():
        zout.writestr(n, d)
    zout.writestr('xl/worksheets/sheet2.xml', sheet2_new.encode('utf-8'))
    zout.writestr('xl/sharedStrings.xml', ss_new.encode('utf-8'))
print("Generated: {} ({} bytes)".format(output_xlsm, os.path.getsize(output_xlsm)))

# ====== Generate Summary XLSX ======
print("=== Generate summary xlsx ===")
first_po = DOMESTIC_MERGED[0]['po'] if DOMESTIC_MERGED else OVERSEAS_ITEMS[0]['po']
output_xlsx = os.path.join(WORKSPACE, "data/output/{}.xlsx".format(first_po))

wb_xl = openpyxl.Workbook()
ws_xl = wb_xl.active
ws_xl.title = "采购订单数据"
headers = ['加载', '业务实体', '类型', '采购订单号', '币种', '采购员', '供应商',
           '供应商地点', '来源子库存', '收货方', '目的子库存', '收单方',
           '付款方式', '内部申请类型', '货贷', '是否报关', '加工费报价OA单据号',
           '摘要', '业务模式', '行号', '行类型', '物料', '物料说明', '单位',
           '数量', '创建日期', '承诺日期', '需求日期', '不含税单价', '含税单价',
           '税率', '品牌/厂商']
ws_xl.append(headers)

# Domestic rows
for item in DOMESTIC_MERGED:
    ws_xl.append([
        'Y', 'SZK', '标准采购订单', item['po'], 'USD', '何宇川,',
        'BITMAIN DEVELOPMENT PTE.  LTD.', 'SG', 'XAP', '1004.Bitmain Shenzhen',
        'SZKXYCL', '1004.Bitmain Shenzhen', '付款方式一', '生产用料销售',
        '', 'Y', '', '手工录入', '', 1, 'BM系列',
        item['material_code'], '{}芯片'.format(item['model']), '个', item['qty'],
        TODAY_DATE, TODAY_DATE, TODAY_DATE, item['price'], item['price'], '0', 'ANTMINER'
    ])
# Overseas rows
for item in OVERSEAS_ITEMS:
    ws_xl.append([
        'Y', 'DPT', '标准采购订单', item['po'], 'USD', '何宇川,',
        'CHANHUA PTE. LTD.', '费用', 'XAP', '1155.BITMAIN DEVELOPMENT PTE. LTD.',
        item['dest_subinv'], '1155.BITMAIN DEVELOPMENT PTE. LTD.', '付款方式一', '生产用料销售',
        '', 'Y', '', '手工录入', '', 1, 'BM系列',
        item['material_code'], '{}芯片'.format(item['model']), '个', item['qty'],
        TODAY_DATE, TODAY_DATE, TODAY_DATE, item['price'], item['price'], '0', 'ANTMINER'
    ])

wb_xl.save(output_xlsx)
wb_xl.close()
print("Generated: {} ({} bytes)".format(output_xlsx, os.path.getsize(output_xlsx)))

# ====== Verify XLSM ======
print("\n=== Verify xlsm ===")
z2 = zipfile.ZipFile(output_xlsm, 'r')
s2 = z2.read('xl/worksheets/sheet2.xml').decode('utf-8')
ss2 = z2.read('xl/sharedStrings.xml').decode('utf-8')
im, _ = add_multiple_ss.__wrapped__ if hasattr(add_multiple_ss, '__wrapped__') else (None, None)
# Quick parse
im2 = {}
for i, si in enumerate(re.findall(r'<si>(.*?)</si>', ss2, re.DOTALL)):
    im2[i] = ''.join(re.findall(r'<t[^>]*>(.*?)</t>', si))
for row in s2.split('</row>'):
    rn_m = re.search(r'r="(\d+)"', row)
    if not rn_m or int(rn_m.group(1)) < 10: continue
    rn = rn_m.group(1)
    pm = re.search(r'<c r="E'+rn+r'"[^>]*t="s"[^>]*><v>(\d+)</v>', row)
    am = re.search(r'<c r="AD'+rn+r'"[^>]*><v>([\d.]+)</v>', row)
    if pm:
        po = im2.get(int(pm.group(1)), '?')
        price = am.group(1) if am else '?'
        if TODAY in po:
            print("  Row {}: {} price={}".format(rn, po, price))
z2.close()

# ====== Send Email ======
print("\n=== Send email ===")
with open(os.path.join(WORKSPACE, "scripts/po_automation_v2.py")) as f:
    m = re.search(r'EMAIL_PASSWORD\s*=\s*"([^"]+)"', f.read())
    pw = m.group(1)

REPORT = ["yuchuan.he@casue.com", "haixia.lu@casue.com", "yunrui.chen@casue.com",
          "na.yang_w@casue.com", "yujia.cheng@casue.com"]
overseas_path = os.path.join(WORKSPACE, "data/output/domestic_statistics.xlsx")

by_supplier = defaultdict(list)
for it in OVERSEAS_ITEMS:
    by_supplier[it['supplier']].append(it)
total_qty = sum(it['qty'] for it in OVERSEAS_ITEMS)

msg = MIMEMultipart()
msg['From'] = formataddr(("PO", "yuchuan.he@casue.com"))
msg['To'] = ", ".join(REPORT)
msg['Subject'] = Header("采购订单数据_{}".format(TODAY), 'utf-8')

body_lines = [
    "各位好，",
    "",
    "附件为{}采购订单数据，请查收：".format(TODAY),
    "",
    "国内进口订单：{}条".format(len(DOMESTIC_MERGED)),
]
for it in DOMESTIC_MERGED:
    body_lines.append("  {}: {} {}PCS @{}USD 供应商={}".format(it['po'], it['model'], it['qty'], it['price'], it['supplier']))

body_lines.extend(["", "海外出口订单：{}条 合计{:,} PCS".format(len(OVERSEAS_ITEMS), total_qty)])
for supplier, items in sorted(by_supplier.items()):
    sqty = sum(it['qty'] for it in items)
    body_lines.append("  [{}] {}条 合计{:,} PCS".format(supplier, len(items), sqty))
    for it in items:
        body_lines.append("    {}: {}({}) {:,}PCS @{}USD {}".format(it['po'], it['model'], it['material_code'], it['qty'], it['price'], it['destination']))

body_lines.extend([
    "",
    "附件：",
    "1. 采购订单_{}.xlsm（WebADI）".format(TODAY),
    "2. {}.xlsx（采购订单数据）".format(first_po),
    "3. 国际进口产品统计表-{}号更新.xlsx".format(TODAY),
    "4. 海外出口统计表-{}号更新.xlsx".format(TODAY),
    "",
    "谢谢！",
])
msg.attach(MIMEText('\n'.join(body_lines), 'plain', 'utf-8'))

attach_list = [
    ("采购订单_{}.xlsm".format(TODAY), output_xlsm),
    ("{}.xlsx".format(first_po), output_xlsx),
    ("国际进口产品统计表-{}号更新.xlsx".format(TODAY), os.path.join(WORKSPACE, "data/output/international_statistics_new.xlsx")),
    ("海外出口统计表-{}号更新.xlsx".format(TODAY), overseas_path),
]

for name, fpath in attach_list:
    if not os.path.exists(fpath):
        print("  SKIP: {} not found".format(name))
        continue
    with open(fpath, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', "attachment; filename*=UTF-8''{}".format(quote(name)))
        msg.attach(part)
    print("  Attached: {} ({} bytes)".format(name, os.path.getsize(fpath)))

server = smtplib.SMTP("smtp.appia.vip", 587)
server.starttls()
server.login("yuchuan.he@casue.com", pw)
server.sendmail("yuchuan.he@casue.com", REPORT, msg.as_string())
server.quit()
print("邮件发送成功!")
print("\nAll done!")
